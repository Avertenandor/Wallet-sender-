#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WalletSender.py — Расширенный скрипт для анализа токенов BSC и распределения вознаграждений с GUI.

🧭 MCP Navigation Tags:
- #MCP:IMPORTS     - Импорты и настройки (строки 1-100)
- #MCP:CONSTANTS   - Константы и конфигурация (строки 70-110)  
- #MCP:UTILS       - Утилиты и декораторы (строки 110-150)
- #MCP:CONFIG      - Управление конфигурацией (строки 150-250)
- #MCP:RPC         - RPC пул соединений (строки 250-350)
- #MCP:CACHE       - Система кэширования (строки 350-380)
- #MCP:DATABASE    - SQLite операции (строки 380-700)
- #MCP:API         - Etherscan API интеграция для BSC (строки 700-900)
- #MCP:GUI         - PyQt5 интерфейс (строки 863-4735)
- #MCP:MAIN        - Точка входа (строки 4730-4735)

🏷️ Development Control Tags:
- #TODO:MCP        - Задачи для MCP интеграции
- #REFACTOR:MCP    - Места для рефакторинга  
- #OPTIMIZE:MCP    - Возможности оптимизации
- #SECURITY:MCP    - Проблемы безопасности
- #DOCS:MCP        - Требуется документация
"""

# ======== #MCP:IMPORTS ========
from typing import Callable, Any, TypeVar, Optional, Dict, List, Union, Tuple, Set, TypedDict
import sys
import json
import time
import threading
import random
import sqlite3
import os  # #MCP:DB_IMPORT
import math
from decimal import Decimal
try:
    from web3 import Web3, HTTPProvider  # type: ignore
    WEB3_AVAILABLE = True
    try:
        from eth_account import Account  # type: ignore
    except Exception:
        Account = None  # type: ignore
        logger.warning("eth_account не установлен – операции с подписями могут не работать")
except Exception:
    WEB3_AVAILABLE = False
    class Web3:  # type: ignore
        @staticmethod
        def is_address(a):
            return isinstance(a, str) and a.startswith('0x') and len(a) == 42
        @staticmethod
        def to_checksum_address(a):
            return a
    class HTTPProvider:  # type: ignore
        def __init__(self, *args, **kwargs):
            pass
    logger.warning("Web3 не установлен – блокчейн функции ограничены")
from pathlib import Path
import functools
import logging
from datetime import datetime, timedelta
from pathlib import Path
from collections import deque

# Импорт внешних библиотек для работы с API, криптографией и GUI
import requests
from cryptography.fernet import Fernet
from PyQt5 import QtWidgets, QtCore, QtGui
import qdarkstyle
from eth_utils.exceptions import ValidationError

# Константы Qt для обхода проблем с типизацией
BOTTOM_DOCK_WIDGET_AREA = 8  # QtCore.Qt.BottomDockWidgetArea
STRETCH_MODE = QtWidgets.QHeaderView.ResizeMode(0)  # QtWidgets.QHeaderView.Stretch
INTERACTIVE_MODE = QtWidgets.QHeaderView.ResizeMode(1)  # QtWidgets.QHeaderView.Interactive  
SELECT_ROWS = QtWidgets.QAbstractItemView.SelectionBehavior(1)  # QtWidgets.QTableView.SelectRows
CUSTOM_CONTEXT_MENU = QtCore.Qt.ContextMenuPolicy(3)  # QtCore.Qt.CustomContextMenu
USER_ROLE = QtCore.Qt.ItemDataRole(256)  # QtCore.Qt.UserRole
ITEM_IS_EDITABLE = QtCore.Qt.ItemFlag(2)  # QtCore.Qt.ItemIsEditable
# Добавленные константы для QMessageBox
QMSGBOX_YES = 0x00004000  # QtWidgets.QMessageBox.Yes
QMSGBOX_NO = 0x00010000  # QtWidgets.QMessageBox.No
# Константа для Qt.QueuedConnection
QUEUED_CONNECTION = 2  # QtCore.Qt.QueuedConnection

# Совместимость с разными версиями Web3.py
try:
    from web3.middleware import geth_poa_middleware  # type: ignore
except ImportError:
    try:
        from web3.middleware.geth_poa import geth_poa_middleware  # type: ignore
    except ImportError:
        # Для самых новых версий Web3.py
        from web3.middleware import ExtraDataToPOAMiddleware as geth_poa_middleware  # type: ignore

# Настройка логирования для отслеживания работы программы
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"wallet_sender_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("WalletSender")

# --- Унифицированная отправка raw транзакций (перенесено из бэкапа) ---
def send_raw_tx(w3, raw_tx_bytes):
    """Отправка raw-транзакции с fallback для разных версий web3.
    Args:
        w3: инстанс Web3
        raw_tx_bytes: подписанные байты транзакции
    Returns:
        HexBytes хеш транзакции
    """
    try:
        return w3.eth.send_raw_transaction(raw_tx_bytes)
    except AttributeError:
        # fallback на альтернативное имя метода (историческая совместимость)
        return getattr(w3.eth, 'send_rawTransaction')(raw_tx_bytes)

# ======== #MCP:CONFIG ========

# ======== #MCP:DATABASE ========
# Константы для базы данных
DB_PATH = Path("wallet_sender.db")
REWARDS_CONFIG_DIR = Path("rewards_configs")
REWARDS_CONFIG_DIR.mkdir(exist_ok=True)

# Константы токенов (если не определены выше)
if 'PLEX_CONTRACT' not in globals():
    PLEX_CONTRACT = '0x07958Be5D12365db62A6535D0a88105944a2E81E'
if 'USDT_CONTRACT' not in globals():
    USDT_CONTRACT = '0x55d398326f99059fF775485246999027B3197955'

# ERC20 ABI минимальный
ERC20_ABI = [
    {"constant":True,"inputs":[],"name":"decimals","outputs":[{"name":"","type":"uint8"}],"type":"function"},
    {"constant":True,"inputs":[{"name":"_owner","type":"address"}],"name":"balanceOf","outputs":[{"name":"balance","type":"uint256"}],"type":"function"},
    {"constant":False,"inputs":[{"name":"_to","type":"address"},{"name":"_value","type":"uint256"}],"name":"transfer","outputs":[{"name":"","type":"bool"}],"type":"function"},
    {"constant":True,"inputs":[],"name":"symbol","outputs":[{"name":"","type":"string"}],"type":"function"},
    {"constant":True,"inputs":[],"name":"name","outputs":[{"name":"","type":"string"}],"type":"function"}
]

def get_db_connection():
    """Получение соединения с базой данных"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Инициализация базы данных и создание таблиц"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Создание таблицы истории
    cur.execute("""
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY,
            ts TEXT,
            token TEXT,
            to_addr TEXT,
            amount REAL,
            tx TEXT,
            status TEXT
        )
    """)
    
    # Создание таблицы найденных транзакций
    cur.execute("""
        CREATE TABLE IF NOT EXISTS found_transactions (
            id INTEGER PRIMARY KEY,
            ts TEXT,
            tx_hash TEXT,
            from_addr TEXT,
            to_addr TEXT,
            token_addr TEXT,
            token_name TEXT,
            amount REAL,
            block INTEGER,
            block_time TEXT,
            search_data TEXT
        )
    """)
    
    # Создание таблицы транзакций отправителей
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sender_transactions (
            id INTEGER PRIMARY KEY,
            search_time TEXT,
            sender_addr TEXT,
            tx_hash TEXT,
            tx_timestamp TEXT,
            token_name TEXT,
            amount REAL,
            block INTEGER,
            rewarded INTEGER DEFAULT 0
        )
    """)
    
    # Создание таблицы наград
    cur.execute("""
        CREATE TABLE IF NOT EXISTS rewards (
            id INTEGER PRIMARY KEY,
            address TEXT,
            plex_amount REAL,
            usdt_amount REAL,
            tx_hash TEXT,
            created_at TEXT,
            sent INTEGER DEFAULT 0
        )
    """)
    
    # Создание таблицы для массовых рассылок
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mass_distributions (
            id INTEGER PRIMARY KEY,
            created_at TEXT,
            name TEXT,
            token_address TEXT,
            token_symbol TEXT,
            amount_per_tx REAL,
            total_addresses INTEGER,
            total_cycles INTEGER,
            interval_seconds INTEGER,
            status TEXT,
            completed_at TEXT
        )
    """)
    
    # Создание таблицы для элементов массовой рассылки
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mass_distribution_items (
            id INTEGER PRIMARY KEY,
            distribution_id INTEGER,
            address TEXT,
            cycle INTEGER,
            tx_hash TEXT,
            status TEXT,
            error_message TEXT,
            sent_at TEXT,
            FOREIGN KEY (distribution_id) REFERENCES mass_distributions(id)
        )
    """)

    # Миграции: добавление поля slot, если отсутствует
    try:
        cur.execute("ALTER TABLE mass_distributions ADD COLUMN slot TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cur.execute("ALTER TABLE mass_distribution_items ADD COLUMN slot TEXT")
    except sqlite3.OperationalError:
        pass
    
    conn.commit()
    conn.close()

# Дополнительные классы и утилиты
class Config:
    """Класс для управления конфигурацией приложения"""
    def __init__(self):
        self.data = {}
        self.config_file = Path("config.json")
        self.load()
    
    def load(self):
        if self.config_file.exists():
            try:
                with open(self.config_file, 'r') as f:
                    self.data = json.load(f)
            except Exception:
                self.data = {}
    
    def save(self):
        with open(self.config_file, 'w') as f:
            json.dump(self.data, f, indent=2)
    
    def get_key(self, slot=None):
        if slot:
            return self.data.get(f'private_key_{slot}')
        return self.data.get('private_key')
    
    def set_key(self, key, slot=None):
        if slot:
            self.data[f'private_key_{slot}'] = key
        else:
            self.data['private_key'] = key
        self.save()
    
    def get_mnemonic(self, slot=None):
        if slot:
            return self.data.get(f'mnemonic_{slot}')
        return self.data.get('mnemonic')
    
    def set_mnemonic(self, mnemonic, slot=None):
        if slot:
            self.data[f'mnemonic_{slot}'] = mnemonic
        else:
            self.data['mnemonic'] = mnemonic
        self.save()
    
    def get_gas_price(self):
        return self.data.get('gas_price', 5.0)
    
    def get_reward_per_tx(self):
        return self.data.get('reward_per_tx', False)
    
    def set_reward_per_tx(self, value):
        self.data['reward_per_tx'] = value
        self.save()

# RPC Pool для работы с блокчейном
RPC_NODES = [
    "https://bsc-dataseed.binance.org/",
    "https://bsc-dataseed1.defibit.io/",
    "https://bsc-dataseed1.ninicoin.io/"
]

class RPCPool:
    """Пул RPC соединений"""
    def __init__(self, nodes):
        self.nodes = nodes
        self.current_index = 0
    
    def web3(self):
        if not WEB3_AVAILABLE:
            raise RuntimeError("Web3 не установлен")
        node = self.nodes[self.current_index]
        self.current_index = (self.current_index + 1) % len(self.nodes)
        w3 = Web3(HTTPProvider(node))
        if hasattr(w3, 'middleware_onion'):
            w3.middleware_onion.inject(geth_poa_middleware, layer=0)
        return w3

# Глобальный флаг доступности блокчейна
blockchain_enabled = WEB3_AVAILABLE

# API ключи для BscScan с ротацией
class APIKeyRotator:
    """Ротатор API ключей для распределения нагрузки"""
    def __init__(self, keys):
        self.keys = keys
        self.current_index = 0
        self.failed_keys = set()
        self.request_counts = {k: 0 for k in keys}
        self.success_counts = {k: 0 for k in keys}
        self.last_used = {k: None for k in keys}
    
    def get_current_key(self):
        """Получить текущий активный ключ"""
        available_keys = [k for k in self.keys if k not in self.failed_keys]
        if not available_keys:
            # Если все ключи заблокированы, сбрасываем блокировки
            self.failed_keys.clear()
            available_keys = self.keys
        
        # Выбираем ключ с наименьшим количеством запросов
        key = min(available_keys, key=lambda k: self.request_counts[k])
        self.request_counts[key] += 1
        self.last_used[key] = datetime.now()
        return key
    
    def mark_key_failed(self, key):
        """Пометить ключ как проблемный"""
        if key in self.keys:
            self.failed_keys.add(key)
            logger.warning(f"API ключ помечен как проблемный: {key[:10]}...")
    
    def mark_key_success(self, key):
        """Пометить ключ как успешный"""
        if key in self.keys:
            self.success_counts[key] += 1
            # Снимаем блокировку при успешном запросе
            if key in self.failed_keys:
                self.failed_keys.discard(key)
    
    def get_statistics(self):
        """Получить статистику использования ключей"""
        stats = {
            'total_keys': len(self.keys),
            'active_keys': len(self.keys) - len(self.failed_keys),
            'failed_keys': len(self.failed_keys),
            'total_requests': sum(self.request_counts.values()),
            'current_key': self.keys[self.current_index][:10] + '...',
            'last_rotation': max([t for t in self.last_used.values() if t], default=None),
            'key_details': {}
        }
        
        for key in self.keys:
            key_masked = key[:10] + '...'
            stats['key_details'][key_masked] = {
                'status': 'FAILED' if key in self.failed_keys else 'ACTIVE',
                'requests': self.request_counts[key],
                'successes': self.success_counts[key],
                'success_rate': (self.success_counts[key] / max(self.request_counts[key], 1)) * 100,
                'last_used': self.last_used[key].strftime('%H:%M:%S') if self.last_used[key] else None
            }
        
        return stats

# Инициализация ротатора API ключей
api_key_rotator = APIKeyRotator([
    "RAI3FTD9W53JPYZ2AHW8IBH9BXUC71NRH1",
    "U89HXHR9Y26CHMWAA9JUZ17YK2AAXS65CZ", 
    "RF1Q8SCFHFD1EVAP5A4WCMIM4DREA7UNUH"
])

# Простой кэш для API запросов
class SimpleCache:
    def __init__(self, ttl_seconds=300):
        self.cache = {}
        self.ttl = ttl_seconds
    
    def get(self, key):
        if key in self.cache:
            data, timestamp = self.cache[key]
            if time.time() - timestamp < self.ttl:
                return data
            del self.cache[key]
        return None
    
    def set(self, key, value):
        self.cache[key] = (value, time.time())

cache = SimpleCache(ttl_seconds=300)

# TypedDict для типизации ответов API
class BSCScanAPIResponse(TypedDict, total=False):
    status: str
    message: str  
    result: Union[List[Dict[str, Any]], str]

# Декоратор для повторных попыток
def retry(times=3, delay=1):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            for i in range(times):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if i == times - 1:
                        raise
                    logger.warning(f"Попытка {i+1}/{times} не удалась: {e}")
                    time.sleep(delay)
            return None
        return wrapper
    return decorator

class LogHandler(logging.Handler):
    """Пользовательский обработчик логов: отправляет сообщения в QTextEdit через сигнал."""
    def __init__(self, text_widget: Any):
        super().__init__()
        self.text_widget = text_widget
        self._proxy = _QtSignalProxy()
        if hasattr(self.text_widget, 'append'):
            self._proxy.append_text.connect(self.text_widget.append)
        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

    def emit(self, record):
        try:
            msg = self.format(record)
            self._proxy.append_text.emit(msg)
        except Exception:
            # Никогда не роняем приложение из-за логов
            pass

# ======== #MCP:GUI ========

def add_history(token: str, to_addr: str, amt: float, tx_hash: str) -> Optional[int]:
    """Добавление записи в историю транзакций"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO history VALUES(NULL,?,?,?,?,?,?)",
        (datetime.now().isoformat(), token, to_addr, amt, tx_hash, 'pending')
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    logger.info(f"Транзакция добавлена в историю: {tx_hash}")
    return row_id

def update_tx_status(tx_hash: str, status: str) -> int:
    """Обновление статуса транзакции в базе данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE history SET status=? WHERE tx=?", (status, tx_hash))
    conn.commit()
    affected_rows = cur.rowcount
    conn.close()
    logger.debug(f"Статус транзакции {tx_hash} обновлен на {status} ({affected_rows} строк)")
    return affected_rows

def fetch_history():
    """Получение истории транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT ts,token,to_addr,amount,tx,status FROM history ORDER BY id DESC LIMIT 100")
    rows = cur.fetchall()
    conn.close()
    return rows

def copy_all_transactions_hashes():
    """Копирование всех хешей транзакций в строку"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT tx FROM history ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return '\n'.join([row['tx'] for row in rows])

def add_found_transaction(tx_data: Dict[str, Any], search_info: Dict[str, Any]) -> Optional[int]:
    """Добавление найденной транзакции в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO found_transactions VALUES(NULL,?,?,?,?,?,?,?,?,?,?)",
        (
            datetime.now().isoformat(),
            tx_data.get('hash', ''),
            tx_data.get('from', ''),
            tx_data.get('to', ''),
            tx_data.get('contractAddress', ''),
            tx_data.get('tokenSymbol', ''),
            float(tx_data.get('value', 0)) / (10 ** int(tx_data.get('tokenDecimal', 18))),
            int(tx_data.get('blockNumber', 0)),
            datetime.fromtimestamp(int(tx_data.get('timeStamp', 0))).isoformat(),
            json.dumps(search_info)
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def add_sender_transaction(sender_addr: str, tx_info: Dict[str, Any], search_time: str) -> Optional[int]:
    """Добавление информации о транзакции отправителя в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Проверяем, не добавлен ли уже хэш транзакции
    cur.execute("SELECT id FROM sender_transactions WHERE tx_hash=?", (tx_info['hash'],))
    if cur.fetchone():
        conn.close()
        return None  # Транзакция уже добавлена
    
    cur.execute(
        "INSERT INTO sender_transactions VALUES(NULL,?,?,?,?,?,?,?,?)",
        (
            search_time,
            sender_addr,
            tx_info['hash'],
            datetime.fromtimestamp(int(tx_info['timestamp'])).isoformat() if tx_info['timestamp'] else None,
            tx_info['token'],
            tx_info['value'],
            tx_info['block'],
            0  # Не награжденная транзакция по умолчанию
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def add_reward(address: str, plex_amount: float, usdt_amount: float, tx_hash: Optional[str] = None) -> Optional[int]:
    """Добавление награды в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO rewards VALUES(NULL,?,?,?,?,?,?)",
        (
            address,
            plex_amount,
            usdt_amount,
            tx_hash,
            datetime.now().isoformat(),
            0  # Не отправленная награда по умолчанию
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def get_rewards():
    """Получение всех наград из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, address, plex_amount, usdt_amount, tx_hash, created_at, sent
        FROM rewards
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def mark_transaction_rewarded(tx_hash: str) -> int:
    """Отметка транзакции как награжденной"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE sender_transactions SET rewarded=1 WHERE tx_hash=?", (tx_hash,))
    conn.commit()
    affected_rows = cur.rowcount
    conn.close()
    return affected_rows

def get_unrewarded_transactions(sender_addr: Optional[str] = None) -> List[sqlite3.Row]:
    """Получение ненагражденных транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    if sender_addr:
        cur.execute("""
            SELECT id, sender_addr, tx_hash, tx_timestamp, token_name, amount
            FROM sender_transactions
            WHERE rewarded=0 AND sender_addr=?
            ORDER BY tx_timestamp DESC
        """, (sender_addr,))
    else:
        cur.execute("""
            SELECT id, sender_addr, tx_hash, tx_timestamp, token_name, amount
            FROM sender_transactions
            WHERE rewarded=0
            ORDER BY sender_addr, tx_timestamp DESC
        """)
    
    rows = cur.fetchall()
    conn.close()
    return rows


def get_transactions_by_sender(sender_addr: str) -> List[sqlite3.Row]:
    """Получение всех транзакций отправителя"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tx_hash, tx_timestamp, token_name, amount, rewarded
        FROM sender_transactions
        WHERE sender_addr=?
        ORDER BY tx_timestamp DESC
    """, (sender_addr,))
    rows = cur.fetchall()
    conn.close()
    return rows

def fetch_found_transactions(limit: int = 1000) -> List[sqlite3.Row]:
    """Получение найденных транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT ts, tx_hash, from_addr, to_addr, token_addr, token_name, amount, block, block_time, search_data
        FROM found_transactions
        ORDER BY id DESC
        LIMIT ?
    """, (limit,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_sender_transaction_counts() -> List[sqlite3.Row]:
    """Получение количества транзакций по отправителям"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sender_addr, COUNT(*) as tx_count, 
               SUM(CASE WHEN rewarded=1 THEN 1 ELSE 0 END) as rewarded_count
        FROM sender_transactions
        GROUP BY sender_addr
        ORDER BY tx_count DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def clear_found_transactions() -> int:
    """Очистка таблицы найденных транзакций"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM found_transactions")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return count

def clear_sender_transactions() -> int:
    """Очистка таблицы транзакций отправителей"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM sender_transactions")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return count

# ======== Управление конфигурацией наград ========
def save_rewards_config(name: str, addresses: List[str], plex_amounts: Dict[int, float], usdt_amounts: Dict[int, float]) -> bool:
    """Сохранение конфигурации наград в файл"""
    config: List[Dict[str, Union[str, float]]] = []
    for i in range(len(addresses)):
        config.append({
            'address': addresses[i],
            'plex': plex_amounts.get(i, 0),
            'usdt': usdt_amounts.get(i, 0)
        })
    
    file_path = REWARDS_CONFIG_DIR / f"{name}.json"
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"Ошибка при сохранении конфигурации наград: {e}")
        return False

def load_rewards_config(name: str) -> Optional[List[Dict[str, Any]]]:
    """Загрузка конфигурации наград из файла"""
    file_path = REWARDS_CONFIG_DIR / f"{name}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка при загрузке конфигурации наград: {e}")
        return None

def get_rewards_configs() -> List[str]:
    """Получение списка доступных конфигураций наград"""
    try:
        return [f.stem for f in REWARDS_CONFIG_DIR.glob('*.json')]
    except Exception as e:
        logger.error(f"Ошибка при получении списка конфигураций наград: {e}")
        return []

# ======== Управление массовыми рассылками ========
def add_mass_distribution(name: str, token_address: str, token_symbol: str, amount_per_tx: float, total_addresses: int, total_cycles: int, interval_seconds: int, slot: Optional[str] = None) -> int:
    """Добавление новой массовой рассылки в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO mass_distributions (id,created_at,name,token_address,token_symbol,amount_per_tx,total_addresses,total_cycles,interval_seconds,status,completed_at,slot) VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?)",
        (
            datetime.now().isoformat(),
            name,
            token_address,
            token_symbol,
            amount_per_tx,
            total_addresses,
            total_cycles,
            interval_seconds,
            'active',
            None,
            slot
        )
    )
    conn.commit()
    distribution_id = cur.lastrowid
    conn.close()
    
    if distribution_id is None:
        raise RuntimeError("Не удалось получить ID созданной рассылки")
    
    logger.info(f"Создана массовая рассылка #{distribution_id}: {name}")
    return distribution_id

def add_mass_distribution_item(distribution_id: int, address: str, cycle: int, tx_hash: str, status: str, error_message: Optional[str] = None, slot: Optional[str] = None) -> Optional[int]:
    """Добавление элемента массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO mass_distribution_items (id,distribution_id,address,cycle,tx_hash,status,error_message,sent_at,slot) VALUES(NULL,?,?,?,?,?,?,?,?)",
        (
            distribution_id,
            address,
            cycle,
            tx_hash,
            status,
            error_message,
            datetime.now().isoformat() if status in ['success', 'error'] else None,
            slot
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def update_mass_distribution_status(distribution_id: int, status: str) -> None:
    """Обновление статуса массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    completed_at = datetime.now().isoformat() if status in ['completed', 'cancelled'] else None
    
    cur.execute(
        "UPDATE mass_distributions SET status=?, completed_at=? WHERE id=?",
        (status, completed_at, distribution_id)
    )
    conn.commit()
    conn.close()
    logger.info(f"Статус рассылки #{distribution_id} обновлен на {status}")

def get_mass_distributions(limit: int = 100) -> List[sqlite3.Row]:
    """Получение списка массовых рассылок"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, created_at, name, token_symbol, amount_per_tx, total_addresses, 
               total_cycles, status, completed_at
        FROM mass_distributions
        ORDER BY id DESC
        LIMIT ?
    """, (limit,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_mass_distribution_items(distribution_id: int) -> List[sqlite3.Row]:
    """Получение элементов массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT address, cycle, tx_hash, status, error_message, sent_at
        FROM mass_distribution_items
        WHERE distribution_id=?
        ORDER BY cycle, address
    """, (distribution_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_mass_distribution_statistics(distribution_id: int) -> Dict[str, Any]:
    """Получение статистики массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Общая статистика
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status='success' THEN 1 ELSE 0 END) as success_count,
            SUM(CASE WHEN status='error' THEN 1 ELSE 0 END) as error_count,
            SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending_count
        FROM mass_distribution_items
        WHERE distribution_id=?
    """, (distribution_id,))
    
    stats = cur.fetchone()
    conn.close()
    
    return {
        'total': stats[0] if stats else 0,
        'success': stats[1] if stats else 0,
        'errors': stats[2] if stats else 0,
        'pending': stats[3] if stats else 0
    }

# ======== #MCP:API ========
@retry(times=3)
def bscscan_request(params: Dict[str, Any], api_key: Optional[str] = None) -> BSCScanAPIResponse:
    """Отправка запроса к API BscScan с поддержкой ротации ключей, кэширования и повторов
    
    OPTIMIZE:MCP - Добавить rate limiting
    REFACTOR:MCP - Выделить в отдельный API клиент
    SECURITY:MCP - Скрыть API ключ из логов
    FEATURES:MCP - Автоматическая ротация API ключей при сбоях
    """
    if not blockchain_enabled:
        raise RuntimeError("Блокчейн отключен")
    
    cache_key = str(params)
    cached = cache.get(cache_key)
    if cached:
        logger.debug(f"Найдено в кэше для запроса BscScan: {params}")
        return cached
    
    # Подбор ключа: если передан явно — используем его, иначе через ротатор
    params_with_key = params.copy()
    if api_key:
        params_with_key['apikey'] = api_key
        current_key = api_key
        use_rotator = False
    else:
        current_key = api_key_rotator.get_current_key()
        params_with_key['apikey'] = current_key
        use_rotator = True
    
    # --- Rate limiting (Free BscScan: ~5 req/sec; делаем мягко 4/сек) ---
    if not hasattr(bscscan_request, '_req_times'):
        bscscan_request._req_times = deque(maxlen=10)  # type: ignore[attr-defined]
    if not hasattr(bscscan_request, '_lock'):
        bscscan_request._lock = threading.Lock()  # type: ignore[attr-defined]

    with bscscan_request._lock:  # type: ignore[attr-defined]
        now = time.time()
        dq = bscscan_request._req_times  # type: ignore[attr-defined]
        # Удаляем устаревшие записи (>1 сек)
        while dq and now - dq[0] > 1.0:
            dq.popleft()
        # Если достигли лимита (>=4 запросов за последнюю секунду) — ждём
        while len(dq) >= 4:
            sleep_for = 1.05 - (now - dq[0])
            if sleep_for > 0:
                logger.debug(f"RateLimit: пауза {sleep_for:.3f}s (запросов за последнюю секунду: {len(dq)})")
                time.sleep(sleep_for)
            now = time.time()
            while dq and now - dq[0] > 1.0:
                dq.popleft()
        dq.append(time.time())

    try:
        # Базовый URL (если ранее не объявлен как константа)
        base_url = globals().get('BSCSCAN_URL', 'https://api.bscscan.com/api')
        response = requests.get(base_url, params=params_with_key, timeout=10)
        if response.status_code != 200:
            logger.warning(f"Ошибка HTTP в BscScan: {response.status_code}")
            # Отмечаем ключ как проблемный при HTTP ошибках
            if use_rotator:
                api_key_rotator.mark_key_failed(current_key)
            raise Exception(f"Ошибка HTTP {response.status_code}")
            
        result: BSCScanAPIResponse = response.json()
        
        # Проверка специфичных ошибок API
        if isinstance(result, dict) and result.get('message') == 'No transactions found':
            api_key_rotator.mark_key_success(current_key)
            return {'status': '1', 'message': 'OK', 'result': []}
            
        if not isinstance(result, dict) or result.get('status') != '1':
            # Собираем подробное сообщение из message + result
            raw_message = result.get('message') if isinstance(result, dict) else None
            raw_result = result.get('result') if isinstance(result, dict) else None
            details = ''
            if isinstance(raw_result, str):
                details = raw_result
            elif raw_result is not None:
                try:
                    details = json.dumps(raw_result, ensure_ascii=False)
                except Exception:
                    details = str(raw_result)
            error_msg_base = raw_message or 'Ошибка Etherscan API'
            detailed_msg = f"{error_msg_base}{(': ' + details) if details else ''}"
            logger.error(f"Ошибка Etherscan API: {detailed_msg}")
            
            # Классификация ошибок для корректной реакции выше по стеку
            low = detailed_msg.lower()
            rate_limit_keywords = [
                'rate limit', 'max rate', 'maximum rate', 'too many request', 'too many requests', 'limit reached'
            ]
            key_error_keywords = [
                'invalid api key', 'apikey invalid', 'missing api key', 'apikey missing', 'forbidden', 'access denied', 'ban'
            ]
            is_rate_limited = any(k in low for k in rate_limit_keywords)
            is_key_error = any(k in low for k in key_error_keywords)
            
            if is_rate_limited or is_key_error:
                # Помечаем текущий ключ проблемным и даём шанс ротатору выбрать другой на следующем вызове
                if use_rotator:
                    api_key_rotator.mark_key_failed(current_key)
                if is_rate_limited and 'rate limit' not in low:
                    # Гарантируем наличие маркера 'rate limit' в Exception, чтобы внешняя логика распознала паузу
                    detailed_msg += ' (rate limit)'
            
            raise Exception(detailed_msg)
            
        # Успешный запрос
        cache.set(cache_key, result)
        if use_rotator:
            api_key_rotator.mark_key_success(current_key)
        
        return result
        
    except requests.RequestException as e:
        logger.error(f"Запрос BscScan не удался: {e}")
        if use_rotator:
            api_key_rotator.mark_key_failed(current_key)
        raise

# ======== Постраничный анализ ========
def search_transactions_paginated(
    wallet_address: str, 
    token_contract: Optional[str] = None, 
    exact_amount: Optional[float] = None, 
    min_amount: Optional[float] = None, 
    max_amount: Optional[float] = None, 
    page_size: int = 1000, 
    max_pages: int = 100, 
    delay_seconds: int = 1, 
    stop_flag: Optional[Any] = None, 
    track_individual_tx: bool = False
) -> Tuple[List[Dict[str, Any]], Dict[str, int], Optional[Dict[str, List[Dict[str, Any]]]]]:
    """
    Постраничный поиск транзакций с указанными параметрами.
    
    OPTIMIZE:MCP - Добавить параллельную обработку страниц
    REFACTOR:MCP - Разделить на более мелкие функции
    TODO:MCP - Добавить прогресс-бар для длительных операций
    
    Args:
        wallet_address (str): Адрес кошелька для поиска
        token_contract (str, optional): Адрес контракта токена для фильтрации
        exact_amount (float, optional): Точная сумма токенов для поиска
        min_amount (float, optional): Минимальная сумма для поиска по диапазону
        max_amount (float, optional): Максимальная сумма для поиска по диапазону
        page_size (int): Количество транзакций на странице (максимум 10000)
        max_pages (int): Максимальное количество страниц для поиска
        delay_seconds (int): Задержка между запросами в секундах
        stop_flag (threading.Event, optional): Флаг для остановки поиска
        track_individual_tx (bool): Если True, возвращает детальную информацию о транзакциях по отправителям
        
    Returns:
        tuple: (список найденных транзакций, счетчик по отправителям, детальная информация по транзакциям)
    """
    if not blockchain_enabled:
        raise RuntimeError("Блокчейн отключен")
    
    # Подготовка базовых параметров запроса
    base_params: Dict[str, Any] = {
        'module': 'account',
        'action': 'tokentx',
        'address': wallet_address,
        'sort': 'desc',
        'offset': page_size
    }
    
    # Добавляем адрес контракта токена, если указан
    if token_contract:
        base_params['contractaddress'] = token_contract
    
    matching_transactions: List[Dict[str, Any]] = []
    sender_counter: Dict[str, int] = {}
    # Словарь для хранения транзакций по отправителям, если нужно детальное отслеживание
    sender_transactions: Optional[Dict[str, List[Dict[str, Any]]]] = {} if track_individual_tx else None
    
    page = 1
    has_more_data = True
    token_decimals: Dict[str, int] = {}  # Кэш для десятичных знаков токенов
    
    logger.info(f"Начинаем постраничный поиск транзакций для адреса {wallet_address}")
    
    # Адаптивная задержка для снижения риска лимита: стартовая = delay_seconds
    adaptive_delay = max(delay_seconds, 0.25)
    consecutive_rate_limits = 0

    while has_more_data and page <= max_pages:
        # Проверка флага остановки
        if stop_flag and stop_flag.is_set():
            logger.info("Поиск остановлен по запросу пользователя")
            break
            
        try:
            params = dict(base_params)
            params['page'] = page
            
            logger.info(f"Запрашиваем страницу {page}...")
            api_response = bscscan_request(params)
            
            # Извлекаем данные транзакций из API ответа
            result = api_response.get('result', []) if isinstance(api_response, dict) else []
            
            if not result:
                logger.info(f"Страница {page} не содержит транзакций. Поиск завершен.")
                break
            
            logger.info(f"Получено {len(result)} транзакций на странице {page}")
            
            for tx in result:
                # Обеспечиваем правильный тип для tx
                if not isinstance(tx, dict):
                    continue
                    
                # Получаем десятичные знаки для токена
                contract_addr = tx.get('contractAddress', '').lower()
                if contract_addr not in token_decimals:
                    token_decimals[contract_addr] = int(tx.get('tokenDecimal', 18))
                
                decimals = token_decimals[contract_addr]
                tx_value = float(tx.get('value', 0)) / (10 ** decimals)
                
                # Проверяем условия
                if wallet_address.lower() == tx.get('to', '').lower():  # Только входящие транзакции
                    # Проверяем, что указаны параметры поиска
                    if exact_amount is not None or (min_amount is not None and max_amount is not None):
                        # Проверка на точную сумму
                        if exact_amount is not None:
                            if abs(tx_value - exact_amount) < 0.0000001:  # Учитываем погрешность float
                                matching_transactions.append(tx)
                                sender = tx.get('from', '').lower()
                                sender_counter[sender] = sender_counter.get(sender, 0) + 1
                                
                                # Сохраняем дополнительную информацию о транзакциях, если нужно
                                if track_individual_tx and sender_transactions is not None:
                                    if sender not in sender_transactions:
                                        sender_transactions[sender] = []
                                    tx_info: Dict[str, Any] = {
                                        'hash': tx.get('hash', ''),
                                        'timestamp': tx.get('timeStamp', ''),
                                        'value': tx_value,
                                        'block': tx.get('blockNumber', ''),
                                        'token': tx.get('tokenSymbol', '')
                                    }
                                    sender_transactions[sender].append(tx_info)
                        
                        # Проверка на диапазон сумм
                        elif min_amount is not None and max_amount is not None:
                            if min_amount <= tx_value <= max_amount:
                                matching_transactions.append(tx)
                                sender = tx.get('from', '').lower()
                                sender_counter[sender] = sender_counter.get(sender, 0) + 1
                                
                                # Сохраняем дополнительную информацию о транзакциях, если нужно
                                if track_individual_tx and sender_transactions is not None:
                                    if sender not in sender_transactions:
                                        sender_transactions[sender] = []
                                    tx_info: Dict[str, Any] = {
                                        'hash': tx.get('hash', ''),
                                        'timestamp': tx.get('timeStamp', ''),
                                        'value': tx_value,
                                        'block': tx.get('blockNumber', ''),
                                        'token': tx.get('tokenSymbol', '')
                                    }
                                    sender_transactions[sender].append(tx_info)
                    else:
                        # Если не указаны параметры поиска, логируем предупреждение и пропускаем транзакцию
                        logger.warning("Не указаны параметры поиска по сумме. Транзакция пропущена.")
                                # Если получили меньше транзакций, чем размер страницы, значит больше данных нет
            if len(result) < page_size:
                has_more_data = False
                logger.info("Достигнут конец данных")
            else:
                # Переходим к следующей странице
                page += 1
                # Задержка (адаптивная)
                time.sleep(adaptive_delay)
        
        except Exception as e:
            err_low = str(e).lower()
            logger.error(f"Ошибка при получении страницы {page}: {e}")
            # Базовая пауза при ошибке
            time.sleep(adaptive_delay * 2)
            if 'rate limit' in err_low or 'too many request' in err_low:
                consecutive_rate_limits += 1
                backoff = min(10, 2 * consecutive_rate_limits)
                logger.warning(f"Rate limit hit (#{consecutive_rate_limits}). Доп. пауза {backoff}s")
                time.sleep(backoff)
                # Увеличиваем адаптивную задержку до макс 3 сек
                adaptive_delay = min(3.0, adaptive_delay * 1.5 + 0.1)
            else:
                consecutive_rate_limits = 0
                # Переходим к следующей странице при не rate-limit ошибке
                page += 1
    
    logger.info(f"Поиск завершен. Найдено {len(matching_transactions)} транзакций от {len(sender_counter)} отправителей")
    return matching_transactions, sender_counter, sender_transactions

def get_token_decimal(token_address: str, api_key: Optional[str] = None) -> int:
    """Получение количества десятичных знаков для токена с учётом API-ключа вкладки"""
    try:
        # Пытаемся через BscScan (с ключом, если дан)
        api_response = bscscan_request({
            'module': 'token',
            'action': 'tokeninfo',
            'contractaddress': token_address
        }, api_key=api_key)
        if isinstance(api_response, dict):
            token_info = api_response.get('result', [])
            if isinstance(token_info, list) and token_info:
                d = token_info[0]
                if isinstance(d, dict):
                    if 'divisor' in d and str(d['divisor']).isdigit():
                        return int(d['divisor'])
                    if 'decimals' in d and str(d['decimals']).isdigit():
                        return int(d['decimals'])
        # Web3 fallback
        w3 = RPCPool(RPC_NODES).web3()
        token_contract = w3.eth.contract(address=Web3.to_checksum_address(token_address), abi=ERC20_ABI)
        return int(token_contract.functions.decimals().call())
    except Exception as e:
        logger.warning(f"Не удалось получить decimals для токена {token_address}: {e}")
        return 18

# ======== GUI ========
class _QtSignalProxy(QtCore.QObject):
    """Прокси-объект для безопасного обновления UI из фоновых потоков"""
    append_text = QtCore.pyqtSignal(str)


class LogHandler(logging.Handler):
    """Пользовательский обработчик логирования: отправляет логи в QTextEdit через сигнал (GUI-поток)."""
    def __init__(self, text_widget: Any):
        super().__init__()
        self.text_widget = text_widget
        self._proxy = _QtSignalProxy()
        if hasattr(self.text_widget, 'append'):
            # queued connection по умолчанию между потоками
            self._proxy.append_text.connect(self.text_widget.append)
        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
    def emit(self, record):
        try:
            msg = self.format(record)
            self._proxy.append_text.emit(msg)
        except Exception:
            # Никогда не роняем приложение из-за логов
            pass

# ======== #MCP:GUI ========

class MassDistributionTab(QtWidgets.QWidget):
    # Сигналы этой вкладки
    update_progress_signal = QtCore.pyqtSignal(str, int)
    update_status_signal = QtCore.pyqtSignal(str)
    update_table_signal = QtCore.pyqtSignal(list, dict, dict)
    update_address_status = QtCore.pyqtSignal(int, str)

    def __init__(self, slot_id: str, slot_title: str, rpc_pool, app_config: Config, api_key: str, parent=None):
        super().__init__(parent)
        self.slot_id = slot_id
        self.slot_title = slot_title
        self.rpc = rpc_pool
        self.cfg = app_config
        self.api_key = api_key
        self._stop = threading.Event()
        self._paused = threading.Event()
        self._active = False
        self._thread: Optional[threading.Thread] = None
        self._w3 = self.rpc.web3()
        self._sender_address: Optional[str] = None

        self._build_ui()
        self._wire_signals()
        self._load_slot_wallet()

    # --- UI построение ---
    def _build_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        header = QtWidgets.QLabel(f"{self.slot_title} [{self.slot_id}]")
        header.setStyleSheet("font-weight:600")
        layout.addWidget(header)

        # Блок кошелька
        wallet_box = QtWidgets.QGroupBox("Кошелёк вкладки")
        wl = QtWidgets.QGridLayout(wallet_box)
        self.addr_label = QtWidgets.QLabel("Адрес: —")
        self.btn_mn = QtWidgets.QPushButton("Ввести сид-фразу")
        self.btn_pk = QtWidgets.QPushButton("Ввести приватный ключ")
        self.btn_refresh = QtWidgets.QPushButton("↻")
        self.balances_label = QtWidgets.QLabel("BNB: — | PLEX: — | USDT: —")
        wl.addWidget(self.addr_label, 0, 0, 1, 3)
        wl.addWidget(self.btn_mn, 1, 0)
        wl.addWidget(self.btn_pk, 1, 1)
        wl.addWidget(self.btn_refresh, 1, 2)
        wl.addWidget(self.balances_label, 2, 0, 1, 3)
        layout.addWidget(wallet_box)

        # Параметры
        params = QtWidgets.QGroupBox("Параметры")
        pl = QtWidgets.QGridLayout(params)
        pl.addWidget(QtWidgets.QLabel("Токен:"), 0, 0)
        self.token_combo = QtWidgets.QComboBox()
        self.token_combo.addItems(["BNB", "PLEX", "USDT", "Другой..."])
        self.token_combo.setCurrentText("PLEX")
        self.token_addr_edit = QtWidgets.QLineEdit()
        self.token_addr_edit.setPlaceholderText("Адрес контракта 0x...")
        self.token_addr_edit.setEnabled(False)
        self.token_combo.currentTextChanged.connect(lambda t: self.token_addr_edit.setEnabled(t == "Другой..."))
        pl.addWidget(self.token_combo, 0, 1)
        pl.addWidget(self.token_addr_edit, 0, 2)
        pl.addWidget(QtWidgets.QLabel("Сумма на адрес:"), 1, 0)
        self.amount_spin = QtWidgets.QDoubleSpinBox()
        self.amount_spin.setRange(0.00000001, 1_000_000)
        self.amount_spin.setDecimals(8)
        self.amount_spin.setValue(0.05)
        pl.addWidget(self.amount_spin, 1, 1)
        pl.addWidget(QtWidgets.QLabel("Интервал (сек):"), 2, 0)
        self.interval_spin = QtWidgets.QSpinBox()
        self.interval_spin.setRange(1, 600)
        self.interval_spin.setValue(5)
        pl.addWidget(self.interval_spin, 2, 1)
        pl.addWidget(QtWidgets.QLabel("Циклы:"), 3, 0)
        self.cycles_spin = QtWidgets.QSpinBox()
        self.cycles_spin.setRange(1, 100)
        self.cycles_spin.setValue(1)
        pl.addWidget(self.cycles_spin, 3, 1)
        layout.addWidget(params)

        # Импорт адресов
        import_box = QtWidgets.QGroupBox("Адреса получателей")
        il = QtWidgets.QVBoxLayout(import_box)
        self.addresses_edit = QtWidgets.QTextEdit()
        self.addresses_edit.setPlaceholderText("Вставьте адреса (через пробел, перевод строки или запятую)")
        btns = QtWidgets.QHBoxLayout()
        self.btn_add = QtWidgets.QPushButton("Добавить")
        self.btn_clear = QtWidgets.QPushButton("Очистить")
        btns.addWidget(self.btn_add)
        btns.addWidget(self.btn_clear)
        il.addWidget(self.addresses_edit)
        il.addLayout(btns)
        layout.addWidget(import_box)

        # Таблица
        self.table = QtWidgets.QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(['Адрес', 'Статус', 'Отправлено', 'Tx', 'Время'])
        self.table.horizontalHeader().setSectionResizeMode(0, STRETCH_MODE)
        layout.addWidget(self.table)

        # Управление
        ctrl = QtWidgets.QHBoxLayout()
        self.btn_start = QtWidgets.QPushButton("Старт")
        self.btn_pause = QtWidgets.QPushButton("Пауза")
        self.btn_resume = QtWidgets.QPushButton("Продолжить")
        self.btn_stop = QtWidgets.QPushButton("Стоп")
        self.btn_pause.setEnabled(False)
        self.btn_resume.setEnabled(False)
        self.btn_stop.setEnabled(False)
        ctrl.addWidget(self.btn_start)
        ctrl.addWidget(self.btn_pause)
        ctrl.addWidget(self.btn_resume)
        ctrl.addWidget(self.btn_stop)
        layout.addLayout(ctrl)

        # Прогресс
        self.progress = QtWidgets.QProgressBar()
        self.progress_label = QtWidgets.QLabel("Готово")
        layout.addWidget(self.progress)
        layout.addWidget(self.progress_label)

    def _wire_signals(self):
        self.btn_mn.clicked.connect(self._enter_mnemonic)
        self.btn_pk.clicked.connect(self._enter_pk)
        self.btn_refresh.clicked.connect(self._refresh_balances)
        self.btn_add.clicked.connect(self._add_addresses)
        self.btn_clear.clicked.connect(self._clear_addresses)
        self.btn_start.clicked.connect(self._start)
        self.btn_pause.clicked.connect(self._pause)
        self.btn_resume.clicked.connect(self._resume)
        self.btn_stop.clicked.connect(self._stop_clicked)

    # --- Кошелёк ---
    def _load_slot_wallet(self):
        pk = self.cfg.get_key(self.slot_id)
        if pk and WEB3_AVAILABLE and Account:
            try:
                acct = Account.from_key(pk)
                self._sender_address = Web3.to_checksum_address(acct.address)
                self.addr_label.setText(f"Адрес: {self._sender_address}")
                self._refresh_balances()
            except Exception as e:
                logger.warning(f"[{self.slot_id}] не удалось загрузить адрес из PK: {e}")

    def _enter_mnemonic(self):
        text, ok = QtWidgets.QInputDialog.getMultiLineText(self, "Сид-фраза", "Введите сид-фразу (m/44'/60'/0'/0/0):")
        if not ok or not text.strip():
            return
        try:
            if not Account:
                raise RuntimeError("eth_account недоступен")
            acct = Account.from_mnemonic(text.strip())
            pk = acct.key.hex()
            self.cfg.set_mnemonic(text.strip(), slot=self.slot_id)
            self.cfg.set_key(pk, slot=self.slot_id)
            self._sender_address = Web3.to_checksum_address(acct.address)
            self.addr_label.setText(f"Адрес: {self._sender_address}")
            self._refresh_balances()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Не удалось импортировать сид-фразу: {e}')

    def _enter_pk(self):
        text, ok = QtWidgets.QInputDialog.getText(self, "Приватный ключ", "Введите приватный ключ (0x...)")
        if not ok or not text.strip():
            return
        pk = text.strip()
        try:
            if pk.startswith('0x') and len(pk) in (66, 64+2):
                acct = Account.from_key(pk)
                self.cfg.set_key(pk, slot=self.slot_id)
                self._sender_address = Web3.to_checksum_address(acct.address)
                self.addr_label.setText(f"Адрес: {self._sender_address}")
                self._refresh_balances()
            else:
                raise ValueError('Неверный формат приватного ключа')
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Не удалось сохранить приватный ключ: {e}')

    def _refresh_balances(self):
        if not self._sender_address or not WEB3_AVAILABLE:
            self.balances_label.setText("BNB: — | PLEX: — | USDT: —")
            return
        try:
            bal_bnb = self._w3.eth.get_balance(self._sender_address) / 10**18
            # PLEX и USDT через контракты
            plex_addr = self.cfg.data.get('plex_address') or PLEX_CONTRACT
            usdt_addr = self.cfg.data.get('usdt_address') or USDT_CONTRACT
            plex = self._w3.eth.contract(address=Web3.to_checksum_address(plex_addr), abi=ERC20_ABI)
            usdt = self._w3.eth.contract(address=Web3.to_checksum_address(usdt_addr), abi=ERC20_ABI)
            plex_dec = plex.functions.decimals().call()
            usdt_dec = usdt.functions.decimals().call()
            plex_bal = plex.functions.balanceOf(self._sender_address).call() / (10**plex_dec)
            usdt_bal = usdt.functions.balanceOf(self._sender_address).call() / (10**usdt_dec)
            self.balances_label.setText(f"BNB: {bal_bnb:.5f} | PLEX: {plex_bal:.4f} | USDT: {usdt_bal:.4f}")
        except Exception as e:
            self.balances_label.setText(f"Баланс: ошибка ({e})")

    # --- Адреса ---
    def _parse_addresses(self, text: str) -> List[str]:
        parts = [p.strip() for p in text.replace('\n', ' ').replace(',', ' ').split(' ') if p.strip()]
        addrs: List[str] = []
        for p in parts:
            if isinstance(p, str) and p.startswith('0x') and len(p) == 42:
                addrs.append(Web3.to_checksum_address(p))
        return list(dict.fromkeys(addrs))

    def _add_addresses(self):
        addrs = self._parse_addresses(self.addresses_edit.toPlainText())
        existing = set(self.table.item(r, 0).text() for r in range(self.table.rowCount()))
        for a in addrs:
            if a in existing:
                continue
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QtWidgets.QTableWidgetItem(a))
            self.table.setItem(row, 1, QtWidgets.QTableWidgetItem("Ожидание"))
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem("0"))
            self.table.setItem(row, 3, QtWidgets.QTableWidgetItem("-"))
            self.table.setItem(row, 4, QtWidgets.QTableWidgetItem("-"))
        self.addresses_edit.clear()

    def _clear_addresses(self):
        self.table.setRowCount(0)

    # --- Управление рассылкой ---
    def _start(self):
        if self._active:
            return
        if not self._sender_address:
            QtWidgets.QMessageBox.warning(self, 'Кошелёк не задан', 'Сначала введите сид-фразу или приватный ключ для этой вкладки')
            return
        if self.table.rowCount() == 0:
            QtWidgets.QMessageBox.information(self, 'Нет адресов', 'Добавьте адреса получателей')
            return
        self._active = True
        self._stop.clear()
        self._paused.clear()
        self.btn_start.setEnabled(False)
        self.btn_pause.setEnabled(True)
        self.btn_stop.setEnabled(True)
        self.progress.setValue(0)
        self.progress_label.setText("Старт")
        self._thread = threading.Thread(target=self._run_loop, daemon=True)
        self._thread.start()

    def _pause(self):
        if not self._active:
            return
        self._paused.set()
        self.btn_pause.setEnabled(False)
        self.btn_resume.setEnabled(True)

    def _resume(self):
        if not self._active:
            return
        self._paused.clear()
        self.btn_pause.setEnabled(True)
        self.btn_resume.setEnabled(False)

    def _stop_clicked(self):
        if not self._active:
            return
        self._stop.set()

    def _run_loop(self):
        try:
            token_label = self.token_combo.currentText()
            token_addr = None
            decimals = 18
            if token_label == 'Другой...':
                token_addr = self.token_addr_edit.text().strip()
                decimals = get_token_decimal(token_addr, api_key=self.api_key)
            elif token_label == 'USDT':
                token_addr = USDT_CONTRACT
                decimals = get_token_decimal(token_addr, api_key=self.api_key)
            elif token_label == 'PLEX':
                token_addr = PLEX_CONTRACT
                decimals = get_token_decimal(token_addr, api_key=self.api_key)

            amount = float(self.amount_spin.value())
            interval = int(self.interval_spin.value())
            cycles = int(self.cycles_spin.value())

            # создаём запись рассылки
            dist_id = add_mass_distribution(
                name=f"{self.slot_title} {datetime.now().strftime('%H:%M:%S')}",
                token_address=token_addr or 'native',
                token_symbol=token_label,
                amount_per_tx=amount,
                total_addresses=self.table.rowCount(),
                total_cycles=cycles,
                interval_seconds=interval,
                slot=self.slot_id
            )

            for cycle in range(cycles):
                if self._stop.is_set():
                    break
                for row in range(self.table.rowCount()):
                    if self._stop.is_set():
                        break
                    while self._paused.is_set() and not self._stop.is_set():
                        time.sleep(0.2)
                    to_addr = self.table.item(row, 0).text()
                    self._set_row_status(row, "⟳ Отправка...")
                    try:
                        if token_addr:
                            tx_hash = self._send_token_slot(to_addr, token_addr, decimals, amount)
                        else:
                            tx_hash = self._send_bnb_slot(to_addr, amount)
                        self._set_row_status(row, "✓ Успешно")
                        self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(tx_hash))
                        self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(int(self.table.item(row, 2).text()) + 1)))
                        self.table.setItem(row, 4, QtWidgets.QTableWidgetItem(datetime.now().strftime('%H:%M:%S')))
                        add_mass_distribution_item(dist_id, to_addr, cycle+1, tx_hash, 'success', None, slot=self.slot_id)
                    except Exception as e:
                        self._set_row_status(row, "✗ Ошибка")
                        add_mass_distribution_item(dist_id, to_addr, cycle+1, '-', 'error', str(e), slot=self.slot_id)
                    # прогресс
                    done = cycle * self.table.rowCount() + row + 1
                    total = cycles * self.table.rowCount()
                    self._set_progress(done, total)
                    time.sleep(interval)

            update_mass_distribution_status(dist_id, 'completed' if not self._stop.is_set() else 'cancelled')
        finally:
            self._active = False
            QtCore.QMetaObject.invokeMethod(self.btn_start, "setEnabled", QtCore.Qt.QueuedConnection, QtCore.Q_ARG(bool, True))
            QtCore.QMetaObject.invokeMethod(self.btn_pause, "setEnabled", QtCore.Qt.QueuedConnection, QtCore.Q_ARG(bool, False))
            QtCore.QMetaObject.invokeMethod(self.btn_resume, "setEnabled", QtCore.Qt.QueuedConnection, QtCore.Q_ARG(bool, False))
            QtCore.QMetaObject.invokeMethod(self.btn_stop, "setEnabled", QtCore.Qt.QueuedConnection, QtCore.Q_ARG(bool, False))

    def _set_progress(self, done: int, total: int):
        pct = int(100 * done / max(total, 1))
        QtCore.QMetaObject.invokeMethod(self.progress, "setValue", QtCore.Qt.QueuedConnection, QtCore.Q_ARG(int, pct))
        QtCore.QMetaObject.invokeMethod(self.progress_label, "setText", QtCore.Qt.QueuedConnection,
                                        QtCore.Q_ARG(str, f"{done}/{total}"))

    def _set_row_status(self, row: int, status: str):
        item = QtWidgets.QTableWidgetItem(status)
        if status.startswith('✓'):
            item.setBackground(QtGui.QColor('#004400'))
        elif status.startswith('✗'):
            item.setBackground(QtGui.QColor('#440000'))
        elif status.startswith('⟳'):
            item.setBackground(QtGui.QColor('#444400'))
        QtCore.QMetaObject.invokeMethod(self.table, "setItem", QtCore.Qt.QueuedConnection,
                                        QtCore.Q_ARG(int, row), QtCore.Q_ARG(int, 1), QtCore.Q_ARG(object, item))

    # --- Отправка ---
    def _get_nonce(self):
        return self._w3.eth.get_transaction_count(self._sender_address)

    def _gas_price(self):
        return int(float(getattr(self.cfg, 'get_gas_price', lambda: 5)()) * 1e9)

    def _send_bnb_slot(self, to_address: str, amount_native: float) -> str:
        pk = self.cfg.get_key(self.slot_id)
        if not pk:
            raise RuntimeError('Не задан приватный ключ для вкладки')
        tx = {
            'nonce': self._get_nonce(),
            'to': Web3.to_checksum_address(to_address),
            'value': int(amount_native * 10**18),
            'gas': 21000,
            'gasPrice': self._gas_price(),
            'chainId': getattr(self._w3.eth, 'chain_id', 56),
        }
        signed = self._w3.eth.account.sign_transaction(tx, pk)
        tx_hash = send_raw_tx(self._w3, signed.rawTransaction).hex()
        return tx_hash

    def _send_token_slot(self, to_address: str, token_addr: str, decimals: int, amount_token: float) -> str:
        pk = self.cfg.get_key(self.slot_id)
        if not pk:
            raise RuntimeError('Не задан приватный ключ для вкладки')
        token = self._w3.eth.contract(address=Web3.to_checksum_address(token_addr), abi=ERC20_ABI)
        amount_wei = int(amount_token * (10 ** decimals))
        nonce = self._get_nonce()
        tx = token.functions.transfer(Web3.to_checksum_address(to_address), amount_wei).build_transaction({
            'from': self._sender_address,
            'nonce': nonce,
            'gasPrice': self._gas_price(),
        })
        # оценка газа с запасом
        gas_est = token.functions.transfer(Web3.to_checksum_address(to_address), amount_wei).estimate_gas({'from': self._sender_address})
        tx['gas'] = int(gas_est * 1.2)
        tx['chainId'] = getattr(self._w3.eth, 'chain_id', 56)
        signed = self._w3.eth.account.sign_transaction(tx, pk)
        tx_hash = send_raw_tx(self._w3, signed.rawTransaction).hex()
        return tx_hash

class MainWindow(QtWidgets.QMainWindow):
    def closeEvent(self, event: QtGui.QCloseEvent) -> None:  # type: ignore
        """Гарантированная остановка фоновых процессов при закрытии окна.
        Останавливает массовую рассылку, трекер транзакций и поисковые/скан потоки.
        """
        try:
            logger.info("Завершение приложения: остановка фоновых потоков…")

            # Массовая рассылка
            if getattr(self, 'mass_distribution_active', False):
                try:
                    self._mass_stop_distribution()
                except Exception as e:
                    logger.warning(f"Ошибка остановки массовой рассылки при закрытии: {e}")

            # Общий трекер транзакций
            if getattr(self, 'tx_tracker', None) and self.tx_tracker.is_alive():  # type: ignore[attr-defined]
                try:
                    self._tx_tracker_stop_flag = True  # если используется флаг
                except Exception:
                    pass

            # Поисковый поток
            if getattr(self, 'search_thread', None) and self.search_thread.is_alive():  # type: ignore[attr-defined]
                try:
                    self._search_stop_flag = True
                except Exception:
                    pass

            # Дополнительно: выставим паузу/стоп для массовой рассылки
            if getattr(self, 'mass_distribution_thread', None) and self.mass_distribution_thread.is_alive():  # type: ignore[attr-defined]
                try:
                    self.mass_distribution_active = False
                except Exception:
                    pass

            # Неблокирующее ожидание небольшое, чтобы потоки успели завершить итерацию
            deadline = time.time() + 2.0
            for th_name in ["mass_distribution_thread", "tx_tracker", "search_thread"]:
                th = getattr(self, th_name, None)
                if th and getattr(th, 'is_alive', lambda: False)():
                    try:
                        remaining = deadline - time.time()
                        if remaining > 0:
                            th.join(timeout=remaining)
                    except Exception:
                        pass

            logger.info("Фоновые потоки остановлены. Закрытие окна.")
        finally:
            event.accept()
    """Главное окно приложения WalletSender
    
    REFACTOR:MCP - Разделить на более мелкие компоненты UI
    OPTIMIZE:MCP - Улучшить производительность обновления таблиц
    TODO:MCP - Добавить горячие клавиши
    DOCS:MCP - Документировать пользовательский интерфейс
    """
    
    # Сигнал для обновления индикаторов прогресса из рабочих потоков
    update_progress_signal = QtCore.pyqtSignal(str, int)
    update_status_signal = QtCore.pyqtSignal(str)
    update_table_signal = QtCore.pyqtSignal(list, dict, dict)
    queue_update_signal = QtCore.pyqtSignal(list)
    queue_item_update_signal = QtCore.pyqtSignal(int, dict)
    completed_item_signal = QtCore.pyqtSignal(dict)
    update_address_status = QtCore.pyqtSignal(int, str)  # row, status для массовой рассылки
    found_tx_added_signal = QtCore.pyqtSignal()  # сигнал для обновления таблицы найденных транзакций
    # Доп. сигналы для потокобезопасного UI
    update_mass_cycle_label = QtCore.pyqtSignal(str)
    mass_distribution_finished_signal = QtCore.pyqtSignal()
    # Сигнал для обновления прогресса адреса (row, current, total)
    address_progress_signal = QtCore.pyqtSignal(int, int, int)
    mass_stats_signal = QtCore.pyqtSignal(int, int, float, float)  # sent, errors, gas_spent, total_amount

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle('Анализатор & Награды PLEX ONE/USDT')
        self.resize(1000, 800)
        self.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())

        # Базовые поля состояния
        self.cfg = Config()
        init_db()
        self.rpc = RPCPool(RPC_NODES)
        self.pk: Optional[str] = self.cfg.get_key()
        self.mn: Optional[str] = self.cfg.get_mnemonic()
        self.sending: bool = False
        self.tx_tracker: Optional[threading.Thread] = None
        self.search_thread: Optional[threading.Thread] = None
        self.is_searching: bool = False
        self.stop_search_event = threading.Event()
        self.last_rewards_state: Optional[List] = None
        self.reward_per_tx_mode = self.cfg.get_reward_per_tx()
        self.queue_paused = False

        # Параметры прямой отправки
        self.direct_send_timer = QtCore.QTimer()
        self.direct_send_timer.timeout.connect(self._direct_send_periodic_send)
        self.direct_send_active = False
        self.direct_send_current_period = 0
        self.direct_send_total_periods = 0
        self.direct_send_params: Dict[str, Any] = {}

        # UI
        self._build_ui()

        # Сигналы
        self.update_progress_signal.connect(self._update_progress)
        self.update_status_signal.connect(self.log)
        self.update_table_signal.connect(self._update_search_results)
        self.queue_update_signal.connect(self._update_queue_table)
        self.queue_item_update_signal.connect(self._update_queue_item)
        self.completed_item_signal.connect(self._add_to_completed_table)
        self.update_address_status.connect(self._update_address_status_ui)
        self.update_mass_cycle_label.connect(self._set_mass_cycle_label)
        self.mass_distribution_finished_signal.connect(self._mass_distribution_finished)
        self.address_progress_signal.connect(self._update_progress_item)
        self.mass_stats_signal.connect(self._update_mass_statistics)

        # Регистрация нестандартных типов (на случай генерации QVector<int> внутри Qt)
        try:
            from PyQt5.QtCore import qRegisterMetaType
            qRegisterMetaType('QVector<int>')  # подавляем предупреждение очереди аргументов
        except Exception:
            pass

        # Логи в UI
        self.log_handler = LogHandler(self.log_area)
        logger.addHandler(self.log_handler)
        logger.setLevel(logging.INFO)
        logger.info("Приложение запущено")
        
    @QtCore.pyqtSlot(int, str)
    def _update_address_status_ui(self, row, status):
        """Обновление статуса адреса в таблице массовой рассылки"""
        try:
            if not hasattr(self, 'mass_table'):
                return
            if row >= self.mass_table.rowCount():
                return
            status_item = QtWidgets.QTableWidgetItem(status)
            # Устанавливаем цвет в зависимости от статуса
            if status == "✓ Успешно":
                status_item.setBackground(QtGui.QColor('#004400'))
            elif status == "✗ Ошибка":
                status_item.setBackground(QtGui.QColor('#440000'))
            elif status == "⟳ Отправка...":
                status_item.setBackground(QtGui.QColor('#444400'))
            self.mass_table.setItem(row, 1, status_item)
        except Exception as e:
            logger.error(f"Ошибка обновления статуса: {e}")

    @QtCore.pyqtSlot(str)
    def _set_mass_cycle_label(self, text: str) -> None:
        if hasattr(self, 'mass_current_cycle_label') and self.mass_current_cycle_label:
            self.mass_current_cycle_label.setText(text)

    def _build_ui(self):
        """Построение пользовательского интерфейса с вкладками"""
        tabs = QtWidgets.QTabWidget()
        tabs.addTab(self._tab_analyze(), 'Анализ')
        tabs.addTab(self._tab_paginated_search(), 'Поиск транзакций')
        tabs.addTab(self._tab_rewards(), 'Награды')
        tabs.addTab(self._tab_tx_rewards(), 'Награды за Tx')
        tabs.addTab(self._tab_direct_send(), 'Прямая отправка')
        # Три независимых вкладки массовой рассылки
        try:
            tabs.addTab(
                MassDistributionTab(slot_id="slot1", slot_title="Массовая рассылка 1", rpc_pool=self.rpc, app_config=self.cfg,
                                    api_key="RAI3FTD9W53JPYZ2AHW8IBH9BXUC71NRH1"),
                'Массовая рассылка 1'
            )
            tabs.addTab(
                MassDistributionTab(slot_id="slot2", slot_title="Массовая рассылка 2", rpc_pool=self.rpc, app_config=self.cfg,
                                    api_key="U89HXHR9Y26CHMWAA9JUZ17YK2AAXS65CZ"),
                'Массовая рассылка 2'
            )
            tabs.addTab(
                MassDistributionTab(slot_id="slot3", slot_title="Массовая рассылка 3", rpc_pool=self.rpc, app_config=self.cfg,
                                    api_key="RF1Q8SCFHFD1EVAP5A4WCMIM4DREA7UNUH"),
                'Массовая рассылка 3'
            )
        except Exception as e:
            logger.error(f"Ошибка инициализации вкладок массовой рассылки: {e}")
        tabs.addTab(self._tab_sending_queue(), 'Очередь отправки')  # Новая вкладка
        tabs.addTab(self._tab_ds(), 'ДС')  # #MCP:DS_TAB - Вкладка дополнительных сервисов
        tabs.addTab(self._tab_history(), 'История')
        tabs.addTab(self._tab_found_tx(), 'Найденные Tx')
        tabs.addTab(self._tab_settings(), 'Настройки')
        self.setCentralWidget(tabs)

        # Область для логов внизу окна
        self.log_area = QtWidgets.QTextEdit()
        self.log_area.setReadOnly(True)
        dock = QtWidgets.QDockWidget('Лог')
        dock.setWidget(self.log_area)
        self.addDockWidget(QtCore.Qt.DockWidgetArea(BOTTOM_DOCK_WIDGET_AREA), dock)

    def _update_progress(self, bar_name: str, value: int) -> None:
        """Обновление индикатора прогресса с защитой от отсутствующих виджетов и ошибок."""
        try:
            mapping = {
                'scan': 'progress',
                'send': 'progress_send',
                'search': 'progress_search',
                'tx_rewards': 'progress_tx_rewards',
                'direct_send': 'direct_send_progress',
                'mass': 'mass_progress'
            }
            attr = mapping.get(bar_name)
            if not attr:
                return
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setValue(int(value))
        except Exception as e:
            logger.error(f"Ошибка _update_progress({bar_name}): {e}")

    def log(self, text: str) -> None:
        """Устаревший метод логирования - перенаправляет в logger"""
        logger.info(text)

    def _toggle_search_mode(self) -> None:
        """Переключает режим поиска между точной суммой и диапазоном"""
        is_exact_mode = self.radio_exact.isChecked()
        
        # Включить/выключить поле для точной суммы
        self.spin_amt.setEnabled(is_exact_mode)
        
        # Включить/выключить поля для диапазона
        self.spin_amt_from.setEnabled(not is_exact_mode)
        self.spin_amt_to.setEnabled(not is_exact_mode)
        
        logger.debug(f"Режим поиска переключен на: {'точная сумма' if is_exact_mode else 'диапазон'}")
        
    # --- Вкладка «Очередь отправки» ---
    def _tab_sending_queue(self):
        """Создание вкладки для отслеживания очереди отправки"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Заголовок и статус текущей отправки
        status_group = QtWidgets.QGroupBox("Статус отправки")
        status_layout = QtWidgets.QVBoxLayout(status_group)
        
        self.queue_status_label = QtWidgets.QLabel("Не активно")
        self.queue_progress_label = QtWidgets.QLabel("0/0 отправлено")
        status_layout.addWidget(self.queue_status_label)
        status_layout.addWidget(self.queue_progress_label)
        
        layout.addWidget(status_group)
        
        # Таблица очереди отправки
        self.queue_table = QtWidgets.QTableWidget(0, 5)
        self.queue_table.setHorizontalHeaderLabels([
            'Статус', 'Адрес получателя', 'Токен', 'Сумма', 'Время'
        ])
        header = self.queue_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(STRETCH_MODE)
        self.queue_table.setSelectionBehavior(SELECT_ROWS)
        layout.addWidget(self.queue_table)
        
        # Таблица завершенных отправок
        completed_group = QtWidgets.QGroupBox("Завершенные отправки")
        completed_layout = QtWidgets.QVBoxLayout(completed_group)
        
        self.completed_table = QtWidgets.QTableWidget(0, 6)
        self.completed_table.setHorizontalHeaderLabels([
            'Статус', 'Адрес получателя', 'Токен', 'Сумма', 'Хэш TX', 'Время'
        ])
        header = self.completed_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(STRETCH_MODE)  # QHeaderView.Stretch
        self.completed_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.completed_table.setContextMenuPolicy(CUSTOM_CONTEXT_MENU)
        self.completed_table.customContextMenuRequested.connect(self._queue_context_menu)
        completed_layout.addWidget(self.completed_table)
        
        layout.addWidget(completed_group)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.pause_queue_btn = QtWidgets.QPushButton("Пауза")
        self.pause_queue_btn.setEnabled(False)
        self.pause_queue_btn.clicked.connect(self._pause_sending_queue)
        buttons_layout.addWidget(self.pause_queue_btn)
        
        self.resume_queue_btn = QtWidgets.QPushButton("Продолжить")
        self.resume_queue_btn.setEnabled(False)
        self.resume_queue_btn.clicked.connect(self._resume_sending_queue)
        buttons_layout.addWidget(self.resume_queue_btn)
        
        self.cancel_queue_btn = QtWidgets.QPushButton("Отменить все")
        self.cancel_queue_btn.setEnabled(False)
        self.cancel_queue_btn.clicked.connect(self._cancel_sending_queue)
        buttons_layout.addWidget(self.cancel_queue_btn)
        
        self.export_results_btn = QtWidgets.QPushButton("Экспорт результатов")
        self.export_results_btn.clicked.connect(self._export_queue_results)
        buttons_layout.addWidget(self.export_results_btn)
        
        # Кнопка для статистики API ключей
        self.api_stats_btn = QtWidgets.QPushButton("Статистика API")
        self.api_stats_btn.clicked.connect(self._show_api_keys_stats)
        buttons_layout.addWidget(self.api_stats_btn)
        
        layout.addLayout(buttons_layout)
        
        return w

    def _pause_sending_queue(self):
        """Приостановка отправки"""
        self.queue_paused = True
        self.pause_queue_btn.setEnabled(False)
        self.resume_queue_btn.setEnabled(True)
        self.queue_status_label.setText("Приостановлено")
        logger.info("Отправка приостановлена")

    def _resume_sending_queue(self):
        """Возобновление отправки"""
        self.queue_paused = False
        self.pause_queue_btn.setEnabled(True)
        self.resume_queue_btn.setEnabled(False)
        self.queue_status_label.setText("Отправка")
        logger.info("Отправка возобновлена")

    def _cancel_sending_queue(self):
        """Отмена всей очереди отправки"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Вы уверены, что хотите отменить все оставшиеся отправки?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.sending = False
            logger.info("Отправка отменена")

    @QtCore.pyqtSlot(list)
    def _update_queue_table(self, items):
        """Обновление таблицы очереди отправки"""
        try:
            self.queue_table.setRowCount(0)
            
            for item in items:
                row = self.queue_table.rowCount()
                self.queue_table.insertRow(row)
                
                # Статус
                status_item = QtWidgets.QTableWidgetItem(item['status'])
                if item['status'] == 'success':
                    status_item.setBackground(QtGui.QColor('#004400'))
                elif item['status'] == 'failed':
                    status_item.setBackground(QtGui.QColor('#440000'))
                elif item['status'] == 'processing':
                    status_item.setBackground(QtGui.QColor('#444400'))
                self.queue_table.setItem(row, 0, status_item)
                
                # Адрес
                self.queue_table.setItem(row, 1, QtWidgets.QTableWidgetItem(item['address']))
                
                # Токен
                self.queue_table.setItem(row, 2, QtWidgets.QTableWidgetItem(item['token']))
                
                # Сумма
                self.queue_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(item['amount'])))
                
                # Время
                time_text = item['time'] if item['time'] else "-"
                self.queue_table.setItem(row, 4, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при обновлении таблицы очереди: {e}")

    @QtCore.pyqtSlot(int, dict)
    def _update_queue_item(self, index, item):
        """Обновление конкретного элемента в очереди"""
        try:
            if index >= self.queue_table.rowCount():
                return
            
            # Статус
            status_item = QtWidgets.QTableWidgetItem(item['status'])
            if item['status'] == 'success':
                status_item.setBackground(QtGui.QColor('#004400'))
            elif item['status'] == 'failed':
                status_item.setBackground(QtGui.QColor('#440000'))
            elif item['status'] == 'processing':
                status_item.setBackground(QtGui.QColor('#444400'))
            self.queue_table.setItem(index, 0, status_item)
            
            # Время
            time_text = item['time'] if item['time'] else "-"
            self.queue_table.setItem(index, 4, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при обновлении элемента очереди: {e}")

    @QtCore.pyqtSlot(dict)
    def _add_to_completed_table(self, item):
        """Добавление элемента в таблицу завершенных отправок"""
        try:
            row = self.completed_table.rowCount()
            self.completed_table.insertRow(row)
            
            # Статус
            status_item = QtWidgets.QTableWidgetItem(item['status'])
            if item['status'] == 'success':
                status_item.setBackground(QtGui.QColor('#004400'))
            elif item['status'] == 'failed':
                status_item.setBackground(QtGui.QColor('#440000'))
            self.completed_table.setItem(row, 0, status_item)
            
            # Адрес
            self.completed_table.setItem(row, 1, QtWidgets.QTableWidgetItem(item['address']))
            
            # Токен
            self.completed_table.setItem(row, 2, QtWidgets.QTableWidgetItem(item['token']))
            
            # Сумма
            self.completed_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(item['amount'])))
            
            # Хэш TX
            tx_item = QtWidgets.QTableWidgetItem(item['tx_hash'][:10] + "..." + item['tx_hash'][-6:] if item['tx_hash'] else "-")
            if item['tx_hash']:
                tx_item.setToolTip(item['tx_hash'])
                tx_item.setData(USER_ROLE, item['tx_hash'])
            self.completed_table.setItem(row, 4, tx_item)
            
            # Время
            time_text = item['time'] if item['time'] else "-"
            self.completed_table.setItem(row, 5, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при добавлении в таблицу завершенных: {e}")

    def _queue_context_menu(self, position):
        """Контекстное меню для таблицы завершенных отправок"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_addr = menu.addAction("Копировать адрес")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        
        action = menu.exec_(self.completed_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.completed_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.completed_table.item(row, 4).data(USER_ROLE)
        address = self.completed_table.item(row, 1).text()
        
        if action == copy_tx and tx_hash:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_addr:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == open_tx and tx_hash:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")

    def _export_queue_results(self):
        """Экспорт результатов очереди в CSV"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить результаты', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('status,address,token,amount,tx_hash,time\n')
                for row in range(self.completed_table.rowCount()):
                    status = self.completed_table.item(row, 0).text()
                    address = self.completed_table.item(row, 1).text()
                    token = self.completed_table.item(row, 2).text()
                    amount = self.completed_table.item(row, 3).text()
                    tx_item = self.completed_table.item(row, 4)
                    tx_hash = tx_item.data(USER_ROLE) if tx_item else ""
                    time = self.completed_table.item(row, 5).text()
                    
                    f.write(f'"{status}","{address}","{token}",{amount},"{tx_hash}","{time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")

    def _show_transaction_results(self, batch_id, transactions):
        """Отображение результатов отправки в отдельном окне для копирования"""
        result_dialog = QtWidgets.QDialog(self)
        result_dialog.setWindowTitle(f"Результаты отправки #{batch_id}")
        result_dialog.resize(800, 600)
        
        layout = QtWidgets.QVBoxLayout(result_dialog)
        
        # Текстовое поле для отображения результатов
        result_text = QtWidgets.QTextEdit()
        result_text.setReadOnly(True)
        result_text.setWordWrapMode(QtGui.QTextOption.NoWrap)
        
        # Формируем текст с транзакциями
        text = "Дата: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "\n\n"
        text += "ID партии: " + str(batch_id) + "\n\n"
        text += "Результаты отправки:\n\n"
        
        for tx in transactions:
            text += f"Адрес: {tx['address']}\n"
            text += f"Токен: {tx['token']}\n"
            text += f"Сумма: {tx['amount']}\n"
            text += f"Хэш TX: {tx['tx_hash']}\n"
            text += f"Статус: {tx['status']}\n"
            text += "-" * 50 + "\n"
        
        result_text.setText(text)
        layout.addWidget(result_text)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        copy_all_btn = QtWidgets.QPushButton("Копировать всё")
        copy_all_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(text))
        buttons_layout.addWidget(copy_all_btn)
        
        copy_hashes_btn = QtWidgets.QPushButton("Копировать только хэши")
        copy_hashes_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(
            "\n".join([tx['tx_hash'] for tx in transactions])
        ))
        buttons_layout.addWidget(copy_hashes_btn)
        
        export_csv_btn = QtWidgets.QPushButton("Экспорт в CSV")
        export_csv_btn.clicked.connect(lambda: self._export_tx_results_to_csv(transactions, batch_id))
        buttons_layout.addWidget(export_csv_btn)
        
        layout.addLayout(buttons_layout)
        
        result_dialog.exec_()

    def _export_tx_results_to_csv(self, transactions, batch_id):
        """Экспорт результатов транзакций в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, f'Сохранить результаты #{batch_id}', f'tx_results_{batch_id}.csv', '*.csv'
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,token,amount,tx_hash,status,time\n')
                for tx in transactions:
                    f.write(f'"{tx["address"]}","{tx["token"]}",{tx["amount"]},"{tx["tx_hash"]}","{tx["status"]}","{tx["time"]}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")

    def _show_api_keys_stats(self) -> None:
        """Отображение статистики API ключей в отдельном окне"""
        stats = api_key_rotator.get_statistics()
        
        stats_dialog = QtWidgets.QDialog(self)
        stats_dialog.setWindowTitle("Статистика API ключей BscScan")
        stats_dialog.resize(700, 500)
        
        layout = QtWidgets.QVBoxLayout(stats_dialog)
        
        # Общая информация
        info_text = QtWidgets.QTextEdit()
        info_text.setReadOnly(True)
        
        info_content = f"""
🔑 СТАТИСТИКА API КЛЮЧЕЙ BSCSCAN

📊 Общая информация:
• Всего ключей: {stats['total_keys']}
• Активных ключей: {stats['active_keys']}
• Заблокированных ключей: {stats['failed_keys']}
• Общее количество запросов: {stats['total_requests']}
• Последняя ротация: {stats['last_rotation']}

🎯 Текущий активный ключ: {stats['current_key']}

📈 Детальная статистика по ключам:
"""
        
        for key_masked, details in stats['key_details'].items():
            status_emoji = "🟢" if details['status'] == 'ACTIVE' else "🔴"
            info_content += f"""
{status_emoji} Ключ: {key_masked}
   • Статус: {details['status']}
   • Запросов: {details['requests']}
   • Успешность: {details['success_rate']:.1f}%
   • Последнее использование: {details['last_used'] or 'Никогда'}
"""
        
        info_content += f"""

⚙️ Управление ключами:
• Ротация происходит автоматически при ошибках
• Заблокированные ключи восстанавливаются при успешных запросах
• При блокировке всех ключей происходит автоматический сброс

🔄 Алгоритм ротации:
1. При ошибке API - блокировка текущего ключа
2. Переключение на следующий активный ключ
3. При блокировке всех ключей - сброс блокировок
4. Автовосстановление при успешных запросах
        """
        
        info_text.setPlainText(info_content)
        layout.addWidget(info_text)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        refresh_btn = QtWidgets.QPushButton("🔄 Обновить")
        refresh_btn.clicked.connect(lambda: self._refresh_api_stats(info_text))
        buttons_layout.addWidget(refresh_btn)
        
        reset_btn = QtWidgets.QPushButton("🔓 Разблокировать все ключи")
        reset_btn.clicked.connect(lambda: self._reset_api_keys(info_text))
        buttons_layout.addWidget(reset_btn)
        
        close_btn = QtWidgets.QPushButton("Закрыть")
        close_btn.clicked.connect(stats_dialog.close)
        buttons_layout.addWidget(close_btn)
        
        layout.addLayout(buttons_layout)
        
        stats_dialog.exec_()
    
    def _refresh_api_stats(self, text_widget: QtWidgets.QTextEdit) -> None:
        """Обновление статистики API ключей"""
        # Повторно вызываем отображение статистики
        stats = api_key_rotator.get_statistics()
        
        # Обновляем содержимое
        info_content = f"""
🔑 СТАТИСТИКА API КЛЮЧЕЙ BSCSCAN (Обновлено: {datetime.now().strftime('%H:%M:%S')})

📊 Общая информация:
• Всего ключей: {stats['total_keys']}
• Активных ключей: {stats['active_keys']}
• Заблокированных ключей: {stats['failed_keys']}
• Общее количество запросов: {stats['total_requests']}
• Последняя ротация: {stats['last_rotation']}

🎯 Текущий активный ключ: {stats['current_key']}

📈 Детальная статистика по ключам:
"""
        
        for key_masked, details in stats['key_details'].items():
            status_emoji = "🟢" if details['status'] == 'ACTIVE' else "🔴"
            info_content += f"""
{status_emoji} Ключ: {key_masked}
   • Статус: {details['status']}
   • Запросов: {details['requests']}
   • Успешность: {details['success_rate']:.1f}%
   • Последнее использование: {details['last_used'] or 'Никогда'}
"""
        
        info_content += """

⚙️ Управление ключами:
• Ротация происходит автоматически при ошибках
• Заблокированные ключи восстанавливаются при успешных запросах
• При блокировке всех ключей происходит автоматический сброс
        """
        
        text_widget.setPlainText(info_content)
    
    def _reset_api_keys(self, text_widget: QtWidgets.QTextEdit) -> None:
        """Сброс блокировок всех API ключей"""
        api_key_rotator.failed_keys.clear()
        api_key_rotator.current_index = 0
        logger.info("Все API ключи разблокированы вручную")
        
        # Обновляем отображение
        self._refresh_api_stats(text_widget)
        
        # Показываем уведомление
        QtWidgets.QMessageBox.information(
            self, 
            "Ключи разблокированы", 
            "Все API ключи успешно разблокированы!\nСброшен счетчик на первый ключ."
        )

    # --- Вкладка «Награды за Tx» ---
    def _tab_tx_rewards(self):
        """Создание вкладки для управления наградами за транзакции"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа настроек награждения
        reward_settings = QtWidgets.QGroupBox("Настройки награждения")
        reward_layout = QtWidgets.QHBoxLayout(reward_settings)
        
        # Переключатель режима награждения
        self.reward_per_tx_checkbox = QtWidgets.QCheckBox("Награждать за каждую транзакцию")
        self.reward_per_tx_checkbox.setChecked(self.reward_per_tx_mode)
        self.reward_per_tx_checkbox.toggled.connect(self._toggle_reward_mode)
        reward_layout.addWidget(self.reward_per_tx_checkbox)
        
        # Кнопка для сканирования транзакций
        self.scan_tx_btn = QtWidgets.QPushButton("Сканировать транзакции")
        self.scan_tx_btn.clicked.connect(self._scan_transactions_for_rewards)
        reward_layout.addWidget(self.scan_tx_btn)
        
        # Кнопка для очистки данных о транзакциях
        self.clear_tx_data_btn = QtWidgets.QPushButton("Очистить данные")
        self.clear_tx_data_btn.clicked.connect(self._clear_tx_data)
        reward_layout.addWidget(self.clear_tx_data_btn)
        
        layout.addWidget(reward_settings)
        
        # Группа настроек размера награды
        reward_size_group = QtWidgets.QGroupBox("Настройки размера награды")
        reward_size_layout = QtWidgets.QHBoxLayout(reward_size_group)

        # Выбор типа токена для награды
        self.reward_token_group = QtWidgets.QButtonGroup(w)
        self.reward_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.reward_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.reward_both_radio = QtWidgets.QRadioButton('Оба (PLEX ONE и USDT)')
        self.reward_plex_radio.setChecked(True)
        self.reward_token_group.addButton(self.reward_plex_radio)
        self.reward_token_group.addButton(self.reward_usdt_radio)
        self.reward_token_group.addButton(self.reward_both_radio)

        token_layout = QtWidgets.QVBoxLayout()
        token_layout.addWidget(QtWidgets.QLabel("Тип токена награды:"))
        token_layout.addWidget(self.reward_plex_radio)
        token_layout.addWidget(self.reward_usdt_radio)
        token_layout.addWidget(self.reward_both_radio)
        reward_size_layout.addLayout(token_layout)

        # Ввод размера награды
        size_layout = QtWidgets.QVBoxLayout()
        size_layout.addWidget(QtWidgets.QLabel("Размер награды:"))

        # PLEX ONE размер
        plex_layout = QtWidgets.QHBoxLayout()
        plex_layout.addWidget(QtWidgets.QLabel("PLEX ONE:"))
        self.reward_plex_size = QtWidgets.QDoubleSpinBox()
        self.reward_plex_size.setRange(0, 1000)
        self.reward_plex_size.setDecimals(2)
        self.reward_plex_size.setValue(3.0)  # По умолчанию 3 PLEX ONE
        plex_layout.addWidget(self.reward_plex_size)
        size_layout.addLayout(plex_layout)

        # USDT размер
        usdt_layout = QtWidgets.QHBoxLayout()
        usdt_layout.addWidget(QtWidgets.QLabel("USDT:"))
        self.reward_usdt_size = QtWidgets.QDoubleSpinBox()
        self.reward_usdt_size.setRange(0, 1000) 
        self.reward_usdt_size.setDecimals(2)
        self.reward_usdt_size.setValue(0.3)  # По умолчанию 0.3 USDT
        usdt_layout.addWidget(self.reward_usdt_size)
        size_layout.addLayout(usdt_layout)

        reward_size_layout.addLayout(size_layout)
        layout.addWidget(reward_size_group)
        
        # Таблица отправителей и их транзакций
        self.tx_senders_table = QtWidgets.QTableWidget(0, 4)
        self.tx_senders_table.setHorizontalHeaderLabels([
            'Адрес отправителя', 'Всего Tx', 'Награждено', 'Осталось наградить'
        ])
        header = self.tx_senders_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(0, STRETCH_MODE)
        self.tx_senders_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.tx_senders_table.setContextMenuPolicy(CUSTOM_CONTEXT_MENU)
        self.tx_senders_table.customContextMenuRequested.connect(self._show_tx_senders_menu)
        layout.addWidget(self.tx_senders_table)
        
        # Таблица транзакций выбранного отправителя
        self.sender_tx_table = QtWidgets.QTableWidget(0, 5)
        self.sender_tx_table.setHorizontalHeaderLabels([
            'Транзакция', 'Время', 'Токен', 'Сумма', 'Награждено'
        ])
        header = self.sender_tx_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(0, STRETCH_MODE)
        self.sender_tx_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.sender_tx_table.setContextMenuPolicy(CUSTOM_CONTEXT_MENU)
        self.sender_tx_table.customContextMenuRequested.connect(self._show_sender_tx_menu)
        layout.addWidget(self.sender_tx_table)
        
        # Кнопки действий
        action_layout = QtWidgets.QHBoxLayout()
        
        # Кнопка для создания наград
        self.create_rewards_btn = QtWidgets.QPushButton("Создать награды")
        self.create_rewards_btn.clicked.connect(self._create_rewards_from_tx)
        action_layout.addWidget(self.create_rewards_btn)
        
        # Кнопка для отправки наград
        self.send_tx_rewards_btn = QtWidgets.QPushButton("Отправить награды")
        self.send_tx_rewards_btn.clicked.connect(self._send_rewards_for_tx)
        action_layout.addWidget(self.send_tx_rewards_btn)
        
        layout.addLayout(action_layout)
        
        # Индикатор прогресса
        self.progress_tx_rewards = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_tx_rewards)
        
        # Загрузка данных при открытии
        QtCore.QTimer.singleShot(100, self._load_tx_senders)
        
        return w

    def _toggle_reward_mode(self, checked):
        """Переключение режима награждения за транзакции"""
        self.reward_per_tx_mode = checked
        self.cfg.set_reward_per_tx(checked)
        logger.info(f"Режим награждения за каждую транзакцию {'включен' if checked else 'выключен'}")
    
    def _scan_transactions_for_rewards(self):
        """Заглушка для сканирования транзакций"""
        logger.info("Сканирование транзакций для наград...")
        self._load_tx_senders()
    
    def _clear_tx_data(self):
        """Очистка данных о транзакциях"""
        clear_sender_transactions()
        self._load_tx_senders()
        logger.info("Данные о транзакциях очищены")
    
    def _show_tx_senders_menu(self, position):
        """Контекстное меню для таблицы отправителей"""
        menu = QtWidgets.QMenu()
        view_tx = menu.addAction("Показать транзакции")
        action = menu.exec_(self.tx_senders_table.viewport().mapToGlobal(position))
        if action == view_tx:
            selected = self.tx_senders_table.selectedItems()
            if selected:
                row = selected[0].row()
                sender = self.tx_senders_table.item(row, 0).text()
                self._show_sender_transactions(sender)
    
    def _show_sender_tx_menu(self, position):
        """Контекстное меню для таблицы транзакций отправителя"""
        menu = QtWidgets.QMenu()
        copy_hash = menu.addAction("Копировать хэш")
        action = menu.exec_(self.sender_tx_table.viewport().mapToGlobal(position))
        if action == copy_hash:
            selected = self.sender_tx_table.selectedItems()
            if selected:
                row = selected[0].row()
                tx_hash = self.sender_tx_table.item(row, 0).text()
                QtWidgets.QApplication.clipboard().setText(tx_hash)
    
    def _create_rewards_from_tx(self):
        """Создание наград на основе транзакций"""
        logger.info("Создание наград на основе транзакций...")
    
    def _send_rewards_for_tx(self):
        """Отправка наград за транзакции"""
        logger.info("Отправка наград за транзакции...")
    
    def _load_tx_senders(self):
        """Загрузка данных об отправителях транзакций"""
        try:
            senders = get_sender_transaction_counts()
            self.tx_senders_table.setRowCount(0)
            for sender_data in senders:
                row = self.tx_senders_table.rowCount()
                self.tx_senders_table.insertRow(row)
                self.tx_senders_table.setItem(row, 0, QtWidgets.QTableWidgetItem(sender_data['sender_addr']))
                self.tx_senders_table.setItem(row, 1, QtWidgets.QTableWidgetItem(str(sender_data['tx_count'])))
                self.tx_senders_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(sender_data['rewarded_count'])))
                unrewarded = sender_data['tx_count'] - sender_data['rewarded_count']
                self.tx_senders_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(unrewarded)))
        except Exception as e:
            logger.error(f"Ошибка при загрузке отправителей: {e}")
    
    def _show_sender_transactions(self, sender_addr):
        """Отображение транзакций конкретного отправителя"""
        try:
            transactions = get_transactions_by_sender(sender_addr)
            self.sender_tx_table.setRowCount(0)
            for tx_data in transactions:
                row = self.sender_tx_table.rowCount()
                self.sender_tx_table.insertRow(row)
                self.sender_tx_table.setItem(row, 0, QtWidgets.QTableWidgetItem(tx_data['tx_hash']))
                self.sender_tx_table.setItem(row, 1, QtWidgets.QTableWidgetItem(tx_data['tx_timestamp'] or '-'))
                self.sender_tx_table.setItem(row, 2, QtWidgets.QTableWidgetItem(tx_data['token_name']))
                self.sender_tx_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(tx_data['amount'])))
                rewarded = "Да" if tx_data['rewarded'] else "Нет"
                self.sender_tx_table.setItem(row, 4, QtWidgets.QTableWidgetItem(rewarded))
        except Exception as e:
            logger.error(f"Ошибка при загрузке транзакций отправителя: {e}")
    
    # Заглушки для остальных вкладок
    def _tab_analyze(self):
        """Вкладка анализа"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка анализа"))
        return w
    
    def _tab_paginated_search(self):
        """Вкладка поиска транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка поиска транзакций"))
        self.radio_exact = QtWidgets.QRadioButton("Точная сумма")
        self.radio_range = QtWidgets.QRadioButton("Диапазон")
        self.spin_amt = QtWidgets.QDoubleSpinBox()
        self.spin_amt_from = QtWidgets.QDoubleSpinBox()
        self.spin_amt_to = QtWidgets.QDoubleSpinBox()
        layout.addWidget(self.radio_exact)
        layout.addWidget(self.radio_range)
        self.progress_search = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_search)
        return w
    
    def _tab_rewards(self):
        """Вкладка наград"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка наград"))
        self.progress_send = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_send)
        return w
    
    def _tab_direct_send(self):
        """Вкладка прямой отправки"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка прямой отправки"))
        self.direct_send_progress = QtWidgets.QProgressBar()
        layout.addWidget(self.direct_send_progress)
        return w
    
    def _direct_send_periodic_send(self):
        """Периодическая отправка"""
        pass
    
    def _tab_ds(self):
        """Вкладка дополнительных сервисов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка дополнительных сервисов"))
        return w
    
    def _tab_history(self):
        """Вкладка истории"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка истории"))
        self.progress = QtWidgets.QProgressBar()
        layout.addWidget(self.progress)
        return w
    
    def _tab_found_tx(self):
        """Вкладка найденных транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка найденных транзакций"))
        return w
    
    def _tab_settings(self):
        """Вкладка настроек"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        layout.addWidget(QtWidgets.QLabel("Вкладка настроек"))
        return w
    
    def _update_search_results(self, transactions, sender_counter, sender_transactions):
        """Обновление результатов поиска"""
        logger.info(f"Найдено {len(transactions)} транзакций")
    
    def _update_progress_item(self, row, current, total):
        """Обновление прогресса элемента"""
        pass
    
    def _update_mass_statistics(self, sent, errors, gas_spent, total_amount):
        """Обновление статистики массовой рассылки"""
        logger.info(f"Отправлено: {sent}, Ошибок: {errors}, Газ: {gas_spent}, Сумма: {total_amount}")
    
    def _mass_distribution_finished(self):
        """Завершение массовой рассылки"""
        logger.info("Массовая рассылка завершена")
    
    def _mass_stop_distribution(self):
        """Остановка массовой рассылки"""
        self.mass_distribution_active = False
        logger.info("Массовая рассылка остановлена")


    def _tab_paginated_search(self):
        """Вкладка поиска транзакций с постраничной загрузкой"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа поиска
        search_group = QtWidgets.QGroupBox("Поиск транзакций")
        search_layout = QtWidgets.QFormLayout(search_group)
        
        # Адрес кошелька
        self.search_wallet_edit = QtWidgets.QLineEdit()
        self.search_wallet_edit.setPlaceholderText("0x...")
        search_layout.addRow("Адрес кошелька:", self.search_wallet_edit)
        
        # Адрес токена
        self.search_token_edit = QtWidgets.QLineEdit()
        self.search_token_edit.setPlaceholderText("0x... (необязательно)")
        search_layout.addRow("Адрес токена:", self.search_token_edit)
        
        # Точная сумма
        self.search_amount_edit = QtWidgets.QLineEdit()
        self.search_amount_edit.setPlaceholderText("Точная сумма (необязательно)")
        search_layout.addRow("Точная сумма:", self.search_amount_edit)
        
        # Кнопки поиска
        search_buttons = QtWidgets.QHBoxLayout()
        self.search_start_btn = QtWidgets.QPushButton("Начать поиск")
        self.search_start_btn.clicked.connect(self._start_paginated_search)
        search_buttons.addWidget(self.search_start_btn)
        
        self.search_stop_btn = QtWidgets.QPushButton("Остановить")
        self.search_stop_btn.clicked.connect(self._stop_paginated_search)
        self.search_stop_btn.setEnabled(False)
        search_buttons.addWidget(self.search_stop_btn)
        
        search_layout.addRow("", search_buttons)
        layout.addWidget(search_group)
        
        # Таблица результатов
        self.search_results_table = QtWidgets.QTableWidget(0, 5)
        self.search_results_table.setHorizontalHeaderLabels([
            'Хэш', 'От', 'К', 'Сумма', 'Время'
        ])
        self.search_results_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        layout.addWidget(self.search_results_table)
        
        # Статус поиска
        self.search_status_label = QtWidgets.QLabel("Готов к поиску")
        layout.addWidget(self.search_status_label)
        
        return w

    def _tab_rewards(self):
        """Вкладка системы наград"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа настроек наград
        settings_group = QtWidgets.QGroupBox("Настройки наград")
        settings_layout = QtWidgets.QFormLayout(settings_group)
        
        # Тип награды
        reward_type_layout = QtWidgets.QHBoxLayout()
        self.reward_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.reward_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.reward_both_radio = QtWidgets.QRadioButton('Оба токена')
        self.reward_both_radio.setChecked(True)
        
        reward_type_layout.addWidget(self.reward_plex_radio)
        reward_type_layout.addWidget(self.reward_usdt_radio)
        reward_type_layout.addWidget(self.reward_both_radio)
        settings_layout.addRow("Тип награды:", reward_type_layout)
        
        # Размеры наград
        self.reward_plex_size = QtWidgets.QDoubleSpinBox()
        self.reward_plex_size.setRange(0.1, 1000)
        self.reward_plex_size.setValue(3.0)
        self.reward_plex_size.setDecimals(1)
        settings_layout.addRow("Размер PLEX:", self.reward_plex_size)
        
        self.reward_usdt_size = QtWidgets.QDoubleSpinBox()
        self.reward_usdt_size.setRange(0.1, 10000)
        self.reward_usdt_size.setValue(1.0)
        self.reward_usdt_size.setDecimals(2)
        settings_layout.addRow("Размер USDT:", self.reward_usdt_size)
        
        layout.addWidget(settings_group)
        
        # Таблица наград
        self.table_rewards = QtWidgets.QTableWidget(0, 3)
        self.table_rewards.setHorizontalHeaderLabels([
            'Адрес получателя', 'PLEX ONE', 'USDT'
        ])
        self.table_rewards.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        layout.addWidget(self.table_rewards)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.send_rewards_btn = QtWidgets.QPushButton("Отправить награды")
        self.send_rewards_btn.clicked.connect(self._send_rewards)
        buttons_layout.addWidget(self.send_rewards_btn)
        
        self.clear_rewards_btn = QtWidgets.QPushButton("Очистить список")
        self.clear_rewards_btn.clicked.connect(self._clear_rewards)
        buttons_layout.addWidget(self.clear_rewards_btn)
        
        layout.addLayout(buttons_layout)
        
        return w

    def _tab_direct_send(self):
        """Вкладка прямой отправки токенов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа отправки
        send_group = QtWidgets.QGroupBox("Прямая отправка")
        send_layout = QtWidgets.QFormLayout(send_group)
        
        # Адрес получателя
        self.direct_address_edit = QtWidgets.QLineEdit()
        self.direct_address_edit.setPlaceholderText("0x...")
        send_layout.addRow("Адрес получателя:", self.direct_address_edit)
        
        # Тип токена
        token_layout = QtWidgets.QHBoxLayout()
        self.direct_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.direct_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.direct_bnb_radio = QtWidgets.QRadioButton('BNB')
        self.direct_plex_radio.setChecked(True)
        
        token_layout.addWidget(self.direct_plex_radio)
        token_layout.addWidget(self.direct_usdt_radio)
        token_layout.addWidget(self.direct_bnb_radio)
        send_layout.addRow("Токен:", token_layout)
        
        # Сумма
        self.direct_amount_edit = QtWidgets.QDoubleSpinBox()
        self.direct_amount_edit.setRange(0.00000001, 1000000)
        self.direct_amount_edit.setValue(1.0)
        self.direct_amount_edit.setDecimals(8)
        send_layout.addRow("Сумма:", self.direct_amount_edit)
        
        # Кнопка отправки
        self.direct_send_btn = QtWidgets.QPushButton("Отправить")
        self.direct_send_btn.clicked.connect(self._direct_send)
        send_layout.addRow("", self.direct_send_btn)
        
        layout.addWidget(send_group)
        
        # Статус отправки
        self.direct_status_label = QtWidgets.QLabel("Готов к отправке")
        layout.addWidget(self.direct_status_label)
        
        return w

    def _tab_ds(self):
        """Вкладка дополнительных сервисов (ДС)"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Заголовок
        title_label = QtWidgets.QLabel("🚀 ДС - Дополнительные Сервисы")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #1976D2;
                padding: 20px;
                text-align: center;
            }
        """)
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Информация о PLEX ONE
        info_group = QtWidgets.QGroupBox("ℹ️ Информация о PLEX ONE")
        info_layout = QtWidgets.QFormLayout(info_group)
        
        info_layout.addRow("Название:", QtWidgets.QLabel("PLEX ONE"))
        info_layout.addRow("Символ:", QtWidgets.QLabel("PLEX"))
        info_layout.addRow("Decimals:", QtWidgets.QLabel("9"))
        info_layout.addRow("Total Supply:", QtWidgets.QLabel("12,600,000 PLEX"))
        info_layout.addRow("Контракт:", QtWidgets.QLabel("0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"))
        info_layout.addRow("Сеть:", QtWidgets.QLabel("Binance Smart Chain"))
        
        layout.addWidget(info_group)
        
        # Быстрые ссылки
        links_group = QtWidgets.QGroupBox("🔗 Быстрые ссылки")
        links_layout = QtWidgets.QVBoxLayout(links_group)
        
        buttons_layout = QtWidgets.QHBoxLayout()
        
        bscscan_btn = QtWidgets.QPushButton("🔍 BSCScan")
        bscscan_btn.clicked.connect(lambda: self._open_url(
            "https://bscscan.com/address/0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
        ))
        buttons_layout.addWidget(bscscan_btn)
        
        pancake_btn = QtWidgets.QPushButton("🥞 PancakeSwap")
        pancake_btn.clicked.connect(lambda: self._open_url(
            "https://pancakeswap.finance/swap"
            "?outputCurrency=0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
            "&inputCurrency=0x55d398326f99059fF775485246999027B3197955"
        ))
        buttons_layout.addWidget(pancake_btn)
        
        chart_btn = QtWidgets.QPushButton("📈 График")
        chart_btn.clicked.connect(lambda: self._open_url(
            "https://www.geckoterminal.com/ru/bsc/pools/0x41d9650faf3341cbf8947fd8063a1fc88dbf1889"
        ))
        buttons_layout.addWidget(chart_btn)
        
        links_layout.addLayout(buttons_layout)
        layout.addWidget(links_group)
        
        return w

    def _tab_history(self):
        """Вкладка истории транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Таблица истории
        self.history_table = QtWidgets.QTableWidget(0, 5)
        self.history_table.setHorizontalHeaderLabels([
            'Тип', 'Адрес', 'Сумма', 'Хэш', 'Время'
        ])
        self.history_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        layout.addWidget(self.history_table)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.refresh_history_btn = QtWidgets.QPushButton("Обновить")
        self.refresh_history_btn.clicked.connect(self._load_history)
        buttons_layout.addWidget(self.refresh_history_btn)
        
        self.export_history_btn = QtWidgets.QPushButton("Экспорт")
        self.export_history_btn.clicked.connect(self._export_history)
        buttons_layout.addWidget(self.export_history_btn)
        
        layout.addLayout(buttons_layout)
        
        return w

    def _tab_found_tx(self):
        """Вкладка найденных транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Таблица найденных транзакций
        self.found_tx_table = QtWidgets.QTableWidget(0, 6)
        self.found_tx_table.setHorizontalHeaderLabels([
            'Хэш', 'Отправитель', 'Получатель', 'Сумма', 'Токен', 'Время'
        ])
        self.found_tx_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        layout.addWidget(self.found_tx_table)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.add_to_rewards_btn = QtWidgets.QPushButton("Добавить в награды")
        self.add_to_rewards_btn.clicked.connect(self._add_found_to_rewards)
        buttons_layout.addWidget(self.add_to_rewards_btn)
        
        self.clear_found_btn = QtWidgets.QPushButton("Очистить")
        self.clear_found_btn.clicked.connect(self._clear_found_tx)
        buttons_layout.addWidget(self.clear_found_btn)
        
        layout.addLayout(buttons_layout)
        
        return w

    def _tab_settings(self):
        """Вкладка настроек приложения"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа общих настроек
        general_group = QtWidgets.QGroupBox("Общие настройки")
        general_layout = QtWidgets.QFormLayout(general_group)
        
        # Цена газа
        self.gas_price_spin = QtWidgets.QDoubleSpinBox()
        self.gas_price_spin.setRange(0.1, 500)
        self.gas_price_spin.setDecimals(1)
        self.gas_price_spin.setSingleStep(0.1)
        self.gas_price_spin.setValue(float(self.cfg.get_gas_price()))
        self.gas_price_spin.valueChanged.connect(self._update_gas_price)
        general_layout.addRow("Цена газа (Gwei):", self.gas_price_spin)
        
        # Количество повторений серии
        self.repeat_count_spin = QtWidgets.QSpinBox()
        self.repeat_count_spin.setRange(1, 100)
        self.repeat_count_spin.setValue(self.cfg.get_repeat_count())
        self.repeat_count_spin.valueChanged.connect(self._update_repeat_count)
        general_layout.addRow("Количество повторений серии:", self.repeat_count_spin)
        
        # Очистка кэша
        btn_clear_cache = QtWidgets.QPushButton("Очистить кэш")
        btn_clear_cache.clicked.connect(self._clear_cache)
        general_layout.addRow("Кэширование:", btn_clear_cache)
        
        layout.addWidget(general_group)
        
        # Группа настроек блокчейна
        blockchain_group = QtWidgets.QGroupBox("Настройки блокчейна")
        blockchain_layout = QtWidgets.QFormLayout(blockchain_group)
        
        # Статус блокчейна
        blockchain_status = "Включен" if blockchain_enabled else "Отключен"
        blockchain_layout.addRow("Статус блокчейна:", QtWidgets.QLabel(blockchain_status))
        
        # Список RPC узлов
        rpc_list = QtWidgets.QListWidget()
        for node in RPC_NODES:
            rpc_list.addItem(node)
        blockchain_layout.addRow("RPC узлы:", rpc_list)
        
        # Тест подключения
        btn_test_conn = QtWidgets.QPushButton("Проверить подключение")
        btn_test_conn.clicked.connect(self._test_blockchain_connection)
        blockchain_layout.addRow("", btn_test_conn)
        
        layout.addWidget(blockchain_group)
        
        return w

    # Реальные методы для вкладок
    def _start_paginated_search(self):
        """Запуск постраничного поиска транзакций"""
        try:
            wallet_addr = self.search_wallet_edit.text().strip()
            token_addr = self.search_token_edit.text().strip()
            amount = self.search_amount_edit.text().strip()
            
            if not wallet_addr:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', 'Введите адрес кошелька')
                return
            
            # Запускаем поиск в отдельном потоке
            self.search_thread = threading.Thread(
                target=self._search_transactions_thread,
                args=(wallet_addr, token_addr, amount),
                daemon=True
            )
            self.search_thread.start()
            
            self.search_start_btn.setEnabled(False)
            self.search_stop_btn.setEnabled(True)
            self.search_status_label.setText("Поиск запущен...")
            
        except Exception as e:
            logger.error(f"Ошибка запуска поиска: {e}")
    
    def _stop_paginated_search(self):
        """Остановка поиска транзакций"""
        self.stop_search_event.set()
        self.search_start_btn.setEnabled(True)
        self.search_stop_btn.setEnabled(False)
        self.search_status_label.setText("Поиск остановлен")
    
    def _search_transactions_thread(self, wallet_addr, token_addr, amount):
        """Поток поиска транзакций"""
        try:
            # Используем функцию поиска из модуля
            transactions, _, _ = search_transactions_paginated(
                wallet_address=wallet_addr,
                token_contract=token_addr if token_addr else None,
                exact_amount=float(amount) if amount else None,
                page_size=100,
                max_pages=10,
                delay_seconds=1,
                stop_flag=self.stop_search_event
            )
            
            # Обновляем UI с результатами
            self.update_search_results.emit(transactions)
            
        except Exception as e:
            logger.error(f"Ошибка поиска транзакций: {e}")
    
    def _send_rewards(self):
        """Отправка наград из таблицы"""
        try:
            if self.table_rewards.rowCount() == 0:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', 'Таблица наград пуста')
                return
            
            # Запрашиваем подтверждение
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение',
                f'Отправить награды для {self.table_rewards.rowCount()} адресов?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No
            )
            
            if reply != QtWidgets.QMessageBox.Yes:
                return
            
            # Собираем данные из таблицы
            rewards_data = []
            for row in range(self.table_rewards.rowCount()):
                address = self.table_rewards.item(row, 0).text()
                plex_amount = float(self.table_rewards.item(row, 1).text() or 0)
                usdt_amount = float(self.table_rewards.item(row, 2).text() or 0)
                
                if plex_amount > 0 or usdt_amount > 0:
                    rewards_data.append((address, plex_amount, usdt_amount))
            
            # Запускаем отправку в отдельном потоке
            threading.Thread(
                target=self._send_rewards_thread,
                args=(rewards_data,),
                daemon=True
            ).start()
            
        except Exception as e:
            logger.error(f"Ошибка отправки наград: {e}")
    
    def _send_rewards_thread(self, rewards_data):
        """Поток отправки наград"""
        try:
            total_sent = 0
            total_errors = 0
            
            for address, plex_amount, usdt_amount in rewards_data:
                try:
                    # Отправляем PLEX
                    if plex_amount > 0:
                        self._send_token(address, plex_amount, 'plex')
                        total_sent += 1
                    
                    # Отправляем USDT
                    if usdt_amount > 0:
                        self._send_token(address, usdt_amount, 'usdt')
                        total_sent += 1
                    
                    time.sleep(2)  # Пауза между отправками
                    
                except Exception as e:
                    total_errors += 1
                    logger.error(f"Ошибка отправки награды на {address}: {e}")
            
            # Обновляем UI
            QtWidgets.QMessageBox.information(
                self, 'Завершено',
                f'Отправлено: {total_sent}, Ошибок: {total_errors}'
            )
            
        except Exception as e:
            logger.error(f"Ошибка в потоке отправки наград: {e}")
    
    def _clear_rewards(self):
        """Очистка таблицы наград"""
        self.table_rewards.setRowCount(0)
        logger.info("Таблица наград очищена")
    
    def _direct_send(self):
        """Прямая отправка токенов"""
        try:
            address = self.direct_address_edit.text().strip()
            amount = self.direct_amount_edit.value()
            
            if not address:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', 'Введите адрес получателя')
                return
            
            # Определяем тип токена
            if self.direct_plex_radio.isChecked():
                token_type = 'plex'
            elif self.direct_usdt_radio.isChecked():
                token_type = 'usdt'
            else:
                token_type = 'bnb'
            
            # Отправляем в отдельном потоке
            threading.Thread(
                target=self._direct_send_thread,
                args=(address, amount, token_type),
                daemon=True
            ).start()
            
        except Exception as e:
            logger.error(f"Ошибка прямой отправки: {e}")
    
    def _direct_send_thread(self, address, amount, token_type):
        """Поток прямой отправки"""
        try:
            self.direct_status_label.setText("Отправка...")
            
            if token_type == 'bnb':
                result = self._send_bnb(address, amount)
            else:
                result = self._send_token(address, amount, token_type)
            
            if result:
                self.direct_status_label.setText("Отправлено успешно!")
                logger.info(f"Прямая отправка успешна: {amount} {token_type} на {address}")
            else:
                self.direct_status_label.setText("Ошибка отправки")
                
        except Exception as e:
            self.direct_status_label.setText("Ошибка отправки")
            logger.error(f"Ошибка прямой отправки: {e}")
    
    def _open_url(self, url):
        """Открытие URL в браузере"""
        import webbrowser
        webbrowser.open(url)
        logger.info(f"Открыт URL: {url}")
    
    def _load_history(self):
        """Загрузка истории транзакций"""
        try:
            # Получаем историю из БД
            history = get_history()
            
            # Обновляем таблицу
            self.history_table.setRowCount(0)
            
            for tx_type, address, amount, tx_hash, timestamp in history:
                row = self.history_table.rowCount()
                self.history_table.insertRow(row)
                
                self.history_table.setItem(row, 0, QtWidgets.QTableWidgetItem(tx_type))
                self.history_table.setItem(row, 1, QtWidgets.QTableWidgetItem(address))
                self.history_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(amount)))
                self.history_table.setItem(row, 3, QtWidgets.QTableWidgetItem(tx_hash))
                self.history_table.setItem(row, 4, QtWidgets.QTableWidgetItem(timestamp))
            
            logger.info(f"Загружено {len(history)} записей истории")
            
        except Exception as e:
            logger.error(f"Ошибка загрузки истории: {e}")
    
    def _export_history(self):
        """Экспорт истории в CSV"""
        try:
            path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, 'Экспорт истории', '', 'CSV Files (*.csv)'
            )
            
            if path:
                history = get_history()
                with open(path, 'w', encoding='utf-8') as f:
                    f.write('Тип,Адрес,Сумма,Хэш,Время\n')
                    for tx_type, address, amount, tx_hash, timestamp in history:
                        f.write(f'{tx_type},{address},{amount},{tx_hash},{timestamp}\n')
                
                logger.info(f"История экспортирована в {path}")
                
        except Exception as e:
            logger.error(f"Ошибка экспорта истории: {e}")
    
    def _add_found_to_rewards(self):
        """Добавление найденных транзакций в награды"""
        try:
            selected_rows = set(item.row() for item in self.found_tx_table.selectedItems())
            
            if not selected_rows:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', 'Выберите транзакции')
                return
            
            for row in selected_rows:
                sender = self.found_tx_table.item(row, 1).text()
                self._add_to_rewards_tab(sender)
            
            logger.info(f"Добавлено {len(selected_rows)} транзакций в награды")
            
        except Exception as e:
            logger.error(f"Ошибка добавления в награды: {e}")
    
    def _clear_found_tx(self):
        """Очистка таблицы найденных транзакций"""
        self.found_tx_table.setRowCount(0)
        logger.info("Таблица найденных транзакций очищена")
    
    def _update_gas_price(self, value):
        """Обновление цены газа"""
        self.cfg.set_gas_price(value)
        logger.info(f"Цена газа установлена: {value} Gwei")
    
    def _update_repeat_count(self, value):
        """Обновление количества повторений"""
        self.cfg.set_repeat_count(value)
        logger.info(f"Количество повторений установлено: {value}")
    
    def _clear_cache(self):
        """Очистка кэша"""
        try:
            count = cache.clear()
            logger.info(f"Очищено {count} элементов кэша")
            QtWidgets.QMessageBox.information(self, 'Кэш очищен', f'Очищено {count} элементов')
        except Exception as e:
            logger.error(f"Ошибка очистки кэша: {e}")
    
    def _test_blockchain_connection(self):
        """Тест подключения к блокчейну"""
        try:
            if not blockchain_enabled:
                QtWidgets.QMessageBox.warning(self, 'Блокчейн отключен', 'Функции блокчейна недоступны')
                return
            
            w3 = self.rpc.web3()
            if w3.is_connected():
                block = w3.eth.block_number
                QtWidgets.QMessageBox.information(
                    self, 'Подключение успешно',
                    f'Подключение к BSC установлено\nТекущий блок: {block}'
                )
            else:
                QtWidgets.QMessageBox.warning(self, 'Ошибка подключения', 'Не удалось подключиться к BSC')
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Ошибка', f'Ошибка тестирования: {e}')
    
    def _add_to_rewards_tab(self, address):
        """Добавление адреса в таблицу наград"""
        try:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес
            self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(address))
            
            # PLEX (по умолчанию)
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT (пусто)
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
            
        except Exception as e:
            logger.error(f"Ошибка добавления в награды: {e}")


# Точка входа в приложение
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('fusion')
    
    # Настройка тёмной темы
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    
    # Создание и отображение главного окна
    window = MainWindow()
    window.show()
    
    # Запуск приложения
                 sys.exit(app.exec_())
        self.plex_min.setDecimals(2)
        self.plex_min.setValue(2.0)
        plex_layout.addWidget(self.plex_min)
        
        plex_layout.addWidget(QtWidgets.QLabel('-'))
        
        self.plex_max = QtWidgets.QDoubleSpinBox()
        self.plex_max.setRange(0, 10000)
        self.plex_max.setDecimals(2)
        self.plex_max.setValue(4.0)
        plex_layout.addWidget(self.plex_max)
        
        self.btn_random_plex = QtWidgets.QPushButton('Случайные PLEX ONE')
        self.btn_random_plex.clicked.connect(self._randomize_plex)
        plex_layout.addWidget(self.btn_random_plex)
        
        random_layout.addLayout(plex_layout)
        
        # Настройки для USDT
        usdt_layout = QtWidgets.QHBoxLayout()
        usdt_layout.addWidget(QtWidgets.QLabel('USDT диапазон:'))
        
        self.usdt_min = QtWidgets.QDoubleSpinBox()
        self.usdt_min.setRange(0, 10000)
        self.usdt_min.setDecimals(2)
        self.usdt_min.setValue(0.1)
        usdt_layout.addWidget(self.usdt_min)
        
        usdt_layout.addWidget(QtWidgets.QLabel('-'))
        
        self.usdt_max = QtWidgets.QDoubleSpinBox()
        self.usdt_max.setRange(0, 10000)
        self.usdt_max.setDecimals(2)
        self.usdt_max.setValue(0.5)
        usdt_layout.addWidget(self.usdt_max)
        
        self.btn_random_usdt = QtWidgets.QPushButton('Случайные USDT')
        self.btn_random_usdt.clicked.connect(self._randomize_usdt)
        usdt_layout.addWidget(self.btn_random_usdt)
        
        random_layout.addLayout(usdt_layout)
        
        layout.addWidget(random_group)
        
        # Группа для сохранения и загрузки конфигураций
        config_group = QtWidgets.QGroupBox("Сохранение и загрузка")
        config_layout = QtWidgets.QHBoxLayout(config_group)
        
        self.reward_config_name = QtWidgets.QLineEdit()
        self.reward_config_name.setPlaceholderText('Имя конфигурации')
        config_layout.addWidget(self.reward_config_name)
        
        self.btn_save_config = QtWidgets.QPushButton('Сохранить')
        self.btn_save_config.clicked.connect(self._save_current_config)
        config_layout.addWidget(self.btn_save_config)
        
        self.configs_combo = QtWidgets.QComboBox()
        self._load_config_list()
        config_layout.addWidget(self.configs_combo)
        
        self.btn_load_config = QtWidgets.QPushButton('Загрузить')
        self.btn_load_config.clicked.connect(self._load_selected_config)
        config_layout.addWidget(self.btn_load_config)
        
        layout.addWidget(config_group)
        
        # Настройки отправки
        send_group = QtWidgets.QGroupBox("Отправка наград")
        send_layout = QtWidgets.QVBoxLayout(send_group)
        
        # Настройки отправки
        send_group = QtWidgets.QGroupBox("Отправка наград")
        send_layout = QtWidgets.QVBoxLayout(send_group)
        
        # Параметры отправки
        params_layout = QtWidgets.QHBoxLayout()
        
        # Интервал между отправками
        self.spin_interval = QtWidgets.QSpinBox()
        self.spin_interval.setRange(1, 600)
        self.spin_interval.setValue(5)  # 5 секунд по умолчанию
        params_layout.addWidget(QtWidgets.QLabel("Интервал (сек):"))
        params_layout.addWidget(self.spin_interval)
        
        # Повторы
        self.btn_repeats = QtWidgets.QPushButton(f"Повторов: {self.cfg.get_repeat_count()}")
        self.btn_repeats.clicked.connect(self._set_repeats)
        params_layout.addWidget(self.btn_repeats)
        
        send_layout.addLayout(params_layout)
        
        # Кнопки отправки
        buttons_send_layout = QtWidgets.QHBoxLayout()
        
        self.btn_start = QtWidgets.QPushButton('Начать отправку')
        self.btn_start.clicked.connect(lambda: self._start_sending())
        buttons_send_layout.addWidget(self.btn_start)
        
        self.stop_btn = QtWidgets.QPushButton('Остановить')
        self.stop_btn.clicked.connect(self._cancel_sending)
        self.stop_btn.setEnabled(False)
        buttons_send_layout.addWidget(self.stop_btn)
        
        send_layout.addLayout(buttons_send_layout)
        
        # Индикатор прогресса
        self.progress_send = QtWidgets.QProgressBar()
        send_layout.addWidget(self.progress_send)
        
        layout.addWidget(send_group)
        
        # Загружаем сохраненные награды, если они есть
        self._load_rewards_from_db()
        
        return w
    
    def _add_rewards_row(self):
        """Добавление новой строки в таблицу наград"""
        row = self.table_rewards.rowCount()
        self.table_rewards.insertRow(row)
        
        # Адрес
        addr_item = QtWidgets.QTableWidgetItem("")
        self.table_rewards.setItem(row, 0, addr_item)
        
        # PLEX - значение по умолчанию
        plex_item = QtWidgets.QTableWidgetItem("3")
        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 1, plex_item)
        
        # USDT - пустое значение по умолчанию
        usdt_item = QtWidgets.QTableWidgetItem("")
        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 2, usdt_item)
    
    def _remove_rewards_row(self):
        """Удаление выбранных строк из таблицы наград"""
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        for row in sorted(selected_rows, reverse=True):
            self.table_rewards.removeRow(row)
        
        logger.info(f"Удалено {len(selected_rows)} строк из таблицы наград")
    def _clear_rewards(self):
        """Очистка всей таблицы наград"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Очистить все данные в таблице наград?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._save_rewards_state()  # Сохраняем текущее состояние для возможности отмены
            self.table_rewards.setRowCount(0)
            logger.info("Таблица наград очищена")
    
    def _randomize_plex(self):
        """Генерация случайных значений PLEX для всех строк таблицы с использованием настраиваемого диапазона"""
        plex_min = self.plex_min.value()
        plex_max = self.plex_max.value()
        
        # Проверка корректности диапазона
        if plex_min >= plex_max:
            logger.error("Минимальное значение PLEX ONE должно быть меньше максимального")
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка', 
                'Минимальное значение PLEX ONE должно быть меньше максимального',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Получаем выбранные строки или все строки, если ничего не выбрано
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        if not selected_rows:
            selected_rows = range(self.table_rewards.rowCount())
        
        for row in selected_rows:
            # Генерируем случайное значение в заданном диапазоне с двумя знаками после запятой
            value = round(random.uniform(plex_min, plex_max), 2)
            self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(str(value)))
        
        logger.info(f"Сгенерированы случайные значения PLEX ONE в диапазоне {plex_min}-{plex_max} для {len(selected_rows)} строк")
    
    def _randomize_usdt(self):
        """Генерация случайных значений USDT для строк таблицы с заданными вероятностями"""
        # Значения и их вероятности
        values = [0.12, 0.18, 0.20, 0.24, 0.38]
        weights = [0.70, 0.10, 0.15, 0.025, 0.025]  # Суммарно 5% на 0.24 и 0.38
        
        # Получаем выбранные строки или все строки, если ничего не выбрано
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        if not selected_rows:
            selected_rows = range(self.table_rewards.rowCount())
        
        for row in selected_rows:
            # Генерируем случайное значение с заданными вероятностями
            value = random.choices(values, weights=weights, k=1)[0]
            self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(str(value)))
        
        logger.info(f"Сгенерированы случайные значения USDT с заданными вероятностями для {len(selected_rows)} строк")
    
    def _save_current_config(self):
        """Сохранение текущей конфигурации наград в файл"""
        name = self.reward_config_name.text().strip()
        if not name:
            return logger.error("Введите имя для сохранения конфигурации")
        
        addresses = []
        plex_amounts = {}
        usdt_amounts = {}
        
        for row in range(self.table_rewards.rowCount()):
            # Получаем адрес и проверяем его наличие
            addr_item = self.table_rewards.item(row, 0)
            if not addr_item or not addr_item.text().strip():
                continue
            
            address = addr_item.text().strip()
            addresses.append(address)
            
            # Получаем значение PLEX ONE
            plex_item = self.table_rewards.item(row, 1)
            if plex_item and plex_item.text().strip():
                try:
                    plex_amounts[row] = float(plex_item.text())
                except ValueError:
                    pass
            
            # Получаем значение USDT
            usdt_item = self.table_rewards.item(row, 2)
            if usdt_item and usdt_item.text().strip():
                try:
                    usdt_amounts[row] = float(usdt_item.text())
                except ValueError:
                    pass
        
        if not addresses:
            return logger.error("Нет адресов для сохранения")
        
        # Сохраняем конфигурацию
        if save_rewards_config(name, addresses, plex_amounts, usdt_amounts):
            logger.info(f"Конфигурация '{name}' успешно сохранена")
            self._load_config_list()
            self.reward_config_name.clear()
        else:
            logger.error(f"Ошибка при сохранении конфигурации '{name}'")
    
    def _load_config_list(self):
        """Загрузка списка доступных конфигураций в комбобокс"""
        self.configs_combo.clear()
        configs = get_rewards_configs()
        self.configs_combo.addItems(configs)
        
        if configs:
            logger.debug(f"Загружено {len(configs)} конфигураций наград")
    
    def _load_selected_config(self):
        """Загрузка выбранной конфигурации наград из файла"""
        config_name = self.configs_combo.currentText()
        if not config_name:
            return logger.warning("Выберите конфигурацию для загрузки")
        
        config = load_rewards_config(config_name)
        if not config:
            return logger.error(f"Не удалось загрузить конфигурацию '{config_name}'")
        
        # Сохраняем текущее состояние для возможности отмены
        self._save_rewards_state()
        
        # Очищаем таблицу
        self.table_rewards.setRowCount(0)
        
        # Заполняем таблицу данными из конфигурации
        for item in config:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес
            self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(item['address']))
            
            # PLEX
            plex_value = str(item.get('plex', '')) if item.get('plex', 0) > 0 else ''
            self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(plex_value))
            
            # USDT
            usdt_value = str(item.get('usdt', '')) if item.get('usdt', 0) > 0 else ''
            self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(usdt_value))
        
        logger.info(f"Конфигурация '{config_name}' загружена: {len(config)} записей")
    
    def _set_repeats(self):
        """Установка количества повторений серии отправок"""
        value, ok = QtWidgets.QInputDialog.getInt(
            self, 'Повторения', 'Количество повторений серии:',
            self.cfg.get_repeat_count(), 1, 100, 1
        )
        
        if ok:
            self.cfg.set_repeat_count(value)
            self.btn_repeats.setText(f"Повторов: {value}")
            logger.info(f"Установлено {value} повторений серии отправок")
    
    def _save_rewards_state(self):
        """Сохранение текущего состояния таблицы наград для возможности отмены"""
        state = []
        for row in range(self.table_rewards.rowCount()):
            row_data = {}
            
            # Адрес
            addr_item = self.table_rewards.item(row, 0)
            if addr_item:
                row_data['address'] = addr_item.text()
                # Если есть хэш транзакции, сохраняем его тоже
                tx_hash = addr_item.data(QtCore.Qt.UserRole)
                if tx_hash:
                    row_data['tx_hash'] = tx_hash
            
            # PLEX
            plex_item = self.table_rewards.item(row, 1)
            if plex_item:
                row_data['plex'] = plex_item.text()
            
            # USDT
            usdt_item = self.table_rewards.item(row, 2)
            if usdt_item:
                row_data['usdt'] = usdt_item.text()
            
            state.append(row_data)
        
        self.last_rewards_state = state
    
    def _restore_rewards_state(self):
        """Восстановление предыдущего состояния таблицы наград"""
        if not self.last_rewards_state:
            return logger.warning("Нет сохраненного состояния для восстановления")
        
        # Очищаем таблицу
        self.table_rewards.setRowCount(0)
        
        # Восстанавливаем данные
        for row_data in self.last_rewards_state:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес
            addr_item = QtWidgets.QTableWidgetItem(row_data.get('address', ''))
            # Если был хэш транзакции, восстанавливаем его тоже
            if 'tx_hash' in row_data:
                addr_item.setData(QtCore.Qt.UserRole, row_data['tx_hash'])
            self.table_rewards.setItem(row, 0, addr_item)
            
            # PLEX
            plex_item = QtWidgets.QTableWidgetItem(row_data.get('plex', ''))
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT
            usdt_item = QtWidgets.QTableWidgetItem(row_data.get('usdt', ''))
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Восстановлено предыдущее состояние: {len(self.last_rewards_state)} записей")
        
        # Очищаем сохраненное состояние
        self.last_rewards_state = None
    def _load_rewards_from_db(self):
        """Загрузка наград из базы данных в таблицу"""
        try:
            rewards = get_rewards()
            
            for reward in rewards:
                row = self.table_rewards.rowCount()
                self.table_rewards.insertRow(row)
                
                # Адрес
                addr_item = QtWidgets.QTableWidgetItem(reward['address'])
                if reward['tx_hash']:
                    addr_item.setData(QtCore.Qt.UserRole, reward['tx_hash'])
                self.table_rewards.setItem(row, 0, addr_item)
                
                # PLEX
                plex_value = str(reward['plex_amount']) if reward['plex_amount'] > 0 else ''
                plex_item = QtWidgets.QTableWidgetItem(plex_value)
                plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 1, plex_item)
                
                # USDT
                usdt_value = str(reward['usdt_amount']) if reward['usdt_amount'] > 0 else ''
                usdt_item = QtWidgets.QTableWidgetItem(usdt_value)
                usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 2, usdt_item)
                
                # Если награда уже отправлена, выделяем строку
                if reward['sent']:
                    for col in range(self.table_rewards.columnCount()):
                        item = self.table_rewards.item(row, col)
                        if item:
                            item.setBackground(QtGui.QColor('#004400'))
            
            logger.info(f"Загружено {len(rewards)} наград из базы данных")
        except Exception as e:
            logger.error(f"Ошибка при загрузке наград из базы данных: {e}")
            
    def _show_rewards_context_menu(self, position):
        """Отображение контекстного меню для таблицы наград"""
        menu = QtWidgets.QMenu()
        
        add_after = menu.addAction("Добавить строку после")
        remove_row = menu.addAction("Удалить строку")
        copy_addr = menu.addAction("Копировать адрес")
        open_bscscan = menu.addAction("Открыть в BscScan")
        
        # Для функции отмены
        undo_menu = menu.addMenu("Отмена/восстановление")
        restore_state = undo_menu.addAction("Восстановить предыдущее состояние")
        
        # Дополнительные опции
        generate_menu = menu.addMenu("Сгенерировать")
        gen_plex = generate_menu.addAction("Случайные PLEX для выбранных")
        gen_usdt = generate_menu.addAction("Случайные USDT для выбранных")
        
        action = menu.exec_(self.table_rewards.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected_rows = list(set(index.row() for index in self.table_rewards.selectedIndexes()))
        
        if not selected_rows and action != restore_state:
            return
            
        if action == add_after:
            # Добавляем строку после последней выбранной
            row = max(selected_rows) + 1
            self.table_rewards.insertRow(row)
            
            # Адрес - пустой
            self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(""))
            
            # PLEX - значение по умолчанию
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT - пустое значение по умолчанию
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
            
            logger.info(f"Добавлена новая строка после строки {row-1}")
        
        elif action == remove_row:
            # Удаляем строки в обратном порядке, чтобы не сдвигались индексы
            for row in sorted(selected_rows, reverse=True):
                self.table_rewards.removeRow(row)
            
            logger.info(f"Удалено {len(selected_rows)} строк")
        
        elif action == copy_addr:
            # Копируем адрес из первой выбранной строки
            row = selected_rows[0]
            addr_item = self.table_rewards.item(row, 0)
            if addr_item and addr_item.text():
                QtWidgets.QApplication.clipboard().setText(addr_item.text())
                logger.info(f"Адрес скопирован: {addr_item.text()}")
        
        elif action == open_bscscan:
            # Открываем адрес в BscScan
            row = selected_rows[0]
            addr_item = self.table_rewards.item(row, 0)
            if addr_item and addr_item.text():
                url = f"https://bscscan.com/address/{addr_item.text()}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открыт адрес в BscScan: {addr_item.text()}")
        
        elif action == restore_state:
            # Восстанавливаем предыдущее состояние таблицы
            self._restore_rewards_state()
        
        elif action == gen_plex:
            # Генерируем случайные значения PLEX для выбранных строк
            for row in selected_rows:
                value = round(random.uniform(2, 4), 2)
                self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(str(value)))
            
            logger.info(f"Сгенерированы случайные значения PLEX ONE для {len(selected_rows)} строк")
        
        elif action == gen_usdt:
            # Генерируем случайные значения USDT для выбранных строк
            for row in selected_rows:
                value = round(random.uniform(0.1, 0.5), 2)
                self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(str(value)))
            
            logger.info(f"Сгенерированы случайные значения USDT для {len(selected_rows)} строк")

    # --- Таб «Прямая отправка» ---
    def _tab_direct_send(self):
        """Создание вкладки для периодической прямой отправки токенов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Форма для параметров отправки
        form = QtWidgets.QFormLayout()
        
        # Адрес получателя
        self.direct_send_address = QtWidgets.QLineEdit()
        self.direct_send_address.setPlaceholderText('0x...')
        form.addRow('Адрес получателя:', self.direct_send_address)
        
        # Выбор типа токена
        self.direct_send_token_group = QtWidgets.QButtonGroup(w)
        self.direct_send_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.direct_send_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.direct_send_plex_radio.setChecked(True)
        self.direct_send_token_group.addButton(self.direct_send_plex_radio)
        self.direct_send_token_group.addButton(self.direct_send_usdt_radio)
        
        token_layout = QtWidgets.QHBoxLayout()
        token_layout.addWidget(self.direct_send_plex_radio)
        token_layout.addWidget(self.direct_send_usdt_radio)
        form.addRow('Тип токена:', token_layout)
        
        # Сумма отправки
        self.direct_send_amount = QtWidgets.QDoubleSpinBox()
        self.direct_send_amount.setRange(0.0001, 1000000)
        self.direct_send_amount.setDecimals(8)
        self.direct_send_amount.setValue(3)  # По умолчанию 3 токена
        form.addRow('Количество токенов:', self.direct_send_amount)
        
        # Интервал отправки
        self.direct_send_interval = QtWidgets.QSpinBox()
        self.direct_send_interval.setRange(1, 3600)
        self.direct_send_interval.setValue(60)  # По умолчанию 60 секунд
        form.addRow('Интервал (сек):', self.direct_send_interval)
        
        # Количество периодов
        self.direct_send_periods = QtWidgets.QSpinBox()
        self.direct_send_periods.setRange(0, 10000)
        self.direct_send_periods.setValue(10)  # По умолчанию 10 периодов
        self.direct_send_periods.setSpecialValueText("Бесконечно")  # Текст для значения 0
        form.addRow('Количество периодов (0 - бесконечно):', self.direct_send_periods)
        
        layout.addLayout(form)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.direct_send_start_btn = QtWidgets.QPushButton('Запустить отправку')
        self.direct_send_start_btn.clicked.connect(self._start_direct_send_ui)
        buttons_layout.addWidget(self.direct_send_start_btn)
        
        self.direct_send_stop_btn = QtWidgets.QPushButton('Остановить отправку')
        self.direct_send_stop_btn.clicked.connect(self._cancel_direct_sending)
        self.direct_send_stop_btn.setEnabled(False)
        buttons_layout.addWidget(self.direct_send_stop_btn)
        
        layout.addLayout(buttons_layout)
        
        # Индикатор прогресса
        self.direct_send_progress = QtWidgets.QProgressBar()
        layout.addWidget(self.direct_send_progress)
        
        # Статус отправки
        self.direct_send_status = QtWidgets.QTextEdit()
        self.direct_send_status.setReadOnly(True)
        self.direct_send_status.setMaximumHeight(150)
        layout.addWidget(self.direct_send_status)
        
        return w
    def _start_direct_send_ui(self):
        """Обработка запуска периодической отправки из интерфейса"""
        if getattr(self, 'direct_send_active', False):
            return logger.warning("Отправка уже запущена")
            
        address = self.direct_send_address.text().strip()
        token_type = 'plex' if self.direct_send_plex_radio.isChecked() else 'usdt'
        amount = self.direct_send_amount.value()
        interval = self.direct_send_interval.value()
        periods = self.direct_send_periods.value()
        
        # Проверка адреса
        if blockchain_enabled and not Web3.is_address(address):
            return logger.error("Некорректный адрес получателя")
            
        # Настройка интерфейса
        self.direct_send_start_btn.setEnabled(False)
        self.direct_send_stop_btn.setEnabled(True)
        
        # Запуск отправки
        self._start_direct_sending(address, token_type, amount, interval, periods)

    def _direct_send_periodic_send(self):
        """
        Метод для периодической отправки токенов по таймеру.
        Обрабатывает каждую итерацию периодической отправки.
        """
        if not self.direct_send_active:
            self.direct_send_timer.stop()
            return
            
        try:
            # Увеличиваем счетчик текущего периода
            self.direct_send_current_period += 1
            
            # Обновляем статус в интерфейсе
            status_text = f"Отправка #{self.direct_send_current_period}"
            if self.direct_send_total_periods > 0:
                status_text += f" из {self.direct_send_total_periods}"
            QtCore.QMetaObject.invokeMethod(
                self.direct_send_status, "append", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(str, status_text)
            )
            
            # Проверяем, не превысили ли общее количество периодов
            if (self.direct_send_total_periods > 0 and 
                self.direct_send_current_period > self.direct_send_total_periods):
                logger.info("Периодическая отправка завершена")
                self.direct_send_active = False
                self.direct_send_timer.stop()
                self.update_progress_signal.emit("direct_send", 100)
                
                # Обновляем интерфейс
                QtCore.QMetaObject.invokeMethod(
                    self.direct_send_start_btn, "setEnabled", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(bool, True)
                )
                QtCore.QMetaObject.invokeMethod(
                    self.direct_send_stop_btn, "setEnabled", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(bool, False)
                )
                return
                
            # Получаем параметры отправки
            params = self.direct_send_params
            to_addr = params.get('to_addr')
            token_type = params.get('token_type', 'plex')
            amount = params.get('amount', 0)
            
            if not to_addr or amount <= 0:
                logger.error("Некорректные параметры для периодической отправки")
                self.direct_send_active = False
                self.direct_send_timer.stop()
                return
                
            # Выполняем отправку токена
            logger.info(f"Периодическая отправка #{self.direct_send_current_period}: "
                      f"{amount} {token_type.upper()} на адрес {to_addr}")
            
            # Отправка PLEX или USDT в зависимости от типа токена через универсальный метод (адрес контракта)
            token_addr = PLEX_CONTRACT if token_type.lower() == 'plex' else USDT_CONTRACT
            self._send_token(to_addr, amount, token_addr)
                
            # Обновляем прогресс
            if self.direct_send_total_periods > 0:
                progress = int(self.direct_send_current_period / self.direct_send_total_periods * 100)
                self.update_progress_signal.emit("direct_send", progress)
            else:
                # Если общее количество не задано, показываем прогресс итераций
                self.update_progress_signal.emit("direct_send", self.direct_send_current_period % 100)
                
        except Exception as e:
            logger.error(f"Ошибка при периодической отправке: {e}")
            self.direct_send_active = False
            self.direct_send_timer.stop()

    def _start_direct_sending(self, to_addr, token_type, amount, interval, periods=0):
        """
        Запуск периодической отправки токенов.
        
        Args:
            to_addr (str): Адрес получателя
            token_type (str): Тип токена ('plex' или 'usdt')
            amount (float): Количество токенов для отправки
            interval (int): Интервал между отправками в секундах
            periods (int, optional): Общее количество периодов отправки, 0 для бесконечной отправки
        """
        if not self.pk:
            return logger.error("Приватный ключ не задан")
            
        if not to_addr or amount <= 0 or interval <= 0:
            return logger.error("Некорректные параметры для периодической отправки")
            
        if self.direct_send_active:
            # Останавливаем текущую периодическую отправку
            self.direct_send_active = False
            self.direct_send_timer.stop()
            
        # Настраиваем параметры новой периодической отправки
        self.direct_send_params = {
            'to_addr': to_addr,
            'token_type': token_type,
            'amount': amount
        }
        self.direct_send_current_period = 0
        self.direct_send_total_periods = periods
        self.direct_send_active = True
        
        # Очищаем область статуса
        self.direct_send_status.clear()
        
        # Настраиваем и запускаем таймер
        self.direct_send_timer.setInterval(interval * 1000)  # Переводим секунды в миллисекунды
        self.direct_send_timer.start()
        
        status_msg = f"Запущена периодическая отправка {amount} {token_type.upper()} " \
                     f"на {to_addr} с интервалом {interval} сек " \
                     f"({periods} периодов)" if periods > 0 else "(без ограничения)"
        
        logger.info(status_msg)
        self.direct_send_status.append(status_msg)

    def _cancel_direct_sending(self):
        """Остановка периодической отправки токенов"""
        if self.direct_send_active:
            self.direct_send_active = False
            self.direct_send_timer.stop()
            logger.info("Периодическая отправка остановлена")
            self.update_progress_signal.emit("direct_send", 100)
            
            # Обновляем интерфейс
            self.direct_send_start_btn.setEnabled(True)
            self.direct_send_stop_btn.setEnabled(False)
            self.direct_send_status.append("Отправка остановлена")

    def _collect_tasks(self):
        """Сбор задач для отправки из таблицы"""
        tasks = []
        for i in range(self.table_rewards.rowCount()):
            addr = self.table_rewards.item(i, 0)
            if not addr or not addr.text():
                continue
                
            plex_item = self.table_rewards.item(i, 1)
            usdt_item = self.table_rewards.item(i, 2)
            
            # Проверка на наличие значений и конвертация в числа
            plex_amount = 0
            if plex_item and plex_item.text():
                try:
                    plex_amount = float(plex_item.text())
                except ValueError:
                    pass
                    
            usdt_amount = 0
            if usdt_item and usdt_item.text():
                try:
                    usdt_amount = float(usdt_item.text())
                except ValueError:
                    pass
                    
            # Добавляем задачу, если есть хотя бы одна ненулевая сумма
            if plex_amount > 0 or usdt_amount > 0:
                tasks.append((addr, plex_amount, usdt_amount))
                
        return tasks
        
    def _start_sending(self, mark_rewarded=False):
        """Запуск процесса отправки наград"""
        if self.sending:
            return logger.warning("Отправка уже запущена")
            
        if not self.pk:
            return logger.error("Приватный ключ не задан")
            
        tasks = self._collect_tasks()
        if not tasks:
            return logger.error("Нет задач для отправки")
            
        # Подтверждение отправки
        total_plex = sum(plex for _, plex, _ in tasks)
        total_usdt = sum(usdt for _, _, usdt in tasks)
        
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение отправки', 
            f'Будет отправлено {len(tasks)} адресам:\n\n'
            f'Всего PLEX: {total_plex}\n'
            f'Всего USDT: {total_usdt}\n\n'
            f'Продолжить?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
            
        # Запуск отправки в отдельном потоке
        self.sending = True
        self.btn_start.setEnabled(False)
        self.stop_btn.setEnabled(True)
        
        self.repeats_remain = self.cfg.get_repeat_count()
        threading.Thread(
            target=self._sending_thread,
            args=(tasks, self.spin_interval.value(), mark_rewarded),
            daemon=True
        ).start()
    
    def _sending_thread(self, tasks, interval, mark_rewarded=False):
        """Поток для отправки токенов с улучшенным отслеживанием"""
        try:
            total = len(tasks)
            batch_id = datetime.now().strftime('%Y%m%d%H%M%S')
            transaction_results = []
            
            # Подготавливаем и отображаем очередь
            queue_items = []
            for _, (addr, plex, usdt) in enumerate(tasks):
                if hasattr(addr, 'text'):
                    to_addr = addr.text()
                    tx_hash = addr.data(QtCore.Qt.UserRole)
                else:
                    to_addr = addr
                    tx_hash = None
                    
                if plex > 0:
                    queue_items.append({
                        'status': 'pending',
                        'address': to_addr,
                        'token': 'PLEX',
                        'amount': plex,
                        'time': None,
                        'tx_hash': None,
                        'original_tx': tx_hash  # Исходная транзакция для награды
                    })
                    
                if usdt > 0:
                    queue_items.append({
                        'status': 'pending',
                        'address': to_addr,
                        'token': 'USDT',
                        'amount': usdt,
                        'time': None,
                        'tx_hash': None,
                        'original_tx': tx_hash  # Исходная транзакция для награды
                    })
                    
            # Создаем копию списка для передачи через сигнал
            queue_items_copy = queue_items.copy() if queue_items else []
            # Запускаем обновление UI через основной поток
            self.update_status_signal.emit("Обновление очереди")
            QtWidgets.QApplication.processEvents()
            self._update_queue_table(queue_items_copy) # Прямой вызов декорированного метода
            
            # Активируем кнопки управления очередью
            QtCore.QMetaObject.invokeMethod(
                self.pause_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.cancel_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            
            # Повторяем серию указанное количество раз
            while self.repeats_remain > 0 and self.sending:
                logger.info(f"Начинаем серию отправок (осталось повторений: {self.repeats_remain})")
                QtCore.QMetaObject.invokeMethod(
                    self.queue_status_label, "setText", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(str, f"Отправка (серия {self.cfg.get_repeat_count() - self.repeats_remain + 1}/{self.cfg.get_repeat_count()})")
                )
                
                # Обрабатываем каждый элемент в очереди
                completed_count = 0
                for i, item in enumerate(queue_items):
                    if not self.sending or self.queue_paused:
                        # Если отправка приостановлена, ждем возобновления
                        while self.queue_paused and self.sending:
                            time.sleep(0.5)
                        
                        if not self.sending:
                            break
                    
                    # Пропускаем уже обработанные элементы
                    if item['status'] != 'pending':
                        continue
                    
                    # Обновляем прогресс
                    completed_count += 1
                    progress = int((completed_count / len(queue_items)) * 100)
                    self.update_progress_signal.emit("send", progress)
                    QtCore.QMetaObject.invokeMethod(
                        self.queue_progress_label, "setText", 
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(str, f"{completed_count}/{len(queue_items)} отправлено")
                    )
                    
                    # Обновляем статус элемента
                    item['status'] = 'processing'
                    self.queue_item_update_signal.emit(i, item)
                    
                    # Отправка токена
                    logger.info(f"Отправка {item['amount']} {item['token']} на адрес {item['address']}")
                    tx_hash = self._send_token(item['address'], item['amount'], item['token'].lower())
                    
                    # Обновляем результат
                    item['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    item['tx_hash'] = tx_hash
                    
                    if tx_hash:
                        item['status'] = 'success'
                        transaction_results.append({
                            'address': item['address'],
                            'token': item['token'],
                            'amount': item['amount'],
                            'tx_hash': tx_hash,
                            'status': 'success',
                            'time': item['time']
                        })
                    else:
                        item['status'] = 'failed'
                    
                    # Если нужно отметить транзакцию как награжденную и есть хэш исходной транзакции
                    if mark_rewarded and item['original_tx'] and item['status'] == 'success':
                        mark_transaction_rewarded(item['original_tx'])
                        logger.info(f"Транзакция {item['original_tx']} отмечена как награжденная")
                    
                    # Обновляем элемент в таблице очереди
                    QtCore.QMetaObject.invokeMethod(
                        self, "_update_queue_item", 
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(int, i),
                        QtCore.Q_ARG(dict, item)
                    )
                    
                    # Добавляем в таблицу завершенных
                    self.completed_item_signal.emit(item)
                    
                    # Задержка между отправками
                    if i < len(queue_items) - 1 and self.sending and not self.queue_paused:
                        time.sleep(interval)
                
                # Уменьшаем счетчик оставшихся повторений
                self.repeats_remain -= 1
                
                # Задержка между сериями
                if self.repeats_remain > 0 and self.sending:
                    logger.info(f"Серия завершена. Ожидание перед следующей серией (осталось: {self.repeats_remain})")
                    time.sleep(interval * 2)  # Увеличенный интервал между сериями
            
            # Завершение отправки
            logger.info("Отправка завершена")
            self.update_progress_signal.emit("send", 100)
            
            # Показываем диалог с результатами
            if transaction_results:
                QtCore.QMetaObject.invokeMethod(
                    self, "_show_transaction_results", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(str, batch_id),
                    QtCore.Q_ARG(list, transaction_results)
                )
            
        except Exception as e:
            logger.error(f"Ошибка при отправке: {e}")
        finally:
            self.sending = False
            self.queue_paused = False
            
            # Обновляем UI в основном потоке
            QtCore.QMetaObject.invokeMethod(
                self.btn_start, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.stop_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.pause_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.resume_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.cancel_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.queue_status_label, "setText", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(str, "Завершено")
            )
    
    def _cancel_sending(self):
        """Остановка процесса отправки"""
        if self.sending:
            self.sending = False
            logger.info("Запрошена остановка отправки")
            self.stop_btn.setEnabled(False)

    # --- Таб «Поиск транзакций» ---
    def _tab_paginated_search(self):
        """Создание вкладки для постраничного поиска транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Форма для ввода параметров поиска
        form = QtWidgets.QFormLayout()
        
        # Адрес кошелька системы
        self.system_address = QtWidgets.QLineEdit()
        self.system_address.setPlaceholderText('0x...')
        form.addRow('Адрес системы (кошелька):', self.system_address)
        
        # Адрес токена
        self.token_address = QtWidgets.QLineEdit()
        self.token_address.setPlaceholderText('0x...')
        self.token_address.setText(PLEX_CONTRACT)
        form.addRow('Адрес токена (контракт):', self.token_address)
        
        # Переключатель режима поиска
        radio_layout = QtWidgets.QHBoxLayout()
        self.search_mode_group = QtWidgets.QButtonGroup(w)
        self.search_exact_radio = QtWidgets.QRadioButton('Точная сумма')
        self.search_range_radio = QtWidgets.QRadioButton('Диапазон сумм')
        self.search_exact_radio.setChecked(True)
        self.search_mode_group.addButton(self.search_exact_radio)
        self.search_mode_group.addButton(self.search_range_radio)
        radio_layout.addWidget(self.search_exact_radio)
        radio_layout.addWidget(self.search_range_radio)
        form.addRow('Режим поиска:', radio_layout)
        
        # Точная сумма
        self.exact_amount = QtWidgets.QDoubleSpinBox()
        self.exact_amount.setRange(0.0, 1000000.0)
        self.exact_amount.setDecimals(8)
        self.exact_amount.setValue(30.0)  # Предустановленное значение 30 PLEX ONE
        form.addRow('Точная сумма:', self.exact_amount)
        
        # Диапазон сумм
        range_layout = QtWidgets.QHBoxLayout()
        self.min_amount = QtWidgets.QDoubleSpinBox()
        self.min_amount.setRange(0.0, 1000000.0)
        self.min_amount.setDecimals(8)
        self.min_amount.setEnabled(False)
        
        self.max_amount = QtWidgets.QDoubleSpinBox()
        self.max_amount.setRange(0.0, 1000000.0)
        self.max_amount.setDecimals(8)
        self.max_amount.setEnabled(False)
        
        range_layout.addWidget(QtWidgets.QLabel('От:'))
        range_layout.addWidget(self.min_amount)
        range_layout.addWidget(QtWidgets.QLabel('До:'))
        range_layout.addWidget(self.max_amount)
        form.addRow('Диапазон:', range_layout)
        
        # Минимальное количество транзакций
        self.min_tx_count = QtWidgets.QSpinBox()
        self.min_tx_count.setRange(1, 1000)
        self.min_tx_count.setValue(1)
        form.addRow('Мин. кол-во транзакций:', self.min_tx_count)
        # Настройки пагинации
        pagination_layout = QtWidgets.QHBoxLayout()
        
        self.page_size = QtWidgets.QSpinBox()
        self.page_size.setRange(10, 10000)
        self.page_size.setValue(1000)
        pagination_layout.addWidget(QtWidgets.QLabel('Размер страницы:'))
        pagination_layout.addWidget(self.page_size)
        
        self.max_pages = QtWidgets.QSpinBox()
        self.max_pages.setRange(1, 1000)
        self.max_pages.setValue(100)
        pagination_layout.addWidget(QtWidgets.QLabel('Макс. страниц:'))
        pagination_layout.addWidget(self.max_pages)
        
        self.api_delay = QtWidgets.QSpinBox()
        self.api_delay.setRange(1, 60)
        self.api_delay.setValue(1)
        pagination_layout.addWidget(QtWidgets.QLabel('Задержка (сек):'))
        pagination_layout.addWidget(self.api_delay)
        
        form.addRow('Настройки API:', pagination_layout)
        
        # Опция для учета отдельных транзакций
        self.track_tx_checkbox = QtWidgets.QCheckBox("Учитывать отдельные транзакции для награждения")
        self.track_tx_checkbox.setChecked(True)
        form.addRow("", self.track_tx_checkbox)
        
        # Группа фильтрации по датам
        date_filter_group = QtWidgets.QGroupBox("Фильтрация по датам")
        date_filter_layout = QtWidgets.QVBoxLayout(date_filter_group)
        
        # Переключатель режима фильтрации по датам
        date_radio_layout = QtWidgets.QHBoxLayout()
        self.date_filter_group = QtWidgets.QButtonGroup(w)
        self.date_all_time_radio = QtWidgets.QRadioButton('За все время')
        self.date_range_radio = QtWidgets.QRadioButton('Диапазон дат')
        self.date_all_time_radio.setChecked(True)
        self.date_filter_group.addButton(self.date_all_time_radio)
        self.date_filter_group.addButton(self.date_range_radio)
        date_radio_layout.addWidget(self.date_all_time_radio)
        date_radio_layout.addWidget(self.date_range_radio)
        date_filter_layout.addLayout(date_radio_layout)
        
        # Выбор диапазона дат
        date_range_layout = QtWidgets.QHBoxLayout()
        date_range_layout.addWidget(QtWidgets.QLabel('От:'))
        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate.currentDate())
        self.date_from.setEnabled(False)
        date_range_layout.addWidget(self.date_from)
        
        date_range_layout.addWidget(QtWidgets.QLabel('До:'))
        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate.currentDate())
        self.date_to.setEnabled(False)
        date_range_layout.addWidget(self.date_to)
        
        # Кнопка "Сегодня"
        self.today_btn = QtWidgets.QPushButton('Сегодня')
        self.today_btn.clicked.connect(self._set_date_to_today)
        self.today_btn.setEnabled(False)
        date_range_layout.addWidget(self.today_btn)
        
        date_filter_layout.addLayout(date_range_layout)
        
        # Подключаем переключение режима фильтрации по датам
        self.date_all_time_radio.toggled.connect(self._toggle_date_filter_mode)
        self.date_range_radio.toggled.connect(self._toggle_date_filter_mode)
        
        layout.addWidget(date_filter_group)
        layout.addLayout(form)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.start_search_btn = QtWidgets.QPushButton('Начать поиск')
        self.start_search_btn.clicked.connect(self._start_paginated_search)
        buttons_layout.addWidget(self.start_search_btn)
        
        self.stop_search_btn = QtWidgets.QPushButton('Остановить')
        self.stop_search_btn.clicked.connect(self._stop_paginated_search)
        self.stop_search_btn.setEnabled(False)
        buttons_layout.addWidget(self.stop_search_btn)
        
        self.clear_results_btn = QtWidgets.QPushButton('Очистить результаты')
        self.clear_results_btn.clicked.connect(self._clear_search_results)
        buttons_layout.addWidget(self.clear_results_btn)
        
        layout.addLayout(buttons_layout)
        
        # Таблица результатов
        self.search_results_table = QtWidgets.QTableWidget(0, 3)
        self.search_results_table.setHorizontalHeaderLabels(['Отправитель', 'Количество TX', 'Последняя TX'])
        header = self.search_results_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(STRETCH_MODE)
        self.search_results_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.search_results_table.setSelectionMode(QtWidgets.QTableView.ExtendedSelection)
        self.search_results_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.search_results_table.customContextMenuRequested.connect(self._show_search_context_menu)
        layout.addWidget(self.search_results_table)
        
        # Кнопки для экспорта
        export_layout = QtWidgets.QHBoxLayout()
        
        self.export_csv_btn = QtWidgets.QPushButton('Экспорт в CSV')
        self.export_csv_btn.clicked.connect(self._export_search_results)
        export_layout.addWidget(self.export_csv_btn)
        
        self.copy_to_clipboard_btn = QtWidgets.QPushButton('Копировать адреса')
        self.copy_to_clipboard_btn.clicked.connect(self._copy_search_results)
        export_layout.addWidget(self.copy_to_clipboard_btn)
        
        self.import_to_rewards_btn = QtWidgets.QPushButton('В раздел наград')
        self.import_to_rewards_btn.clicked.connect(self._import_to_rewards)
        export_layout.addWidget(self.import_to_rewards_btn)
        
        layout.addLayout(export_layout)
        
        # Индикатор прогресса
        self.progress_search = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_search)
        
        # Подключаем переключение режима поиска
        self.search_exact_radio.toggled.connect(self._toggle_search_mode_paginated)
        self.search_range_radio.toggled.connect(self._toggle_search_mode_paginated)
        
        return w

    def _tab_history(self):
        """Простая заглушка вкладки История, чтобы не падало приложение (можно расширить позже)."""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        lbl = QtWidgets.QLabel("История пока недоступна. Функционал в разработке.")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)
        layout.addStretch()
        return w
        
    def _toggle_search_mode_paginated(self):
        """Переключение режима поиска в разделе постраничного поиска"""
        is_exact = self.search_exact_radio.isChecked()
        
        self.exact_amount.setEnabled(is_exact)
        self.min_amount.setEnabled(not is_exact)
        self.max_amount.setEnabled(not is_exact)
        
        logger.debug(f"Режим поиска переключен на: {'точная сумма' if is_exact else 'диапазон'}")
    
    def _toggle_date_filter_mode(self):
        """Переключает режим фильтрации по датам"""
        is_range_mode = self.date_range_radio.isChecked()
        
        # Включить/выключить поля для диапазона дат
        self.date_from.setEnabled(is_range_mode)
        self.date_to.setEnabled(is_range_mode)
        self.today_btn.setEnabled(is_range_mode)
        
        logger.debug(f"Режим фильтрации по датам переключен на: {'диапазон дат' if is_range_mode else 'за все время'}")
    
    def _set_date_to_today(self):
        """Устанавливает дату 'До' на сегодняшний день"""
        today = QtCore.QDate.currentDate()
        self.date_to.setDate(today)
        logger.debug(f"Дата 'До' установлена на сегодня: {today.toString('yyyy-MM-dd')}")
        
    def _start_paginated_search(self):
        """Запуск процесса постраничного поиска транзакций"""
        if self.is_searching:
            return logger.warning("Поиск уже запущен")
        
        # Получаем параметры поиска
        wallet_address = self.system_address.text().strip()
        token_address = self.token_address.text().strip()
        
        # Проверяем корректность адресов
        if blockchain_enabled and not Web3.is_address(wallet_address):
            return logger.error("Некорректный адрес кошелька")
        
        if token_address and blockchain_enabled and not Web3.is_address(token_address):
            return logger.error("Некорректный адрес токена")
        
        # Определяем режим поиска и параметры
        exact_amount = None
        min_amount = None
        max_amount = None
        
        if self.search_exact_radio.isChecked():
            exact_amount = self.exact_amount.value()
            if exact_amount <= 0:
                return logger.error("Точная сумма должна быть больше 0")
        else:
            min_amount = self.min_amount.value()
            max_amount = self.max_amount.value()
            if min_amount < 0:
                return logger.error("Минимальная сумма должна быть не меньше 0")
            if max_amount <= min_amount:
                return logger.error("Максимальная сумма должна быть больше минимальной")
        
        # Дополнительная проверка: убеждаемся, что указаны параметры поиска
        if exact_amount is None and (min_amount is None or max_amount is None):
            return logger.error("Необходимо указать либо точную сумму, либо диапазон сумм для поиска")
        
        # Получаем настройки пагинации
        page_size = self.page_size.value()
        max_pages = self.max_pages.value()
        delay_seconds = self.api_delay.value()
        min_tx_count = self.min_tx_count.value()
        track_individual_tx = self.track_tx_checkbox.isChecked()
        
        # Получаем параметры фильтрации по датам
        date_from = None
        date_to = None
        if self.date_range_radio.isChecked():
            date_from = self.date_from.date().toPyDate()
            date_to = self.date_to.date().toPyDate()
            
            # Проверяем корректность диапазона дат
            if date_from > date_to:
                return logger.error("Дата 'От' должна быть раньше или равна дате 'До'")
            
            logger.info(f"Фильтрация по датам: с {date_from} по {date_to}")
        
        # Запускаем поиск в отдельном потоке
        self.is_searching = True
        self.start_search_btn.setEnabled(False)
        self.stop_search_btn.setEnabled(True)
        self.stop_search_event.clear()  # Сбрасываем флаг остановки
        
        # Очищаем предыдущие результаты в таблице
        self.search_results_table.setRowCount(0)
        
        self.search_thread = threading.Thread(
            target=self._paginated_search_thread, 
            args=(wallet_address, token_address, exact_amount, min_amount, max_amount, 
                  page_size, max_pages, delay_seconds, min_tx_count, track_individual_tx,
                  date_from, date_to),
            daemon=True
        )
        self.search_thread.start()
        
    def _paginated_search_thread(self, wallet_address, token_address, exact_amount, min_amount, max_amount, 
                               page_size, max_pages, delay_seconds, min_tx_count, track_individual_tx,
                               date_from=None, date_to=None):
        """Поток для выполнения постраничного поиска транзакций"""
        try:
            logger.info(f"Начинаем поиск транзакций для адреса {wallet_address}")
            
            # Сохраняем параметры поиска для записи в базу данных
            search_params = {
                'wallet_address': wallet_address,
                'token_address': token_address,
                'exact_amount': exact_amount,
                'min_amount': min_amount,
                'max_amount': max_amount,
                'min_tx_count': min_tx_count,
                'date_from': date_from.isoformat() if date_from else None,
                'date_to': date_to.isoformat() if date_to else None,
                'search_time': datetime.now().isoformat()
            }
            
            # Выполняем поиск с постраничной обработкой
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=wallet_address,
                token_contract=token_address if token_address else None,
                exact_amount=exact_amount,
                min_amount=min_amount,
                max_amount=max_amount,
                page_size=page_size,
                max_pages=max_pages,
                delay_seconds=delay_seconds,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Фильтруем транзакции по датам, если указан диапазон
            filtered_txs = matching_txs
            if date_from or date_to:
                filtered_txs = []
                for tx in matching_txs:
                    try:
                        tx_time = datetime.fromtimestamp(int(tx.get('timeStamp', 0)))
                        tx_date = tx_time.date()
                        
                        # Проверяем фильтр по датам
                        if date_from and tx_date < date_from:
                            continue
                        if date_to and tx_date > date_to:
                            continue
                        
                        filtered_txs.append(tx)
                    except Exception as e:
                        logger.error(f"Ошибка фильтрации транзакции по дате: {e}")
                
                logger.info(f"Фильтрация по датам: {len(matching_txs)} -> {len(filtered_txs)} транзакций")
            
            # Сохраняем найденные транзакции в базу данных (только отфильтрованные)
            for tx in filtered_txs:
                add_found_transaction(tx, search_params)
            
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу (только отфильтрованные)
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    # Фильтруем транзакции отправителя по датам
                    filtered_tx_list = []
                    for tx_info in tx_list:
                        try:
                            tx_time = datetime.fromtimestamp(int(tx_info.get('timestamp', 0)))
                            tx_date = tx_time.date()
                            
                            # Проверяем фильтр по датам
                            if date_from and tx_date < date_from:
                                continue
                            if date_to and tx_date > date_to:
                                continue
                            
                            filtered_tx_list.append(tx_info)
                        except Exception as e:
                            logger.error(f"Ошибка фильтрации транзакции отправителя по дате: {e}")
                    
                    # Сохраняем только отфильтрованные транзакции
                    for tx_info in filtered_tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
            
            # Пересчитываем счетчики отправителей после фильтрации по датам
            sender_counter_filtered = {}
            for tx in filtered_txs:
                sender = tx.get('from', '').lower()
                sender_counter_filtered[sender] = sender_counter_filtered.get(sender, 0) + 1
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {addr: count for addr, count in sender_counter_filtered.items() 
                              if count >= min_tx_count}
            
            # Формируем информацию о последних транзакциях
            sender_last_tx = {}
            for tx in filtered_txs:
                try:
                    sender = tx.get('from', '').lower()
                    if sender in filtered_senders:
                        tx_time = datetime.fromtimestamp(int(tx.get('timeStamp', 0)))
                        if sender not in sender_last_tx or tx_time > sender_last_tx[sender][0]:
                            sender_last_tx[sender] = (tx_time, tx.get('hash', ''))
                except Exception as e:
                    logger.error(f"Ошибка обработки транзакции: {e}")
      
            # Создаем список результатов для обновления таблицы
            results = []
            for sender, count in filtered_senders.items():
                last_tx_info = sender_last_tx.get(sender, (datetime.now(), ''))
                results.append((sender, count, last_tx_info[0].strftime('%Y-%m-%d %H:%M:%S'), last_tx_info[1]))
            
            # Создаем отфильтрованный словарь транзакций отправителей
            filtered_sender_transactions = {}
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    if sender in filtered_senders:  # Только для отфильтрованных отправителей
                        filtered_tx_list = []
                        for tx_info in tx_list:
                            try:
                                tx_time = datetime.fromtimestamp(int(tx_info.get('timestamp', 0)))
                                tx_date = tx_time.date()
                                
                                # Проверяем фильтр по датам
                                if date_from and tx_date < date_from:
                                    continue
                                if date_to and tx_date > date_to:
                                    continue
                                
                                filtered_tx_list.append(tx_info)
                            except Exception as e:
                                logger.error(f"Ошибка фильтрации транзакции отправителя по дате: {e}")
                        
                        if filtered_tx_list:  # Добавляем только если есть отфильтрованные транзакции
                            filtered_sender_transactions[sender] = filtered_tx_list
            
            # Отправляем результаты в основной поток для обновления UI
            self.update_table_signal.emit(results, filtered_senders, filtered_sender_transactions)
            
            # Формируем сообщение о завершении поиска
            completion_message = f"Поиск завершен. Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями"
            if date_from or date_to:
                date_range = f"с {date_from} по {date_to}" if date_from and date_to else f"с {date_from}" if date_from else f"по {date_to}"
                completion_message += f" (фильтр по датам: {date_range})"
            
            logger.info(completion_message)
            
            # Отправляем сигнал для обновления таблицы найденных транзакций
            self.found_tx_added_signal.emit()
            
        except Exception as e:
            logger.error(f"Ошибка при выполнении поиска: {e}")
        finally:
            self.is_searching = False
            # Обновляем состояние кнопок в основном потоке
            QtCore.QMetaObject.invokeMethod(
                self.start_search_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.stop_search_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            self.update_progress_signal.emit("search", 100)
    def _update_search_results(self, results, sender_counter, sender_transactions):
        """Обновление таблицы результатов поиска"""
        self.search_results_table.setRowCount(0)
        
        for sender, count, last_tx_time, last_tx_hash in results:
            row = self.search_results_table.rowCount()
            self.search_results_table.insertRow(row)
            
            # Адрес отправителя
            sender_item = QtWidgets.QTableWidgetItem(sender)
            sender_item.setData(QtCore.Qt.UserRole, last_tx_hash)  # Сохраняем хэш последней транзакции
            
            # Если включен режим награждения за каждую транзакцию и у отправителя есть транзакции,
            # сохраняем их в данных элемента
            if self.reward_per_tx_mode and sender in sender_transactions:
                sender_item.setData(QtCore.Qt.UserRole + 1, sender_transactions[sender])
                
            self.search_results_table.setItem(row, 0, sender_item)
            
            # Количество транзакций
            count_item = QtWidgets.QTableWidgetItem(str(count))
            self.search_results_table.setItem(row, 1, count_item)
            
            # Время последней транзакции
            time_item = QtWidgets.QTableWidgetItem(last_tx_time)
            self.search_results_table.setItem(row, 2, time_item)
    
    def _stop_paginated_search(self):
        """Остановка процесса поиска транзакций"""
        if not self.is_searching:
            return
        
        self.stop_search_event.set()  # Установка флага остановки
        logger.info("Отправлен запрос на остановку поиска транзакций...")
        self.stop_search_btn.setEnabled(False)
    
    def _clear_search_results(self):
        """Очистка результатов поиска"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Очистить все результаты поиска?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.search_results_table.setRowCount(0)
            count = clear_found_transactions()
            # Также очищаем таблицу найденных транзакций
            if hasattr(self, 'found_tx_table'):
                self.found_tx_table.setRowCount(0)
            logger.info(f"Очищено {count} записей из истории поиска")
    
    def _show_search_context_menu(self, position):
        """Отображение контекстного меню для таблицы результатов поиска"""
        menu = QtWidgets.QMenu()
        
        copy_action = menu.addAction("Копировать адрес")
        open_bscscan = menu.addAction("Открыть в BscScan")
        add_to_rewards = menu.addAction("Добавить в награды")
        
        # Если включен режим награждения за каждую транзакцию, добавляем соответствующий пункт
        if self.reward_per_tx_mode:
            add_all_tx_to_rewards = menu.addAction("Добавить все транзакции в награды")
        
        action = menu.exec_(self.search_results_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected_indexes = self.search_results_table.selectedIndexes()
        if not selected_indexes:
            return
            
        # Получаем данные выбранных строк
        selected_row = selected_indexes[0].row()
        sender_address = self.search_results_table.item(selected_row, 0).text()
        tx_hash = self.search_results_table.item(selected_row, 0).data(QtCore.Qt.UserRole)
        
        if action == copy_action:
            QtWidgets.QApplication.clipboard().setText(sender_address)
            logger.info(f"Скопирован адрес: {sender_address}")
        
        elif action == open_bscscan:
            if tx_hash:
                url = f"https://bscscan.com/tx/{tx_hash}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открываем транзакцию в BscScan: {tx_hash}")
            else:
                url = f"https://bscscan.com/address/{sender_address}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открываем адрес в BscScan: {sender_address}")
        
        elif action == add_to_rewards:
            self._add_to_rewards_tab(sender_address)
        
        elif self.reward_per_tx_mode and action == add_all_tx_to_rewards:
            # Получаем список транзакций из данных элемента
            tx_list = self.search_results_table.item(selected_row, 0).data(QtCore.Qt.UserRole + 1)
            if tx_list:
                self._add_tx_list_to_rewards(sender_address, tx_list)
            else:
                logger.warning(f"Нет данных о транзакциях для {sender_address}")
    
    def _add_tx_list_to_rewards(self, sender_addr, tx_list):
        """Добавление списка транзакций в таблицу наград"""
        if not tx_list:
            return logger.warning(f"Список транзакций пуст для {sender_addr}")
        
        # Сохраняем текущее состояние таблицы наград
        self._save_rewards_state()
        
        # Добавляем каждую транзакцию в таблицу наград
        for tx_info in tx_list:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес отправителя
            addr_item = QtWidgets.QTableWidgetItem(sender_addr)
            addr_item.setData(QtCore.Qt.UserRole, tx_info['hash'])
            self.table_rewards.setItem(row, 0, addr_item)
            
            # PLEX - значение по умолчанию
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT - пустое значение по умолчанию
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Добавлено {len(tx_list)} транзакций отправителя {sender_addr} в таблицу наград")
        
        # Предлагаем сгенерировать случайные значения
        reply = QtWidgets.QMessageBox.question(
            self, 'Рандомизация', 
            'Сгенерировать случайные значения PLEX и USDT для транзакций?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.Yes
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._randomize_plex()
            self._randomize_usdt()
            
    def _export_search_results(self):
        """Экспорт результатов поиска в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,tx_count,last_tx_time\n')
                for row in range(self.search_results_table.rowCount()):
                    address = self.search_results_table.item(row, 0).text()
                    count = self.search_results_table.item(row, 1).text()
                    time = self.search_results_table.item(row, 2).text()
                    f.write(f'"{address}",{count},"{time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте в CSV: {e}")
    
    def _copy_search_results(self):
        """Копирование адресов из результатов поиска в буфер обмена"""
        addresses = []
        
        # Если есть выбранные строки, копируем только их
        selected_rows = set(index.row() for index in self.search_results_table.selectedIndexes())
        if selected_rows:
            for row in selected_rows:
                addresses.append(self.search_results_table.item(row, 0).text())
        else:
            # Иначе копируем все адреса
            for row in range(self.search_results_table.rowCount()):
                addresses.append(self.search_results_table.item(row, 0).text())
        
        if addresses:
            QtWidgets.QApplication.clipboard().setText('\n'.join(addresses))
            logger.info(f"Скопировано {len(addresses)} адресов в буфер обмена")
        else:
            logger.warning("Нет адресов для копирования")
    
    def _import_to_rewards(self):
        """Импорт выбранных адресов в таблицу наград"""
        selected_rows = set(index.row() for index in self.search_results_table.selectedIndexes())
        
        # Если ничего не выбрано, импортируем все
        if not selected_rows:
            selected_rows = range(self.search_results_table.rowCount())
        
        # Запрашиваем режим импорта, если включен режим награждения за каждую транзакцию
        import_mode = "single"
        if self.reward_per_tx_mode:
            reply = QtWidgets.QMessageBox.question(
                self, 'Режим импорта', 
                'Выберите режим импорта:\n\n'
                'Нажмите "Да" для импорта каждой транзакции (одна транзакция - одна награда)\n'
                'Нажмите "Нет" для импорта только адресов (один адрес - одна награда)',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            import_mode = "per_tx" if reply == QtWidgets.QMessageBox.Yes else "single"
        
        count = 0
        if import_mode == "per_tx":
            # Импортируем каждую транзакцию
            for row in selected_rows:
                sender_addr = self.search_results_table.item(row, 0).text()
                tx_list = self.search_results_table.item(row, 0).data(QtCore.Qt.UserRole + 1)
                if tx_list:
                    # Перед добавлением сохраняем текущее состояние
                    if count == 0:
                        self._save_rewards_state()
                    count += len(tx_list)
                    for tx_info in tx_list:
                        self._add_tx_to_rewards(sender_addr, tx_info['hash'])
                else:
                    # Если нет данных о транзакциях, добавляем просто адрес
                    self._add_to_rewards_tab(sender_addr)
                    count += 1
        else:
            # Обычный импорт - один адрес, одна награда
            for row in selected_rows:
                sender_addr = self.search_results_table.item(row, 0).text()
                self._add_to_rewards_tab(sender_addr)
                count += 1
        
        logger.info(f"Импортировано {count} записей в таблицу наград")
        
        # Предлагаем сгенерировать случайные значения
        if count > 0:
            reply = QtWidgets.QMessageBox.question(
                self, 'Рандомизация', 
                'Сгенерировать случайные значения для импортированных адресов?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            if reply == QtWidgets.QMessageBox.Yes:
                self._randomize_plex()
                self._randomize_usdt()
    # --- Таб «Анализ» ---
    def _tab_analyze(self):
        """Создание вкладки для анализа транзакций"""
        w = QtWidgets.QWidget()
        l = QtWidgets.QVBoxLayout(w)

        # Форма ввода данных
        form = QtWidgets.QFormLayout()
        self.addr_in = QtWidgets.QLineEdit()
        self.addr_in.setPlaceholderText('0x...')
        form.addRow('Адрес для анализа:', self.addr_in)

        btn_seed = QtWidgets.QPushButton('Ввести seed-фразу')
        btn_seed.clicked.connect(self._enter_mnemonic)
        form.addRow('Seed-фраза:', btn_seed)

        btn_pk = QtWidgets.QPushButton('Ввести приватный ключ')
        btn_pk.clicked.connect(self._enter_pk)
        form.addRow('PrivKey:', btn_pk)

        self.sys_in = QtWidgets.QLineEdit()
        self.sys_in.setPlaceholderText('Системный адрес')
        form.addRow('Системный адрес:', self.sys_in)

        self.spin_amt = QtWidgets.QDoubleSpinBox()
        self.spin_amt.setRange(0, 1e12)
        self.spin_amt.setDecimals(8)
        self.spin_amt.setValue(30.0)  # Предустановка на 30 PLEX
        
        # Минимальное количество транзакций
        self.spin_tx_count = QtWidgets.QSpinBox()
        self.spin_tx_count.setRange(1, 1000)
        self.spin_tx_count.setValue(1)
        form.addRow('Мин. кол-во транзакций:', self.spin_tx_count)
        
        # Переключатель режима поиска
        self.search_mode_group = QtWidgets.QButtonGroup(self)
        self.radio_exact = QtWidgets.QRadioButton('Точная сумма')
        self.radio_range = QtWidgets.QRadioButton('Диапазон сумм')
        self.radio_exact.setChecked(True)
        self.search_mode_group.addButton(self.radio_exact)
        self.search_mode_group.addButton(self.radio_range)
        
        # Горизонтальный layout для радиокнопок
        radio_layout = QtWidgets.QHBoxLayout()
        radio_layout.addWidget(self.radio_exact)
        radio_layout.addWidget(self.radio_range)
        form.addRow('Режим поиска:', radio_layout)
        
        # Поля для точной суммы
        form.addRow('Точная сумма:', self.spin_amt)
        
        # Поля для диапазона сумм
        self.spin_amt_from = QtWidgets.QDoubleSpinBox()
        self.spin_amt_from.setRange(0, 1e12)
        self.spin_amt_from.setDecimals(8)
        self.spin_amt_from.setEnabled(False)
        
        self.spin_amt_to = QtWidgets.QDoubleSpinBox()
        self.spin_amt_to.setRange(0, 1e12)
        self.spin_amt_to.setDecimals(8)
        self.spin_amt_to.setEnabled(False)
        
        # Layout для полей диапазона
        range_layout = QtWidgets.QHBoxLayout()
        range_layout.addWidget(QtWidgets.QLabel('От:'))
        range_layout.addWidget(self.spin_amt_from)
        range_layout.addWidget(QtWidgets.QLabel('До:'))
        range_layout.addWidget(self.spin_amt_to)
        form.addRow('Диапазон:', range_layout)
        
        # Опция для учета отдельных транзакций
        self.track_tx_checkbox_analyze = QtWidgets.QCheckBox("Учитывать отдельные транзакции для награждения")
        self.track_tx_checkbox_analyze.setChecked(True)
        form.addRow("", self.track_tx_checkbox_analyze)
        
        # Соединение сигналов для переключения режима
        self.radio_exact.toggled.connect(self._toggle_search_mode)
        self.radio_range.toggled.connect(self._toggle_search_mode)

        l.addLayout(form)

        # Кнопки действий
        h = QtWidgets.QHBoxLayout()
        btn_scan = QtWidgets.QPushButton('Сканировать токены')
        btn_scan.clicked.connect(self._scan_tokens)
        btn_search = QtWidgets.QPushButton('Поиск Tx')
        btn_search.clicked.connect(self._search_txs)
        h.addWidget(btn_scan); h.addWidget(btn_search)
        l.addLayout(h)
        
        # Таблица токенов
        self.tokens_table = QtWidgets.QTableWidget(0, 4)
        self.tokens_table.setHorizontalHeaderLabels(["Токен","Контракт","Баланс","Decimals"])
        header = self.tokens_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.tokens_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.tokens_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tokens_table.customContextMenuRequested.connect(self._tokens_context_menu)
        l.addWidget(self.tokens_table)

        # Таблица результатов
        self.table_res = QtWidgets.QTableWidget(0, 2)  # Увеличиваем до 2 столбцов (адрес и количество)
        self.table_res.setHorizontalHeaderLabels(['Отправитель', 'Кол-во TX'])
        result_header = self.table_res.horizontalHeader()
        if result_header is not None:
            result_header.setStretchLastSection(True)
        self.table_res.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table_res.customContextMenuRequested.connect(self._analysis_context_menu)
        l.addWidget(self.table_res)

        # Кнопки экспорта
        h2 = QtWidgets.QHBoxLayout()
        btn_csv = QtWidgets.QPushButton('Экспорт CSV')
        btn_csv.clicked.connect(self._export_csv)
        h2.addWidget(btn_csv)
        
        btn_clipboard = QtWidgets.QPushButton('Копировать в буфер')
        btn_clipboard.clicked.connect(self._copy_to_clipboard)
        h2.addWidget(btn_clipboard)
        
        btn_to_rewards = QtWidgets.QPushButton('В раздел наград')
        btn_to_rewards.clicked.connect(self._import_from_analysis)
        h2.addWidget(btn_to_rewards)
        
        l.addLayout(h2)
        
        # Индикатор прогресса
        self.progress = QtWidgets.QProgressBar()
        l.addWidget(self.progress)

        return w
    
    def _tokens_context_menu(self, position):
        """Контекстное меню для таблицы токенов"""
        if self.tokens_table.rowCount() == 0:
            return
            
        menu = QtWidgets.QMenu()
        copy_contract = menu.addAction("Копировать адрес контракта")
        view_token = menu.addAction("Открыть токен в BscScan")
        use_as_search = menu.addAction("Использовать для поиска")
        
        action = menu.exec_(self.tokens_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.tokens_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        contract = self.tokens_table.item(row, 1).text()
        
        if action == copy_contract:
            QtWidgets.QApplication.clipboard().setText(contract)
            logger.info(f"Адрес контракта скопирован: {contract}")
            
        elif action == view_token:
            url = f"https://bscscan.com/token/{contract}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт токен в BscScan: {contract}")
            
        elif action == use_as_search:
            # Переключиться на вкладку поиска и установить токен
            tabs = self.centralWidget()
            tabs.setCurrentIndex(1)  # Переключаемся на вкладку поиска
            self.token_address.setText(contract)
            logger.info(f"Токен {contract} установлен для поиска")
    
    def _analysis_context_menu(self, position):
        """Контекстное меню для таблицы результатов анализа"""
        if self.table_res.rowCount() == 0:
            return
            
        menu = QtWidgets.QMenu()
        copy_address = menu.addAction("Копировать адрес")
        view_address = menu.addAction("Открыть адрес в BscScan")
        add_to_rewards = menu.addAction("Добавить в награды")
        
        action = menu.exec_(self.table_res.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.table_res.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        address = self.table_res.item(row, 0).text()
        
        if action == copy_address:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == view_address:
            url = f"https://bscscan.com/address/{address}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт адрес в BscScan: {address}")
            
        elif action == add_to_rewards:
            self._add_to_rewards_tab(address)
    def _copy_to_clipboard(self):
        """Копирование выбранных адресов в буфер обмена"""
        addresses = []
        for idx in self.table_res.selectedIndexes():
            if idx.column() == 0:
                addresses.append(self.table_res.item(idx.row(), 0).text())
        
        if not addresses:
            # Если ничего не выбрано, копировать все
            addresses = [self.table_res.item(row, 0).text() 
                      for row in range(self.table_res.rowCount())]
                      
        if addresses:
            clipboard = QtWidgets.QApplication.clipboard()
            clipboard.setText('\n'.join(addresses))
            self.log(f"Скопировано адресов: {len(addresses)}")
        else:
            self.log("Нет адресов для копирования")
            
    def _enter_mnemonic(self):
        """Диалог для ввода seed-фразы"""
        m, ok = QtWidgets.QInputDialog.getText(self, 'Seed', 'Введите seed-фразу')
        if not (ok and m):
            return

        # Убираем переносы строк и лишние пробелы
        clean = ' '.join(m.strip().split())

        # Сохраняем зашифрованную версию "как есть"
        self.cfg.set_mnemonic(clean)

        if blockchain_enabled:
            try:
                # Включаем небезопасные HD фичи web3 (требует подтверждения пользователя)
                from eth_account import Account as _AccountModule
                try:
                    # Включаем один раз (idempotent)
                    if hasattr(_AccountModule, 'enable_unaudited_hdwallet_features'):
                        _AccountModule.enable_unaudited_hdwallet_features()
                except Exception as ee:
                    logger.warning(f"Не удалось включить HD wallet фичи: {ee}")

                acct = Account.from_mnemonic(clean, account_path="m/44'/60'/0'/0/0")
            except ValidationError as e:
                logger.error(f"Неверная seed-фраза: {e}")
                return
            except Exception as e:
                logger.error(f"Ошибка при обработке seed: {e}")
                return

            # если всё OK - сохраняем PK и выводим адрес
            pk_hex = acct.key.hex()
            self.cfg.set_key(pk_hex)
            self.pk = pk_hex
            # Маскируем часть адреса в логах
            short_addr = f"{acct.address[:6]}...{acct.address[-4:]}"
            logger.info(f"Seed сохранён, адрес: {short_addr}")
        else:
            logger.info("Seed сохранён (blockchain отключён)")

    def _enter_pk(self):
        """Диалог для ввода приватного ключа"""
        p, ok = QtWidgets.QInputDialog.getText(
            self, 'PrivKey', 'Введите приватный ключ:', QtWidgets.QLineEdit.Password)
        if ok and p:
            # Проверка формата приватного ключа
            if blockchain_enabled:
                try:
                    # Проверка действительности приватного ключа путем создания аккаунта
                    acct = Account.from_key(p.strip())
                    addr = acct.address
                    logger.info(f"Ключ валиден, адрес: {addr}")
                except Exception as e:
                    logger.error(f"Неверный формат ключа: {e}")
                    return
            
            self.cfg.set_key(p.strip())
            self.pk = p.strip()
            logger.info("PrivKey сохранен")

    def _scan_tokens(self):
        """Запуск сканирования токенов для выбранного адреса"""
        addr = self.addr_in.text().strip()
        if blockchain_enabled and not Web3.is_address(addr):
            return logger.error("Некорректный адрес")
        self.table_res.setRowCount(0)
        threading.Thread(target=self._scan_thread, args=(addr,), daemon=True).start()

    def _scan_thread(self, addr):
        """Поток для сканирования токенов на адресе"""
        logger.info("Сканирование токенов...")
        contracts, page = set(), 1
        
        try:
            # Используем постраничную обработку для получения всех токенов
            while True:
                try:
                    params = {
                        'module': 'account',
                        'action': 'tokentx',
                        'address': addr,
                        'page': page,
                        'offset': 100,
                        'sort': 'asc'
                    }
                    result = bscscan_request(params)
                except Exception as e:
                    logger.error(f"Ошибка BscScan на странице {page}: {e}")
                    break
                    
                if not result:
                    break
                
                for tx in result:
                    contracts.add((tx['contractAddress'], tx['tokenSymbol'], int(tx['tokenDecimal'])))
                
                # Обновляем прогресс и переходим к следующей странице
                logger.info(f"Обработана страница {page}, найдено {len(contracts)} уникальных токенов")
                page += 1
                self.update_progress_signal.emit("scan", min(30, (page % 31)))
                time.sleep(1)  # Задержка 1 секунда между запросами
        
            tokens = []
            total = len(contracts)
            for i, (c, s, d) in enumerate(contracts, start=1):
                self.update_progress_signal.emit("scan", 30 + int(i / total * 70))
                try:
                    bal = int(bscscan_request({
                        'module':'account','action':'tokenbalance',
                        'contractaddress': c,'address': addr,'tag':'latest'
                    })) / 10**d
                    if bal > 0:
                        tokens.append((c, s, d, bal))
                except Exception as e:
                    logger.warning(f"Ошибка при получении баланса {s}: {e}")
                time.sleep(1)  # Задержка 1 секунда между запросами

            # Обновляем таблицу токенов
            self._update_tokens_table(tokens)
            
            self.combo = QtWidgets.QComboBox()
            for c, s, d, v in tokens:
                self.combo.addItem(f"{s} ({v})", userData=(c, d))
                
            logger.info(f"Найдено {len(tokens)} токенов c положительным балансом")
            self.update_progress_signal.emit("scan", 100)
        except Exception as e:
            logger.error(f"Ошибка при сканировании токенов: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _update_tokens_table(self, tokens):
        """Обновить таблицу токенов: список tuples (contract, symbol, decimals, balance)"""
        self.tokens_table.setRowCount(0)
        for c, s, d, v in tokens:
            row = self.tokens_table.rowCount()
            self.tokens_table.insertRow(row)
            self.tokens_table.setItem(row, 0, QtWidgets.QTableWidgetItem(s))
            self.tokens_table.setItem(row, 1, QtWidgets.QTableWidgetItem(c))
            self.tokens_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(v)))
            self.tokens_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(d)))

    def _search_txs(self):
        """Поиск транзакций по заданным параметрам"""
        sys_addr = self.sys_in.text().strip()
        if blockchain_enabled and not Web3.is_address(sys_addr):
            return logger.error("Неверный системный адрес")
        
        data = getattr(self, 'combo', None)
        if data is None or data.currentIndex() < 0:
            return logger.error("Выберите токен")
        
        contract, dec = data.currentData()
        track_individual_tx = self.track_tx_checkbox_analyze.isChecked()
        
        # Определяем режим поиска
        if self.radio_exact.isChecked():
            # Режим точной суммы
            amt = self.spin_amt.value()
            if amt <= 0:
                return logger.error("Сумма должна быть > 0")
            
            # Показываем диалог подтверждения
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение', 
                f'Искать входящие транзакции с суммой {amt} токенов на адрес {sys_addr}?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            
            if reply == QtWidgets.QMessageBox.Yes:
                threading.Thread(
                    target=self._search_thread,
                    args=(sys_addr, contract, dec, amt, track_individual_tx),
                    daemon=True
                ).start()
        else:
            # Режим диапазона сумм
            min_amt = self.spin_amt_from.value()
            max_amt = self.spin_amt_to.value()
            
            if min_amt < 0:
                return logger.error("Минимальная сумма должна быть >= 0")
            
            if max_amt <= min_amt:
                return logger.error("Максимальная сумма должна быть больше минимальной")
            
            # Показываем диалог подтверждения
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение', 
                f'Искать входящие транзакции с суммой от {min_amt} до {max_amt} токенов на адрес {sys_addr}?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            
            if reply == QtWidgets.QMessageBox.Yes:
                threading.Thread(
                    target=self._search_range_thread, 
                    args=(sys_addr, contract, dec, min_amt, max_amt, track_individual_tx),
                    daemon=True
                ).start()
    def _search_thread(self, sys_addr, contract, dec, amt, track_individual_tx):
        """Поток для поиска транзакций с точным значением"""
        self.table_res.setRowCount(0)
        logger.info(f"Поиск Tx для адреса {sys_addr}, токен {contract}, сумма {amt}...")
        min_tx_count = self.spin_tx_count.value()
        
        # Используем новую функцию для постраничного поиска
        try:
            self.stop_search_event.clear()  # Сбрасываем флаг остановки
            
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=sys_addr,
                token_contract=contract,
                exact_amount=amt,
                page_size=1000,
                max_pages=100,
                delay_seconds=1,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {sender: count for sender, count in sender_counter.items() 
                              if count >= min_tx_count}
            
            # Заполняем таблицу результатов
            for row, (sender, count) in enumerate(sorted(filtered_senders.items(), key=lambda x: x[1], reverse=True)):
                self.table_res.insertRow(row)
                sender_item = QtWidgets.QTableWidgetItem(sender)
                
                # Если включен режим награждения за каждую транзакцию и есть данные о транзакциях,
                # сохраняем их в данных элемента
                if track_individual_tx and sender in sender_transactions:
                    sender_item.setData(QtCore.Qt.UserRole, sender_transactions[sender])
                
                self.table_res.setItem(row, 0, sender_item)
                self.table_res.setItem(row, 1, QtWidgets.QTableWidgetItem(str(count)))

            logger.info(f"Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями")
            self.update_progress_signal.emit("scan", 100)
            
            # Сохраняем найденные транзакции в базу данных
            search_info = {
                'wallet_address': sys_addr,
                'token_contract': contract,
                'exact_amount': amt,
                'min_tx_count': min_tx_count,
                'search_time': datetime.now().isoformat()
            }
            
            for tx in matching_txs:
                add_found_transaction(tx, search_info)
            
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
            
            # Отправляем сигнал для обновления таблицы найденных транзакций
            self.found_tx_added_signal.emit()
                
        except Exception as e:
            logger.error(f"Ошибка при поиске транзакций: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _search_range_thread(self, sys_addr, contract, dec, min_amt, max_amt, track_individual_tx):
        """Поток для поиска транзакций в диапазоне сумм"""
        self.table_res.setRowCount(0)
        logger.info(f"Поиск Tx для адреса {sys_addr}, токен {contract}, диапазон {min_amt}-{max_amt}...")
        min_tx_count = self.spin_tx_count.value()
        
        try:
            self.stop_search_event.clear()  # Сбрасываем флаг остановки
            
            # Используем новую функцию для постраничного поиска с диапазоном
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=sys_addr,
                token_contract=contract,
                min_amount=min_amt,
                max_amount=max_amt,
                page_size=1000,
                max_pages=100,
                delay_seconds=1,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {sender: count for sender, count in sender_counter.items() 
                              if count >= min_tx_count}
            
            # Заполняем таблицу результатов
            for row, (sender, count) in enumerate(sorted(filtered_senders.items(), key=lambda x: x[1], reverse=True)):
                self.table_res.insertRow(row)
                sender_item = QtWidgets.QTableWidgetItem(sender)
                
                # Если включен режим награждения за каждую транзакцию и есть данные о транзакциях,
                # сохраняем их в данных элемента
                if track_individual_tx and sender in sender_transactions:
                    sender_item.setData(QtCore.Qt.UserRole, sender_transactions[sender])
                
                self.table_res.setItem(row, 0, sender_item)
                self.table_res.setItem(row, 1, QtWidgets.QTableWidgetItem(str(count)))

            logger.info(f"Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями")
            self.update_progress_signal.emit("scan", 100)
            
            # Сохраняем найденные транзакции в базу данных
            search_info = {
                'wallet_address': sys_addr,
                'token_contract': contract,
                'min_amount': min_amt,
                'max_amount': max_amt,
                'min_tx_count': min_tx_count,
                'search_time': datetime.now().isoformat()
            }
            
            for tx in matching_txs:
                add_found_transaction(tx, search_info)
                
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
            
            # Отправляем сигнал для обновления таблицы найденных транзакций
            self.found_tx_added_signal.emit()
                
        except Exception as e:
            logger.error(f"Ошибка при поиске транзакций по диапазону: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _export_csv(self):
        """Экспорт результатов в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('from_address,tx_count\n')
                for i in range(self.table_res.rowCount()):
                    address = self.table_res.item(i, 0).text()
                    count = self.table_res.item(i, 1).text() if self.table_res.columnCount() > 1 else "1"
                    f.write(f'{address},{count}\n')
            logger.info(f"Экспортировано {self.table_res.rowCount()} адресов в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте CSV: {e}")

    def _add_to_rewards_tab(self, address):
        """Добавление адреса в таблицу наград"""
        # Проверка, есть ли уже такой адрес в таблице
        for row in range(self.table_rewards.rowCount()):
            if self.table_rewards.item(row, 0).text() == address:
                logger.warning(f"Адрес {address} уже добавлен в таблицу наград")
                return
        
        # Добавляем новую строку
        row = self.table_rewards.rowCount()
        self.table_rewards.insertRow(row)
        
        # Адрес
        self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(address))
        
        # PLEX - пустое значение по умолчанию
        plex_item = QtWidgets.QTableWidgetItem("")
        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 1, plex_item)
        
        # USDT - пустое значение по умолчанию
        usdt_item = QtWidgets.QTableWidgetItem("")
        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Адрес {address} добавлен в таблицу наград")
        
    def _import_from_analysis(self):
        """Импорт адресов из вкладки анализа в таблицу наград"""
        if self.table_res.rowCount() == 0:
            return logger.warning("Нет данных в таблице анализа")
            
        # Запрашиваем режим импорта, если включен режим награждения за каждую транзакцию
        import_mode = "single"
        if self.reward_per_tx_mode:
            reply = QtWidgets.QMessageBox.question(
                self, 'Режим импорта', 
                'Выберите режим импорта:\n\n'
                'Нажмите "Да" для импорта каждой транзакции (одна транзакция - одна награда)\n'
                'Нажмите "Нет" для импорта только адресов (один адрес - одна награда)',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            import_mode = "per_tx" if reply == QtWidgets.QMessageBox.Yes else "single"
        
        # Сохраняем текущее состояние для возможности отмены
        self._save_rewards_state()
        
        count = 0
        if import_mode == "per_tx":
            # Импортируем каждую транзакцию
            for row in range(self.table_res.rowCount()):
                sender_item = self.table_res.item(row, 0)
                sender = sender_item.text()
                # Получаем данные о транзакциях из элемента
                tx_list = sender_item.data(QtCore.Qt.UserRole) if hasattr(sender_item, 'data') else None
                
                if tx_list:
                    for tx_info in tx_list:
                        r = self.table_rewards.rowCount()
                        self.table_rewards.insertRow(r)
                        
                        addr_item = QtWidgets.QTableWidgetItem(sender)
                        addr_item.setData(QtCore.Qt.UserRole, tx_info['hash'])
                        self.table_rewards.setItem(r, 0, addr_item)
                        
                        plex_item = QtWidgets.QTableWidgetItem("3")
                        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                        self.table_rewards.setItem(r, 1, plex_item)
                        
                        usdt_item = QtWidgets.QTableWidgetItem("")
                        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                        self.table_rewards.setItem(r, 2, usdt_item)
                        
                        count += 1
                else:
                    # Если нет данных о транзакциях, добавляем просто адрес
                    self._add_to_rewards_tab(sender)
                    count += 1
        else:
            # Обычный импорт - один адрес, одна награда
            for row in range(self.table_res.rowCount()):
                address = self.table_res.item(row, 0).text()
                self._add_to_rewards_tab(address)
                count += 1
        
        logger.info(f"Импортировано {count} записей из таблицы анализа")
        
        # Предлагаем сгенерировать случайные значения
        reply = QtWidgets.QMessageBox.question(
            self, 'Генерация значений', 
            'Сгенерировать значения PLEX и USDT для импортированных адресов?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.Yes
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._randomize_plex()
            self._randomize_usdt()

    def _load_history(self):
        """Загрузка истории отправок из БД"""
        try:
            self.history_table.setRowCount(0)
            rows = fetch_history(limit=500)
            for row_data in rows:
                row = self.history_table.rowCount()
                self.history_table.insertRow(row)
                # Время
                self.history_table.setItem(row, 0, QtWidgets.QTableWidgetItem(row_data['time']))
                # Токен
                self.history_table.setItem(row, 1, QtWidgets.QTableWidgetItem(row_data['token']))
                # Получатель
                self.history_table.setItem(row, 2, QtWidgets.QTableWidgetItem(row_data['to_addr']))
                # Сумма
                self.history_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(row_data['amount'])))
                # TX Hash
                tx_item = QtWidgets.QTableWidgetItem(row_data['tx'][:10] + "..." + row_data['tx'][-6:])
                tx_item.setToolTip(row_data['tx'])
                tx_item.setData(QtCore.Qt.UserRole, row_data['tx'])
                self.history_table.setItem(row, 4, tx_item)
                # Статус
                status_item = QtWidgets.QTableWidgetItem(row_data['status'])
                if row_data['status'] == 'success':
                    status_item.setBackground(QtGui.QColor('#004400'))
                elif row_data['status'] == 'failed':
                    status_item.setBackground(QtGui.QColor('#440000'))
                self.history_table.setItem(row, 5, status_item)
            logger.info(f"Загружено {self.history_table.rowCount()} записей истории")
        except Exception as e:
            logger.error(f"Ошибка при загрузке истории: {e}")
    
    def _history_context_menu(self, position):
        """Отображение контекстного меню для таблицы истории"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_addr = menu.addAction("Копировать адрес получателя")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        check_status = menu.addAction("Проверить статус")
        
        action = menu.exec_(self.history_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.history_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.history_table.item(row, 4).data(QtCore.Qt.UserRole)
        to_addr = self.history_table.item(row, 2).text()
        
        if action == copy_tx:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_addr:
            QtWidgets.QApplication.clipboard().setText(to_addr)
            logger.info(f"Адрес получателя скопирован: {to_addr}")
            
        elif action == open_tx:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")
            
        elif action == check_status:
            threading.Thread(
                target=self._check_tx_status,
                args=(tx_hash,),
                daemon=True
            ).start()
    
    def _check_tx_status(self, tx_hash):
        """Проверка статуса транзакции"""
        try:
            if not blockchain_enabled:
                logger.warning("Блокчейн отключен, невозможно проверить статус")
                return
                
            w3 = self.rpc.web3()
            
            logger.info(f"Проверка статуса транзакции {tx_hash}...")
            try:
                receipt = w3.eth.get_transaction_receipt(tx_hash)
                
                if receipt and receipt.get('status') == 1:
                    status = 'success'
                    logger.info(f"Транзакция {tx_hash} успешно выполнена")
                else:
                    status = 'failed'
                    logger.info(f"Транзакция {tx_hash} не выполнена")
                
                # Обновляем статус в базе данных
                update_tx_status(tx_hash, status)
                
                # Обновляем таблицу
                self._refresh_history()
                
            except Exception as e:
                logger.warning(f"Не удалось получить чек транзакции {tx_hash}: {e}")
                
        except Exception as e:
            logger.error(f"Ошибка при проверке статуса транзакции: {e}")
    
    def _track_tx_statuses(self):
        """Запуск отслеживания статусов всех транзакций в отдельном потоке"""
        if self.tx_tracker and self.tx_tracker.is_alive():
            return logger.warning("Отслеживание уже запущено")
            
        self.tx_tracker = threading.Thread(
            target=self._tx_tracker_thread,
            daemon=True
        )
        self.tx_tracker.start()
        logger.info("Запущено отслеживание статусов транзакций")
    
    def _tx_tracker_thread(self):
        """Поток для отслеживания статусов транзакций"""
        if not blockchain_enabled:
            return logger.warning("Блокчейн отключен, невозможно отслеживать статусы")
            
        try:
            w3 = self.rpc.web3()
            
            # Получаем все транзакции со статусом pending
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT tx FROM history WHERE status='pending'")
            pending_txs = [row['tx'] for row in cur.fetchall()]
            conn.close()
            
            logger.info(f"Найдено {len(pending_txs)} ожидающих транзакций для проверки")
            
            for tx_hash in pending_txs:
                try:
                    receipt = w3.eth.get_transaction_receipt(tx_hash)
                    
                    if receipt:
                        if receipt.get('status') == 1:
                            status = 'success'
                            logger.info(f"Транзакция {tx_hash} успешно выполнена")
                        else:
                            status = 'failed'
                            logger.info(f"Транзакция {tx_hash} не выполнена")
                        
                        # Обновляем статус в базе данных
                        update_tx_status(tx_hash, status)
                except Exception as e:
                    logger.warning(f"Не удалось проверить транзакцию {tx_hash}: {e}")
                
                time.sleep(0.5)  # Небольшая задержка между запросами
            
            # Обновляем таблицу
            QtCore.QMetaObject.invokeMethod(
                self, "_refresh_history", 
                QtCore.Qt.QueuedConnection
            )
            
            logger.info("Проверка статусов транзакций завершена")
            
        except Exception as e:
            logger.error(f"Ошибка при отслеживании статусов: {e}")
    def _copy_all_tx_hashes(self):
        """Копирование всех хэшей транзакций в буфер обмена"""
        try:
            tx_hashes = copy_all_transactions_hashes()
            if tx_hashes:
                QtWidgets.QApplication.clipboard().setText(tx_hashes)
                logger.info(f"Скопированы все хэши транзакций")
            else:
                logger.warning("Нет транзакций для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании хэшей: {e}")

    # --- Таб «Найденные Tx» ---
    def _tab_found_tx(self):
        """Создание вкладки для просмотра найденных транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа фильтрации
        filter_group = QtWidgets.QGroupBox("Фильтрация")
        filter_layout = QtWidgets.QHBoxLayout(filter_group)
        
        # Поле для фильтрации по адресу кошелька
        filter_layout.addWidget(QtWidgets.QLabel("Адрес кошелька:"))
        self.found_tx_wallet_filter = QtWidgets.QLineEdit()
        self.found_tx_wallet_filter.setPlaceholderText("Введите адрес кошелька для фильтрации")
        self.found_tx_wallet_filter.textChanged.connect(self._refresh_found_tx)
        filter_layout.addWidget(self.found_tx_wallet_filter)
        
        # Кнопка очистки фильтра
        btn_clear_filter = QtWidgets.QPushButton("Очистить фильтр")
        btn_clear_filter.clicked.connect(self._clear_found_tx_filter)
        filter_layout.addWidget(btn_clear_filter)
        
        layout.addWidget(filter_group)
        
        # Таблица найденных транзакций
        self.found_tx_table = QtWidgets.QTableWidget(0, 8)
        self.found_tx_table.setHorizontalHeaderLabels([
            'Время поиска', 'Адрес кошелька', 'От кого', 'Кому', 'Токен', 'Сумма', 'Хэш Tx', 'Время Tx'
        ])
        header = self.found_tx_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
            header.setSectionResizeMode(6, QtWidgets.QHeaderView.Interactive)
        self.found_tx_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.found_tx_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.found_tx_table.customContextMenuRequested.connect(self._found_tx_context_menu)
        layout.addWidget(self.found_tx_table)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        btn_refresh_found = QtWidgets.QPushButton('Обновить')
        btn_refresh_found.clicked.connect(self._refresh_found_tx)
        buttons_layout.addWidget(btn_refresh_found)
        
        btn_export_found = QtWidgets.QPushButton('Экспорт в CSV')
        btn_export_found.clicked.connect(self._export_found_tx)
        buttons_layout.addWidget(btn_export_found)
        
        btn_clear_found = QtWidgets.QPushButton('Очистить')
        btn_clear_found.clicked.connect(self._clear_found_tx)
        buttons_layout.addWidget(btn_clear_found)
        
        layout.addLayout(buttons_layout)
        
        # Кнопки копирования данных
        copy_buttons_layout = QtWidgets.QHBoxLayout()
        
        btn_copy_all_addresses = QtWidgets.QPushButton('Копировать все адреса')
        btn_copy_all_addresses.clicked.connect(self._copy_all_found_addresses)
        copy_buttons_layout.addWidget(btn_copy_all_addresses)
        
        btn_copy_senders = QtWidgets.QPushButton('Копировать отправителей')
        btn_copy_senders.clicked.connect(self._copy_found_senders)
        copy_buttons_layout.addWidget(btn_copy_senders)
        
        btn_copy_receivers = QtWidgets.QPushButton('Копировать получателей')
        btn_copy_receivers.clicked.connect(self._copy_found_receivers)
        copy_buttons_layout.addWidget(btn_copy_receivers)
        
        btn_copy_tx_hashes = QtWidgets.QPushButton('Копировать хэши транзакций')
        btn_copy_tx_hashes.clicked.connect(self._copy_found_tx_hashes)
        copy_buttons_layout.addWidget(btn_copy_tx_hashes)
        
        layout.addLayout(copy_buttons_layout)
        
        # Не загружаем данные автоматически - только при явной команде обновления
        # QtCore.QTimer.singleShot(100, self._refresh_found_tx)
        
        # Подключаем сигнал для автоматического обновления при добавлении новых транзакций
        self.found_tx_added_signal.connect(self._refresh_found_tx)
        
        return w
    
    def _refresh_found_tx(self):
        """Обновление таблицы найденных транзакций"""
        try:
            # Получаем данные из БД
            found_txs = fetch_found_transactions()
            
            # Применяем фильтр по адресу кошелька (ищем в from_addr и to_addr)
            wallet_filter = self.found_tx_wallet_filter.text().strip().lower()
            if wallet_filter:
                filtered_txs = []
                for tx in found_txs:
                    # Парсим search_data для получения wallet_address
                    try:
                        search_data = json.loads(tx['search_data']) if tx['search_data'] else {}
                        wallet_address = search_data.get('wallet_address', '')
                    except:
                        wallet_address = ''
                    
                    if (wallet_filter in tx['from_addr'].lower() or 
                        wallet_filter in tx['to_addr'].lower() or
                        wallet_filter in wallet_address.lower()):
                        filtered_txs.append(tx)
                found_txs = filtered_txs
            
            # Очищаем таблицу
            self.found_tx_table.setRowCount(0)
            
            # Заполняем таблицу
            for tx in found_txs:
                row = self.found_tx_table.rowCount()
                self.found_tx_table.insertRow(row)
                
                # Время поиска
                ts = datetime.fromisoformat(tx['ts'])
                self.found_tx_table.setItem(row, 0, QtWidgets.QTableWidgetItem(ts.strftime('%Y-%m-%d %H:%M:%S')))
                
                # Адрес кошелька (из search_data)
                try:
                    search_data = json.loads(tx['search_data']) if tx['search_data'] else {}
                    wallet_address = search_data.get('wallet_address', tx['from_addr'])
                except:
                    wallet_address = tx['from_addr']
                self.found_tx_table.setItem(row, 1, QtWidgets.QTableWidgetItem(wallet_address))
                
                # От кого
                from_item = QtWidgets.QTableWidgetItem(tx['from_addr'])
                self.found_tx_table.setItem(row, 2, from_item)
                
                # Кому
                to_item = QtWidgets.QTableWidgetItem(tx['to_addr'])
                self.found_tx_table.setItem(row, 3, to_item)
                
                # Токен
                self.found_tx_table.setItem(row, 4, QtWidgets.QTableWidgetItem(tx['token_name'] or 'N/A'))
                
                # Сумма
                self.found_tx_table.setItem(row, 5, QtWidgets.QTableWidgetItem(str(tx['amount'])))
                
                # Хэш Tx
                tx_hash = tx['tx_hash']
                if len(tx_hash) > 16:
                    tx_display = tx_hash[:10] + "..." + tx_hash[-6:]
                else:
                    tx_display = tx_hash
                tx_item = QtWidgets.QTableWidgetItem(tx_display)
                tx_item.setToolTip(tx_hash)
                tx_item.setData(QtCore.Qt.UserRole, tx_hash)
                self.found_tx_table.setItem(row, 6, tx_item)
                
                # Время Tx
                if tx['block_time']:
                    try:
                        block_time = datetime.fromisoformat(tx['block_time'])
                        self.found_tx_table.setItem(row, 7, QtWidgets.QTableWidgetItem(block_time.strftime('%Y-%m-%d %H:%M:%S')))
                    except:
                        self.found_tx_table.setItem(row, 7, QtWidgets.QTableWidgetItem(str(tx['block_time'])))
                else:
                    self.found_tx_table.setItem(row, 7, QtWidgets.QTableWidgetItem("-"))
            
            logger.info(f"Загружено {self.found_tx_table.rowCount()} найденных транзакций")
            
        except Exception as e:
            logger.error(f"Ошибка при обновлении найденных транзакций: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
    
    def _found_tx_context_menu(self, position):
        """Отображение контекстного меню для таблицы найденных транзакций"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_from = menu.addAction("Копировать адрес отправителя")
        copy_to = menu.addAction("Копировать адрес получателя")
        copy_wallet = menu.addAction("Копировать адрес кошелька")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        add_from_to_rewards = menu.addAction("Добавить отправителя в награды")
        
        action = menu.exec_(self.found_tx_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.found_tx_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.found_tx_table.item(row, 6).data(QtCore.Qt.UserRole)
        wallet_addr = self.found_tx_table.item(row, 1).text()
        from_addr = self.found_tx_table.item(row, 2).text()
        to_addr = self.found_tx_table.item(row, 3).text()
        
        if action == copy_tx:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_from:
            QtWidgets.QApplication.clipboard().setText(from_addr)
            logger.info(f"Адрес отправителя скопирован: {from_addr}")
            
        elif action == copy_to:
            QtWidgets.QApplication.clipboard().setText(to_addr)
            logger.info(f"Адрес получателя скопирован: {to_addr}")
            
        elif action == copy_wallet:
            QtWidgets.QApplication.clipboard().setText(wallet_addr)
            logger.info(f"Адрес кошелька скопирован: {wallet_addr}")
            
        elif action == open_tx:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")
            
        elif action == add_from_to_rewards:
            self._add_to_rewards_tab(from_addr)
    def _export_found_tx(self):
        """Экспорт найденных транзакций в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('timestamp,wallet_address,from_addr,to_addr,token,amount,tx_hash,block_time\n')
                for row in range(self.found_tx_table.rowCount()):
                    timestamp = self.found_tx_table.item(row, 0).text()
                    wallet_addr = self.found_tx_table.item(row, 1).text()
                    from_addr = self.found_tx_table.item(row, 2).text()
                    to_addr = self.found_tx_table.item(row, 3).text()
                    token = self.found_tx_table.item(row, 4).text()
                    amount = self.found_tx_table.item(row, 5).text()
                    tx_hash = self.found_tx_table.item(row, 6).data(QtCore.Qt.UserRole)
                    block_time = self.found_tx_table.item(row, 7).text()
                    
                    f.write(f'"{timestamp}","{wallet_addr}","{from_addr}","{to_addr}","{token}",{amount},"{tx_hash}","{block_time}"\n')
            
            logger.info(f"Найденные транзакции экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте найденных транзакций: {e}")
    
    def _clear_found_tx_filter(self):
        """Очистка фильтра найденных транзакций"""
        self.found_tx_wallet_filter.clear()
        self._refresh_found_tx()
        logger.info("Фильтр найденных транзакций очищен")
    
    def _copy_all_found_addresses(self):
        """Копирование всех адресов из найденных транзакций"""
        try:
            addresses = set()
            for row in range(self.found_tx_table.rowCount()):
                # Добавляем адрес кошелька
                wallet_addr = self.found_tx_table.item(row, 1).text()
                addresses.add(wallet_addr)
                # Добавляем адрес отправителя
                from_addr = self.found_tx_table.item(row, 2).text()
                addresses.add(from_addr)
                # Добавляем адрес получателя
                to_addr = self.found_tx_table.item(row, 3).text()
                addresses.add(to_addr)
            
            if addresses:
                addresses_text = '\n'.join(sorted(addresses))
                QtWidgets.QApplication.clipboard().setText(addresses_text)
                logger.info(f"Скопировано {len(addresses)} уникальных адресов")
            else:
                logger.warning("Нет адресов для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании адресов: {e}")
    
    def _copy_found_senders(self):
        """Копирование адресов отправителей из найденных транзакций"""
        try:
            senders = set()
            for row in range(self.found_tx_table.rowCount()):
                from_addr = self.found_tx_table.item(row, 2).text()
                senders.add(from_addr)
            
            if senders:
                senders_text = '\n'.join(sorted(senders))
                QtWidgets.QApplication.clipboard().setText(senders_text)
                logger.info(f"Скопировано {len(senders)} адресов отправителей")
            else:
                logger.warning("Нет отправителей для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании отправителей: {e}")
    
    def _copy_found_receivers(self):
        """Копирование адресов получателей из найденных транзакций"""
        try:
            receivers = set()
            for row in range(self.found_tx_table.rowCount()):
                to_addr = self.found_tx_table.item(row, 3).text()
                receivers.add(to_addr)
            
            if receivers:
                receivers_text = '\n'.join(sorted(receivers))
                QtWidgets.QApplication.clipboard().setText(receivers_text)
                logger.info(f"Скопировано {len(receivers)} адресов получателей")
            else:
                logger.warning("Нет получателей для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании получателей: {e}")
    
    def _copy_found_tx_hashes(self):
        """Копирование хэшей транзакций из найденных транзакций"""
        try:
            tx_hashes = []
            for row in range(self.found_tx_table.rowCount()):
                tx_hash = self.found_tx_table.item(row, 6).data(QtCore.Qt.UserRole)
                if tx_hash:
                    tx_hashes.append(tx_hash)
            
            if tx_hashes:
                tx_hashes_text = '\n'.join(tx_hashes)
                QtWidgets.QApplication.clipboard().setText(tx_hashes_text)
                logger.info(f"Скопировано {len(tx_hashes)} хэшей транзакций")
            else:
                logger.warning("Нет хэшей транзакций для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании хэшей транзакций: {e}")
    
    def _clear_found_tx(self):
        """Очистка таблицы найденных транзакций"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Вы уверены, что хотите очистить все найденные транзакции?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            count = clear_found_transactions()
            self.found_tx_table.setRowCount(0)
            logger.info(f"Очищено {count} найденных транзакций")

    # (Удалено: устаревший монолитный UI и логика старой массовой рассылки)

    def _mass_update_wallet_info(self):
        """Обновление информации о кошельке и балансах для массовой рассылки"""
        try:
            if not self.pk:
                self.mass_sender_address_label.setText("Адрес: (ключ не задан)")
                self.mass_balance_plex_label.setText("PLEX: -")
                self.mass_balance_bnb_label.setText("BNB: -")
                self.mass_balance_usdt_label.setText("USDT: -")
                return

            w3 = self.rpc.web3()
            account = Account.from_key(self.pk)
            addr = w3.to_checksum_address(account.address)
            self.mass_sender_address_label.setText(f"Адрес: {addr[:10]}...{addr[-6:]}")
            self.mass_sender_address_label.setToolTip(addr)

            # BNB баланс
            try:
                balance_bnb = w3.from_wei(w3.eth.get_balance(addr), 'ether')
                self.mass_balance_bnb_label.setText(f"BNB: {float(balance_bnb):.5f}")
            except Exception as e_bnb:
                self.mass_balance_bnb_label.setText("BNB: ошибка")
                logger.warning(f"Не удалось получить баланс BNB: {e_bnb}")

            # PLEX баланс
            try:
                plex_contract = w3.eth.contract(address=w3.to_checksum_address(PLEX_CONTRACT), abi=ERC20_ABI)
                plex_dec = plex_contract.functions.decimals().call()
                plex_raw = plex_contract.functions.balanceOf(addr).call()
                plex_val = plex_raw / (10 ** plex_dec)
                self.mass_balance_plex_label.setText(f"PLEX: {plex_val:.4f}")
            except Exception as e_plex:
                self.mass_balance_plex_label.setText("PLEX: ошибка")
                logger.warning(f"Не удалось получить баланс PLEX: {e_plex}")

            # USDT баланс
            try:
                usdt_contract = w3.eth.contract(address=w3.to_checksum_address(USDT_CONTRACT), abi=ERC20_ABI)
                usdt_dec = usdt_contract.functions.decimals().call()
                usdt_raw = usdt_contract.functions.balanceOf(addr).call()
                usdt_val = usdt_raw / (10 ** usdt_dec)
                self.mass_balance_usdt_label.setText(f"USDT: {usdt_val:.4f}")
            except Exception as e_usdt:
                self.mass_balance_usdt_label.setText("USDT: ошибка")
                logger.warning(f"Не удалось получить баланс USDT: {e_usdt}")

        except Exception as e:
            logger.error(f"Ошибка обновления информации о кошельке массовой рассылки: {e}")
    
    def _mass_paste_addresses(self):
        """Вставка адресов из буфера обмена"""
        clipboard = QtWidgets.QApplication.clipboard()
        text = clipboard.text()
        if text:
            self.mass_addresses_text.setText(text)
            logger.info("Адреса вставлены из буфера обмена")
    
    def _mass_import_excel(self):
        """Импорт адресов из Excel файла"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Выберите Excel файл', '', 'Excel Files (*.xlsx *.xls);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            # Попытка импорта с использованием pandas
            try:
                import pandas as pd
                df = pd.read_excel(path)
                
                # Ищем столбец с адресами
                addresses = []
                for col in df.columns:
                    col_data = df[col].dropna().astype(str)
                    # Проверяем, похожи ли данные на адреса BSC
                    potential_addresses = [addr for addr in col_data if addr.startswith('0x') and len(addr) == 42]
                    if potential_addresses:
                        addresses.extend(potential_addresses)
                
                if addresses:
                    self.mass_addresses_text.setText(' '.join(addresses))
                    logger.info(f"Импортировано {len(addresses)} адресов из Excel")
                else:
                    logger.warning("Не найдено адресов в Excel файле")
                    
            except ImportError:
                # Если pandas не установлен, используем openpyxl напрямую
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    
                    addresses = []
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell and isinstance(cell, str) and cell.startswith('0x') and len(cell) == 42:
                                addresses.append(cell)
                    
                    if addresses:
                        self.mass_addresses_text.setText(' '.join(addresses))
                        logger.info(f"Импортировано {len(addresses)} адресов из Excel")
                    else:
                        logger.warning("Не найдено адресов в Excel файле")
                        
                except ImportError:
                    logger.error("Для импорта Excel необходимо установить pandas или openpyxl")
                    QtWidgets.QMessageBox.warning(
                        self, 'Ошибка',
                        'Для импорта Excel файлов необходимо установить библиотеку pandas или openpyxl:\n'
                        'pip install pandas openpyxl',
                        QtWidgets.QMessageBox.Ok
                    )
                    
        except Exception as e:
            logger.error(f"Ошибка при импорте Excel: {e}")
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка импорта',
                f'Не удалось импортировать адреса из Excel:\n{str(e)}',
                QtWidgets.QMessageBox.Ok
            )
    
    def _mass_add_addresses(self):
        """Добавление адресов из текстового поля в таблицу"""
        text = self.mass_addresses_text.toPlainText().strip()
        if not text:
            return logger.warning("Нет адресов для добавления")
        
        # Разбиваем текст на адреса по различным разделителям
        import re
        addresses = re.split(r'[\s,;]+', text)
        
        # Фильтруем и валидируем адреса
        valid_addresses = []
        for addr in addresses:
            addr = addr.strip()
            if addr and self._validate_address(addr):
                valid_addresses.append(addr)
        
        if not valid_addresses:
            return logger.warning("Не найдено валидных адресов")
        
        # Получаем текущие адреса из таблицы
        current_addresses = set()
        for row in range(self.mass_table.rowCount()):
            current_addresses.add(self.mass_table.item(row, 0).text())
        
        # Добавляем только уникальные адреса
        added_count = 0
        for addr in valid_addresses:
            if addr not in current_addresses:
                row = self.mass_table.rowCount()
                self.mass_table.insertRow(row)
                
                # Адрес
                self.mass_table.setItem(row, 0, QtWidgets.QTableWidgetItem(addr))
                # Статус
                self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem("Ожидание"))
                # Отправлено раз
                self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem("0"))
                # Последний хэш
                self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem("-"))
                # Время
                self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem("-"))
                
                current_addresses.add(addr)
                added_count += 1
        
        # Обновляем статистику
        self._mass_update_statistics()
        
        # Очищаем поле ввода
        self.mass_addresses_text.clear()
        
        logger.info(f"Добавлено {added_count} уникальных адресов")
    
    def _validate_address(self, address):
        """Валидация BSC адреса"""
        if blockchain_enabled:
            try:
                Web3.to_checksum_address(address)
                return True
            except:
                return False
        else:
            # Простая проверка формата без Web3
            return address.startswith('0x') and len(address) == 42
    
    def _mass_clear_addresses(self):
        """Очистка списка адресов"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            'Очистить весь список адресов?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.mass_table.setRowCount(0)
            self._mass_update_statistics()
            logger.info("Список адресов очищен")
    
    def _mass_update_statistics(self):
        """Обновление статистики"""
        total = self.mass_table.rowCount()
        unique = len(set(self.mass_table.item(row, 0).text() for row in range(total)))
        
        sent_count = 0
        error_count = 0
        
        for row in range(total):
            status = self.mass_table.item(row, 1).text()
            if status == "Успешно":
                sent_count += 1
            elif status == "Ошибка":
                error_count += 1
        
        self.mass_total_addresses_label.setText(f"Всего адресов: {total}")
        self.mass_unique_addresses_label.setText(f"Уникальных: {unique}")
        self.mass_sent_count_label.setText(f"Отправлено: {sent_count}")
        self.mass_error_count_label.setText(f"Ошибок: {error_count}")
    
    def _mass_token_changed(self, token):
        """Обработка изменения выбранного токена"""
        self.mass_custom_token_edit.setEnabled(token == "Другой...")
        if token == "Другой...":
            self.mass_custom_token_edit.setFocus()
    
    def _mass_table_context_menu(self, position):
        """Контекстное меню для таблицы адресов"""
        menu = QtWidgets.QMenu()
        
        remove_action = menu.addAction("Удалить адрес")
        copy_action = menu.addAction("Копировать адрес")
        copy_hash_action = menu.addAction("Копировать хэш транзакции")
        open_bscscan_action = menu.addAction("Открыть в BscScan")
        menu.addSeparator()
        reset_status_action = menu.addAction("Сбросить статус")
        
        action = menu.exec_(self.mass_table.viewport().mapToGlobal(position))
        
        if not action:
            return
        
        selected = self.mass_table.selectedItems()
        if not selected:
            return
        
        row = selected[0].row()
        address = self.mass_table.item(row, 0).text()
        tx_hash = self.mass_table.item(row, 3).text()
        
        if action == remove_action:
            self.mass_table.removeRow(row)
            self._mass_update_statistics()
            logger.info(f"Адрес {address} удален из списка")
            
        elif action == copy_action:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == copy_hash_action and tx_hash != "-":
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == open_bscscan_action:
            url = f"https://bscscan.com/address/{address}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт адрес в BscScan: {address}")
            
        elif action == reset_status_action:
            self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem("Ожидание"))
            self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem("0"))
            self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem("-"))
            self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem("-"))
            self._mass_update_statistics()
            logger.info(f"Статус адреса {address} сброшен")
    
    def _mass_estimate_cost(self):
        """Оценка стоимости рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для оценки")
        
        addresses_count = self.mass_table.rowCount()
        cycles = self.mass_cycles_spin.value()
        amount = self.mass_amount_spin.value()
        token = self.mass_token_combo.currentText()
        
        total_transactions = addresses_count * cycles
        total_amount = total_transactions * amount
        
        # Оценка газа (примерная)
        estimated_gas_per_tx = 0.0003  # BNB (примерно)
        total_gas = total_transactions * estimated_gas_per_tx
        
        message = f"""
Оценка стоимости рассылки:

Адресов: {addresses_count}
Циклов: {cycles}
Сумма за транзакцию: {amount} {token}

Всего транзакций: {total_transactions}
Всего токенов: {total_amount} {token}
Примерная стоимость газа: {total_gas:.4f} BNB

Убедитесь, что на кошельке достаточно средств!
        """
        
        QtWidgets.QMessageBox.information(self, "Оценка стоимости", message)
        logger.info(f"Оценка: {total_transactions} транзакций, {total_amount} {token}, ~{total_gas:.4f} BNB газа")
    
    def _mass_start_distribution(self):
        """Начало массовой рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для рассылки")
        
        if not self.pk:
            return logger.error("Не настроен приватный ключ")
        
        # Получаем параметры
        token = self.mass_token_combo.currentText()
        if token == "Другой...":
            token_address = self.mass_custom_token_edit.text().strip()
            if not token_address or not self._validate_address(token_address):
                return logger.error("Некорректный адрес токена")
        elif token == "PLEX":
            token_address = PLEX_CONTRACT
        elif token == "USDT":
            token_address = USDT_CONTRACT
        elif token == "BNB":
            token_address = None  # Нативный токен
        
        amount = self.mass_amount_spin.value()
        interval = self.mass_interval_spin.value()
        cycles = self.mass_cycles_spin.value()
        mode = self.mass_mode_combo.currentText()
        
        # Подтверждение
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            f'Начать рассылку {amount} {token} на {self.mass_table.rowCount()} адресов?\n'
            f'Циклов: {cycles}, интервал: {interval} сек',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
        
        # Запуск рассылки
        self.mass_distribution_active = True
        self.mass_distribution_paused = False
        self.mass_current_cycle = 0
        self.mass_total_sent = 0
        self.mass_errors_count = 0
        
        # Обновление UI
        self.mass_start_btn.setEnabled(False)
        self.mass_pause_btn.setEnabled(True)
        self.mass_stop_btn.setEnabled(True)
        self.mass_progress.setValue(0)
        
        # Запуск в отдельном потоке
        self.mass_distribution_thread = threading.Thread(
            target=self._mass_distribution_worker,
            args=(token_address, token, amount, interval, cycles, mode),
            daemon=True
        )
        self.mass_distribution_thread.start()
        
        logger.info(f"Начата массовая рассылка: {amount} {token}, {cycles} циклов")
    
    def _mass_distribution_worker(self, token_address, token_name, amount, interval, cycles, mode):
        """Рабочий поток для массовой рассылки"""
        try:
            addresses = []
            for row in range(self.mass_table.rowCount()):
                addresses.append(self.mass_table.item(row, 0).text())
            
            total_transactions = len(addresses) * cycles
            completed_transactions = 0
            
            # Создаем запись в базе данных
            distribution_name = f"Массовая рассылка {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            distribution_id = add_mass_distribution(
                name=distribution_name,
                token_address=token_address or "BNB",
                token_symbol=token_name,
                amount_per_tx=amount,
                total_addresses=len(addresses),
                total_cycles=cycles,
                interval_seconds=interval
            )
            
            for cycle in range(cycles):
                if not self.mass_distribution_active:
                    break
                
                self.mass_current_cycle = cycle + 1
                logger.info(f"Начат цикл {self.mass_current_cycle} из {cycles}")
                
                for idx, address in enumerate(addresses):
                    if not self.mass_distribution_active:
                        break
                    
                    # Проверка паузы
                    while self.mass_distribution_paused and self.mass_distribution_active:
                        time.sleep(0.1)
                    
                    # Обновление прогресса
                    progress = int((completed_transactions / total_transactions) * 100)
                    QtCore.QMetaObject.invokeMethod(
                        self.mass_progress, "setValue",
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(int, progress)
                    )
                    
                    QtCore.QMetaObject.invokeMethod(
                        self.mass_progress_label, "setText",
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(str, f"Цикл {self.mass_current_cycle}/{cycles}, адрес {idx+1}/{len(addresses)}")
                    )
                    
                    # Обновление статуса в таблице
                    self._mass_update_table_status(idx, "Отправка...", None, None)
                    
                    # Отправка транзакции
                    try:
                        if token_address:  # Токен
                            tx_hash = self._send_token_transaction(address, amount, token_address)
                        else:  # BNB
                            tx_hash = self._send_bnb_transaction(address, amount)
                        
                        if tx_hash:
                            self.mass_total_sent += 1
                            sent_count = int(self.mass_table.item(idx, 2).text()) + 1
                            self._mass_update_table_status(
                                idx, "Успешно", tx_hash, 
                                datetime.now().strftime('%H:%M:%S')
                            )
                            self._mass_update_sent_count(idx, sent_count)
                            
                            # Сохраняем в базу данных
                            add_mass_distribution_item(
                                distribution_id=distribution_id,
                                address=address,
                                cycle=self.mass_current_cycle,
                                tx_hash=tx_hash,
                                status='success'
                            )
                            
                            logger.info(f"Отправлено {amount} {token_name} на {address}, tx: {tx_hash}")

                            # Пытаемся получить подтверждение (receipt) для дополнительного доказательства выплаты
                            try:
                                w3 = self.rpc.web3()
                                receipt = w3.eth.wait_for_transaction_receipt(tx_hash, timeout=60)
                                if receipt and receipt.status == 1:
                                    logger.info(f"Подтверждена транзакция (receipt) для {address}: {tx_hash}")
                                else:
                                    logger.warning(f"Транзакция не подтверждена (status!=1) пока: {tx_hash}")
                            except Exception as rec_e:
                                logger.warning(f"Не удалось получить receipt для {tx_hash}: {rec_e}")

                            # Отправляем данные в вкладку 'Очередь отправки' (таблица завершенных)
                            try:
                                completed_item = {
                                    'status': 'success',
                                    'address': address,
                                    'token': token_name,
                                    'amount': amount,
                                    'tx_hash': tx_hash,
                                    'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                }
                                # Используем уже существующий сигнал для добавления строки в completed_table
                                self.completed_item_signal.emit(completed_item)
                            except Exception as emit_e:
                                logger.error(f"Не удалось отправить элемент массовой рассылки в очередь: {emit_e}")
                        else:
                            raise Exception("Не удалось получить хэш транзакции")
                            
                    except Exception as e:
                        self.mass_errors_count += 1
                        self._mass_update_table_status(
                            idx, "Ошибка", None,
                            datetime.now().strftime('%H:%M:%S')
                        )
                        
                        # Сохраняем ошибку в базу данных
                        add_mass_distribution_item(
                            distribution_id=distribution_id,
                            address=address,
                            cycle=self.mass_current_cycle,
                            tx_hash=None,
                            status='error',
                            error_message=str(e)
                        )
                        
                        logger.error(f"Ошибка отправки на {address}: {e}")

                        # Также отражаем неуспешную попытку в completed_table очереди
                        try:
                            failed_item = {
                                'status': 'failed',
                                'address': address,
                                'token': token_name,
                                'amount': amount,
                                'tx_hash': None,
                                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            self.completed_item_signal.emit(failed_item)
                        except Exception as emit_e:
                            logger.error(f"Не удалось отправить ошибочный элемент массовой рассылки в очередь: {emit_e}")
                    
                    completed_transactions += 1
                    
                    # Задержка между транзакциями
                    if idx < len(addresses) - 1 and self.mass_distribution_active:
                        time.sleep(interval)
                
                # Задержка между циклами
                if cycle < cycles - 1 and self.mass_distribution_active:
                    logger.info(f"Цикл {self.mass_current_cycle} завершен, ожидание перед следующим...")
                    time.sleep(interval)
            
            # Обновляем статус рассылки
            if self.mass_distribution_active:
                update_mass_distribution_status(distribution_id, 'completed')
            else:
                update_mass_distribution_status(distribution_id, 'cancelled')
            
            logger.info(f"Массовая рассылка завершена: {self.mass_total_sent} успешных, {self.mass_errors_count} ошибок")
            
        except Exception as e:
            logger.error(f"Критическая ошибка в массовой рассылке: {e}")
        finally:
            self.mass_distribution_active = False
            QtCore.QMetaObject.invokeMethod(
                self, "_mass_distribution_finished",
                QtCore.Qt.QueuedConnection
            )
    
    def _mass_update_table_status(self, row, status, tx_hash, time_str):
        """Обновление статуса в таблице"""
        try:
            # Используем сигнал для обновления UI из потока
            if status == "Отправка...":
                self.update_address_status.emit(row, "⟳ Отправка...")
            elif status == "Успешно":
                self.update_address_status.emit(row, "✓ Успешно")
            elif status == "Ошибка":
                self.update_address_status.emit(row, "✗ Ошибка")
            else:
                self.update_address_status.emit(row, status)
            
            # Обновляем статистику
            self._mass_update_statistics()
            
            # Обновляем данные через слот в основном потоке
            QtCore.QMetaObject.invokeMethod(
                self, 
                "_update_table_item_data",
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(int, row),
                QtCore.Q_ARG(str, status),
                QtCore.Q_ARG(str, tx_hash or ""),
                QtCore.Q_ARG(str, time_str or "")
            )
            
        except Exception as e:
            logger.error(f"Ошибка обновления статуса: {e}")
    
    @QtCore.pyqtSlot(int, str, str, str)
    def _update_table_item_data(self, row, status, tx_hash, time_str):
        """Слот для обновления данных таблицы в основном потоке"""
        try:
            if row < self.mass_table.rowCount():
                # Обновляем счетчик отправок
                current_count = int(self.mass_table.item(row, 2).text().split('/')[0])
                new_count = current_count + 1
                count_item = QtWidgets.QTableWidgetItem(f"{new_count}/{self.mass_cycles_spin.value()}")
                self.mass_table.setItem(row, 2, count_item)
                
                # Обновляем хэш транзакции
                if tx_hash:
                    hash_item = QtWidgets.QTableWidgetItem(tx_hash[:10] + "..." + tx_hash[-6:])
                    hash_item.setToolTip(tx_hash)
                    self.mass_table.setItem(row, 3, hash_item)
                
                # Обновляем время
                if time_str:
                    time_item = QtWidgets.QTableWidgetItem(time_str)
                    self.mass_table.setItem(row, 4, time_item)
                    
        except Exception as e:
            logger.error(f"Ошибка обновления данных таблицы: {e}")
    
    def _mass_update_sent_count(self, row, count):
        """Обновление счетчика отправок"""
        try:
            QtCore.QMetaObject.invokeMethod(
                self, 
                "_update_sent_count_slot",
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(int, row),
                QtCore.Q_ARG(int, count)
            )
        except Exception as e:
            logger.error(f"Ошибка обновления счетчика отправок: {e}")
    
    @QtCore.pyqtSlot(int, int)
    def _update_sent_count_slot(self, row, count):
        """Слот для обновления счетчика отправок в основном потоке"""
        try:
            if row < self.mass_table.rowCount():
                count_item = QtWidgets.QTableWidgetItem(str(count))
                self.mass_table.setItem(row, 2, count_item)
        except Exception as e:
            logger.error(f"Ошибка обновления счетчика в слоте: {e}")
    
    # Удалено: старая версия _mass_distribution_finished (заменена в другом месте или в новой вкладке)
    
    def _mass_save_list(self):
        """Сохранение списка адресов"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для сохранения")
        
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Сохранить список адресов', '', 'JSON Files (*.json);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            data = {
                'addresses': [],
                'settings': {
                    'token': self.mass_token_combo.currentText(),
                    'custom_token': self.mass_custom_token_edit.text() if self.mass_token_combo.currentText() == "Другой..." else "",
                    'amount': self.mass_amount_spin.value(),
                    'interval': self.mass_interval_spin.value(),
                    'cycles': self.mass_cycles_spin.value()
                }
            }
            
            for row in range(self.mass_table.rowCount()):
                data['addresses'].append({
                    'address': self.mass_table.item(row, 0).text(),
                    'status': self.mass_table.item(row, 1).text(),
                    'sent_count': int(self.mass_table.item(row, 2).text()),
                    'last_hash': self.mass_table.item(row, 3).text(),
                    'last_time': self.mass_table.item(row, 4).text()
                })
            
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Список сохранен в {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении списка: {e}")
    
    def _mass_load_list(self):
        """Загрузка списка адресов"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Загрузить список адресов', '', 'JSON Files (*.json);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Очищаем текущую таблицу
            self.mass_table.setRowCount(0)
            
            # Загружаем адреса
            for addr_data in data.get('addresses', []):
                row = self.mass_table.rowCount()
                self.mass_table.insertRow(row)
                
                self.mass_table.setItem(row, 0, QtWidgets.QTableWidgetItem(addr_data['address']))
                self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem(addr_data.get('status', 'Ожидание')))
                self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(addr_data.get('sent_count', 0))))
                self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem(addr_data.get('last_hash', '-')))
                self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem(addr_data.get('last_time', '-')))
            
            # Загружаем настройки
            settings = data.get('settings', {})
            if settings.get('token'):
                self.mass_token_combo.setCurrentText(settings['token'])
            if settings.get('custom_token'):
                self.mass_custom_token_edit.setText(settings['custom_token'])
            if 'amount' in settings:
                self.mass_amount_spin.setValue(settings['amount'])
            if 'interval' in settings:
                self.mass_interval_spin.setValue(settings['interval'])
            if 'cycles' in settings:
                self.mass_cycles_spin.setValue(settings['cycles'])
            
            self._mass_update_statistics()
            logger.info(f"Загружено {self.mass_table.rowCount()} адресов из {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке списка: {e}")
    
    def _mass_export_results(self):
        """Экспорт результатов рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет данных для экспорта")
        
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Экспорт результатов', '', 'CSV Files (*.csv);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,status,sent_count,last_hash,last_time\n')
                
                for row in range(self.mass_table.rowCount()):
                    address = self.mass_table.item(row, 0).text()
                    status = self.mass_table.item(row, 1).text()
                    sent_count = self.mass_table.item(row, 2).text()
                    last_hash = self.mass_table.item(row, 3).text()
                    last_time = self.mass_table.item(row, 4).text()
                    
                    f.write(f'"{address}","{status}",{sent_count},"{last_hash}","{last_time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")
    
    def _send_token_transaction(self, to_address, amount, token_address):
        """Отправка токенов на адрес"""
        try:
            w3 = self.rpc.web3()
            account = Account.from_key(self.pk)
            
            # Конвертируем адреса в checksum формат
            to_address = w3.to_checksum_address(to_address)
            token_address = w3.to_checksum_address(token_address)
            
            # Проверяем баланс
            token_contract = w3.eth.contract(address=token_address, abi=ERC20_ABI)
            decimals = token_contract.functions.decimals().call()
            balance = token_contract.functions.balanceOf(account.address).call()
            amount_wei = int(amount * (10 ** decimals))
            
            if balance < amount_wei:
                raise Exception(f"Недостаточно токенов: {balance / (10 ** decimals)} < {amount}")
            
            # Подготовка транзакции
            nonce = w3.eth.get_transaction_count(account.address)
            # Используем настроенную цену газа вместо автоматической
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            tx = token_contract.functions.transfer(
                to_address,
                amount_wei
            ).build_transaction({
                'from': account.address,
                'nonce': nonce,
                'gas': 100000,
                'gasPrice': gas_price
            })
            
            # Подписание и отправка
            signed_tx = w3.eth.account.sign_transaction(tx, self.pk)
            tx_hash = send_raw_tx(w3, signed_tx.rawTransaction)
            
            return tx_hash.hex()
            
        except Exception as e:
            logger.error(f"Ошибка при отправке токенов: {e}")
            raise
    
    def _send_bnb_transaction(self, to_address, amount):
        """Отправка BNB на адрес"""
        try:
            w3 = self.rpc.web3()
            account = Account.from_key(self.pk)
            
            # Проверяем баланс
            balance = w3.eth.get_balance(account.address)
            amount_wei = Web3.to_wei(amount, 'ether')
            
            if balance < amount_wei:
                raise Exception(f"Недостаточно BNB: {Web3.from_wei(balance, 'ether')} < {amount}")
            
            # Подготовка транзакции
            nonce = w3.eth.get_transaction_count(account.address)
            # Используем настроенную цену газа вместо автоматической
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            tx = {
                'from': account.address,
                'to': Web3.to_checksum_address(to_address),
                'value': amount_wei,
                'nonce': nonce,
                'gas': 21000,
                'gasPrice': gas_price
            }
            
            # Подписание и отправка
            signed_tx = w3.eth.account.sign_transaction(tx, self.pk)
            tx_hash = send_raw_tx(w3, signed_tx.rawTransaction)
            
            return tx_hash.hex()
            
        except Exception as e:
            logger.error(f"Ошибка при отправке BNB: {e}")
            raise

    # --- #MCP:DS_TAB - Таб «ДС» (Дополнительные Сервисы) ---
    def _tab_ds(self):
        """Создание вкладки ДС (Дополнительные Сервисы) для работы с PLEX ONE
        
        TODO:MCP - Добавить сохранение состояния вкладки ДС
        OPTIMIZE:MCP - Кэшировать данные компонентов ДС
        REFACTOR:MCP - Вынести в отдельный модуль ds_tab
        """
        try:
            # Попытка импорта компонентов ДС
            import sys
            import os
            
            # Добавляем путь к MCP компонентам ДС
            ds_components_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "ds_components"
            )
            ds_components_path = os.path.normpath(ds_components_path)
            
            if os.path.exists(ds_components_path):
                sys.path.insert(0, ds_components_path)
                
                try:
                    from ds_tab_controller import DSTabController  # type: ignore
                    
                    # Получаем BSCScan API ключ из конфигурации или устанавливаем по умолчанию
                    bscscan_api_key = ""
                    if hasattr(self, 'cfg') and self.cfg:
                        bscscan_api_key = getattr(self.cfg, 'bscscan_api_key', "")
                    
                    # Создаем контроллер ДС с передачей всех необходимых параметров
                    ds_controller = DSTabController(
                        rpc_pool=self.rpc_pool if hasattr(self, 'rpc_pool') else None,
                        config=self.cfg if hasattr(self, 'cfg') else None,
                        bscscan_api_key=bscscan_api_key
                    )
                    
                    # Подключаем сигналы
                    ds_controller.status_update.connect(self._update_ds_status)
                    ds_controller.transaction_ready.connect(self._handle_ds_transaction)
                    ds_controller.balance_alert.connect(self._handle_balance_alert)
                    ds_controller.analysis_completed.connect(self._handle_analysis_completed)
                    
                    logger.info("✅ Вкладка ДС успешно инициализирована с полным функционалом")
                    return ds_controller
                    
                except ImportError:
                    # Тихо переходим к fallback без лишнего спама
                    pass
                    
            else:
                logger.warning(f"Путь к компонентам ДС не найден: {ds_components_path}")
                
        except Exception as e:
            logger.error(f"Ошибка инициализации вкладки ДС: {e}")
        
        # Fallback - создаем базовую заглушку
        return self._create_ds_fallback()
    
    def _create_ds_fallback(self):
        """Создание fallback интерфейса для вкладки ДС"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Заголовок
        title_label = QtWidgets.QLabel("🚀 ДС - Дополнительные Сервисы")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #1976D2;
                padding: 20px;
                text-align: center;
            }
        """)
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Информация о PLEX ONE
        info_group = QtWidgets.QGroupBox("ℹ️ Информация о PLEX ONE")
        info_layout = QtWidgets.QFormLayout(info_group)
        
        info_layout.addRow("Название:", QtWidgets.QLabel("PLEX ONE"))
        info_layout.addRow("Символ:", QtWidgets.QLabel("PLEX"))
        info_layout.addRow("Decimals:", QtWidgets.QLabel("9"))
        info_layout.addRow("Total Supply:", QtWidgets.QLabel("12,600,000 PLEX"))
        info_layout.addRow("Контракт:", QtWidgets.QLabel("0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"))
        info_layout.addRow("Сеть:", QtWidgets.QLabel("Binance Smart Chain"))
        
        layout.addWidget(info_group)
        
        # Быстрые ссылки
        links_group = QtWidgets.QGroupBox("🔗 Быстрые ссылки")
        links_layout = QtWidgets.QVBoxLayout(links_group)
        
        # Кнопки для основных действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        # BSCScan
        bscscan_btn = QtWidgets.QPushButton("🔍 BSCScan")
        bscscan_btn.clicked.connect(lambda: self._open_url(
            "https://bscscan.com/address/0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
        ))
        buttons_layout.addWidget(bscscan_btn)
        
        # PancakeSwap
        pancake_btn = QtWidgets.QPushButton("🥞 PancakeSwap")
        pancake_btn.clicked.connect(lambda: self._open_url(
            "https://pancakeswap.finance/swap"
            "?outputCurrency=0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
            "&inputCurrency=0x55d398326f99059fF775485246999027B3197955"
        ))
        buttons_layout.addWidget(pancake_btn)
        
        # График
        chart_btn = QtWidgets.QPushButton("📈 График")
        chart_btn.clicked.connect(lambda: self._open_url(
            "https://www.geckoterminal.com/ru/bsc/pools/0x41d9650faf3341cbf8947fd8063a1fc88dbf1889"
        ))
        buttons_layout.addWidget(chart_btn)
        
        links_layout.addLayout(buttons_layout)
        layout.addWidget(links_group)
        
        # Статус загрузки компонентов
        status_group = QtWidgets.QGroupBox("📊 Статус")
        status_layout = QtWidgets.QVBoxLayout(status_group)
        
        status_text = QtWidgets.QLabel("""
<b>Статус компонентов ДС:</b><br>
• ⚠️ Полный функционал ДС не загружен<br>
• 🔗 Базовые ссылки доступны<br>
• 📁 Проверьте путь к MCP компонентам<br>
• 🔄 Перезапустите приложение после установки
        """)
        status_text.setWordWrap(True)
        status_layout.addWidget(status_text)
        
        # Кнопка перезагрузки компонентов
        reload_btn = QtWidgets.QPushButton("🔄 Перезагрузить компоненты ДС")
        reload_btn.clicked.connect(self._reload_ds_components)
        reload_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        status_layout.addWidget(reload_btn)
        
        # Кнопка настройки API ключа
        api_btn = QtWidgets.QPushButton("🔑 Настроить BSCScan API")
        api_btn.clicked.connect(self._setup_bscscan_api)
        api_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        status_layout.addWidget(api_btn)
        
        layout.addWidget(status_group)
        
        layout.addStretch()
        
        return w
    
    def _open_url(self, url):
        """Открытие URL в браузере"""
        import webbrowser
        webbrowser.open(url)
        logger.info(f"Открыт URL: {url}")
    
    def _reload_ds_components(self):
        """Перезагрузка компонентов ДС"""
        try:
            # Можно добавить логику перезагрузки
            QtWidgets.QMessageBox.information(
                self, 'Перезагрузка компонентов',
                'Для применения изменений перезапустите приложение',
                QtWidgets.QMessageBox.Ok
            )
        except Exception as e:
            logger.error(f"Ошибка перезагрузки компонентов ДС: {e}")
    
    def _setup_bscscan_api(self):
        """Настройка BSCScan API ключа"""
        text, ok = QtWidgets.QInputDialog.getText(
            self, 'BSCScan API Key',
            'Введите ваш BSCScan API ключ:\n(Получить можно на bscscan.com)',
            QtWidgets.QLineEdit.Normal
        )
        
        if ok and text.strip():
            api_key = text.strip()
            # Сохраняем в конфигурации (можно расширить Config класс)
            logger.info(f"BSCScan API ключ установлен: {api_key[:10]}...")
            
            QtWidgets.QMessageBox.information(
                self, 'API ключ установлен',
                'BSCScan API ключ сохранен.\nПерезапустите приложение для применения изменений.',
                QtWidgets.QMessageBox.Ok
            )
    
    def _update_ds_status(self, message):
        """Обновление статуса от вкладки ДС"""
        logger.info(f"ДС статус: {message}")
    
    def _handle_ds_transaction(self, tx_data):
        """Обработка готовой транзакции от ДС"""
        tx_type = tx_data.get('type', 'unknown')
        amount = tx_data.get('amount', 0)
        logger.info(f"ДС транзакция готова: {tx_type} на сумму {amount}")
        
        # Можно добавить в очередь отправки или обработать другим способом
    
    def _handle_balance_alert(self, address, balance):
        """Обработка алерта по балансу от ДС"""
        short_address = f"{address[:6]}...{address[-4:]}"
        logger.warning(f"ДС алерт: большой баланс у {short_address}: {balance} PLEX")
    
    def _handle_analysis_completed(self, results):
        """Обработка завершения анализа торговой активности"""
        all_traders = results.get('all_traders', {})
        holders = results.get('holders', {})
        analysis_type = results.get('analysis_type', 'unknown')
        
        logger.info(f"ДС анализ завершен ({analysis_type}): {len(all_traders)} трейдеров, {len(holders)} держателей")
        
        # Можно показать уведомление или обновить статистику
        if len(holders) > 0:
            logger.info(f"Найдено {len(holders)} держателей PLEX ONE")

    # --- Таб «Настройки» ---
    def _tab_settings(self):
        """Создание вкладки для настроек приложения"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа общих настроек
        general_group = QtWidgets.QGroupBox("Общие настройки")
        general_layout = QtWidgets.QFormLayout(general_group)
        
        # Цена газа
        self.gas_price_spin = QtWidgets.QDoubleSpinBox()
        self.gas_price_spin.setRange(0.1, 500)
        self.gas_price_spin.setDecimals(1)
        self.gas_price_spin.setSingleStep(0.1)
        self.gas_price_spin.setValue(float(self.cfg.get_gas_price()))
        self.gas_price_spin.valueChanged.connect(self._update_gas_price)
        general_layout.addRow("Цена газа (Gwei):", self.gas_price_spin)
        
        # Количество повторений серии
        self.repeat_count_spin = QtWidgets.QSpinBox()
        self.repeat_count_spin.setRange(1, 100)
        self.repeat_count_spin.setValue(self.cfg.get_repeat_count())
        self.repeat_count_spin.valueChanged.connect(self._update_repeat_count)
        general_layout.addRow("Количество повторений серии:", self.repeat_count_spin)
        
        # Очистка кэша
        btn_clear_cache = QtWidgets.QPushButton("Очистить кэш")
        btn_clear_cache.clicked.connect(self._clear_cache)
        general_layout.addRow("Кэширование:", btn_clear_cache)
        
        layout.addWidget(general_group)
        
        # Группа настроек блокчейна
        blockchain_group = QtWidgets.QGroupBox("Настройки блокчейна")
        blockchain_layout = QtWidgets.QFormLayout(blockchain_group)
        
        # Статус блокчейна
        blockchain_status = "Включен" if blockchain_enabled else "Отключен"
        blockchain_layout.addRow("Статус блокчейна:", QtWidgets.QLabel(blockchain_status))
        
        # Список RPC узлов
        rpc_list = QtWidgets.QListWidget()
        for node in RPC_NODES:
            rpc_list.addItem(node)
        blockchain_layout.addRow("RPC узлы:", rpc_list)
        
        # Тест подключения
        btn_test_conn = QtWidgets.QPushButton("Проверить подключение")
        btn_test_conn.clicked.connect(self._test_blockchain_connection)
        blockchain_layout.addRow("", btn_test_conn)
        
        # Информация о кошельке
        wallet_info = QtWidgets.QLabel("Не настроен")
        if blockchain_enabled and self.pk:
            try:
                address = Web3().eth.account.from_key(self.pk).address
                wallet_info.setText(address)
            except Exception:
                wallet_info.setText("Ошибка получения адреса")
        blockchain_layout.addRow("Текущий адрес:", wallet_info)
        
        layout.addWidget(blockchain_group)
        
        # Группа информации о приложении
        about_group = QtWidgets.QGroupBox("О приложении")
        about_layout = QtWidgets.QVBoxLayout(about_group)
        
        about_text = QtWidgets.QLabel(
            "WalletSender - приложение для анализа токенов BSC и распределения вознаграждений.\n"
            "Версия: 2.0\n"
            "© 2023-2025"
        )
        about_text.setWordWrap(True)
        about_layout.addWidget(about_text)
        
        layout.addWidget(about_group)
        
        return w
    
    def _update_gas_price(self, value):
        """Обновление цены газа в конфигурации"""
        self.cfg.set_gas_price(value)
        logger.info(f"Цена газа установлена на {value} Gwei")
        # Обновляем отображение в массовой рассылке, если виджет существует
        if hasattr(self, 'mass_gas_price_label'):
            self.mass_gas_price_label.setText(f"Цена газа: {value} Gwei")
    
    def _update_repeat_count(self, value):
        """Обновление количества повторений в конфигурации"""
        self.cfg.set_repeat_count(value)
        logger.info(f"Количество повторений установлено на {value}")
    
    def _clear_cache(self):
        """Очистка кэша"""
        count = cache.clear()
        logger.info(f"Очищено {count} элементов кэша")
    
    def _test_blockchain_connection(self):
        """Проверка подключения к блокчейну"""
        if not blockchain_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Блокчейн отключен',
                'Функции блокчейна отключены. Установите необходимые библиотеки.',
                QtWidgets.QMessageBox.Ok
            )
            return
            
        # Проверяем подключение в отдельном потоке
        threading.Thread(
            target=self._test_connection_thread,
            daemon=True
        ).start()
    
    def _test_connection_thread(self):
        """Поток для проверки подключения к блокчейну"""
        try:
            logger.info("Проверка подключения к блокчейну...")
            
            w3 = self.rpc.web3()
            
            if w3.is_connected():
                block = w3.eth.block_number
                used_node = self.rpc.get_healthy_node()
                logger.info(f"Подключение успешно! Узел: {used_node}, Блок: {block}")
                
                # Отправляем сигнал для обновления UI
                self.update_status_signal.emit(f"Подключение успешно! Узел: {used_node}, Блок: {block}")
                
            else:
                logger.warning("Не удалось подключиться к блокчейну")
                self.update_status_signal.emit("Не удалось подключиться к блокчейну")
        
        except Exception as e:
            logger.error(f"Ошибка при проверке подключения: {e}")
            self.update_status_signal.emit(f"Ошибка подключения: {e}")

    def _show_successful_connection(self, details):
        """Отображение успешного подключения"""
        QtWidgets.QMessageBox.information(
            self, 'Подключение успешно',
            f"Подключение к блокчейну BSC успешно установлено\n\n{details}",
            QtWidgets.QMessageBox.Ok
        )

    def _show_failed_connection(self, error_message=""):
        """Отображение ошибки подключения"""
        message = "Не удалось подключиться к блокчейну BSC"
        if error_message:
            message += f"\n\nПричина: {error_message}"
        
        QtWidgets.QMessageBox.warning(
            self, 'Ошибка подключения',
            message,
            QtWidgets.QMessageBox.Ok
        )
                
    def _show_connection_result(self, success, message):
        """Отображение результата проверки подключения"""
        if success:
            QtWidgets.QMessageBox.information(
                self, 'Подключение успешно',
                message,
                QtWidgets.QMessageBox.Ok
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка подключения',
                message,
                QtWidgets.QMessageBox.Ok
            )
            
    # --- Вкладка «Массовая рассылка» ---
    def _tab_mass_distribution(self):
        """Создание вкладки для массовой рассылки токенов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)

        # Блок информации о кошельке отправителя (адрес + балансы)
        wallet_group = QtWidgets.QGroupBox("Кошелек отправителя")
        wallet_layout = QtWidgets.QHBoxLayout(wallet_group)
        self.mass_sender_address_label = QtWidgets.QLabel("Адрес: -")
        self.mass_balance_plex_label = QtWidgets.QLabel("PLEX: -")
        self.mass_balance_bnb_label = QtWidgets.QLabel("BNB: -")
        self.mass_balance_usdt_label = QtWidgets.QLabel("USDT: -")
        wallet_layout.addWidget(self.mass_sender_address_label)
        wallet_layout.addWidget(self.mass_balance_plex_label)
        wallet_layout.addWidget(self.mass_balance_bnb_label)
        wallet_layout.addWidget(self.mass_balance_usdt_label)
        self.mass_refresh_wallet_btn = QtWidgets.QPushButton("↻")
        self.mass_refresh_wallet_btn.setFixedWidth(30)
        self.mass_refresh_wallet_btn.setToolTip("Обновить балансы")
        self.mass_refresh_wallet_btn.clicked.connect(self._mass_update_wallet_info)
        wallet_layout.addWidget(self.mass_refresh_wallet_btn)
        layout.addWidget(wallet_group)
        QtCore.QTimer.singleShot(300, self._mass_update_wallet_info)

        # Группа импорта адресов
        import_group = QtWidgets.QGroupBox("Импорт адресов")
        import_layout = QtWidgets.QVBoxLayout(import_group)

        # Кнопки импорта
        import_buttons_layout = QtWidgets.QHBoxLayout()
        self.mass_import_clipboard_btn = QtWidgets.QPushButton("Из буфера обмена")
        self.mass_import_clipboard_btn.clicked.connect(self._mass_import_from_clipboard)
        import_buttons_layout.addWidget(self.mass_import_clipboard_btn)
        self.mass_import_excel_btn = QtWidgets.QPushButton("Из Excel")
        self.mass_import_excel_btn.clicked.connect(self._mass_import_from_excel)
        import_buttons_layout.addWidget(self.mass_import_excel_btn)
        self.mass_clear_addresses_btn = QtWidgets.QPushButton("Очистить список")
        self.mass_clear_addresses_btn.clicked.connect(self._mass_clear_addresses)
        import_buttons_layout.addWidget(self.mass_clear_addresses_btn)
        import_layout.addLayout(import_buttons_layout)

        # Таблица адресов
        self.mass_table = QtWidgets.QTableWidget(0, 4)
        self.mass_table.setHorizontalHeaderLabels(['№', 'Адрес', 'Статус', 'Прогресс'])
        header = self.mass_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.mass_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        import_layout.addWidget(self.mass_table)
        # Информация
        self.mass_import_info = QtWidgets.QLabel("Адресов загружено: 0")
        import_layout.addWidget(self.mass_import_info)
        layout.addWidget(import_group)

        # Настройки
        settings_group = QtWidgets.QGroupBox("Настройки рассылки")
        settings_layout = QtWidgets.QFormLayout(settings_group)
        token_layout = QtWidgets.QHBoxLayout()
        self.mass_token_group = QtWidgets.QButtonGroup(w)
        self.mass_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.mass_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.mass_bnb_radio = QtWidgets.QRadioButton('BNB')
        self.mass_custom_radio = QtWidgets.QRadioButton('Другой токен')
        self.mass_plex_radio.setChecked(True)
        for btn in [self.mass_plex_radio, self.mass_usdt_radio, self.mass_bnb_radio, self.mass_custom_radio]:
            self.mass_token_group.addButton(btn)
            token_layout.addWidget(btn)
        settings_layout.addRow("Токен:", token_layout)
        self.mass_custom_token_address = QtWidgets.QLineEdit()
        self.mass_custom_token_address.setPlaceholderText("Адрес контракта токена (0x...)")
        self.mass_custom_token_address.setEnabled(False)
        self.mass_custom_radio.toggled.connect(self.mass_custom_token_address.setEnabled)
        settings_layout.addRow("Адрес токена:", self.mass_custom_token_address)
        self.mass_amount = QtWidgets.QDoubleSpinBox()
        self.mass_amount.setRange(0.00000001, 1000000)
        self.mass_amount.setDecimals(8)
        self.mass_amount.setValue(0.05)
        settings_layout.addRow("Сумма на адрес:", self.mass_amount)
        self.mass_interval = QtWidgets.QSpinBox()
        self.mass_interval.setRange(1, 600)
        self.mass_interval.setValue(5)
        settings_layout.addRow("Интервал (сек):", self.mass_interval)
        self.mass_cycles = QtWidgets.QSpinBox()
        self.mass_cycles.setRange(1, 100)
        self.mass_cycles.setValue(10)
        settings_layout.addRow("Количество циклов:", self.mass_cycles)
        layout.addWidget(settings_group)

        # Управление
        control_group = QtWidgets.QGroupBox("Управление рассылкой")
        control_layout = QtWidgets.QVBoxLayout(control_group)
        buttons = QtWidgets.QHBoxLayout()
        self.mass_start_btn = QtWidgets.QPushButton("Начать рассылку")
        self.mass_start_btn.clicked.connect(self._mass_start_distribution)
        buttons.addWidget(self.mass_start_btn)
        self.mass_pause_btn = QtWidgets.QPushButton("Пауза")
        self.mass_pause_btn.setEnabled(False)
        self.mass_pause_btn.clicked.connect(self._mass_pause_distribution)
        buttons.addWidget(self.mass_pause_btn)
        self.mass_resume_btn = QtWidgets.QPushButton("Продолжить")
        self.mass_resume_btn.setEnabled(False)
        self.mass_resume_btn.clicked.connect(self._mass_resume_distribution)
        buttons.addWidget(self.mass_resume_btn)
        self.mass_stop_btn = QtWidgets.QPushButton("Остановить")
        self.mass_stop_btn.setEnabled(False)
        self.mass_stop_btn.clicked.connect(self._mass_stop_distribution)
        buttons.addWidget(self.mass_stop_btn)
        control_layout.addLayout(buttons)
        self.mass_progress = QtWidgets.QProgressBar()
        control_layout.addWidget(self.mass_progress)
        self.mass_status_label = QtWidgets.QLabel("Готов к рассылке")
        control_layout.addWidget(self.mass_status_label)
        info_layout = QtWidgets.QHBoxLayout()
        self.mass_current_cycle_label = QtWidgets.QLabel("Цикл: 0/0")
        self.mass_sent_count_label = QtWidgets.QLabel("Отправлено: 0")
        self.mass_errors_label = QtWidgets.QLabel("Ошибок: 0")
        for wgt in [self.mass_current_cycle_label, self.mass_sent_count_label, self.mass_errors_label]:
            info_layout.addWidget(wgt)
        control_layout.addLayout(info_layout)
        layout.addWidget(control_group)

        stats_group = QtWidgets.QGroupBox("Статистика")
        stats_layout = QtWidgets.QFormLayout(stats_group)
        self.mass_total_sent_label = QtWidgets.QLabel("0")
        self.mass_total_amount_label = QtWidgets.QLabel("0")
        self.mass_gas_spent_label = QtWidgets.QLabel("0 BNB")
        self.mass_gas_price_label = QtWidgets.QLabel(f"Цена газа: {self.cfg.get_gas_price()} Gwei")
        stats_layout.addRow("Всего транзакций:", self.mass_total_sent_label)
        stats_layout.addRow("Всего отправлено:", self.mass_total_amount_label)
        stats_layout.addRow("Потрачено на газ:", self.mass_gas_spent_label)
        stats_layout.addRow("", self.mass_gas_price_label)
        layout.addWidget(stats_group)

        self.mass_distribution_active = False
        self.mass_distribution_paused = False
        self.mass_distribution_thread = None
        self.mass_addresses = []
        self.mass_current_distribution_id = None
        return w
    
    def _mass_import_from_clipboard(self):
        """Импорт адресов из буфера обмена"""
        clipboard = QtWidgets.QApplication.clipboard()
        text = clipboard.text()
        
        if not text:
            logger.warning("Буфер обмена пуст")
            return
        
        # Разделяем текст по различным разделителям
        import re
        addresses = re.split(r'[\s,;\n]+', text)
        
        # Фильтруем и валидируем адреса
        valid_addresses = []
        for addr in addresses:
            addr = addr.strip()
            if addr and blockchain_enabled:
                try:
                    if Web3.is_address(addr):
                        # Приводим к checksum формату
                        checksum_addr = Web3.to_checksum_address(addr)
                        if checksum_addr not in valid_addresses:
                            valid_addresses.append(checksum_addr)
                except Exception as e:
                    logger.warning(f"Некорректный адрес: {addr}, ошибка: {e}")
            elif addr and not blockchain_enabled:
                # Если блокчейн отключен, принимаем любые строки похожие на адреса
                if addr.startswith('0x') and len(addr) == 42:
                    if addr not in valid_addresses:
                        valid_addresses.append(addr)
        
        if valid_addresses:
            self._mass_add_addresses(valid_addresses)
            logger.info(f"Импортировано {len(valid_addresses)} адресов из буфера обмена")
        else:
            logger.warning("Не найдено корректных адресов в буфере обмена")
    
    def _mass_import_from_excel(self):
        """Импорт адресов из Excel файла"""
        if not excel_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Excel не поддерживается',
                'Для импорта из Excel необходимо установить библиотеку openpyxl:\npip install openpyxl',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Диалог выбора файла
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Выберите Excel файл', '', 'Excel Files (*.xlsx *.xls)'
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            # Читаем первый лист
            sheet = workbook.active
            
            valid_addresses = []
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        addr = cell.strip()
                        if addr and blockchain_enabled:
                            try:
                                if Web3.is_address(addr):
                                    checksum_addr = Web3.to_checksum_address(addr)
                                    if checksum_addr not in valid_addresses:
                                        valid_addresses.append(checksum_addr)
                            except Exception:
                                pass
                        elif addr and not blockchain_enabled:
                            if addr.startswith('0x') and len(addr) == 42:
                                if addr not in valid_addresses:
                                    valid_addresses.append(addr)
            
            workbook.close()
            
            if valid_addresses:
                self._mass_add_addresses(valid_addresses)
                logger.info(f"Импортировано {len(valid_addresses)} адресов из Excel файла")
            else:
                logger.warning("Не найдено корректных адресов в Excel файле")
        
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла: {e}")
            QtWidgets.QMessageBox.critical(
                self, 'Ошибка',
                f'Не удалось прочитать Excel файл:\n{str(e)}',
                QtWidgets.QMessageBox.Ok
            )
    
    def _mass_add_addresses(self, addresses):
        """Добавление адресов в таблицу"""
        # Очищаем текущий список
        self.mass_table.setRowCount(0)
        self.mass_addresses = []
        # Добавляем адреса в таблицу
        for i, addr in enumerate(addresses):
            if not addr:
                continue
            row = self.mass_table.rowCount()
            self.mass_table.insertRow(row)
            # Номер
            self.mass_table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(i + 1)))
            # Адрес
            self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem(addr))
            # Статус
            status_item = QtWidgets.QTableWidgetItem("Ожидание")
            self.mass_table.setItem(row, 2, status_item)
            # Прогресс
            self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem("0/0"))
        # Обновляем список и инфо
        self.mass_addresses = [a for a in addresses if a]
        if hasattr(self, 'mass_import_info'):
            self.mass_import_info.setText(f"Адресов загружено: {len(self.mass_addresses)}")
        logger.info(f"Импортировано {len(self.mass_addresses)} адресов в таблицу")
    
    def _mass_clear_addresses(self):
        """Очистка списка адресов"""
        self.mass_table.setRowCount(0)
        self.mass_addresses = []
        self.mass_import_info.setText("Адресов загружено: 0")
        logger.info("Список адресов очищен")
    
    def _mass_start_distribution(self):
        """Запуск массовой рассылки"""
        if not self.mass_addresses:
            QtWidgets.QMessageBox.warning(
                self, 'Нет адресов',
                'Сначала импортируйте адреса для рассылки',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        if not blockchain_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Блокчейн отключен',
                'Функции блокчейна отключены. Проверьте настройки.',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        if not self.pk:
            QtWidgets.QMessageBox.warning(
                self, 'Кошелек не настроен',
                'Сначала настройте кошелек в настройках',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Получаем параметры рассылки
        token_type = None
        token_address = None
        
        if self.mass_plex_radio.isChecked():
            token_type = 'PLEX ONE'
            token_address = PLEX_CONTRACT
        elif self.mass_usdt_radio.isChecked():
            token_type = 'USDT'
            token_address = USDT_CONTRACT
        elif self.mass_bnb_radio.isChecked():
            token_type = 'BNB'
            token_address = None  # Нативный токен
        elif self.mass_custom_radio.isChecked():
            custom_addr = self.mass_custom_token_address.text().strip()
            if not custom_addr or not Web3.is_address(custom_addr):
                QtWidgets.QMessageBox.warning(
                    self, 'Некорректный адрес',
                    'Введите корректный адрес токена',
                    QtWidgets.QMessageBox.Ok
                )
                return
            token_type = 'CUSTOM'
            token_address = Web3.to_checksum_address(custom_addr)
        
        amount = self.mass_amount.value()
        interval = self.mass_interval.value()
        cycles = self.mass_cycles.value()
        
        # Подтверждение
        total_transactions = len(self.mass_addresses) * cycles
        total_amount = amount * total_transactions
        
        msg = f"Вы собираетесь отправить:\n\n"
        msg += f"Токен: {token_type}\n"
        msg += f"Сумма на адрес: {amount}\n"
        msg += f"Адресов: {len(self.mass_addresses)}\n"
        msg += f"Циклов: {cycles}\n"
        msg += f"Всего транзакций: {total_transactions}\n"
        msg += f"Общая сумма: {total_amount}\n\n"
        msg += f"Продолжить?"
        
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение рассылки',
            msg,
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
        
        # Создаем запись в БД
        self.mass_current_distribution_id = add_mass_distribution(
            name=f"Рассылка {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            token_address=token_address or '0x0',
            token_symbol=token_type,
            amount_per_tx=amount,
            total_addresses=len(self.mass_addresses),
            total_cycles=cycles,
            interval_seconds=interval
        )
        
        # Запускаем рассылку
        self.mass_distribution_active = True
        self.mass_distribution_paused = False
        
        # Обновляем UI
        self.mass_start_btn.setEnabled(False)
        self.mass_pause_btn.setEnabled(True)
        self.mass_stop_btn.setEnabled(True)
        self.mass_import_clipboard_btn.setEnabled(False)
        self.mass_import_excel_btn.setEnabled(False)
        self.mass_clear_addresses_btn.setEnabled(False)
        
        # Запускаем в отдельном потоке
        self.mass_distribution_thread = threading.Thread(
            target=self._mass_distribution_worker,
            args=(token_type, token_address, amount, interval, cycles),
            daemon=False  # Изменено: убираем daemon=True чтобы приложение не завершалось
        )
        self.mass_distribution_thread.start()
        
        logger.info(f"Запущена массовая рассылка: {token_type}, {amount} x {cycles} циклов")
    
    def _mass_distribution_worker(self, token_type: str, token_address: str, amount: float, interval: int, cycles: int) -> None:
        """Рабочий поток для массовой рассылки"""
        logger.info("🔄 Начат рабочий поток массовой рассылки")
        try:
            total_sent = 0
            total_errors = 0
            total_gas_spent: float = 0.0
            
            for cycle in range(1, cycles + 1):
                if not self.mass_distribution_active:
                    logger.info(f"⏹️ Массовая рассылка остановлена на цикле {cycle}")
                    break
                
                # Обновляем информацию о цикле
                self.update_status_signal.emit(f"Цикл {cycle}/{cycles}")
                # Безопасно обновляем метку в GUI-потоке
                self.update_mass_cycle_label.emit(f"Цикл: {cycle}/{cycles}")
                
                for i, address in enumerate(self.mass_addresses):
                    if not self.mass_distribution_active:
                        break
                    
                    # Проверка паузы
                    while self.mass_distribution_paused and self.mass_distribution_active:
                        time.sleep(0.5)
                    
                    if not self.mass_distribution_active:
                        break
                    
                    # Обновляем статус адреса
                    self._update_address_status(i, "Отправка...", QtGui.QColor('#FFA500'))
                    
                    try:
                        # Отправляем транзакцию
                        if token_type == 'BNB':
                            result = self._send_bnb(address, amount)
                            tx_hash = str(result.get('tx_hash', ''))
                            gas_used = float(result.get('gas_used', 0.0))
                        else:
                            result = self._send_token(address, amount, token_address)
                            tx_hash = str(result.get('tx_hash', ''))
                            gas_used = float(result.get('gas_used', 0.0))
                        
                        if tx_hash and self.mass_current_distribution_id is not None:
                            total_sent += 1
                            total_gas_spent += gas_used
                            
                            # Сохраняем в БД
                            add_mass_distribution_item(
                                self.mass_current_distribution_id,
                                address,
                                cycle,
                                tx_hash,
                                'success'
                            )
                            
                            # Обновляем статус
                            self._update_address_status(i, "Успешно", QtGui.QColor('#00FF00'))
                            # Отправляем прогресс через сигнал (без прямого вызова UI метода)
                            self.address_progress_signal.emit(i, cycle, cycles)
                            
                            logger.info(f"Отправлено {amount} {token_type} на {address}, цикл {cycle}")
                        else:
                            raise Exception("Не удалось получить хэш транзакции")
                    
                    except Exception as e:
                        total_errors += 1
                        
                        # Сохраняем ошибку в БД
                        if self.mass_current_distribution_id is not None:
                            add_mass_distribution_item(
                                self.mass_current_distribution_id,
                                address,
                                cycle,
                                '',
                                'error',
                                str(e)
                            )
                        
                        # Обновляем статус
                        self._update_address_status(i, f"Ошибка: {str(e)[:30]}...", QtGui.QColor('#FF0000'))
                        
                        logger.error(f"Ошибка отправки на {address}: {e}")
                    
                    # Обновляем общую статистику
                    self.mass_stats_signal.emit(total_sent, total_errors, total_gas_spent, amount * total_sent)
                    
                    # Прогресс
                    progress = int(((cycle - 1) * len(self.mass_addresses) + i + 1) / (cycles * len(self.mass_addresses)) * 100)
                    self.update_progress_signal.emit("mass", progress)
                    
                    # Интервал между отправками
                    if i < len(self.mass_addresses) - 1 and self.mass_distribution_active:
                        time.sleep(interval)
                
                # Интервал между циклами
                if cycle < cycles and self.mass_distribution_active:
                    logger.info(f"Цикл {cycle} завершен, ожидание перед следующим циклом...")
                    time.sleep(interval * 2)  # Двойной интервал между циклами
            
            # Завершение рассылки
            if self.mass_current_distribution_id:
                status = 'completed' if self.mass_distribution_active else 'cancelled'
                update_mass_distribution_status(self.mass_current_distribution_id, status)
            
            logger.info(f"✅ Массовая рассылка завершена. Отправлено: {total_sent}, ошибок: {total_errors}")
        
        except Exception as e:
            logger.error(f"🔴 Критическая ошибка в массовой рассылке: {e}")
            import traceback
            logger.error(f"Полный traceback: {traceback.format_exc()}")
        
        finally:
            logger.info("🔧 Восстанавливаем UI после массовой рассылки")
            # Сбрасываем флаг состояния
            self.mass_distribution_active = False
            # Инициируем завершение рассылки в GUI-потоке
            try:
                self.mass_distribution_finished_signal.emit()
                logger.info("✅ UI восстановление инициировано")
            except Exception as e:
                logger.error(f"🔴 Ошибка при отправке сигнала завершения UI: {e}")
    
    @QtCore.pyqtSlot()
    def _mass_distribution_finished(self):
        """Завершение массовой рассылки - обновление UI"""
        self.mass_start_btn.setEnabled(True)
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(False)
        self.mass_stop_btn.setEnabled(False)
        self.mass_import_clipboard_btn.setEnabled(True)
        self.mass_import_excel_btn.setEnabled(True)
        self.mass_clear_addresses_btn.setEnabled(True)
        self.mass_status_label.setText("Рассылка завершена")
    
    def _update_address_status(self, row: int, status: str, color: Optional[str] = None) -> None:
        """Обновление статуса адреса в таблице"""
        # Используем сигнал для обновления UI из потока
        if '✓' in status or 'Успешно' in status:
            self.update_address_status.emit(row, "✓ Успешно")
        elif '✗' in status or 'Ошибка' in status:
            self.update_address_status.emit(row, "✗ Ошибка")
        elif '⟳' in status or 'Отправка' in status:
            self.update_address_status.emit(row, "⟳ Отправка...")
        else:
            self.update_address_status.emit(row, status)
    
    def _update_address_progress(self, row: int, current: int, total: int) -> None:
        """(DEPRECATED) Оставлено для совместимости; теперь используем address_progress_signal напрямую"""
        self.address_progress_signal.emit(row, current, total)
    
    @QtCore.pyqtSlot(int, int, int)
    def _update_progress_item(self, row: int, current: int, total: int) -> None:
        """Слот для обновления прогресса в основном потоке"""
        try:
            if row < self.mass_table.rowCount():
                progress_item = QtWidgets.QTableWidgetItem(f"{current}/{total}")
                self.mass_table.setItem(row, 3, progress_item)
        except Exception as e:
            logger.error(f"Ошибка обновления элемента прогресса: {e}")
    
    @QtCore.pyqtSlot(int, int, float, float)
    def _update_mass_statistics(self, sent: int, errors: int, gas_spent: float, total_amount: float) -> None:
        """Слот: обновление статистики рассылки (GUI-поток)"""
        try:
            if getattr(self, 'mass_sent_count_label', None):
                self.mass_sent_count_label.setText(f"Отправлено: {sent}")
            if getattr(self, 'mass_errors_label', None):
                self.mass_errors_label.setText(f"Ошибок: {errors}")
            if getattr(self, 'mass_total_sent_label', None):
                self.mass_total_sent_label.setText(str(sent))
            if getattr(self, 'mass_total_amount_label', None):
                self.mass_total_amount_label.setText(f"{total_amount:.8f}")
            if getattr(self, 'mass_gas_spent_label', None):
                self.mass_gas_spent_label.setText(f"{gas_spent:.8f} BNB")
        except Exception as e:
            logger.error(f"Ошибка обновления статистики: {e}")
    
    def _mass_pause_distribution(self):
        """Приостановка массовой рассылки"""
        self.mass_distribution_paused = True
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(True)
        self.mass_status_label.setText("Рассылка приостановлена")
        logger.info("Массовая рассылка приостановлена")
    
    def _mass_resume_distribution(self):
        """Возобновление массовой рассылки"""
        self.mass_distribution_paused = False
        self.mass_pause_btn.setEnabled(True)
        self.mass_resume_btn.setEnabled(False)
        self.mass_status_label.setText("Рассылка продолжается")
        logger.info("Массовая рассылка возобновлена")
    
    def _mass_stop_distribution(self):
        """Остановка массовой рассылки"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите остановить рассылку?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.mass_distribution_active = False
            self.mass_distribution_paused = False
            self.mass_status_label.setText("Рассылка остановлена")
            logger.info("Массовая рассылка остановлена пользователем")
    
    def _send_bnb(self, to_address: str, amount: float) -> Dict[str, Union[str, float, bool]]:
        """Отправка BNB"""
        try:
            w3 = self.rpc.web3()
            account = w3.eth.account.from_key(self.pk)
            
            # Получаем nonce
            nonce = w3.eth.get_transaction_count(account.address)
            
            # Подготавливаем транзакцию
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            value = w3.to_wei(amount, 'ether')
            
            tx = {
                'nonce': nonce,
                'to': to_address,
                'value': value,
                'gas': 21000,
                'gasPrice': gas_price,
                'chainId': 56  # BSC mainnet
            }
            
            # Подписываем и отправляем
            signed = account.sign_transaction(tx)
            tx_hash = send_raw_tx(w3, signed.rawTransaction)
            
            # Ждем подтверждения
            receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
            
            gas_used = float(receipt['gasUsed']) * gas_price / 10**18
            
            return {
                'tx_hash': tx_hash.hex(),
                'gas_used': gas_used,
                'success': True
            }
        
        except Exception as e:
            logger.error(f"Ошибка отправки BNB: {e}")
            raise
    
    def _send_token(self, to_address: str, amount: float, token_address: str) -> Dict[str, Union[str, float, bool]]:
        """Отправка токена"""
        try:
            w3 = self.rpc.web3()
            account = w3.eth.account.from_key(self.pk)
            
            # Конвертируем адреса в checksum формат
            to_address = w3.to_checksum_address(to_address)
            token_address = w3.to_checksum_address(token_address)
            
            # Получаем контракт токена
            token_contract = w3.eth.contract(address=token_address, abi=ERC20_ABI)
            
            # Получаем decimals
            decimals = token_contract.functions.decimals().call()
            
            # Конвертируем сумму
            amount_wei = int(amount * 10**decimals)
            
            # Получаем nonce
            nonce = w3.eth.get_transaction_count(account.address)
            
            # Подготавливаем транзакцию
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            # Создаем транзакцию transfer
            tx = token_contract.functions.transfer(to_address, amount_wei).build_transaction({
                'from': account.address,
                'nonce': nonce,
                'gasPrice': gas_price,
                'gas': 100000  # Примерная оценка газа
            })
            
            # Подписываем и отправляем
            signed = account.sign_transaction(tx)
            tx_hash = send_raw_tx(w3, signed.rawTransaction)
            
            # Ждем подтверждения
            receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
            
            gas_used = float(receipt['gasUsed']) * gas_price / 10**18
            
            return {
                'tx_hash': tx_hash.hex(),
                'gas_used': gas_used,
                'success': True
            }
        
        except Exception as e:
            logger.error(f"Ошибка отправки токена: {e}")
            raise

# ======== #MCP:MAIN ========
# Entry point
if __name__ == "__main__":
    """Точка входа в приложение
    
    TODO:MCP - Добавить обработку аргументов командной строки
    REFACTOR:MCP - Выделить инициализацию в отдельную функцию
    """
    app = QtWidgets.QApplication(sys.argv)
    
    # Регистрация Qt мета-типов для решения предупреждений при работе с сигналами
    # Совместимость с PyQt5/PyQt6
    try:
        QtCore.qRegisterMetaType('QVector<int>')
        QtCore.qRegisterMetaType('QTextCursor')
    except AttributeError:
        # В PyQt5 qRegisterMetaType может отсутствовать - это не критично
        pass
    
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
