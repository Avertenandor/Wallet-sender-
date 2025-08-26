#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WalletSender.py — Расширенный скрипт для анализа токенов BSC и распределения вознаграждений с GUI.

🧭 MCP Navigation Tags:
- #MCP:IMPORTS     - Импорты и настройки (строки 1-100)
- #MCP:CONSTANTS   - Константы и конфигурация (строки 70-110)  
- #MCP:UTILS       - Утилиты и декораторы (строки 110-150)
- #MCP:CONFIG      - Управление конфигурацией (строки 150-250)
- #MCP:RPC         - RPC пул соединений (строки 250-350)
- #MCP:CACHE       - Система кэширования (строки 350-380)
- #MCP:DATABASE    - SQLite операции (строки 380-700)
- #MCP:API         - BscScan API интеграция (строки 700-900)
- #MCP:GUI         - PyQt5 интерфейс (строки 863-4735)
- #MCP:MAIN        - Точка входа (строки 4730-4735)

🏷️ Development Control Tags:
- #TODO:MCP        - Задачи для MCP интеграции
- #REFACTOR:MCP    - Места для рефакторинга  
- #OPTIMIZE:MCP    - Возможности оптимизации
- #SECURITY:MCP    - Проблемы безопасности
- #DOCS:MCP        - Требуется документация
"""

# ======== #MCP:IMPORTS ========
# Импорт стандартных библиотек для работы с файлами, потоками и т.д.
import sys
import os
import json
import time
import threading
import random
import sqlite3
import functools
import logging
from datetime import datetime, timedelta
from pathlib import Path

# Импорт внешних библиотек для работы с API, криптографией и GUI
import requests
from mnemonic import Mnemonic
from cryptography.fernet import Fernet
from PyQt5 import QtWidgets, QtCore, QtGui
import qdarkstyle
from eth_utils.exceptions import ValidationError
import csv  # Для работы с CSV файлами

# Совместимость с разными версиями Web3.py
try:
    from web3.middleware import geth_poa_middleware
except ImportError:
    try:
        from web3.middleware.geth_poa import geth_poa_middleware  # type: ignore
    except ImportError:
        # Для самых новых версий Web3.py
        from web3.middleware import ExtraDataToPOAMiddleware as geth_poa_middleware

# Настройка логирования для отслеживания работы программы
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"wallet_sender_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("WalletSender")

# Импорт для работы с Excel файлами
try:
    import openpyxl
    excel_enabled = True
except ImportError:
    logger.warning("openpyxl не установлен - импорт из Excel будет недоступен")
    excel_enabled = False

# Попытка импортировать web3 и eth_account для работы с блокчейном
blockchain_enabled = True
try:
    from web3 import Web3, HTTPProvider, middleware
    # Совместимость с разными версиями Web3.py для газовых стратегий
    try:
        from web3.gas_strategies.time_based import medium_gas_price_strategy
    except ImportError:
        try:
            from web3.gas_strategies.rpc import rpc_gas_price_strategy as medium_gas_price_strategy
        except ImportError:
            medium_gas_price_strategy = None
            logger.warning("Не найдена стратегия цены газа, будет использована базовая")
    
    # geth_poa_middleware уже импортирован выше с проверкой совместимости
    from eth_account import Account
    Account.enable_unaudited_hdwallet_features()
except ImportError as e:
    logger.warning(f"web3 или eth_account не установлены—функции блокчейна отключены. Ошибка: {e}")
    blockchain_enabled = False
    medium_gas_price_strategy = None

# ======== #MCP:CONSTANTS ========
# Список RPC-узлов для подключения к блокчейну BSC
RPC_NODES = [
    "https://bsc-dataseed1.binance.org/",
    "https://bsc-dataseed.nariox.org/",
    "https://bsc-dataseed.defibit.io/",
    "https://bsc.publicnode.com"
]
BSCSCAN_URL   = "https://api.bscscan.com/api"
BSCSCAN_KEY   = "ARA9FYMNCIZHTB2PPBSWF686GID9F99P41"  # SECURITY:MCP - Лучше хранить в переменных окружения
PLEX_CONTRACT = "0xdf179b6cadbc61ffd86a3d2e55f6d6e083ade6c1"  # Адрес контракта токена PLEX ONE
USDT_CONTRACT = "0x55d398326f99059ff775485246999027b3197955"  # Адрес контракта USDT на BSC
CONFIG_FILE   = "config.json"  # Файл для хранения конфигурации
DB_FILE       = "history.sqlite"  # База данных для истории транзакций
CACHE_TTL     = timedelta(minutes=10)  # Время жизни кэша
REWARDS_CONFIG_DIR = Path("rewards_configs")  # Директория для сохранения конфигураций наград
REWARDS_CONFIG_DIR.mkdir(exist_ok=True)

# Минимальный ABI для взаимодействия с токенами стандарта ERC20
ERC20_ABI = [
    { "constant": False, "inputs":[{"name":"_to","type":"address"},{"name":"_value","type":"uint256"}],
      "name":"transfer","outputs":[{"name":"","type":"bool"}],"type":"function" },  # Функция перевода токенов
    { "constant": True, "inputs":[{"name":"_owner","type":"address"}],
      "name":"balanceOf","outputs":[{"name":"balance","type":"uint256"}],"type":"function" },  # Проверка баланса
    { "constant": True, "inputs":[],
      "name":"decimals","outputs":[{"name":"","type":"uint8"}],"type":"function" },  # Количество десятичных знаков
    { "constant": True, "inputs":[],
      "name":"symbol","outputs":[{"name":"","type":"string"}],"type":"function" }  # Символ токена
]

# ======== #MCP:UTILS ========
def retry(times=3, delay=0.5):
    """Декоратор: повторять вызов при исключении.
    
    TODO:MCP - Добавить экспоненциальный backoff
    OPTIMIZE:MCP - Улучшить обработку различных типов исключений
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(times):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    logger.warning(f"Попытка {attempt+1}/{times} не удалась: {e}")
                    time.sleep(delay * (attempt + 1))  # Увеличение задержки с каждой попыткой
            logger.error(f"Все {times} попыток не удались. Последняя ошибка: {last_exception}")
            raise last_exception
        return wrapper
    return decorator

# ======== #MCP:CONFIG ========
class Config:
    """Шифрование/хранение приватного ключа и seed-фразы.
    
    SECURITY:MCP - Проверить стойкость шифрования
    REFACTOR:MCP - Добавить валидацию конфигурации
    DOCS:MCP - Документировать формат конфигурации
    """
    def __init__(self):
        self.config_path = Path(CONFIG_FILE)
        if self.config_path.exists():
            try:
                self.data = json.loads(self.config_path.read_text(encoding='utf-8'))
            except json.JSONDecodeError:
                logger.error(f"Неверный JSON в {CONFIG_FILE}, создаю новую конфигурацию")
                self.data = {}
        else:
            self.data = {}
        
        # Генерация ключа шифрования, если он отсутствует
        if 'fernet_key' not in self.data:
            self.data['fernet_key'] = Fernet.generate_key().decode()
            self._save()
        self.fernet = Fernet(self.data['fernet_key'].encode())

    def _save(self):
        """Сохранение конфигурации в файл"""
        try:
            self.config_path.write_text(
                json.dumps(self.data, ensure_ascii=False, indent=2),
                encoding='utf-8'
            )
            logger.debug("Конфигурация сохранена")
        except Exception as e:
            logger.error(f"Не удалось сохранить конфигурацию: {e}")

    def set_key(self, raw):
        """Сохранение зашифрованного приватного ключа"""
        self.data['pk'] = self.fernet.encrypt(raw.encode()).decode()
        self._save()

    def get_key(self):
        """Получение расшифрованного приватного ключа"""
        return (self.fernet.decrypt(self.data['pk'].encode()).decode()
                if 'pk' in self.data else None)

    def set_mnemonic(self, m):
        """Сохранение зашифрованной мнемонической фразы"""
        self.data['mn'] = self.fernet.encrypt(m.encode()).decode()
        self._save()

    def get_mnemonic(self):
        """Получение расшифрованной мнемонической фразы"""
        return (self.fernet.decrypt(self.data['mn'].encode()).decode()
                if 'mn' in self.data else None)
    
    def set_gas_price(self, price_gwei):
        """Сохранение предпочтительной цены газа в Gwei"""
        self.data['gas_price'] = price_gwei
        self._save()
    
    def get_gas_price(self):
        """Получение предпочтительной цены газа, по умолчанию 5 Gwei"""
        return self.data.get('gas_price', 5)
        
    def get_repeat_count(self):
        """Получение количества повторений серии отправок"""
        return self.data.get('repeat_count', 1)
        
    def set_repeat_count(self, count):
        """Сохранение количества повторений серии отправок"""
        self.data['repeat_count'] = count
        self._save()
        
    def get_reward_per_tx(self):
        """Получение режима награждения за каждую транзакцию"""
        return self.data.get('reward_per_tx', False)
        
    def set_reward_per_tx(self, value):
        """Сохранение режима награждения за каждую транзакцию"""
        self.data['reward_per_tx'] = value
        self._save()

# ======== #MCP:RPC ========
class RPCPool:
    """Отказоустойчивый Web3-подключатель.
    
    OPTIMIZE:MCP - Добавить метрики производительности узлов
    REFACTOR:MCP - Выделить в отдельный модуль
    TODO:MCP - Добавить автоматическое обновление списка узлов
    """
    def __init__(self, nodes):
        self.nodes = nodes
        self.idx = 0
        self.node_health = {node: True for node in nodes}  # Статус здоровья узла
        self.last_check = {node: datetime.min for node in nodes}  # Время последней проверки
        self.check_interval = timedelta(minutes=5)  # Интервал между проверками

    def _check_node(self, node_url):
        """Проверка, работает ли узел и отвечает ли он"""
        try:
            w3 = Web3(HTTPProvider(node_url, request_kwargs={'timeout': 10}))
            if w3.is_connected():
                block = w3.eth.block_number
                if block > 0:
                    self.node_health[node_url] = True
                    self.last_check[node_url] = datetime.now()
                    logger.debug(f"Узел {node_url} работает (блок {block})")
                    return True
        except Exception as e:
            logger.warning(f"Проверка узла {node_url} не удалась: {e}")
            self.node_health[node_url] = False
            self.last_check[node_url] = datetime.now()
        return False

    def get_healthy_node(self):
        """Возвращает URL работающего узла"""
        # Сначала проверяем текущий узел, если он был недавно проверен
        current = self.nodes[self.idx]
        if (self.node_health[current] and 
            datetime.now() - self.last_check[current] < self.check_interval):
            return current
            
        # В противном случае ищем любой работающий узел
        for _ in range(len(self.nodes)):
            self.idx = (self.idx + 1) % len(self.nodes)
            node = self.nodes[self.idx]
            if self._check_node(node):
                return node
                
        # Если все узлы не работают, пробуем текущий как последнее средство
        return self.nodes[self.idx]

    def web3(self):
        """Получение экземпляра web3, подключенного к работающему узлу"""
        if not blockchain_enabled:
            raise RuntimeError("Функции блокчейна отключены")
        
        node_url = self.get_healthy_node()
        
        # Создаем Web3 с уже внедренным geth_poa_middleware
        w3 = Web3(HTTPProvider(node_url, request_kwargs={'timeout': 10}))
        
        # Важно указать явное внедрение middleware для POA сетей
        # Используем уже импортированный geth_poa_middleware
        try:
            w3.middleware_onion.inject(geth_poa_middleware, layer=0)
        except (AttributeError, TypeError):
            # Для новых версий Web3.py
            try:
                w3.middleware_onion.add(geth_poa_middleware)
            except Exception as e:
                logger.warning(f"Не удалось добавить POA middleware: {e}")
        
        # Добавление middleware для стратегии цены газа
        if medium_gas_price_strategy:
            try:
                w3.eth.set_gas_price_strategy(medium_gas_price_strategy)
            except Exception as e:
                logger.warning(f"Не удалось установить стратегию цены газа: {e}")
        
        if not w3.is_connected():
            raise ConnectionError(f"Не удалось подключиться к узлу: {node_url}")
            
        logger.debug(f"Подключено к {node_url}")
        return w3
# ======== #MCP:CACHE ========
class SimpleCache:
    """Кэширование результатов с TTL.
    
    OPTIMIZE:MCP - Добавить LRU стратегию
    REFACTOR:MCP - Использовать готовую библиотеку кэширования
    TODO:MCP - Добавить персистентное кэширование
    """
    def __init__(self):
        self.store = {}  # Хранилище кэша

    def get(self, key):
        """Получение значения из кэша, если оно не устарело"""
        entry = self.store.get(key)
        if entry and datetime.now() - entry[1] < CACHE_TTL:
            return entry[0]
        return None

    def set(self, key, val):
        """Сохранение значения в кэш с текущей временной меткой"""
        self.store[key] = (val, datetime.now())
        
    def clear(self):
        """Очистка всех кэшированных данных"""
        old_count = len(self.store)
        self.store = {}
        return old_count

cache = SimpleCache()

# ======== #MCP:DATABASE ========
def get_db_connection():
    """Получение соединения с базой данных SQLite с нужными настройками
    
    OPTIMIZE:MCP - Добавить пул соединений
    SECURITY:MCP - Проверить SQL injection защиту
    """
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row  # Для именованных столбцов
    return conn

def init_db():
    """Инициализация базы данных - создание таблиц, если они не существуют
    
    REFACTOR:MCP - Использовать миграции для изменения схемы
    DOCS:MCP - Документировать схему базы данных
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS history(
            id      INTEGER PRIMARY KEY,
            ts      TEXT,
            token   TEXT,
            to_addr TEXT,
            amount  REAL,
            tx      TEXT,
            status  TEXT DEFAULT 'pending'
        )
    """)
    # Добавление столбца status, если его нет
    try:
        cur.execute("ALTER TABLE history ADD COLUMN status TEXT DEFAULT 'pending'")
    except sqlite3.OperationalError:
        pass  # Столбец уже существует
    
    # Создание таблицы для сохранения найденных транзакций
    cur.execute("""
        CREATE TABLE IF NOT EXISTS found_transactions(
            id          INTEGER PRIMARY KEY,
            ts          TEXT,
            tx_hash     TEXT,
            from_addr   TEXT,
            to_addr     TEXT,
            token_addr  TEXT,
            token_name  TEXT,
            amount      REAL,
            block       INTEGER,
            block_time  TEXT,
            search_data TEXT
        )
    """)
    
    # Создание таблицы для хранения детальной информации о транзакциях по отправителям
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sender_transactions(
            id           INTEGER PRIMARY KEY,
            search_time  TEXT,
            sender_addr  TEXT,
            tx_hash      TEXT,
            tx_timestamp TEXT,
            token_name   TEXT,
            amount       REAL,
            block        INTEGER,
            rewarded     INTEGER DEFAULT 0
        )
    """)
    
    # Создание таблицы для хранения наград
    cur.execute("""
        CREATE TABLE IF NOT EXISTS rewards(
            id           INTEGER PRIMARY KEY,
            address      TEXT,
            plex_amount  REAL,
            usdt_amount  REAL,
            tx_hash      TEXT,
            created_at   TEXT,
            sent         INTEGER DEFAULT 0
        )
    """)
    
    # Создание таблицы для массовых рассылок
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mass_distributions (
            id INTEGER PRIMARY KEY,
            created_at TEXT,
            name TEXT,
            token_address TEXT,
            token_symbol TEXT,
            amount_per_tx REAL,
            total_addresses INTEGER,
            total_cycles INTEGER,
            interval_seconds INTEGER,
            status TEXT,
            completed_at TEXT
        )
    """)
    
    # Создание таблицы для элементов массовой рассылки
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mass_distribution_items (
            id INTEGER PRIMARY KEY,
            distribution_id INTEGER,
            address TEXT,
            cycle INTEGER,
            tx_hash TEXT,
            status TEXT,
            error_message TEXT,
            sent_at TEXT,
            FOREIGN KEY (distribution_id) REFERENCES mass_distributions(id)
        )
    """)
    
    conn.commit()
    conn.close()

def add_history(token, to_addr, amt, tx_hash):
    """Добавление записи в историю транзакций"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO history VALUES(NULL,?,?,?,?,?,?)",
        (datetime.now().isoformat(), token, to_addr, amt, tx_hash, 'pending')
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    logger.info(f"Транзакция добавлена в историю: {tx_hash}")
    return row_id

def update_tx_status(tx_hash, status):
    """Обновление статуса транзакции в базе данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE history SET status=? WHERE tx=?", (status, tx_hash))
    conn.commit()
    affected_rows = cur.rowcount
    conn.close()
    logger.debug(f"Статус транзакции {tx_hash} обновлен на {status} ({affected_rows} строк)")
    return affected_rows

def fetch_history():
    """Получение истории транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT ts,token,to_addr,amount,tx,status FROM history ORDER BY id DESC LIMIT 100")
    rows = cur.fetchall()
    conn.close()
    return rows

def copy_all_transactions_hashes():
    """Копирование всех хешей транзакций в строку"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT tx FROM history ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return '\n'.join([row['tx'] for row in rows])

def add_found_transaction(tx_data, search_info):
    """Добавление найденной транзакции в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO found_transactions VALUES(NULL,?,?,?,?,?,?,?,?,?,?)",
        (
            datetime.now().isoformat(),
            tx_data.get('hash', ''),
            tx_data.get('from', ''),
            tx_data.get('to', ''),
            tx_data.get('contractAddress', ''),
            tx_data.get('tokenSymbol', ''),
            float(tx_data.get('value', 0)) / (10 ** int(tx_data.get('tokenDecimal', 18))),
            int(tx_data.get('blockNumber', 0)),
            datetime.fromtimestamp(int(tx_data.get('timeStamp', 0))).isoformat(),
            json.dumps(search_info)
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def add_sender_transaction(sender_addr, tx_info, search_time):
    """Добавление информации о транзакции отправителя в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Проверяем, не добавлен ли уже хэш транзакции
    cur.execute("SELECT id FROM sender_transactions WHERE tx_hash=?", (tx_info['hash'],))
    if cur.fetchone():
        conn.close()
        return None  # Транзакция уже добавлена
    
    cur.execute(
        "INSERT INTO sender_transactions VALUES(NULL,?,?,?,?,?,?,?,?)",
        (
            search_time,
            sender_addr,
            tx_info['hash'],
            datetime.fromtimestamp(int(tx_info['timestamp'])).isoformat() if tx_info['timestamp'] else None,
            tx_info['token'],
            tx_info['value'],
            tx_info['block'],
            0  # Не награжденная транзакция по умолчанию
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def add_reward(address, plex_amount, usdt_amount, tx_hash=None):
    """Добавление награды в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO rewards VALUES(NULL,?,?,?,?,?,?)",
        (
            address,
            plex_amount,
            usdt_amount,
            tx_hash,
            datetime.now().isoformat(),
            0  # Не отправленная награда по умолчанию
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def get_rewards():
    """Получение всех наград из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, address, plex_amount, usdt_amount, tx_hash, created_at, sent
        FROM rewards
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def mark_transaction_rewarded(tx_hash):
    """Отметка транзакции как награжденной"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE sender_transactions SET rewarded=1 WHERE tx_hash=?", (tx_hash,))
    conn.commit()
    affected_rows = cur.rowcount
    conn.close()
    return affected_rows

def get_unrewarded_transactions(sender_addr=None):
    """Получение ненагражденных транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    if sender_addr:
        cur.execute("""
            SELECT id, sender_addr, tx_hash, tx_timestamp, token_name, amount
            FROM sender_transactions
            WHERE rewarded=0 AND sender_addr=?
            ORDER BY tx_timestamp DESC
        """, (sender_addr,))
    else:
        cur.execute("""
            SELECT id, sender_addr, tx_hash, tx_timestamp, token_name, amount
            FROM sender_transactions
            WHERE rewarded=0
            ORDER BY sender_addr, tx_timestamp DESC
        """)
    
    rows = cur.fetchall()
    conn.close()
    return rows


def get_transactions_by_sender(sender_addr):
    """Получение всех транзакций отправителя"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tx_hash, tx_timestamp, token_name, amount, rewarded
        FROM sender_transactions
        WHERE sender_addr=?
        ORDER BY tx_timestamp DESC
    """, (sender_addr,))
    rows = cur.fetchall()
    conn.close()
    return rows

def fetch_found_transactions(limit=1000):
    """Получение найденных транзакций из базы данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT ts, tx_hash, from_addr, to_addr, token_name, amount, block_time
        FROM found_transactions
        ORDER BY id DESC
        LIMIT ?
    """, (limit,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_sender_transaction_counts():
    """Получение количества транзакций по отправителям"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sender_addr, COUNT(*) as tx_count, 
               SUM(CASE WHEN rewarded=1 THEN 1 ELSE 0 END) as rewarded_count
        FROM sender_transactions
        GROUP BY sender_addr
        ORDER BY tx_count DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def clear_found_transactions():
    """Очистка таблицы найденных транзакций"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM found_transactions")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return count

def clear_sender_transactions():
    """Очистка таблицы транзакций отправителей"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM sender_transactions")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return count

# ======== Управление конфигурацией наград ========
def save_rewards_config(name, addresses, plex_amounts, usdt_amounts):
    """Сохранение конфигурации наград в файл"""
    config = []
    for i in range(len(addresses)):
        config.append({
            'address': addresses[i],
            'plex': plex_amounts.get(i, 0),
            'usdt': usdt_amounts.get(i, 0)
        })
    
    file_path = REWARDS_CONFIG_DIR / f"{name}.json"
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"Ошибка при сохранении конфигурации наград: {e}")
        return False

def load_rewards_config(name):
    """Загрузка конфигурации наград из файла"""
    file_path = REWARDS_CONFIG_DIR / f"{name}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка при загрузке конфигурации наград: {e}")
        return None

def get_rewards_configs():
    """Получение списка доступных конфигураций наград"""
    try:
        return [f.stem for f in REWARDS_CONFIG_DIR.glob('*.json')]
    except Exception as e:
        logger.error(f"Ошибка при получении списка конфигураций наград: {e}")
        return []

# ======== Управление массовыми рассылками ========
def add_mass_distribution(name, token_address, token_symbol, amount_per_tx, total_addresses, total_cycles, interval_seconds):
    """Добавление новой массовой рассылки в базу данных"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO mass_distributions VALUES(NULL,?,?,?,?,?,?,?,?,?,?)",
        (
            datetime.now().isoformat(),
            name,
            token_address,
            token_symbol,
            amount_per_tx,
            total_addresses,
            total_cycles,
            interval_seconds,
            'active',
            None
        )
    )
    conn.commit()
    distribution_id = cur.lastrowid
    conn.close()
    logger.info(f"Создана массовая рассылка #{distribution_id}: {name}")
    return distribution_id

def add_mass_distribution_item(distribution_id, address, cycle, tx_hash, status, error_message=None):
    """Добавление элемента массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO mass_distribution_items VALUES(NULL,?,?,?,?,?,?,?)",
        (
            distribution_id,
            address,
            cycle,
            tx_hash,
            status,
            error_message,
            datetime.now().isoformat() if status in ['success', 'error'] else None
        )
    )
    conn.commit()
    row_id = cur.lastrowid
    conn.close()
    return row_id

def update_mass_distribution_status(distribution_id, status):
    """Обновление статуса массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    completed_at = datetime.now().isoformat() if status in ['completed', 'cancelled'] else None
    
    cur.execute(
        "UPDATE mass_distributions SET status=?, completed_at=? WHERE id=?",
        (status, completed_at, distribution_id)
    )
    conn.commit()
    conn.close()
    logger.info(f"Статус рассылки #{distribution_id} обновлен на {status}")

def get_mass_distributions(limit=100):
    """Получение списка массовых рассылок"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, created_at, name, token_symbol, amount_per_tx, total_addresses, 
               total_cycles, status, completed_at
        FROM mass_distributions
        ORDER BY id DESC
        LIMIT ?
    """, (limit,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_mass_distribution_items(distribution_id):
    """Получение элементов массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT address, cycle, tx_hash, status, error_message, sent_at
        FROM mass_distribution_items
        WHERE distribution_id=?
        ORDER BY cycle, address
    """, (distribution_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_mass_distribution_statistics(distribution_id):
    """Получение статистики массовой рассылки"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Общая статистика
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status='success' THEN 1 ELSE 0 END) as success_count,
            SUM(CASE WHEN status='error' THEN 1 ELSE 0 END) as error_count,
            SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending_count
        FROM mass_distribution_items
        WHERE distribution_id=?
    """, (distribution_id,))
    
    stats = cur.fetchone()
    conn.close()
    
    return {
        'total': stats[0] if stats else 0,
        'success': stats[1] if stats else 0,
        'errors': stats[2] if stats else 0,
        'pending': stats[3] if stats else 0
    }

# ======== #MCP:API ========
@retry(times=3)
def bscscan_request(params):
    """Отправка запроса к API BscScan с поддержкой кэширования и повторов
    
    OPTIMIZE:MCP - Добавить rate limiting
    REFACTOR:MCP - Выделить в отдельный API клиент
    SECURITY:MCP - Скрыть API ключ из логов
    """
    if not blockchain_enabled:
        raise RuntimeError("Блокчейн отключен")
    
    cache_key = str(params)
    cached = cache.get(cache_key)
    if cached:
        logger.debug(f"Найдено в кэше для запроса BscScan: {params}")
        return cached
    
    params.update({'apikey': BSCSCAN_KEY})
    try:
        response = requests.get(BSCSCAN_URL, params=params, timeout=10)
        if response.status_code != 200:
            logger.warning(f"Ошибка HTTP в BscScan: {response.status_code}")
            raise Exception(f"Ошибка HTTP {response.status_code}")
            
        result = response.json()
        
        if result.get('message') == 'No transactions found':
            return []
        if result.get('status') != '1':
            error_msg = result.get('message') or 'Ошибка API BscScan'
            logger.error(f"Ошибка API BscScan: {error_msg}")
            raise Exception(error_msg)
            
        data = result.get('result', [])
        cache.set(cache_key, data)
        return data
    except requests.RequestException as e:
        logger.error(f"Запрос BscScan не удался: {e}")
        raise

# ======== Постраничный анализ ========
def search_transactions_paginated(wallet_address, token_contract=None, exact_amount=None, min_amount=None, max_amount=None, page_size=1000, max_pages=100, delay_seconds=1, stop_flag=None, track_individual_tx=False):
    """
    Постраничный поиск транзакций с указанными параметрами.
    
    OPTIMIZE:MCP - Добавить параллельную обработку страниц
    REFACTOR:MCP - Разделить на более мелкие функции
    TODO:MCP - Добавить прогресс-бар для длительных операций
    
    Args:
        wallet_address (str): Адрес кошелька для поиска
        token_contract (str, optional): Адрес контракта токена для фильтрации
        exact_amount (float, optional): Точная сумма токенов для поиска
        min_amount (float, optional): Минимальная сумма для поиска по диапазону
        max_amount (float, optional): Максимальная сумма для поиска по диапазону
        page_size (int): Количество транзакций на странице (максимум 10000)
        max_pages (int): Максимальное количество страниц для поиска
        delay_seconds (int): Задержка между запросами в секундах
        stop_flag (threading.Event, optional): Флаг для остановки поиска
        track_individual_tx (bool): Если True, возвращает детальную информацию о транзакциях по отправителям
        
    Returns:
        tuple: (список найденных транзакций, счетчик по отправителям, детальная информация по транзакциям)
    """
    if not blockchain_enabled:
        raise RuntimeError("Блокчейн отключен")
    
    # Подготовка базовых параметров запроса
    base_params = {
        'module': 'account',
        'action': 'tokentx',
        'address': wallet_address,
        'sort': 'desc',
        'offset': page_size
    }
    
    # Добавляем адрес контракта токена, если указан
    if token_contract:
        base_params['contractaddress'] = token_contract
    
    matching_transactions = []
    sender_counter = {}
    # Словарь для хранения транзакций по отправителям, если нужно детальное отслеживание
    sender_transactions = {} if track_individual_tx else None
    
    page = 1
    has_more_data = True
    token_decimals = {}  # Кэш для десятичных знаков токенов
    
    logger.info(f"Начинаем постраничный поиск транзакций для адреса {wallet_address}")
    
    while has_more_data and page <= max_pages:
        # Проверка флага остановки
        if stop_flag and stop_flag.is_set():
            logger.info("Поиск остановлен по запросу пользователя")
            break
            
        try:
            params = dict(base_params)
            params['page'] = page
            
            logger.info(f"Запрашиваем страницу {page}...")
            result = bscscan_request(params)
            
            if not result:
                logger.info(f"Страница {page} не содержит транзакций. Поиск завершен.")
                break
            
            logger.info(f"Получено {len(result)} транзакций на странице {page}")
            
            for tx in result:
                # Получаем десятичные знаки для токена
                contract_addr = tx.get('contractAddress', '').lower()
                if contract_addr not in token_decimals:
                    token_decimals[contract_addr] = int(tx.get('tokenDecimal', 18))
                
                decimals = token_decimals[contract_addr]
                tx_value = float(tx.get('value', 0)) / (10 ** decimals)
                
                # Проверяем условия
                if wallet_address.lower() == tx.get('to', '').lower():  # Только входящие транзакции
                    # Проверка на точную сумму
                    if exact_amount is not None:
                        if abs(tx_value - exact_amount) < 0.0000001:  # Учитываем погрешность float
                            matching_transactions.append(tx)
                            sender = tx.get('from', '').lower()
                            sender_counter[sender] = sender_counter.get(sender, 0) + 1
                            
                            # Сохраняем дополнительную информацию о транзакциях, если нужно
                            if track_individual_tx:
                                if sender not in sender_transactions:
                                    sender_transactions[sender] = []
                                tx_info = {
                                    'hash': tx.get('hash', ''),
                                    'timestamp': tx.get('timeStamp', ''),
                                    'value': tx_value,
                                    'block': tx.get('blockNumber', ''),
                                    'token': tx.get('tokenSymbol', '')
                                }
                                sender_transactions[sender].append(tx_info)
                    
                    # Проверка на диапазон сумм
                    elif min_amount is not None and max_amount is not None:
                        if min_amount <= tx_value <= max_amount:
                            matching_transactions.append(tx)
                            sender = tx.get('from', '').lower()
                            sender_counter[sender] = sender_counter.get(sender, 0) + 1
                            
                            # Сохраняем дополнительную информацию о транзакциях, если нужно
                            if track_individual_tx:
                                if sender not in sender_transactions:
                                    sender_transactions[sender] = []
                                tx_info = {
                                    'hash': tx.get('hash', ''),
                                    'timestamp': tx.get('timeStamp', ''),
                                    'value': tx_value,
                                    'block': tx.get('blockNumber', ''),
                                    'token': tx.get('tokenSymbol', '')
                                }
                                sender_transactions[sender].append(tx_info)
                                # Если получили меньше транзакций, чем размер страницы, значит больше данных нет
            if len(result) < page_size:
                has_more_data = False
                logger.info("Достигнут конец данных")
            else:
                # Переходим к следующей странице
                page += 1
                # Задержка для соблюдения ограничений API
                time.sleep(delay_seconds)
        
        except Exception as e:
            logger.error(f"Ошибка при получении страницы {page}: {e}")
            time.sleep(delay_seconds * 2)  # Увеличенная задержка при ошибке
            
            # Если ошибка связана с превышением лимита, делаем более длительную паузу
            if str(e).lower().find('rate limit') >= 0:
                logger.warning("Превышен лимит запросов API. Пауза 10 секунд...")
                time.sleep(10)
            else:
                # Для других ошибок переходим к следующей странице
                page += 1
    
    logger.info(f"Поиск завершен. Найдено {len(matching_transactions)} транзакций от {len(sender_counter)} отправителей")
    return matching_transactions, sender_counter, sender_transactions

def get_token_decimal(token_address):
    """Получение количества десятичных знаков для токена"""
    if not blockchain_enabled:
        return 18  # Значение по умолчанию
    
    try:
        # Сначала пробуем через BSCScan API
        token_info = bscscan_request({
            'module': 'token',
            'action': 'tokeninfo',
            'contractaddress': token_address
        })
        
        if token_info and isinstance(token_info, list) and len(token_info) > 0:
            return int(token_info[0].get('divisor', 18))
        
        # Если не удалось получить через API, пробуем через Web3
        w3 = RPCPool(RPC_NODES).web3()
        token_contract = w3.eth.contract(address=Web3.to_checksum_address(token_address), abi=ERC20_ABI)
        return token_contract.functions.decimals().call()
    
    except Exception as e:
        logger.warning(f"Не удалось получить decimals для токена {token_address}: {e}")
        return 18  # Значение по умолчанию

# ======== GUI ========
class LogHandler(logging.Handler):
    """Пользовательский обработчик логирования, отправляющий логи в QTextEdit"""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
    def emit(self, record):
        msg = self.format(record)
        # Используем signal/slot для безопасного обновления UI из любого потока
        QtCore.QMetaObject.invokeMethod(
            self.text_widget, 
            "append", 
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, msg)
        )

# ======== #MCP:GUI ========
class MainWindow(QtWidgets.QMainWindow):
    """Главное окно приложения WalletSender
    
    REFACTOR:MCP - Разделить на более мелкие компоненты UI
    OPTIMIZE:MCP - Улучшить производительность обновления таблиц
    TODO:MCP - Добавить горячие клавиши
    DOCS:MCP - Документировать пользовательский интерфейс
    """
    
    # Сигнал для обновления индикаторов прогресса из рабочих потоков
    update_progress_signal = QtCore.pyqtSignal(str, int)
    update_status_signal = QtCore.pyqtSignal(str)
    update_table_signal = QtCore.pyqtSignal(list, dict, dict)
    queue_update_signal = QtCore.pyqtSignal(list)
    queue_item_update_signal = QtCore.pyqtSignal(int, dict)
    completed_item_signal = QtCore.pyqtSignal(dict)
    update_address_status = QtCore.pyqtSignal(int, str)  # row, status для массовой рассылки

    def __init__(self):
        super().__init__()
        self.setWindowTitle('Анализатор & Награды PLEX ONE/USDT')
        self.resize(1000, 800)  # Увеличил высоту окна
        self.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())  # Тёмная тема

        self.cfg = Config()  # Инициализация конфигурации
        init_db()  # Инициализация базы данных
        self.rpc = RPCPool(RPC_NODES)  # Создание пула RPC узлов
        self.pk = self.cfg.get_key()  # Получение приватного ключа из конфигурации
        self.mn = self.cfg.get_mnemonic()  # Получение мнемонической фразы
        self.sending = False  # Флаг активного процесса отправки
        self.tx_tracker = None  # Поток для отслеживания транзакций
        self.search_thread = None  # Поток для поиска транзакций
        self.is_searching = False  # Флаг активного поиска
        self.stop_search_event = threading.Event()  # Событие для остановки поиска
        self.last_rewards_state = None  # Хранение предыдущего состояния таблицы наград
        self.reward_per_tx_mode = self.cfg.get_reward_per_tx()  # Режим награждения за каждую транзакцию
        self.queue_paused = False  # Флаг паузы очереди отправки
        
        
        # Инициализация таймера для периодической прямой отправки
        self.direct_send_timer = QtCore.QTimer()
        self.direct_send_timer.timeout.connect(self._direct_send_periodic_send)
        self.direct_send_active = False
        self.direct_send_current_period = 0
        self.direct_send_total_periods = 0
        self.direct_send_params = {}  # Параметры для отправки: адрес, тип токена, сумма

        self._build_ui()  # Построение пользовательского интерфейса
        
        # Подключение сигналов
        self.update_progress_signal.connect(self._update_progress)
        self.update_status_signal.connect(self.log)
        self.update_table_signal.connect(self._update_search_results)
        self.queue_update_signal.connect(self._update_queue_table)
        self.queue_item_update_signal.connect(self._update_queue_item)
        self.completed_item_signal.connect(self._add_to_completed_table)
        self.update_address_status.connect(self._update_address_status_ui)
        
        # Настройка пользовательского обработчика логов
        self.log_handler = LogHandler(self.log_area)
        logger.addHandler(self.log_handler)
        logger.setLevel(logging.INFO)
        
        logger.info("Приложение запущено")

    def _build_ui(self):
        """Построение пользовательского интерфейса с вкладками"""
        tabs = QtWidgets.QTabWidget()
        tabs.addTab(self._tab_analyze(), 'Анализ')
        tabs.addTab(self._tab_paginated_search(), 'Поиск транзакций')
        tabs.addTab(self._tab_rewards(), 'Награды')
        tabs.addTab(self._tab_tx_rewards(), 'Награды за Tx')
        tabs.addTab(self._tab_direct_send(), 'Прямая отправка')
        tabs.addTab(self._tab_mass_distribution(), 'Массовая рассылка')  # #MCP:MASS_DIST - Вкладка массовой рассылки
        tabs.addTab(self._tab_sending_queue(), 'Очередь отправки')  # Новая вкладка
        tabs.addTab(self._tab_ds(), 'ДС')  # #MCP:DS_TAB - Вкладка дополнительных сервисов
        tabs.addTab(self._tab_history(), 'История')
        tabs.addTab(self._tab_found_tx(), 'Найденные Tx')
        tabs.addTab(self._tab_settings(), 'Настройки')
        self.setCentralWidget(tabs)

        # Область для логов внизу окна
        self.log_area = QtWidgets.QTextEdit(readOnly=True)
        dock = QtWidgets.QDockWidget('Лог')
        dock.setWidget(self.log_area)
        self.addDockWidget(QtCore.Qt.BottomDockWidgetArea, dock)

    def _update_progress(self, bar_name, value):
        """Обновление указанного индикатора прогресса значением"""
        if bar_name == "scan":
            self.progress.setValue(value)
        elif bar_name == "send":
            self.progress_send.setValue(value)
        elif bar_name == "search":
            self.progress_search.setValue(value)
        elif bar_name == "tx_rewards":
            self.progress_tx_rewards.setValue(value)
        elif bar_name == "direct_send":
            self.direct_send_progress.setValue(value)
        elif bar_name == "mass":
            self.mass_progress.setValue(value)

    def log(self, text):
        """Устаревший метод логирования - перенаправляет в logger"""
        logger.info(text)

    def _toggle_search_mode(self):
        """Переключает режим поиска между точной суммой и диапазоном"""
        is_exact_mode = self.radio_exact.isChecked()
        
        # Включить/выключить поле для точной суммы
        self.spin_amt.setEnabled(is_exact_mode)
        
        # Включить/выключить поля для диапазона
        self.spin_amt_from.setEnabled(not is_exact_mode)
        self.spin_amt_to.setEnabled(not is_exact_mode)
        
        logger.debug(f"Режим поиска переключен на: {'точная сумма' if is_exact_mode else 'диапазон'}")
    # --- Вкладка «Очередь отправки» ---
    def _tab_sending_queue(self):
        """Создание вкладки для отслеживания очереди отправки"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Заголовок и статус текущей отправки
        status_group = QtWidgets.QGroupBox("Статус отправки")
        status_layout = QtWidgets.QVBoxLayout(status_group)
        
        self.queue_status_label = QtWidgets.QLabel("Не активно")
        self.queue_progress_label = QtWidgets.QLabel("0/0 отправлено")
        status_layout.addWidget(self.queue_status_label)
        status_layout.addWidget(self.queue_progress_label)
        
        layout.addWidget(status_group)
        
        # Таблица очереди отправки
        self.queue_table = QtWidgets.QTableWidget(0, 5)
        self.queue_table.setHorizontalHeaderLabels([
            'Статус', 'Адрес получателя', 'Токен', 'Сумма', 'Время'
        ])
        self.queue_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.queue_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        layout.addWidget(self.queue_table)
        
        # Таблица завершенных отправок
        completed_group = QtWidgets.QGroupBox("Завершенные отправки")
        completed_layout = QtWidgets.QVBoxLayout(completed_group)
        
        self.completed_table = QtWidgets.QTableWidget(0, 6)
        self.completed_table.setHorizontalHeaderLabels([
            'Статус', 'Адрес получателя', 'Токен', 'Сумма', 'Хэш TX', 'Время'
        ])
        self.completed_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.completed_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.completed_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.completed_table.customContextMenuRequested.connect(self._queue_context_menu)
        completed_layout.addWidget(self.completed_table)
        
        layout.addWidget(completed_group)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.pause_queue_btn = QtWidgets.QPushButton("Пауза")
        self.pause_queue_btn.setEnabled(False)
        self.pause_queue_btn.clicked.connect(self._pause_sending_queue)
        buttons_layout.addWidget(self.pause_queue_btn)
        
        self.resume_queue_btn = QtWidgets.QPushButton("Продолжить")
        self.resume_queue_btn.setEnabled(False)
        self.resume_queue_btn.clicked.connect(self._resume_sending_queue)
        buttons_layout.addWidget(self.resume_queue_btn)
        
        self.cancel_queue_btn = QtWidgets.QPushButton("Отменить все")
        self.cancel_queue_btn.setEnabled(False)
        self.cancel_queue_btn.clicked.connect(self._cancel_sending_queue)
        buttons_layout.addWidget(self.cancel_queue_btn)
        
        self.export_results_btn = QtWidgets.QPushButton("Экспорт результатов")
        self.export_results_btn.clicked.connect(self._export_queue_results)
        buttons_layout.addWidget(self.export_results_btn)
        
        layout.addLayout(buttons_layout)
        
        return w

    def _pause_sending_queue(self):
        """Приостановка отправки"""
        self.queue_paused = True
        self.pause_queue_btn.setEnabled(False)
        self.resume_queue_btn.setEnabled(True)
        self.queue_status_label.setText("Приостановлено")
        logger.info("Отправка приостановлена")

    def _resume_sending_queue(self):
        """Возобновление отправки"""
        self.queue_paused = False
        self.pause_queue_btn.setEnabled(True)
        self.resume_queue_btn.setEnabled(False)
        self.queue_status_label.setText("Отправка")
        logger.info("Отправка возобновлена")

    def _cancel_sending_queue(self):
        """Отмена всей очереди отправки"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Вы уверены, что хотите отменить все оставшиеся отправки?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.sending = False
            logger.info("Отправка отменена")

    @QtCore.pyqtSlot(list)
    def _update_queue_table(self, items):
        """Обновление таблицы очереди отправки"""
        try:
            self.queue_table.setRowCount(0)
            
            for item in items:
                row = self.queue_table.rowCount()
                self.queue_table.insertRow(row)
                
                # Статус
                status_item = QtWidgets.QTableWidgetItem(item['status'])
                if item['status'] == 'success':
                    status_item.setBackground(QtGui.QColor('#004400'))
                elif item['status'] == 'failed':
                    status_item.setBackground(QtGui.QColor('#440000'))
                elif item['status'] == 'processing':
                    status_item.setBackground(QtGui.QColor('#444400'))
                self.queue_table.setItem(row, 0, status_item)
                
                # Адрес
                self.queue_table.setItem(row, 1, QtWidgets.QTableWidgetItem(item['address']))
                
                # Токен
                self.queue_table.setItem(row, 2, QtWidgets.QTableWidgetItem(item['token']))
                
                # Сумма
                self.queue_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(item['amount'])))
                
                # Время
                time_text = item['time'] if item['time'] else "-"
                self.queue_table.setItem(row, 4, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при обновлении таблицы очереди: {e}")

    @QtCore.pyqtSlot(int, dict)
    def _update_queue_item(self, index, item):
        """Обновление конкретного элемента в очереди"""
        try:
            if index >= self.queue_table.rowCount():
                return
            
            # Статус
            status_item = QtWidgets.QTableWidgetItem(item['status'])
            if item['status'] == 'success':
                status_item.setBackground(QtGui.QColor('#004400'))
            elif item['status'] == 'failed':
                status_item.setBackground(QtGui.QColor('#440000'))
            elif item['status'] == 'processing':
                status_item.setBackground(QtGui.QColor('#444400'))
            self.queue_table.setItem(index, 0, status_item)
            
            # Время
            time_text = item['time'] if item['time'] else "-"
            self.queue_table.setItem(index, 4, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при обновлении элемента очереди: {e}")

    @QtCore.pyqtSlot(dict)
    def _add_to_completed_table(self, item):
        """Добавление элемента в таблицу завершенных отправок"""
        try:
            row = self.completed_table.rowCount()
            self.completed_table.insertRow(row)
            
            # Статус
            status_item = QtWidgets.QTableWidgetItem(item['status'])
            if item['status'] == 'success':
                status_item.setBackground(QtGui.QColor('#004400'))
            elif item['status'] == 'failed':
                status_item.setBackground(QtGui.QColor('#440000'))
            self.completed_table.setItem(row, 0, status_item)
            
            # Адрес
            self.completed_table.setItem(row, 1, QtWidgets.QTableWidgetItem(item['address']))
            
            # Токен
            self.completed_table.setItem(row, 2, QtWidgets.QTableWidgetItem(item['token']))
            
            # Сумма
            self.completed_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(item['amount'])))
            
            # Хэш TX
            tx_item = QtWidgets.QTableWidgetItem(item['tx_hash'][:10] + "..." + item['tx_hash'][-6:] if item['tx_hash'] else "-")
            if item['tx_hash']:
                tx_item.setToolTip(item['tx_hash'])
                tx_item.setData(QtCore.Qt.UserRole, item['tx_hash'])
            self.completed_table.setItem(row, 4, tx_item)
            
            # Время
            time_text = item['time'] if item['time'] else "-"
            self.completed_table.setItem(row, 5, QtWidgets.QTableWidgetItem(time_text))
        except Exception as e:
            logger.error(f"Ошибка при добавлении в таблицу завершенных: {e}")

    def _queue_context_menu(self, position):
        """Контекстное меню для таблицы завершенных отправок"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_addr = menu.addAction("Копировать адрес")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        
        action = menu.exec_(self.completed_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.completed_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.completed_table.item(row, 4).data(QtCore.Qt.UserRole)
        address = self.completed_table.item(row, 1).text()
        
        if action == copy_tx and tx_hash:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_addr:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == open_tx and tx_hash:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")

    def _export_queue_results(self):
        """Экспорт результатов очереди в CSV"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить результаты', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('status,address,token,amount,tx_hash,time\n')
                for row in range(self.completed_table.rowCount()):
                    status = self.completed_table.item(row, 0).text()
                    address = self.completed_table.item(row, 1).text()
                    token = self.completed_table.item(row, 2).text()
                    amount = self.completed_table.item(row, 3).text()
                    tx_item = self.completed_table.item(row, 4)
                    tx_hash = tx_item.data(QtCore.Qt.UserRole) if tx_item else ""
                    time = self.completed_table.item(row, 5).text()
                    
                    f.write(f'"{status}","{address}","{token}",{amount},"{tx_hash}","{time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")

    def _show_transaction_results(self, batch_id, transactions):
        """Отображение результатов отправки в отдельном окне для копирования"""
        result_dialog = QtWidgets.QDialog(self)
        result_dialog.setWindowTitle(f"Результаты отправки #{batch_id}")
        result_dialog.resize(800, 600)
        
        layout = QtWidgets.QVBoxLayout(result_dialog)
        
        # Текстовое поле для отображения результатов
        result_text = QtWidgets.QTextEdit()
        result_text.setReadOnly(True)
        result_text.setWordWrapMode(QtGui.QTextOption.NoWrap)
        
        # Формируем текст с транзакциями
        text = "Дата: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "\n\n"
        text += "ID партии: " + str(batch_id) + "\n\n"
        text += "Результаты отправки:\n\n"
        
        for tx in transactions:
            text += f"Адрес: {tx['address']}\n"
            text += f"Токен: {tx['token']}\n"
            text += f"Сумма: {tx['amount']}\n"
            text += f"Хэш TX: {tx['tx_hash']}\n"
            text += f"Статус: {tx['status']}\n"
            text += "-" * 50 + "\n"
        
        result_text.setText(text)
        layout.addWidget(result_text)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        copy_all_btn = QtWidgets.QPushButton("Копировать всё")
        copy_all_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(text))
        buttons_layout.addWidget(copy_all_btn)
        
        copy_hashes_btn = QtWidgets.QPushButton("Копировать только хэши")
        copy_hashes_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(
            "\n".join([tx['tx_hash'] for tx in transactions])
        ))
        buttons_layout.addWidget(copy_hashes_btn)
        
        export_csv_btn = QtWidgets.QPushButton("Экспорт в CSV")
        export_csv_btn.clicked.connect(lambda: self._export_tx_results_to_csv(transactions, batch_id))
        buttons_layout.addWidget(export_csv_btn)
        
        layout.addLayout(buttons_layout)
        
        result_dialog.exec_()

    def _export_tx_results_to_csv(self, transactions, batch_id):
        """Экспорт результатов транзакций в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, f'Сохранить результаты #{batch_id}', f'tx_results_{batch_id}.csv', '*.csv'
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,token,amount,tx_hash,status,time\n')
                for tx in transactions:
                    f.write(f'"{tx["address"]}","{tx["token"]}",{tx["amount"]},"{tx["tx_hash"]}","{tx["status"]}","{tx["time"]}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")

    # --- Вкладка «Награды за Tx» ---
    def _tab_tx_rewards(self):
        """Создание вкладки для управления наградами за транзакции"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа настроек награждения
        reward_settings = QtWidgets.QGroupBox("Настройки награждения")
        reward_layout = QtWidgets.QHBoxLayout(reward_settings)
        
        # Переключатель режима награждения
        self.reward_per_tx_checkbox = QtWidgets.QCheckBox("Награждать за каждую транзакцию")
        self.reward_per_tx_checkbox.setChecked(self.reward_per_tx_mode)
        self.reward_per_tx_checkbox.toggled.connect(self._toggle_reward_mode)
        reward_layout.addWidget(self.reward_per_tx_checkbox)
        
        # Кнопка для сканирования транзакций
        self.scan_tx_btn = QtWidgets.QPushButton("Сканировать транзакции")
        self.scan_tx_btn.clicked.connect(self._scan_transactions_for_rewards)
        reward_layout.addWidget(self.scan_tx_btn)
        
        # Кнопка для очистки данных о транзакциях
        self.clear_tx_data_btn = QtWidgets.QPushButton("Очистить данные")
        self.clear_tx_data_btn.clicked.connect(self._clear_tx_data)
        reward_layout.addWidget(self.clear_tx_data_btn)
        
        layout.addWidget(reward_settings)
        
        # Группа настроек размера награды
        reward_size_group = QtWidgets.QGroupBox("Настройки размера награды")
        reward_size_layout = QtWidgets.QHBoxLayout(reward_size_group)

        # Выбор типа токена для награды
        self.reward_token_group = QtWidgets.QButtonGroup(w)
        self.reward_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.reward_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.reward_both_radio = QtWidgets.QRadioButton('Оба (PLEX ONE и USDT)')
        self.reward_plex_radio.setChecked(True)
        self.reward_token_group.addButton(self.reward_plex_radio)
        self.reward_token_group.addButton(self.reward_usdt_radio)
        self.reward_token_group.addButton(self.reward_both_radio)

        token_layout = QtWidgets.QVBoxLayout()
        token_layout.addWidget(QtWidgets.QLabel("Тип токена награды:"))
        token_layout.addWidget(self.reward_plex_radio)
        token_layout.addWidget(self.reward_usdt_radio)
        token_layout.addWidget(self.reward_both_radio)
        reward_size_layout.addLayout(token_layout)

        # Ввод размера награды
        size_layout = QtWidgets.QVBoxLayout()
        size_layout.addWidget(QtWidgets.QLabel("Размер награды:"))

        # PLEX ONE размер
        plex_layout = QtWidgets.QHBoxLayout()
        plex_layout.addWidget(QtWidgets.QLabel("PLEX ONE:"))
        self.reward_plex_size = QtWidgets.QDoubleSpinBox()
        self.reward_plex_size.setRange(0, 1000)
        self.reward_plex_size.setDecimals(2)
        self.reward_plex_size.setValue(3.0)  # По умолчанию 3 PLEX ONE
        plex_layout.addWidget(self.reward_plex_size)
        size_layout.addLayout(plex_layout)

        # USDT размер
        usdt_layout = QtWidgets.QHBoxLayout()
        usdt_layout.addWidget(QtWidgets.QLabel("USDT:"))
        self.reward_usdt_size = QtWidgets.QDoubleSpinBox()
        self.reward_usdt_size.setRange(0, 1000) 
        self.reward_usdt_size.setDecimals(2)
        self.reward_usdt_size.setValue(0.3)  # По умолчанию 0.3 USDT
        usdt_layout.addWidget(self.reward_usdt_size)
        size_layout.addLayout(usdt_layout)

        reward_size_layout.addLayout(size_layout)
        layout.addWidget(reward_size_group)
        
        # Таблица отправителей и их транзакций
        self.tx_senders_table = QtWidgets.QTableWidget(0, 4)
        self.tx_senders_table.setHorizontalHeaderLabels([
            'Адрес отправителя', 'Всего Tx', 'Награждено', 'Осталось наградить'
        ])
        self.tx_senders_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        self.tx_senders_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.tx_senders_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tx_senders_table.customContextMenuRequested.connect(self._show_tx_senders_menu)
        layout.addWidget(self.tx_senders_table)
        
        # Таблица транзакций выбранного отправителя
        self.sender_tx_table = QtWidgets.QTableWidget(0, 5)
        self.sender_tx_table.setHorizontalHeaderLabels([
            'Транзакция', 'Время', 'Токен', 'Сумма', 'Награждено'
        ])
        self.sender_tx_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        self.sender_tx_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.sender_tx_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.sender_tx_table.customContextMenuRequested.connect(self._show_sender_tx_menu)
        layout.addWidget(self.sender_tx_table)
        
        # Кнопки действий
        action_layout = QtWidgets.QHBoxLayout()
        
        # Кнопка для создания наград
        self.create_rewards_btn = QtWidgets.QPushButton("Создать награды")
        self.create_rewards_btn.clicked.connect(self._create_rewards_from_tx)
        action_layout.addWidget(self.create_rewards_btn)
        
        # Кнопка для отправки наград
        self.send_tx_rewards_btn = QtWidgets.QPushButton("Отправить награды")
        self.send_tx_rewards_btn.clicked.connect(self._send_rewards_for_tx)
        action_layout.addWidget(self.send_tx_rewards_btn)
        
        layout.addLayout(action_layout)
        
        # Индикатор прогресса
        self.progress_tx_rewards = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_tx_rewards)
        
        # Загрузка данных при открытии
        QtCore.QTimer.singleShot(100, self._load_tx_senders)
        
        return w

    def _toggle_reward_mode(self, checked):
        """Переключение режима награждения за транзакции"""
        self.reward_per_tx_mode = checked
        self.cfg.set_reward_per_tx(checked)
        logger.info(f"Режим награждения за каждую транзакцию {'включен' if checked else 'выключен'}")
    def _scan_transactions_for_rewards(self):
        """Запуск сканирования транзакций для награждения"""
        if self.is_searching:
            return logger.warning("Поиск уже запущен")
            
        # Диалог для выбора параметров сканирования
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Сканирование транзакций")
        dialog_layout = QtWidgets.QVBoxLayout(dialog)
        
        form = QtWidgets.QFormLayout()
        
        # Адрес системы
        system_addr = QtWidgets.QLineEdit()
        system_addr.setPlaceholderText("Адрес системы (0x...)")
        form.addRow("Адрес системы:", system_addr)
        
        # Адрес токена
        token_addr = QtWidgets.QLineEdit()
        token_addr.setPlaceholderText("Адрес токена (0x...)")
        token_addr.setText(PLEX_CONTRACT)
        form.addRow("Адрес токена:", token_addr)
        
        # Точная сумма
        exact_amount = QtWidgets.QDoubleSpinBox()
        exact_amount.setRange(0, 1000000)
        exact_amount.setDecimals(8)
        exact_amount.setValue(30)
        form.addRow("Точная сумма:", exact_amount)
        
        dialog_layout.addLayout(form)
        
        # Кнопки
        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        dialog_layout.addWidget(buttons)
        
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Получаем параметры
            sys_addr = system_addr.text().strip()
            token = token_addr.text().strip()
            amount = exact_amount.value()
            
            # Проверяем адреса
            if blockchain_enabled:
                if not Web3.is_address(sys_addr):
                    return logger.error("Некорректный адрес системы")
                if not Web3.is_address(token):
                    return logger.error("Некорректный адрес токена")
            
            # Запускаем поиск в отдельном потоке
            self.is_searching = True
            self.stop_search_event.clear()
            self.update_progress_signal.emit("tx_rewards", 0)
            
            # Обновляем UI
            self.scan_tx_btn.setEnabled(False)
            self.clear_tx_data_btn.setEnabled(False)
            
            threading.Thread(
                target=self._scan_tx_thread,
                args=(sys_addr, token, amount),
                daemon=True
            ).start()
            
            logger.info(f"Запущено сканирование транзакций: {sys_addr}, {token}, {amount}")
    
    def _scan_tx_thread(self, system_addr, token_addr, amount):
        """Поток для сканирования транзакций для награждения"""
        try:
            logger.info(f"Начинаем сканирование транзакций: адрес={system_addr}, токен={token_addr}, сумма={amount}")
            
            # Используем расширенную функцию поиска
            transactions, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=system_addr,
                token_contract=token_addr,
                exact_amount=amount,
                page_size=1000,
                max_pages=100,
                delay_seconds=1,
                stop_flag=self.stop_search_event,
                track_individual_tx=True  # Включаем отслеживание отдельных транзакций
            )
            
            # Сохраняем данные о транзакциях в базу данных
            search_time = datetime.now().isoformat()
            total_tx = 0
            
            if sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
                        total_tx += 1
            
            logger.info(f"Сканирование завершено. Сохранено {total_tx} транзакций от {len(sender_counter)} отправителей")
            
            # Обновляем UI в основном потоке через сигнал
            self.update_status_signal.emit("Загрузка отправителей")
            QtWidgets.QApplication.processEvents()
            self._load_tx_senders()  # Прямой вызов, так как метод теперь декорирован как слот
            
        except Exception as e:
            logger.error(f"Ошибка при сканировании транзакций: {e}")
        finally:
            self.is_searching = False
            self.update_progress_signal.emit("tx_rewards", 100)
    
    @QtCore.pyqtSlot()
    def _load_tx_senders(self):
        """Загрузка данных об отправителях и их транзакциях"""
        try:
            # Получаем список отправителей и статистику их транзакций
            senders = get_sender_transaction_counts()
            
            # Обновляем таблицу
            self.tx_senders_table.setRowCount(0)
            
            for sender, tx_count, rewarded_count in senders:
                row = self.tx_senders_table.rowCount()
                self.tx_senders_table.insertRow(row)
                
                # Адрес отправителя
                self.tx_senders_table.setItem(row, 0, QtWidgets.QTableWidgetItem(sender))
                
                # Всего транзакций
                self.tx_senders_table.setItem(row, 1, QtWidgets.QTableWidgetItem(str(tx_count)))
                
                # Награждено транзакций
                self.tx_senders_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(rewarded_count)))
                
                # Осталось наградить
                self.tx_senders_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(tx_count - rewarded_count)))
            
            logger.info(f"Загружено {self.tx_senders_table.rowCount()} отправителей")
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке данных об отправителях: {e}")
    
    def _show_tx_senders_menu(self, position):
        """Показ контекстного меню для таблицы отправителей"""
        menu = QtWidgets.QMenu()
        
        view_tx = menu.addAction("Посмотреть транзакции")
        add_to_rewards = menu.addAction("Добавить в награды")
        copy_address = menu.addAction("Копировать адрес")
        open_bscscan = menu.addAction("Открыть в BscScan")
        
        action = menu.exec_(self.tx_senders_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.tx_senders_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        sender = self.tx_senders_table.item(row, 0).text()
        
        if action == view_tx:
            self._load_sender_transactions(sender)
        
        elif action == add_to_rewards:
            if self.reward_per_tx_mode:
                # В режиме награждения за каждую транзакцию
                self._add_sender_tx_to_rewards(sender)
            else:
                # В обычном режиме - просто добавляем адрес
                self._add_to_rewards_tab(sender)
        
        elif action == copy_address:
            QtWidgets.QApplication.clipboard().setText(sender)
            logger.info(f"Адрес скопирован: {sender}")
        
        elif action == open_bscscan:
            url = f"https://bscscan.com/address/{sender}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт адрес в BscScan: {sender}")
    
    def _load_sender_transactions(self, sender_addr):
        """Загрузка транзакций выбранного отправителя"""
        try:
            # Получаем транзакции отправителя
            transactions = get_transactions_by_sender(sender_addr)
            
            # Обновляем таблицу
            self.sender_tx_table.setRowCount(0)
            
            for tx_id, tx_hash, timestamp, token, amount, rewarded in transactions:
                row = self.sender_tx_table.rowCount()
                self.sender_tx_table.insertRow(row)
                
                # Хэш транзакции
                tx_item = QtWidgets.QTableWidgetItem(tx_hash[:10] + "..." + tx_hash[-6:])
                tx_item.setToolTip(tx_hash)
                tx_item.setData(QtCore.Qt.UserRole, tx_hash)
                self.sender_tx_table.setItem(row, 0, tx_item)
                
                # Время транзакции
                self.sender_tx_table.setItem(row, 1, QtWidgets.QTableWidgetItem(timestamp))
                
                # Токен
                self.sender_tx_table.setItem(row, 2, QtWidgets.QTableWidgetItem(token))
                
                # Сумма
                self.sender_tx_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(amount)))
                
                # Статус награждения
                status_item = QtWidgets.QTableWidgetItem("Да" if rewarded else "Нет")
                status_item.setData(QtCore.Qt.UserRole, rewarded)
                if rewarded:
                    status_item.setBackground(QtGui.QColor('#004400'))
                self.sender_tx_table.setItem(row, 4, status_item)
            
            logger.info(f"Загружено {self.sender_tx_table.rowCount()} транзакций отправителя {sender_addr}")
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке транзакций отправителя: {e}")
    
    def _show_sender_tx_menu(self, position):
        """Показ контекстного меню для таблицы транзакций отправителя"""
        menu = QtWidgets.QMenu()
        
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        copy_tx = menu.addAction("Копировать хэш")
        mark_rewarded = menu.addAction("Отметить как награжденную")
        add_tx_to_rewards = menu.addAction("Добавить в награды")
        
        action = menu.exec_(self.sender_tx_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.sender_tx_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.sender_tx_table.item(row, 0).data(QtCore.Qt.UserRole)
        is_rewarded = self.sender_tx_table.item(row, 4).data(QtCore.Qt.UserRole)
        
        if action == open_tx:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")
        
        elif action == copy_tx:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
        
        elif action == mark_rewarded:
            if not is_rewarded:
                mark_transaction_rewarded(tx_hash)
                logger.info(f"Транзакция {tx_hash} отмечена как награжденная")
                
                # Обновляем таблицу
                status_item = QtWidgets.QTableWidgetItem("Да")
                status_item.setData(QtCore.Qt.UserRole, 1)
                status_item.setBackground(QtGui.QColor('#004400'))
                self.sender_tx_table.setItem(row, 4, status_item)
                
                # Обновляем данные об отправителях
                selected_row = self.tx_senders_table.currentRow()
                if selected_row >= 0:
                    sender = self.tx_senders_table.item(selected_row, 0).text()
                    self._load_tx_senders()
                    
                    # Пытаемся восстановить выделение
                    for i in range(self.tx_senders_table.rowCount()):
                        if self.tx_senders_table.item(i, 0).text() == sender:
                            self.tx_senders_table.selectRow(i)
                            break
        
        elif action == add_tx_to_rewards:
            # Получаем адрес отправителя из текущего выбора в таблице отправителей
            selected_row = self.tx_senders_table.currentRow()
            if selected_row >= 0:
                sender = self.tx_senders_table.item(selected_row, 0).text()
                self._add_tx_to_rewards(sender, tx_hash)
    
    def _add_sender_tx_to_rewards(self, sender_addr):
        """Добавление транзакций отправителя в таблицу наград"""
        try:
            # Получаем ненагражденные транзакции отправителя
            transactions = get_unrewarded_transactions(sender_addr)
            
            if not transactions:
                return logger.warning(f"Нет ненагражденных транзакций для {sender_addr}")
            
            # Сохраняем текущее состояние таблицы наград
            self._save_rewards_state()
            
            # Добавляем каждую транзакцию в таблицу наград
            for tx_id, sender, tx_hash, timestamp, token, amount in transactions:
                row = self.table_rewards.rowCount()
                self.table_rewards.insertRow(row)
                
                # Адрес отправителя
                addr_item = QtWidgets.QTableWidgetItem(sender)
                addr_item.setData(QtCore.Qt.UserRole, tx_hash)
                self.table_rewards.setItem(row, 0, addr_item)
                
                # PLEX - значение по умолчанию
                plex_item = QtWidgets.QTableWidgetItem("3")
                plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 1, plex_item)
                
                # USDT - пустое значение по умолчанию
                usdt_item = QtWidgets.QTableWidgetItem("")
                usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 2, usdt_item)
            
            logger.info(f"Добавлено {len(transactions)} транзакций отправителя {sender_addr} в таблицу наград")
            
        except Exception as e:
            logger.error(f"Ошибка при добавлении транзакций в награды: {e}")
    
    def _add_tx_to_rewards(self, sender_addr, tx_hash):
        """Добавление одной конкретной транзакции в таблицу наград"""
        try:
            # Проверяем, не добавлена ли уже эта транзакция в таблицу наград
            for row in range(self.table_rewards.rowCount()):
                item = self.table_rewards.item(row, 0)
                if item and item.data(QtCore.Qt.UserRole) == tx_hash:
                    logger.warning(f"Транзакция {tx_hash} уже добавлена в таблицу наград")
                    return
            
            # Добавляем новую строку
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес отправителя
            addr_item = QtWidgets.QTableWidgetItem(sender_addr)
            addr_item.setData(QtCore.Qt.UserRole, tx_hash)
            self.table_rewards.setItem(row, 0, addr_item)
            
            # PLEX - значение по умолчанию
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT - пустое значение по умолчанию
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
            
            logger.info(f"Транзакция {tx_hash} добавлена в таблицу наград")
            
        except Exception as e:
            logger.error(f"Ошибка при добавлении транзакции в награды: {e}")
    def _create_rewards_from_tx(self):
        """Создание наград для всех ненагражденных транзакций"""
        try:
            # Получаем ненагражденные транзакции
            transactions = get_unrewarded_transactions()
            
            if not transactions:
                return logger.warning("Нет ненагражденных транзакций")
            
            # Запрашиваем подтверждение
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение', 
                f'Создать награды для {len(transactions)} ненагражденных транзакций?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No
            )
            
            if reply != QtWidgets.QMessageBox.Yes:
                return
            
            # Определяем тип и размер награды
            use_plex = self.reward_plex_radio.isChecked() or self.reward_both_radio.isChecked()
            use_usdt = self.reward_usdt_radio.isChecked() or self.reward_both_radio.isChecked()
            plex_size = self.reward_plex_size.value() if use_plex else 0
            usdt_size = self.reward_usdt_size.value() if use_usdt else 0
            
            # Сохраняем текущее состояние таблицы наград
            self._save_rewards_state()
            
            # Очищаем текущую таблицу наград
            self.table_rewards.setRowCount(0)
            
            # Добавляем каждую транзакцию в таблицу наград
            for tx_id, sender, tx_hash, timestamp, token, amount in transactions:
                row = self.table_rewards.rowCount()
                self.table_rewards.insertRow(row)
                
                # Адрес отправителя
                addr_item = QtWidgets.QTableWidgetItem(sender)
                addr_item.setData(QtCore.Qt.UserRole, tx_hash)
                self.table_rewards.setItem(row, 0, addr_item)
                
                # PLEX - значение из настроек
                plex_item = QtWidgets.QTableWidgetItem(str(plex_size) if plex_size > 0 else "")
                plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 1, plex_item)
                
                # USDT - значение из настроек
                usdt_item = QtWidgets.QTableWidgetItem(str(usdt_size) if usdt_size > 0 else "")
                usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 2, usdt_item)
            
            logger.info(f"Создано {len(transactions)} наград для ненагражденных транзакций")
            
        except Exception as e:
            logger.error(f"Ошибка при создании наград: {e}")
    
    def _send_rewards_for_tx(self):
        """Отправка наград для выбранных транзакций"""
        if not self.table_rewards.rowCount():
            return logger.warning("Нет наград для отправки")
        
        # Проверяем наличие транзакций в таблице наград
        has_tx_data = False
        for row in range(self.table_rewards.rowCount()):
            tx_hash = self.table_rewards.item(row, 0).data(QtCore.Qt.UserRole)
            if tx_hash:
                has_tx_data = True
                break
        
        if not has_tx_data:
            return logger.warning("В таблице нет наград за транзакции")
        
        # Запускаем стандартную отправку
        self._start_sending(mark_rewarded=True)
    
    def _clear_tx_data(self):
        """Очистка всех данных о транзакциях"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Очистить все данные о транзакциях отправителей?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            count = clear_sender_transactions()
            self.tx_senders_table.setRowCount(0)
            self.sender_tx_table.setRowCount(0)
            logger.info(f"Очищены данные о {count} транзакциях")

    # --- Вкладка «Награды» ---
    def _tab_rewards(self):
        """Создание вкладки для управления наградами"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Таблица наград
        self.table_rewards = QtWidgets.QTableWidget(0, 3)
        self.table_rewards.setHorizontalHeaderLabels(['Адрес', 'PLEX ONE', 'USDT'])
        self.table_rewards.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table_rewards.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.table_rewards.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table_rewards.customContextMenuRequested.connect(self._show_rewards_context_menu)
        layout.addWidget(self.table_rewards)
        
        # Кнопки действий с наградами
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.btn_add_row = QtWidgets.QPushButton('Добавить строку')
        self.btn_add_row.clicked.connect(self._add_rewards_row)
        buttons_layout.addWidget(self.btn_add_row)
        
        self.btn_remove_row = QtWidgets.QPushButton('Удалить выбранное')
        self.btn_remove_row.clicked.connect(self._remove_rewards_row)
        buttons_layout.addWidget(self.btn_remove_row)
        
        self.btn_clear = QtWidgets.QPushButton('Очистить всё')
        self.btn_clear.clicked.connect(self._clear_rewards)
        buttons_layout.addWidget(self.btn_clear)
        
        layout.addLayout(buttons_layout)
        
        # Группа для случайной генерации сумм с настраиваемыми диапазонами
        random_group = QtWidgets.QGroupBox("Генерация случайных сумм")
        random_layout = QtWidgets.QVBoxLayout(random_group)
        
        # Настройки для PLEX
        plex_layout = QtWidgets.QHBoxLayout()
        plex_layout.addWidget(QtWidgets.QLabel('PLEX ONE диапазон:'))
        
        self.plex_min = QtWidgets.QDoubleSpinBox()
        self.plex_min.setRange(0, 10000)
        self.plex_min.setDecimals(2)
        self.plex_min.setValue(2.0)
        plex_layout.addWidget(self.plex_min)
        
        plex_layout.addWidget(QtWidgets.QLabel('-'))
        
        self.plex_max = QtWidgets.QDoubleSpinBox()
        self.plex_max.setRange(0, 10000)
        self.plex_max.setDecimals(2)
        self.plex_max.setValue(4.0)
        plex_layout.addWidget(self.plex_max)
        
        self.btn_random_plex = QtWidgets.QPushButton('Случайные PLEX ONE')
        self.btn_random_plex.clicked.connect(self._randomize_plex)
        plex_layout.addWidget(self.btn_random_plex)
        
        random_layout.addLayout(plex_layout)
        
        # Настройки для USDT
        usdt_layout = QtWidgets.QHBoxLayout()
        usdt_layout.addWidget(QtWidgets.QLabel('USDT диапазон:'))
        
        self.usdt_min = QtWidgets.QDoubleSpinBox()
        self.usdt_min.setRange(0, 10000)
        self.usdt_min.setDecimals(2)
        self.usdt_min.setValue(0.1)
        usdt_layout.addWidget(self.usdt_min)
        
        usdt_layout.addWidget(QtWidgets.QLabel('-'))
        
        self.usdt_max = QtWidgets.QDoubleSpinBox()
        self.usdt_max.setRange(0, 10000)
        self.usdt_max.setDecimals(2)
        self.usdt_max.setValue(0.5)
        usdt_layout.addWidget(self.usdt_max)
        
        self.btn_random_usdt = QtWidgets.QPushButton('Случайные USDT')
        self.btn_random_usdt.clicked.connect(self._randomize_usdt)
        usdt_layout.addWidget(self.btn_random_usdt)
        
        random_layout.addLayout(usdt_layout)
        
        layout.addWidget(random_group)
        
        # Группа для сохранения и загрузки конфигураций
        config_group = QtWidgets.QGroupBox("Сохранение и загрузка")
        config_layout = QtWidgets.QHBoxLayout(config_group)
        
        self.reward_config_name = QtWidgets.QLineEdit()
        self.reward_config_name.setPlaceholderText('Имя конфигурации')
        config_layout.addWidget(self.reward_config_name)
        
        self.btn_save_config = QtWidgets.QPushButton('Сохранить')
        self.btn_save_config.clicked.connect(self._save_current_config)
        config_layout.addWidget(self.btn_save_config)
        
        self.configs_combo = QtWidgets.QComboBox()
        self._load_config_list()
        config_layout.addWidget(self.configs_combo)
        
        self.btn_load_config = QtWidgets.QPushButton('Загрузить')
        self.btn_load_config.clicked.connect(self._load_selected_config)
        config_layout.addWidget(self.btn_load_config)
        
        layout.addWidget(config_group)
        
        # Настройки отправки
        send_group = QtWidgets.QGroupBox("Отправка наград")
        send_layout = QtWidgets.QVBoxLayout(send_group)
        
        # Настройки отправки
        send_group = QtWidgets.QGroupBox("Отправка наград")
        send_layout = QtWidgets.QVBoxLayout(send_group)
        
        # Параметры отправки
        params_layout = QtWidgets.QHBoxLayout()
        
        # Интервал между отправками
        self.spin_interval = QtWidgets.QSpinBox()
        self.spin_interval.setRange(1, 600)
        self.spin_interval.setValue(5)  # 5 секунд по умолчанию
        params_layout.addWidget(QtWidgets.QLabel("Интервал (сек):"))
        params_layout.addWidget(self.spin_interval)
        
        # Повторы
        self.btn_repeats = QtWidgets.QPushButton(f"Повторов: {self.cfg.get_repeat_count()}")
        self.btn_repeats.clicked.connect(self._set_repeats)
        params_layout.addWidget(self.btn_repeats)
        
        send_layout.addLayout(params_layout)
        
        # Кнопки отправки
        buttons_send_layout = QtWidgets.QHBoxLayout()
        
        self.btn_start = QtWidgets.QPushButton('Начать отправку')
        self.btn_start.clicked.connect(lambda: self._start_sending())
        buttons_send_layout.addWidget(self.btn_start)
        
        self.stop_btn = QtWidgets.QPushButton('Остановить')
        self.stop_btn.clicked.connect(self._cancel_sending)
        self.stop_btn.setEnabled(False)
        buttons_send_layout.addWidget(self.stop_btn)
        
        send_layout.addLayout(buttons_send_layout)
        
        # Индикатор прогресса
        self.progress_send = QtWidgets.QProgressBar()
        send_layout.addWidget(self.progress_send)
        
        layout.addWidget(send_group)
        
        # Загружаем сохраненные награды, если они есть
        self._load_rewards_from_db()
        
        return w
    
    def _add_rewards_row(self):
        """Добавление новой строки в таблицу наград"""
        row = self.table_rewards.rowCount()
        self.table_rewards.insertRow(row)
        
        # Адрес
        addr_item = QtWidgets.QTableWidgetItem("")
        self.table_rewards.setItem(row, 0, addr_item)
        
        # PLEX - значение по умолчанию
        plex_item = QtWidgets.QTableWidgetItem("3")
        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 1, plex_item)
        
        # USDT - пустое значение по умолчанию
        usdt_item = QtWidgets.QTableWidgetItem("")
        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 2, usdt_item)
    
    def _remove_rewards_row(self):
        """Удаление выбранных строк из таблицы наград"""
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        for row in sorted(selected_rows, reverse=True):
            self.table_rewards.removeRow(row)
        
        logger.info(f"Удалено {len(selected_rows)} строк из таблицы наград")
    def _clear_rewards(self):
        """Очистка всей таблицы наград"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Очистить все данные в таблице наград?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._save_rewards_state()  # Сохраняем текущее состояние для возможности отмены
            self.table_rewards.setRowCount(0)
            logger.info("Таблица наград очищена")
    
    def _randomize_plex(self):
        """Генерация случайных значений PLEX для всех строк таблицы с использованием настраиваемого диапазона"""
        plex_min = self.plex_min.value()
        plex_max = self.plex_max.value()
        
        # Проверка корректности диапазона
        if plex_min >= plex_max:
            logger.error("Минимальное значение PLEX ONE должно быть меньше максимального")
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка', 
                'Минимальное значение PLEX ONE должно быть меньше максимального',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Получаем выбранные строки или все строки, если ничего не выбрано
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        if not selected_rows:
            selected_rows = range(self.table_rewards.rowCount())
        
        for row in selected_rows:
            # Генерируем случайное значение в заданном диапазоне с двумя знаками после запятой
            value = round(random.uniform(plex_min, plex_max), 2)
            self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(str(value)))
        
        logger.info(f"Сгенерированы случайные значения PLEX ONE в диапазоне {plex_min}-{plex_max} для {len(selected_rows)} строк")
    
    def _randomize_usdt(self):
        """Генерация случайных значений USDT для строк таблицы с заданными вероятностями"""
        # Значения и их вероятности
        values = [0.12, 0.18, 0.20, 0.24, 0.38]
        weights = [0.70, 0.10, 0.15, 0.025, 0.025]  # Суммарно 5% на 0.24 и 0.38
        
        # Получаем выбранные строки или все строки, если ничего не выбрано
        selected_rows = set(index.row() for index in self.table_rewards.selectedIndexes())
        if not selected_rows:
            selected_rows = range(self.table_rewards.rowCount())
        
        for row in selected_rows:
            # Генерируем случайное значение с заданными вероятностями
            value = random.choices(values, weights=weights, k=1)[0]
            self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(str(value)))
        
        logger.info(f"Сгенерированы случайные значения USDT с заданными вероятностями для {len(selected_rows)} строк")
    
    def _save_current_config(self):
        """Сохранение текущей конфигурации наград в файл"""
        name = self.reward_config_name.text().strip()
        if not name:
            return logger.error("Введите имя для сохранения конфигурации")
        
        addresses = []
        plex_amounts = {}
        usdt_amounts = {}
        
        for row in range(self.table_rewards.rowCount()):
            # Получаем адрес и проверяем его наличие
            addr_item = self.table_rewards.item(row, 0)
            if not addr_item or not addr_item.text().strip():
                continue
            
            address = addr_item.text().strip()
            addresses.append(address)
            
            # Получаем значение PLEX ONE
            plex_item = self.table_rewards.item(row, 1)
            if plex_item and plex_item.text().strip():
                try:
                    plex_amounts[row] = float(plex_item.text())
                except ValueError:
                    pass
            
            # Получаем значение USDT
            usdt_item = self.table_rewards.item(row, 2)
            if usdt_item and usdt_item.text().strip():
                try:
                    usdt_amounts[row] = float(usdt_item.text())
                except ValueError:
                    pass
        
        if not addresses:
            return logger.error("Нет адресов для сохранения")
        
        # Сохраняем конфигурацию
        if save_rewards_config(name, addresses, plex_amounts, usdt_amounts):
            logger.info(f"Конфигурация '{name}' успешно сохранена")
            self._load_config_list()
            self.reward_config_name.clear()
        else:
            logger.error(f"Ошибка при сохранении конфигурации '{name}'")
    
    def _load_config_list(self):
        """Загрузка списка доступных конфигураций в комбобокс"""
        self.configs_combo.clear()
        configs = get_rewards_configs()
        self.configs_combo.addItems(configs)
        
        if configs:
            logger.debug(f"Загружено {len(configs)} конфигураций наград")
    
    def _load_selected_config(self):
        """Загрузка выбранной конфигурации наград из файла"""
        config_name = self.configs_combo.currentText()
        if not config_name:
            return logger.warning("Выберите конфигурацию для загрузки")
        
        config = load_rewards_config(config_name)
        if not config:
            return logger.error(f"Не удалось загрузить конфигурацию '{config_name}'")
        
        # Сохраняем текущее состояние для возможности отмены
        self._save_rewards_state()
        
        # Очищаем таблицу
        self.table_rewards.setRowCount(0)
        
        # Заполняем таблицу данными из конфигурации
        for item in config:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес
            self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(item['address']))
            
            # PLEX
            plex_value = str(item.get('plex', '')) if item.get('plex', 0) > 0 else ''
            self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(plex_value))
            
            # USDT
            usdt_value = str(item.get('usdt', '')) if item.get('usdt', 0) > 0 else ''
            self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(usdt_value))
        
        logger.info(f"Конфигурация '{config_name}' загружена: {len(config)} записей")
    
    def _set_repeats(self):
        """Установка количества повторений серии отправок"""
        value, ok = QtWidgets.QInputDialog.getInt(
            self, 'Повторения', 'Количество повторений серии:',
            self.cfg.get_repeat_count(), 1, 100, 1
        )
        
        if ok:
            self.cfg.set_repeat_count(value)
            self.btn_repeats.setText(f"Повторов: {value}")
            logger.info(f"Установлено {value} повторений серии отправок")
    
    def _save_rewards_state(self):
        """Сохранение текущего состояния таблицы наград для возможности отмены"""
        state = []
        for row in range(self.table_rewards.rowCount()):
            row_data = {}
            
            # Адрес
            addr_item = self.table_rewards.item(row, 0)
            if addr_item:
                row_data['address'] = addr_item.text()
                # Если есть хэш транзакции, сохраняем его тоже
                tx_hash = addr_item.data(QtCore.Qt.UserRole)
                if tx_hash:
                    row_data['tx_hash'] = tx_hash
            
            # PLEX
            plex_item = self.table_rewards.item(row, 1)
            if plex_item:
                row_data['plex'] = plex_item.text()
            
            # USDT
            usdt_item = self.table_rewards.item(row, 2)
            if usdt_item:
                row_data['usdt'] = usdt_item.text()
            
            state.append(row_data)
        
        self.last_rewards_state = state
    
    def _restore_rewards_state(self):
        """Восстановление предыдущего состояния таблицы наград"""
        if not self.last_rewards_state:
            return logger.warning("Нет сохраненного состояния для восстановления")
        
        # Очищаем таблицу
        self.table_rewards.setRowCount(0)
        
        # Восстанавливаем данные
        for row_data in self.last_rewards_state:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес
            addr_item = QtWidgets.QTableWidgetItem(row_data.get('address', ''))
            # Если был хэш транзакции, восстанавливаем его тоже
            if 'tx_hash' in row_data:
                addr_item.setData(QtCore.Qt.UserRole, row_data['tx_hash'])
            self.table_rewards.setItem(row, 0, addr_item)
            
            # PLEX
            plex_item = QtWidgets.QTableWidgetItem(row_data.get('plex', ''))
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT
            usdt_item = QtWidgets.QTableWidgetItem(row_data.get('usdt', ''))
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Восстановлено предыдущее состояние: {len(self.last_rewards_state)} записей")
        
        # Очищаем сохраненное состояние
        self.last_rewards_state = None
    def _load_rewards_from_db(self):
        """Загрузка наград из базы данных в таблицу"""
        try:
            rewards = get_rewards()
            
            for reward in rewards:
                row = self.table_rewards.rowCount()
                self.table_rewards.insertRow(row)
                
                # Адрес
                addr_item = QtWidgets.QTableWidgetItem(reward['address'])
                if reward['tx_hash']:
                    addr_item.setData(QtCore.Qt.UserRole, reward['tx_hash'])
                self.table_rewards.setItem(row, 0, addr_item)
                
                # PLEX
                plex_value = str(reward['plex_amount']) if reward['plex_amount'] > 0 else ''
                plex_item = QtWidgets.QTableWidgetItem(plex_value)
                plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 1, plex_item)
                
                # USDT
                usdt_value = str(reward['usdt_amount']) if reward['usdt_amount'] > 0 else ''
                usdt_item = QtWidgets.QTableWidgetItem(usdt_value)
                usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                self.table_rewards.setItem(row, 2, usdt_item)
                
                # Если награда уже отправлена, выделяем строку
                if reward['sent']:
                    for col in range(self.table_rewards.columnCount()):
                        item = self.table_rewards.item(row, col)
                        if item:
                            item.setBackground(QtGui.QColor('#004400'))
            
            logger.info(f"Загружено {len(rewards)} наград из базы данных")
        except Exception as e:
            logger.error(f"Ошибка при загрузке наград из базы данных: {e}")
            
    def _show_rewards_context_menu(self, position):
        """Отображение контекстного меню для таблицы наград"""
        menu = QtWidgets.QMenu()
        
        add_after = menu.addAction("Добавить строку после")
        remove_row = menu.addAction("Удалить строку")
        copy_addr = menu.addAction("Копировать адрес")
        open_bscscan = menu.addAction("Открыть в BscScan")
        
        # Для функции отмены
        undo_menu = menu.addMenu("Отмена/восстановление")
        restore_state = undo_menu.addAction("Восстановить предыдущее состояние")
        
        # Дополнительные опции
        generate_menu = menu.addMenu("Сгенерировать")
        gen_plex = generate_menu.addAction("Случайные PLEX для выбранных")
        gen_usdt = generate_menu.addAction("Случайные USDT для выбранных")
        
        action = menu.exec_(self.table_rewards.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected_rows = list(set(index.row() for index in self.table_rewards.selectedIndexes()))
        
        if not selected_rows and action != restore_state:
            return
            
        if action == add_after:
            # Добавляем строку после последней выбранной
            row = max(selected_rows) + 1
            self.table_rewards.insertRow(row)
            
            # Адрес - пустой
            self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(""))
            
            # PLEX - значение по умолчанию
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT - пустое значение по умолчанию
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
            
            logger.info(f"Добавлена новая строка после строки {row-1}")
        
        elif action == remove_row:
            # Удаляем строки в обратном порядке, чтобы не сдвигались индексы
            for row in sorted(selected_rows, reverse=True):
                self.table_rewards.removeRow(row)
            
            logger.info(f"Удалено {len(selected_rows)} строк")
        
        elif action == copy_addr:
            # Копируем адрес из первой выбранной строки
            row = selected_rows[0]
            addr_item = self.table_rewards.item(row, 0)
            if addr_item and addr_item.text():
                QtWidgets.QApplication.clipboard().setText(addr_item.text())
                logger.info(f"Адрес скопирован: {addr_item.text()}")
        
        elif action == open_bscscan:
            # Открываем адрес в BscScan
            row = selected_rows[0]
            addr_item = self.table_rewards.item(row, 0)
            if addr_item and addr_item.text():
                url = f"https://bscscan.com/address/{addr_item.text()}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открыт адрес в BscScan: {addr_item.text()}")
        
        elif action == restore_state:
            # Восстанавливаем предыдущее состояние таблицы
            self._restore_rewards_state()
        
        elif action == gen_plex:
            # Генерируем случайные значения PLEX для выбранных строк
            for row in selected_rows:
                value = round(random.uniform(2, 4), 2)
                self.table_rewards.setItem(row, 1, QtWidgets.QTableWidgetItem(str(value)))
            
            logger.info(f"Сгенерированы случайные значения PLEX ONE для {len(selected_rows)} строк")
        
        elif action == gen_usdt:
            # Генерируем случайные значения USDT для выбранных строк
            for row in selected_rows:
                value = round(random.uniform(0.1, 0.5), 2)
                self.table_rewards.setItem(row, 2, QtWidgets.QTableWidgetItem(str(value)))
            
            logger.info(f"Сгенерированы случайные значения USDT для {len(selected_rows)} строк")

    # --- Таб «Прямая отправка» ---
    def _tab_direct_send(self):
        """Создание вкладки для периодической прямой отправки токенов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Форма для параметров отправки
        form = QtWidgets.QFormLayout()
        
        # Адрес получателя
        self.direct_send_address = QtWidgets.QLineEdit()
        self.direct_send_address.setPlaceholderText('0x...')
        form.addRow('Адрес получателя:', self.direct_send_address)
        
        # Выбор типа токена
        self.direct_send_token_group = QtWidgets.QButtonGroup(w)
        self.direct_send_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.direct_send_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.direct_send_plex_radio.setChecked(True)
        self.direct_send_token_group.addButton(self.direct_send_plex_radio)
        self.direct_send_token_group.addButton(self.direct_send_usdt_radio)
        
        token_layout = QtWidgets.QHBoxLayout()
        token_layout.addWidget(self.direct_send_plex_radio)
        token_layout.addWidget(self.direct_send_usdt_radio)
        form.addRow('Тип токена:', token_layout)
        
        # Сумма отправки
        self.direct_send_amount = QtWidgets.QDoubleSpinBox()
        self.direct_send_amount.setRange(0.0001, 1000000)
        self.direct_send_amount.setDecimals(8)
        self.direct_send_amount.setValue(3)  # По умолчанию 3 токена
        form.addRow('Количество токенов:', self.direct_send_amount)
        
        # Интервал отправки
        self.direct_send_interval = QtWidgets.QSpinBox()
        self.direct_send_interval.setRange(1, 3600)
        self.direct_send_interval.setValue(60)  # По умолчанию 60 секунд
        form.addRow('Интервал (сек):', self.direct_send_interval)
        
        # Количество периодов
        self.direct_send_periods = QtWidgets.QSpinBox()
        self.direct_send_periods.setRange(0, 10000)
        self.direct_send_periods.setValue(10)  # По умолчанию 10 периодов
        self.direct_send_periods.setSpecialValueText("Бесконечно")  # Текст для значения 0
        form.addRow('Количество периодов (0 - бесконечно):', self.direct_send_periods)
        
        layout.addLayout(form)
        
        # Кнопки управления
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.direct_send_start_btn = QtWidgets.QPushButton('Запустить отправку')
        self.direct_send_start_btn.clicked.connect(self._start_direct_send_ui)
        buttons_layout.addWidget(self.direct_send_start_btn)
        
        self.direct_send_stop_btn = QtWidgets.QPushButton('Остановить отправку')
        self.direct_send_stop_btn.clicked.connect(self._cancel_direct_sending)
        self.direct_send_stop_btn.setEnabled(False)
        buttons_layout.addWidget(self.direct_send_stop_btn)
        
        layout.addLayout(buttons_layout)
        
        # Индикатор прогресса
        self.direct_send_progress = QtWidgets.QProgressBar()
        layout.addWidget(self.direct_send_progress)
        
        # Статус отправки
        self.direct_send_status = QtWidgets.QTextEdit()
        self.direct_send_status.setReadOnly(True)
        self.direct_send_status.setMaximumHeight(150)
        layout.addWidget(self.direct_send_status)
        
        return w
    def _start_direct_send_ui(self):
        """Обработка запуска периодической отправки из интерфейса"""
        if getattr(self, 'direct_send_active', False):
            return logger.warning("Отправка уже запущена")
            
        address = self.direct_send_address.text().strip()
        token_type = 'plex' if self.direct_send_plex_radio.isChecked() else 'usdt'
        amount = self.direct_send_amount.value()
        interval = self.direct_send_interval.value()
        periods = self.direct_send_periods.value()
        
        # Проверка адреса
        if blockchain_enabled and not Web3.is_address(address):
            return logger.error("Некорректный адрес получателя")
            
        # Настройка интерфейса
        self.direct_send_start_btn.setEnabled(False)
        self.direct_send_stop_btn.setEnabled(True)
        
        # Запуск отправки
        self._start_direct_sending(address, token_type, amount, interval, periods)

    def _direct_send_periodic_send(self):
        """
        Метод для периодической отправки токенов по таймеру.
        Обрабатывает каждую итерацию периодической отправки.
        """
        if not self.direct_send_active:
            self.direct_send_timer.stop()
            return
            
        try:
            # Увеличиваем счетчик текущего периода
            self.direct_send_current_period += 1
            
            # Обновляем статус в интерфейсе
            status_text = f"Отправка #{self.direct_send_current_period}"
            if self.direct_send_total_periods > 0:
                status_text += f" из {self.direct_send_total_periods}"
            QtCore.QMetaObject.invokeMethod(
                self.direct_send_status, "append", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(str, status_text)
            )
            
            # Проверяем, не превысили ли общее количество периодов
            if (self.direct_send_total_periods > 0 and 
                self.direct_send_current_period > self.direct_send_total_periods):
                logger.info("Периодическая отправка завершена")
                self.direct_send_active = False
                self.direct_send_timer.stop()
                self.update_progress_signal.emit("direct_send", 100)
                
                # Обновляем интерфейс
                QtCore.QMetaObject.invokeMethod(
                    self.direct_send_start_btn, "setEnabled", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(bool, True)
                )
                QtCore.QMetaObject.invokeMethod(
                    self.direct_send_stop_btn, "setEnabled", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(bool, False)
                )
                return
                
            # Получаем параметры отправки
            params = self.direct_send_params
            to_addr = params.get('to_addr')
            token_type = params.get('token_type', 'plex')
            amount = params.get('amount', 0)
            
            if not to_addr or amount <= 0:
                logger.error("Некорректные параметры для периодической отправки")
                self.direct_send_active = False
                self.direct_send_timer.stop()
                return
                
            # Выполняем отправку токена
            logger.info(f"Периодическая отправка #{self.direct_send_current_period}: "
                      f"{amount} {token_type.upper()} на адрес {to_addr}")
            
            # Отправка PLEX или USDT в зависимости от типа токена
            self._send_token(to_addr, amount, token_type)
                
            # Обновляем прогресс
            if self.direct_send_total_periods > 0:
                progress = int(self.direct_send_current_period / self.direct_send_total_periods * 100)
                self.update_progress_signal.emit("direct_send", progress)
            else:
                # Если общее количество не задано, показываем прогресс итераций
                self.update_progress_signal.emit("direct_send", self.direct_send_current_period % 100)
                
        except Exception as e:
            logger.error(f"Ошибка при периодической отправке: {e}")
            self.direct_send_active = False
            self.direct_send_timer.stop()

    def _start_direct_sending(self, to_addr, token_type, amount, interval, periods=0):
        """
        Запуск периодической отправки токенов.
        
        Args:
            to_addr (str): Адрес получателя
            token_type (str): Тип токена ('plex' или 'usdt')
            amount (float): Количество токенов для отправки
            interval (int): Интервал между отправками в секундах
            periods (int, optional): Общее количество периодов отправки, 0 для бесконечной отправки
        """
        if not self.pk:
            return logger.error("Приватный ключ не задан")
            
        if not to_addr or amount <= 0 or interval <= 0:
            return logger.error("Некорректные параметры для периодической отправки")
            
        if self.direct_send_active:
            # Останавливаем текущую периодическую отправку
            self.direct_send_active = False
            self.direct_send_timer.stop()
            
        # Настраиваем параметры новой периодической отправки
        self.direct_send_params = {
            'to_addr': to_addr,
            'token_type': token_type,
            'amount': amount
        }
        self.direct_send_current_period = 0
        self.direct_send_total_periods = periods
        self.direct_send_active = True
        
        # Очищаем область статуса
        self.direct_send_status.clear()
        
        # Настраиваем и запускаем таймер
        self.direct_send_timer.setInterval(interval * 1000)  # Переводим секунды в миллисекунды
        self.direct_send_timer.start()
        
        status_msg = f"Запущена периодическая отправка {amount} {token_type.upper()} " \
                     f"на {to_addr} с интервалом {interval} сек " \
                     f"({periods} периодов)" if periods > 0 else "(без ограничения)"
        
        logger.info(status_msg)
        self.direct_send_status.append(status_msg)

    def _cancel_direct_sending(self):
        """Остановка периодической отправки токенов"""
        if self.direct_send_active:
            self.direct_send_active = False
            self.direct_send_timer.stop()
            logger.info("Периодическая отправка остановлена")
            self.update_progress_signal.emit("direct_send", 100)
            
            # Обновляем интерфейс
            self.direct_send_start_btn.setEnabled(True)
            self.direct_send_stop_btn.setEnabled(False)
            self.direct_send_status.append("Отправка остановлена")

    def _send_token(self, to_addr, amount, token_type):
        """
        Отправка токена (PLEX или USDT)
        
        Args:
            to_addr (str): Адрес получателя
            amount (float): Количество токенов
            token_type (str): Тип токена ('plex' или 'usdt')
        """
        try:
            if not blockchain_enabled:
                logger.warning("Блокчейн отключен. Имитация отправки...")
                time.sleep(1)
                tx_hash = "0x" + ''.join(random.choices('0123456789abcdef', k=64))
                add_history(token_type.upper(), to_addr, amount, tx_hash)
                return tx_hash
                
            token_addr = PLEX_CONTRACT if token_type.lower() == 'plex' else USDT_CONTRACT
            w3 = self.rpc.web3()  # Получаем экземпляр web3 с уже внедренным POA middleware
            
            # Преобразование адресов в формат checksummed
            from_addr = w3.eth.account.from_key(self.pk).address
            to_addr = w3.to_checksum_address(to_addr)
            token_addr = w3.to_checksum_address(token_addr)
            
            # Получение контракта токена
            token_contract = w3.eth.contract(address=token_addr, abi=ERC20_ABI)
            
            # Получение decimals токена
            decimals = token_contract.functions.decimals().call()
            
            # Преобразование суммы в wei
            amount_wei = int(amount * 10**decimals)
            
            # Установка цены газа
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            # Создание транзакции
            tx_params = {
                'from': from_addr,
                'nonce': w3.eth.get_transaction_count(from_addr),
                'gasPrice': gas_price
            }
            
            # Оценка газа для транзакции
            gas_estimate = token_contract.functions.transfer(
                to_addr, amount_wei
            ).estimate_gas(tx_params)
            
            tx_params['gas'] = int(gas_estimate * 1.2)  # Добавляем 20% запаса для газа
            
            # Создание подписанной транзакции
            tx = token_contract.functions.transfer(
                to_addr, amount_wei
            ).build_transaction(tx_params)
            
            signed_tx = w3.eth.account.sign_transaction(tx, self.pk)
            
            # Отправка транзакции
            tx_hash = w3.eth.send_raw_transaction(signed_tx.raw_transaction)
            tx_hash_hex = tx_hash.hex()
            
            # Добавление в историю
            add_history(token_type.upper(), to_addr, amount, tx_hash_hex)
            
            logger.info(f"Токен {token_type.upper()} отправлен: {amount} на {to_addr}, TX: {tx_hash_hex}")
            return tx_hash_hex
                
        except Exception as e:
            logger.error(f"Ошибка при отправке {token_type.upper()}: {e}")
            return None
    def _collect_tasks(self):
        """Сбор задач для отправки из таблицы"""
        tasks = []
        for i in range(self.table_rewards.rowCount()):
            addr = self.table_rewards.item(i, 0)
            if not addr or not addr.text():
                continue
                
            plex_item = self.table_rewards.item(i, 1)
            usdt_item = self.table_rewards.item(i, 2)
            
            # Проверка на наличие значений и конвертация в числа
            plex_amount = 0
            if plex_item and plex_item.text():
                try:
                    plex_amount = float(plex_item.text())
                except ValueError:
                    pass
                    
            usdt_amount = 0
            if usdt_item and usdt_item.text():
                try:
                    usdt_amount = float(usdt_item.text())
                except ValueError:
                    pass
                    
            # Добавляем задачу, если есть хотя бы одна ненулевая сумма
            if plex_amount > 0 or usdt_amount > 0:
                tasks.append((addr, plex_amount, usdt_amount))
                
        return tasks
        
    def _start_sending(self, mark_rewarded=False):
        """Запуск процесса отправки наград"""
        if self.sending:
            return logger.warning("Отправка уже запущена")
            
        if not self.pk:
            return logger.error("Приватный ключ не задан")
            
        tasks = self._collect_tasks()
        if not tasks:
            return logger.error("Нет задач для отправки")
            
        # Подтверждение отправки
        total_plex = sum(plex for _, plex, _ in tasks)
        total_usdt = sum(usdt for _, _, usdt in tasks)
        
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение отправки', 
            f'Будет отправлено {len(tasks)} адресам:\n\n'
            f'Всего PLEX: {total_plex}\n'
            f'Всего USDT: {total_usdt}\n\n'
            f'Продолжить?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
            
        # Запуск отправки в отдельном потоке
        self.sending = True
        self.btn_start.setEnabled(False)
        self.stop_btn.setEnabled(True)
        
        self.repeats_remain = self.cfg.get_repeat_count()
        threading.Thread(
            target=self._sending_thread,
            args=(tasks, self.spin_interval.value(), mark_rewarded),
            daemon=True
        ).start()
    
    def _sending_thread(self, tasks, interval, mark_rewarded=False):
        """Поток для отправки токенов с улучшенным отслеживанием"""
        try:
            total = len(tasks)
            batch_id = datetime.now().strftime('%Y%m%d%H%M%S')
            transaction_results = []
            
            # Подготавливаем и отображаем очередь
            queue_items = []
            for i, (addr, plex, usdt) in enumerate(tasks):
                if hasattr(addr, 'text'):
                    to_addr = addr.text()
                    tx_hash = addr.data(QtCore.Qt.UserRole)
                else:
                    to_addr = addr
                    tx_hash = None
                    
                if plex > 0:
                    queue_items.append({
                        'status': 'pending',
                        'address': to_addr,
                        'token': 'PLEX',
                        'amount': plex,
                        'time': None,
                        'tx_hash': None,
                        'original_tx': tx_hash  # Исходная транзакция для награды
                    })
                    
                if usdt > 0:
                    queue_items.append({
                        'status': 'pending',
                        'address': to_addr,
                        'token': 'USDT',
                        'amount': usdt,
                        'time': None,
                        'tx_hash': None,
                        'original_tx': tx_hash  # Исходная транзакция для награды
                    })
                    
            # Создаем копию списка для передачи через сигнал
            queue_items_copy = queue_items.copy() if queue_items else []
            # Запускаем обновление UI через основной поток
            self.update_status_signal.emit("Обновление очереди")
            QtWidgets.QApplication.processEvents()
            self._update_queue_table(queue_items_copy) # Прямой вызов декорированного метода
            
            # Активируем кнопки управления очередью
            QtCore.QMetaObject.invokeMethod(
                self.pause_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.cancel_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            
            # Повторяем серию указанное количество раз
            while self.repeats_remain > 0 and self.sending:
                logger.info(f"Начинаем серию отправок (осталось повторений: {self.repeats_remain})")
                QtCore.QMetaObject.invokeMethod(
                    self.queue_status_label, "setText", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(str, f"Отправка (серия {self.cfg.get_repeat_count() - self.repeats_remain + 1}/{self.cfg.get_repeat_count()})")
                )
                
                # Обрабатываем каждый элемент в очереди
                completed_count = 0
                for i, item in enumerate(queue_items):
                    if not self.sending or self.queue_paused:
                        # Если отправка приостановлена, ждем возобновления
                        while self.queue_paused and self.sending:
                            time.sleep(0.5)
                        
                        if not self.sending:
                            break
                    
                    # Пропускаем уже обработанные элементы
                    if item['status'] != 'pending':
                        continue
                    
                    # Обновляем прогресс
                    completed_count += 1
                    progress = int((completed_count / len(queue_items)) * 100)
                    self.update_progress_signal.emit("send", progress)
                    QtCore.QMetaObject.invokeMethod(
                        self.queue_progress_label, "setText", 
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(str, f"{completed_count}/{len(queue_items)} отправлено")
                    )
                    
                    # Обновляем статус элемента
                    item['status'] = 'processing'
                    self.queue_item_update_signal.emit(i, item)
                    
                    # Отправка токена
                    logger.info(f"Отправка {item['amount']} {item['token']} на адрес {item['address']}")
                    tx_hash = self._send_token(item['address'], item['amount'], item['token'].lower())
                    
                    # Обновляем результат
                    item['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    item['tx_hash'] = tx_hash
                    
                    if tx_hash:
                        item['status'] = 'success'
                        transaction_results.append({
                            'address': item['address'],
                            'token': item['token'],
                            'amount': item['amount'],
                            'tx_hash': tx_hash,
                            'status': 'success',
                            'time': item['time']
                        })
                    else:
                        item['status'] = 'failed'
                    
                    # Если нужно отметить транзакцию как награжденную и есть хэш исходной транзакции
                    if mark_rewarded and item['original_tx'] and item['status'] == 'success':
                        mark_transaction_rewarded(item['original_tx'])
                        logger.info(f"Транзакция {item['original_tx']} отмечена как награжденная")
                    
                    # Обновляем элемент в таблице очереди
                    QtCore.QMetaObject.invokeMethod(
                        self, "_update_queue_item", 
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(int, i),
                        QtCore.Q_ARG(dict, item)
                    )
                    
                    # Добавляем в таблицу завершенных
                    self.completed_item_signal.emit(item)
                    
                    # Задержка между отправками
                    if i < len(queue_items) - 1 and self.sending and not self.queue_paused:
                        time.sleep(interval)
                
                # Уменьшаем счетчик оставшихся повторений
                self.repeats_remain -= 1
                
                # Задержка между сериями
                if self.repeats_remain > 0 and self.sending:
                    logger.info(f"Серия завершена. Ожидание перед следующей серией (осталось: {self.repeats_remain})")
                    time.sleep(interval * 2)  # Увеличенный интервал между сериями
            
            # Завершение отправки
            logger.info("Отправка завершена")
            self.update_progress_signal.emit("send", 100)
            
            # Показываем диалог с результатами
            if transaction_results:
                QtCore.QMetaObject.invokeMethod(
                    self, "_show_transaction_results", 
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(str, batch_id),
                    QtCore.Q_ARG(list, transaction_results)
                )
            
        except Exception as e:
            logger.error(f"Ошибка при отправке: {e}")
        finally:
            self.sending = False
            self.queue_paused = False
            
            # Обновляем UI в основном потоке
            QtCore.QMetaObject.invokeMethod(
                self.btn_start, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.stop_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.pause_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.resume_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.cancel_queue_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            QtCore.QMetaObject.invokeMethod(
                self.queue_status_label, "setText", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(str, "Завершено")
            )
    
    def _cancel_sending(self):
        """Остановка процесса отправки"""
        if self.sending:
            self.sending = False
            logger.info("Запрошена остановка отправки")
            self.stop_btn.setEnabled(False)

    # --- Таб «Поиск транзакций» ---
    def _tab_paginated_search(self):
        """Создание вкладки для постраничного поиска транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Форма для ввода параметров поиска
        form = QtWidgets.QFormLayout()
        
        # Адрес кошелька системы
        self.system_address = QtWidgets.QLineEdit()
        self.system_address.setPlaceholderText('0x...')
        form.addRow('Адрес системы (кошелька):', self.system_address)
        
        # Адрес токена
        self.token_address = QtWidgets.QLineEdit()
        self.token_address.setPlaceholderText('0x...')
        self.token_address.setText(PLEX_CONTRACT)
        form.addRow('Адрес токена (контракт):', self.token_address)
        
        # Переключатель режима поиска
        radio_layout = QtWidgets.QHBoxLayout()
        self.search_mode_group = QtWidgets.QButtonGroup(w)
        self.search_exact_radio = QtWidgets.QRadioButton('Точная сумма')
        self.search_range_radio = QtWidgets.QRadioButton('Диапазон сумм')
        self.search_exact_radio.setChecked(True)
        self.search_mode_group.addButton(self.search_exact_radio)
        self.search_mode_group.addButton(self.search_range_radio)
        radio_layout.addWidget(self.search_exact_radio)
        radio_layout.addWidget(self.search_range_radio)
        form.addRow('Режим поиска:', radio_layout)
        
        # Точная сумма
        self.exact_amount = QtWidgets.QDoubleSpinBox()
        self.exact_amount.setRange(0.0, 1000000.0)
        self.exact_amount.setDecimals(8)
        self.exact_amount.setValue(30.0)  # Предустановленное значение 30 PLEX ONE
        form.addRow('Точная сумма:', self.exact_amount)
        
        # Диапазон сумм
        range_layout = QtWidgets.QHBoxLayout()
        self.min_amount = QtWidgets.QDoubleSpinBox()
        self.min_amount.setRange(0.0, 1000000.0)
        self.min_amount.setDecimals(8)
        self.min_amount.setEnabled(False)
        
        self.max_amount = QtWidgets.QDoubleSpinBox()
        self.max_amount.setRange(0.0, 1000000.0)
        self.max_amount.setDecimals(8)
        self.max_amount.setEnabled(False)
        
        range_layout.addWidget(QtWidgets.QLabel('От:'))
        range_layout.addWidget(self.min_amount)
        range_layout.addWidget(QtWidgets.QLabel('До:'))
        range_layout.addWidget(self.max_amount)
        form.addRow('Диапазон:', range_layout)
        
        # Минимальное количество транзакций
        self.min_tx_count = QtWidgets.QSpinBox()
        self.min_tx_count.setRange(1, 1000)
        self.min_tx_count.setValue(1)
        form.addRow('Мин. кол-во транзакций:', self.min_tx_count)
        # Настройки пагинации
        pagination_layout = QtWidgets.QHBoxLayout()
        
        self.page_size = QtWidgets.QSpinBox()
        self.page_size.setRange(10, 10000)
        self.page_size.setValue(1000)
        pagination_layout.addWidget(QtWidgets.QLabel('Размер страницы:'))
        pagination_layout.addWidget(self.page_size)
        
        self.max_pages = QtWidgets.QSpinBox()
        self.max_pages.setRange(1, 1000)
        self.max_pages.setValue(100)
        pagination_layout.addWidget(QtWidgets.QLabel('Макс. страниц:'))
        pagination_layout.addWidget(self.max_pages)
        
        self.api_delay = QtWidgets.QSpinBox()
        self.api_delay.setRange(1, 60)
        self.api_delay.setValue(1)
        pagination_layout.addWidget(QtWidgets.QLabel('Задержка (сек):'))
        pagination_layout.addWidget(self.api_delay)
        
        form.addRow('Настройки API:', pagination_layout)
        
        # Опция для учета отдельных транзакций
        self.track_tx_checkbox = QtWidgets.QCheckBox("Учитывать отдельные транзакции для награждения")
        self.track_tx_checkbox.setChecked(True)
        form.addRow("", self.track_tx_checkbox)
        
        layout.addLayout(form)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        self.start_search_btn = QtWidgets.QPushButton('Начать поиск')
        self.start_search_btn.clicked.connect(self._start_paginated_search)
        buttons_layout.addWidget(self.start_search_btn)
        
        self.stop_search_btn = QtWidgets.QPushButton('Остановить')
        self.stop_search_btn.clicked.connect(self._stop_paginated_search)
        self.stop_search_btn.setEnabled(False)
        buttons_layout.addWidget(self.stop_search_btn)
        
        self.clear_results_btn = QtWidgets.QPushButton('Очистить результаты')
        self.clear_results_btn.clicked.connect(self._clear_search_results)
        buttons_layout.addWidget(self.clear_results_btn)
        
        layout.addLayout(buttons_layout)
        
        # Таблица результатов
        self.search_results_table = QtWidgets.QTableWidget(0, 3)
        self.search_results_table.setHorizontalHeaderLabels(['Отправитель', 'Количество TX', 'Последняя TX'])
        self.search_results_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.search_results_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.search_results_table.setSelectionMode(QtWidgets.QTableView.ExtendedSelection)
        self.search_results_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.search_results_table.customContextMenuRequested.connect(self._show_search_context_menu)
        layout.addWidget(self.search_results_table)
        
        # Кнопки для экспорта
        export_layout = QtWidgets.QHBoxLayout()
        
        self.export_csv_btn = QtWidgets.QPushButton('Экспорт в CSV')
        self.export_csv_btn.clicked.connect(self._export_search_results)
        export_layout.addWidget(self.export_csv_btn)
        
        self.copy_to_clipboard_btn = QtWidgets.QPushButton('Копировать адреса')
        self.copy_to_clipboard_btn.clicked.connect(self._copy_search_results)
        export_layout.addWidget(self.copy_to_clipboard_btn)
        
        self.import_to_rewards_btn = QtWidgets.QPushButton('В раздел наград')
        self.import_to_rewards_btn.clicked.connect(self._import_to_rewards)
        export_layout.addWidget(self.import_to_rewards_btn)
        
        layout.addLayout(export_layout)
        
        # Индикатор прогресса
        self.progress_search = QtWidgets.QProgressBar()
        layout.addWidget(self.progress_search)
        
        # Подключаем переключение режима поиска
        self.search_exact_radio.toggled.connect(self._toggle_search_mode_paginated)
        self.search_range_radio.toggled.connect(self._toggle_search_mode_paginated)
        
        return w
        
    def _toggle_search_mode_paginated(self):
        """Переключение режима поиска в разделе постраничного поиска"""
        is_exact = self.search_exact_radio.isChecked()
        
        self.exact_amount.setEnabled(is_exact)
        self.min_amount.setEnabled(not is_exact)
        self.max_amount.setEnabled(not is_exact)
        
        logger.debug(f"Режим поиска переключен на: {'точная сумма' if is_exact else 'диапазон'}")
        
    def _start_paginated_search(self):
        """Запуск процесса постраничного поиска транзакций"""
        if self.is_searching:
            return logger.warning("Поиск уже запущен")
        
        # Получаем параметры поиска
        wallet_address = self.system_address.text().strip()
        token_address = self.token_address.text().strip()
        
        # Проверяем корректность адресов
        if blockchain_enabled and not Web3.is_address(wallet_address):
            return logger.error("Некорректный адрес кошелька")
        
        if token_address and blockchain_enabled and not Web3.is_address(token_address):
            return logger.error("Некорректный адрес токена")
        
        # Определяем режим поиска и параметры
        exact_amount = None
        min_amount = None
        max_amount = None
        
        if self.search_exact_radio.isChecked():
            exact_amount = self.exact_amount.value()
            if exact_amount <= 0:
                return logger.error("Точная сумма должна быть больше 0")
        else:
            min_amount = self.min_amount.value()
            max_amount = self.max_amount.value()
            if min_amount < 0:
                return logger.error("Минимальная сумма должна быть не меньше 0")
            if max_amount <= min_amount:
                return logger.error("Максимальная сумма должна быть больше минимальной")
        
        # Получаем настройки пагинации
        page_size = self.page_size.value()
        max_pages = self.max_pages.value()
        delay_seconds = self.api_delay.value()
        min_tx_count = self.min_tx_count.value()
        track_individual_tx = self.track_tx_checkbox.isChecked()
        
        # Запускаем поиск в отдельном потоке
        self.is_searching = True
        self.start_search_btn.setEnabled(False)
        self.stop_search_btn.setEnabled(True)
        self.stop_search_event.clear()  # Сбрасываем флаг остановки
        
        # Очищаем предыдущие результаты в таблице
        self.search_results_table.setRowCount(0)
        
        self.search_thread = threading.Thread(
            target=self._paginated_search_thread, 
            args=(wallet_address, token_address, exact_amount, min_amount, max_amount, 
                  page_size, max_pages, delay_seconds, min_tx_count, track_individual_tx),
            daemon=True
        )
        self.search_thread.start()
        
    def _paginated_search_thread(self, wallet_address, token_address, exact_amount, min_amount, max_amount, 
                               page_size, max_pages, delay_seconds, min_tx_count, track_individual_tx):
        """Поток для выполнения постраничного поиска транзакций"""
        try:
            logger.info(f"Начинаем поиск транзакций для адреса {wallet_address}")
            
            # Сохраняем параметры поиска для записи в базу данных
            search_params = {
                'wallet_address': wallet_address,
                'token_address': token_address,
                'exact_amount': exact_amount,
                'min_amount': min_amount,
                'max_amount': max_amount,
                'min_tx_count': min_tx_count,
                'search_time': datetime.now().isoformat()
            }
            
            # Выполняем поиск с постраничной обработкой
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=wallet_address,
                token_contract=token_address if token_address else None,
                exact_amount=exact_amount,
                min_amount=min_amount,
                max_amount=max_amount,
                page_size=page_size,
                max_pages=max_pages,
                delay_seconds=delay_seconds,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Сохраняем найденные транзакции в базу данных
            for tx in matching_txs:
                add_found_transaction(tx, search_params)
            
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {addr: count for addr, count in sender_counter.items() 
                              if count >= min_tx_count}
            
               # Формируем информацию о последних транзакциях
            sender_last_tx = {}
            for tx in matching_txs:
                try:
                    sender = tx.get('from', '').lower()
                    if sender in filtered_senders:
                        tx_time = datetime.fromtimestamp(int(tx.get('timeStamp', 0)))
                        if sender not in sender_last_tx or tx_time > sender_last_tx[sender][0]:
                            sender_last_tx[sender] = (tx_time, tx.get('hash', ''))
                except Exception as e:
                    logger.error(f"Ошибка обработки транзакции: {e}")
      
            # Создаем список результатов для обновления таблицы
            results = []
            for sender, count in filtered_senders.items():
                last_tx_info = sender_last_tx.get(sender, (datetime.now(), ''))
                results.append((sender, count, last_tx_info[0].strftime('%Y-%m-%d %H:%M:%S'), last_tx_info[1]))
            
            # Отправляем результаты в основной поток для обновления UI
            self.update_table_signal.emit(results, filtered_senders, sender_transactions or {})
            logger.info(f"Поиск завершен. Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями")
            
        except Exception as e:
            logger.error(f"Ошибка при выполнении поиска: {e}")
        finally:
            self.is_searching = False
            # Обновляем состояние кнопок в основном потоке
            QtCore.QMetaObject.invokeMethod(
                self.start_search_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, True)
            )
            QtCore.QMetaObject.invokeMethod(
                self.stop_search_btn, "setEnabled", 
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(bool, False)
            )
            self.update_progress_signal.emit("search", 100)
    def _update_search_results(self, results, sender_counter, sender_transactions):
        """Обновление таблицы результатов поиска"""
        self.search_results_table.setRowCount(0)
        
        for sender, count, last_tx_time, last_tx_hash in results:
            row = self.search_results_table.rowCount()
            self.search_results_table.insertRow(row)
            
            # Адрес отправителя
            sender_item = QtWidgets.QTableWidgetItem(sender)
            sender_item.setData(QtCore.Qt.UserRole, last_tx_hash)  # Сохраняем хэш последней транзакции
            
            # Если включен режим награждения за каждую транзакцию и у отправителя есть транзакции,
            # сохраняем их в данных элемента
            if self.reward_per_tx_mode and sender in sender_transactions:
                sender_item.setData(QtCore.Qt.UserRole + 1, sender_transactions[sender])
                
            self.search_results_table.setItem(row, 0, sender_item)
            
            # Количество транзакций
            count_item = QtWidgets.QTableWidgetItem(str(count))
            self.search_results_table.setItem(row, 1, count_item)
            
            # Время последней транзакции
            time_item = QtWidgets.QTableWidgetItem(last_tx_time)
            self.search_results_table.setItem(row, 2, time_item)
    
    def _stop_paginated_search(self):
        """Остановка процесса поиска транзакций"""
        if not self.is_searching:
            return
        
        self.stop_search_event.set()  # Установка флага остановки
        logger.info("Отправлен запрос на остановку поиска транзакций...")
        self.stop_search_btn.setEnabled(False)
    
    def _clear_search_results(self):
        """Очистка результатов поиска"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Очистить все результаты поиска?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.search_results_table.setRowCount(0)
            count = clear_found_transactions()
            logger.info(f"Очищено {count} записей из истории поиска")
    
    def _show_search_context_menu(self, position):
        """Отображение контекстного меню для таблицы результатов поиска"""
        menu = QtWidgets.QMenu()
        
        copy_action = menu.addAction("Копировать адрес")
        open_bscscan = menu.addAction("Открыть в BscScan")
        add_to_rewards = menu.addAction("Добавить в награды")
        
        # Если включен режим награждения за каждую транзакцию, добавляем соответствующий пункт
        if self.reward_per_tx_mode:
            add_all_tx_to_rewards = menu.addAction("Добавить все транзакции в награды")
        
        action = menu.exec_(self.search_results_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected_indexes = self.search_results_table.selectedIndexes()
        if not selected_indexes:
            return
            
        # Получаем данные выбранных строк
        selected_row = selected_indexes[0].row()
        sender_address = self.search_results_table.item(selected_row, 0).text()
        tx_hash = self.search_results_table.item(selected_row, 0).data(QtCore.Qt.UserRole)
        
        if action == copy_action:
            QtWidgets.QApplication.clipboard().setText(sender_address)
            logger.info(f"Скопирован адрес: {sender_address}")
        
        elif action == open_bscscan:
            if tx_hash:
                url = f"https://bscscan.com/tx/{tx_hash}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открываем транзакцию в BscScan: {tx_hash}")
            else:
                url = f"https://bscscan.com/address/{sender_address}"
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
                logger.info(f"Открываем адрес в BscScan: {sender_address}")
        
        elif action == add_to_rewards:
            self._add_to_rewards_tab(sender_address)
        
        elif self.reward_per_tx_mode and action == add_all_tx_to_rewards:
            # Получаем список транзакций из данных элемента
            tx_list = self.search_results_table.item(selected_row, 0).data(QtCore.Qt.UserRole + 1)
            if tx_list:
                self._add_tx_list_to_rewards(sender_address, tx_list)
            else:
                logger.warning(f"Нет данных о транзакциях для {sender_address}")
    
    def _add_tx_list_to_rewards(self, sender_addr, tx_list):
        """Добавление списка транзакций в таблицу наград"""
        if not tx_list:
            return logger.warning(f"Список транзакций пуст для {sender_addr}")
        
        # Сохраняем текущее состояние таблицы наград
        self._save_rewards_state()
        
        # Добавляем каждую транзакцию в таблицу наград
        for tx_info in tx_list:
            row = self.table_rewards.rowCount()
            self.table_rewards.insertRow(row)
            
            # Адрес отправителя
            addr_item = QtWidgets.QTableWidgetItem(sender_addr)
            addr_item.setData(QtCore.Qt.UserRole, tx_info['hash'])
            self.table_rewards.setItem(row, 0, addr_item)
            
            # PLEX - значение по умолчанию
            plex_item = QtWidgets.QTableWidgetItem("3")
            plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 1, plex_item)
            
            # USDT - пустое значение по умолчанию
            usdt_item = QtWidgets.QTableWidgetItem("")
            usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
            self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Добавлено {len(tx_list)} транзакций отправителя {sender_addr} в таблицу наград")
        
        # Предлагаем сгенерировать случайные значения
        reply = QtWidgets.QMessageBox.question(
            self, 'Рандомизация', 
            'Сгенерировать случайные значения PLEX и USDT для транзакций?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.Yes
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._randomize_plex()
            self._randomize_usdt()
            
    def _export_search_results(self):
        """Экспорт результатов поиска в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,tx_count,last_tx_time\n')
                for row in range(self.search_results_table.rowCount()):
                    address = self.search_results_table.item(row, 0).text()
                    count = self.search_results_table.item(row, 1).text()
                    time = self.search_results_table.item(row, 2).text()
                    f.write(f'"{address}",{count},"{time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте в CSV: {e}")
    
    def _copy_search_results(self):
        """Копирование адресов из результатов поиска в буфер обмена"""
        addresses = []
        
        # Если есть выбранные строки, копируем только их
        selected_rows = set(index.row() for index in self.search_results_table.selectedIndexes())
        if selected_rows:
            for row in selected_rows:
                addresses.append(self.search_results_table.item(row, 0).text())
        else:
            # Иначе копируем все адреса
            for row in range(self.search_results_table.rowCount()):
                addresses.append(self.search_results_table.item(row, 0).text())
        
        if addresses:
            QtWidgets.QApplication.clipboard().setText('\n'.join(addresses))
            logger.info(f"Скопировано {len(addresses)} адресов в буфер обмена")
        else:
            logger.warning("Нет адресов для копирования")
    
    def _import_to_rewards(self):
        """Импорт выбранных адресов в таблицу наград"""
        selected_rows = set(index.row() for index in self.search_results_table.selectedIndexes())
        
        # Если ничего не выбрано, импортируем все
        if not selected_rows:
            selected_rows = range(self.search_results_table.rowCount())
        
        # Запрашиваем режим импорта, если включен режим награждения за каждую транзакцию
        import_mode = "single"
        if self.reward_per_tx_mode:
            reply = QtWidgets.QMessageBox.question(
                self, 'Режим импорта', 
                'Выберите режим импорта:\n\n'
                'Нажмите "Да" для импорта каждой транзакции (одна транзакция - одна награда)\n'
                'Нажмите "Нет" для импорта только адресов (один адрес - одна награда)',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            import_mode = "per_tx" if reply == QtWidgets.QMessageBox.Yes else "single"
        
        count = 0
        if import_mode == "per_tx":
            # Импортируем каждую транзакцию
            for row in selected_rows:
                sender_addr = self.search_results_table.item(row, 0).text()
                tx_list = self.search_results_table.item(row, 0).data(QtCore.Qt.UserRole + 1)
                if tx_list:
                    # Перед добавлением сохраняем текущее состояние
                    if count == 0:
                        self._save_rewards_state()
                    count += len(tx_list)
                    for tx_info in tx_list:
                        self._add_tx_to_rewards(sender_addr, tx_info['hash'])
                else:
                    # Если нет данных о транзакциях, добавляем просто адрес
                    self._add_to_rewards_tab(sender_addr)
                    count += 1
        else:
            # Обычный импорт - один адрес, одна награда
            for row in selected_rows:
                sender_addr = self.search_results_table.item(row, 0).text()
                self._add_to_rewards_tab(sender_addr)
                count += 1
        
        logger.info(f"Импортировано {count} записей в таблицу наград")
        
        # Предлагаем сгенерировать случайные значения
        if count > 0:
            reply = QtWidgets.QMessageBox.question(
                self, 'Рандомизация', 
                'Сгенерировать случайные значения для импортированных адресов?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            if reply == QtWidgets.QMessageBox.Yes:
                self._randomize_plex()
                self._randomize_usdt()
    # --- Таб «Анализ» ---
    def _tab_analyze(self):
        """Создание вкладки для анализа транзакций"""
        w = QtWidgets.QWidget()
        l = QtWidgets.QVBoxLayout(w)

        # Форма ввода данных
        form = QtWidgets.QFormLayout()
        self.addr_in = QtWidgets.QLineEdit()
        self.addr_in.setPlaceholderText('0x...')
        form.addRow('Адрес для анализа:', self.addr_in)

        btn_seed = QtWidgets.QPushButton('Ввести seed-фразу')
        btn_seed.clicked.connect(self._enter_mnemonic)
        form.addRow('Seed-фраза:', btn_seed)

        btn_pk = QtWidgets.QPushButton('Ввести приватный ключ')
        btn_pk.clicked.connect(self._enter_pk)
        form.addRow('PrivKey:', btn_pk)

        self.sys_in = QtWidgets.QLineEdit()
        self.sys_in.setPlaceholderText('Системный адрес')
        form.addRow('Системный адрес:', self.sys_in)

        self.spin_amt = QtWidgets.QDoubleSpinBox()
        self.spin_amt.setRange(0, 1e12)
        self.spin_amt.setDecimals(8)
        self.spin_amt.setValue(30.0)  # Предустановка на 30 PLEX
        
        # Минимальное количество транзакций
        self.spin_tx_count = QtWidgets.QSpinBox()
        self.spin_tx_count.setRange(1, 1000)
        self.spin_tx_count.setValue(1)
        form.addRow('Мин. кол-во транзакций:', self.spin_tx_count)
        
        # Переключатель режима поиска
        self.search_mode_group = QtWidgets.QButtonGroup(self)
        self.radio_exact = QtWidgets.QRadioButton('Точная сумма')
        self.radio_range = QtWidgets.QRadioButton('Диапазон сумм')
        self.radio_exact.setChecked(True)
        self.search_mode_group.addButton(self.radio_exact)
        self.search_mode_group.addButton(self.radio_range)
        
        # Горизонтальный layout для радиокнопок
        radio_layout = QtWidgets.QHBoxLayout()
        radio_layout.addWidget(self.radio_exact)
        radio_layout.addWidget(self.radio_range)
        form.addRow('Режим поиска:', radio_layout)
        
        # Поля для точной суммы
        form.addRow('Точная сумма:', self.spin_amt)
        
        # Поля для диапазона сумм
        self.spin_amt_from = QtWidgets.QDoubleSpinBox()
        self.spin_amt_from.setRange(0, 1e12)
        self.spin_amt_from.setDecimals(8)
        self.spin_amt_from.setEnabled(False)
        
        self.spin_amt_to = QtWidgets.QDoubleSpinBox()
        self.spin_amt_to.setRange(0, 1e12)
        self.spin_amt_to.setDecimals(8)
        self.spin_amt_to.setEnabled(False)
        
        # Layout для полей диапазона
        range_layout = QtWidgets.QHBoxLayout()
        range_layout.addWidget(QtWidgets.QLabel('От:'))
        range_layout.addWidget(self.spin_amt_from)
        range_layout.addWidget(QtWidgets.QLabel('До:'))
        range_layout.addWidget(self.spin_amt_to)
        form.addRow('Диапазон:', range_layout)
        
        # Опция для учета отдельных транзакций
        self.track_tx_checkbox_analyze = QtWidgets.QCheckBox("Учитывать отдельные транзакции для награждения")
        self.track_tx_checkbox_analyze.setChecked(True)
        form.addRow("", self.track_tx_checkbox_analyze)
        
        # Соединение сигналов для переключения режима
        self.radio_exact.toggled.connect(self._toggle_search_mode)
        self.radio_range.toggled.connect(self._toggle_search_mode)

        l.addLayout(form)

        # Кнопки действий
        h = QtWidgets.QHBoxLayout()
        btn_scan = QtWidgets.QPushButton('Сканировать токены')
        btn_scan.clicked.connect(self._scan_tokens)
        btn_search = QtWidgets.QPushButton('Поиск Tx')
        btn_search.clicked.connect(self._search_txs)
        h.addWidget(btn_scan); h.addWidget(btn_search)
        l.addLayout(h)
        
        # Таблица токенов
        self.tokens_table = QtWidgets.QTableWidget(0, 4)
        self.tokens_table.setHorizontalHeaderLabels(["Токен","Контракт","Баланс","Decimals"])
        self.tokens_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.tokens_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.tokens_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tokens_table.customContextMenuRequested.connect(self._tokens_context_menu)
        l.addWidget(self.tokens_table)

        # Таблица результатов
        self.table_res = QtWidgets.QTableWidget(0, 2)  # Увеличиваем до 2 столбцов (адрес и количество)
        self.table_res.setHorizontalHeaderLabels(['Отправитель', 'Кол-во TX'])
        self.table_res.horizontalHeader().setStretchLastSection(True)
        self.table_res.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table_res.customContextMenuRequested.connect(self._analysis_context_menu)
        l.addWidget(self.table_res)

        # Кнопки экспорта
        h2 = QtWidgets.QHBoxLayout()
        btn_csv = QtWidgets.QPushButton('Экспорт CSV')
        btn_csv.clicked.connect(self._export_csv)
        h2.addWidget(btn_csv)
        
        btn_clipboard = QtWidgets.QPushButton('Копировать в буфер')
        btn_clipboard.clicked.connect(self._copy_to_clipboard)
        h2.addWidget(btn_clipboard)
        
        btn_to_rewards = QtWidgets.QPushButton('В раздел наград')
        btn_to_rewards.clicked.connect(self._import_from_analysis)
        h2.addWidget(btn_to_rewards)
        
        l.addLayout(h2)
        
        # Индикатор прогресса
        self.progress = QtWidgets.QProgressBar()
        l.addWidget(self.progress)

        return w
    
    def _tokens_context_menu(self, position):
        """Контекстное меню для таблицы токенов"""
        if self.tokens_table.rowCount() == 0:
            return
            
        menu = QtWidgets.QMenu()
        copy_contract = menu.addAction("Копировать адрес контракта")
        view_token = menu.addAction("Открыть токен в BscScan")
        use_as_search = menu.addAction("Использовать для поиска")
        
        action = menu.exec_(self.tokens_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.tokens_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        contract = self.tokens_table.item(row, 1).text()
        
        if action == copy_contract:
            QtWidgets.QApplication.clipboard().setText(contract)
            logger.info(f"Адрес контракта скопирован: {contract}")
            
        elif action == view_token:
            url = f"https://bscscan.com/token/{contract}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт токен в BscScan: {contract}")
            
        elif action == use_as_search:
            # Переключиться на вкладку поиска и установить токен
            tabs = self.centralWidget()
            tabs.setCurrentIndex(1)  # Переключаемся на вкладку поиска
            self.token_address.setText(contract)
            logger.info(f"Токен {contract} установлен для поиска")
    
    def _analysis_context_menu(self, position):
        """Контекстное меню для таблицы результатов анализа"""
        if self.table_res.rowCount() == 0:
            return
            
        menu = QtWidgets.QMenu()
        copy_address = menu.addAction("Копировать адрес")
        view_address = menu.addAction("Открыть адрес в BscScan")
        add_to_rewards = menu.addAction("Добавить в награды")
        
        action = menu.exec_(self.table_res.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.table_res.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        address = self.table_res.item(row, 0).text()
        
        if action == copy_address:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == view_address:
            url = f"https://bscscan.com/address/{address}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт адрес в BscScan: {address}")
            
        elif action == add_to_rewards:
            self._add_to_rewards_tab(address)
    def _copy_to_clipboard(self):
        """Копирование выбранных адресов в буфер обмена"""
        addresses = []
        for idx in self.table_res.selectedIndexes():
            if idx.column() == 0:
                addresses.append(self.table_res.item(idx.row(), 0).text())
        
        if not addresses:
            # Если ничего не выбрано, копировать все
            addresses = [self.table_res.item(row, 0).text() 
                      for row in range(self.table_res.rowCount())]
                      
        if addresses:
            clipboard = QtWidgets.QApplication.clipboard()
            clipboard.setText('\n'.join(addresses))
            self.log(f"Скопировано адресов: {len(addresses)}")
        else:
            self.log("Нет адресов для копирования")
            
    def _enter_mnemonic(self):
        """Диалог для ввода seed-фразы"""
        m, ok = QtWidgets.QInputDialog.getText(self, 'Seed', 'Введите seed-фразу')
        if not (ok and m):
            return

        # Убираем переносы строк и лишние пробелы
        clean = ' '.join(m.strip().split())

        # Сохраняем зашифрованную версию "как есть"
        self.cfg.set_mnemonic(clean)

        if blockchain_enabled:
            try:
                acct = Account.from_mnemonic(clean, account_path="m/44'/60'/0'/0/0")
            except ValidationError as e:
                logger.error(f"Неверная seed-фраза: {e}")
                return
            except Exception as e:
                logger.error(f"Ошибка при обработке seed: {e}")
                return

            # если всё OK - сохраняем PK и выводим адрес
            self.cfg.set_key(acct.key.hex())
            self.pk = acct.key.hex()
            logger.info(f"Seed сохранён, адрес: {acct.address}")
        else:
            logger.info("Seed сохранён (blockchain отключён)")

    def _enter_pk(self):
        """Диалог для ввода приватного ключа"""
        p, ok = QtWidgets.QInputDialog.getText(
            self, 'PrivKey', 'Введите приватный ключ:', QtWidgets.QLineEdit.Password)
        if ok and p:
            # Проверка формата приватного ключа
            if blockchain_enabled:
                try:
                    # Проверка действительности приватного ключа путем создания аккаунта
                    acct = Account.from_key(p.strip())
                    addr = acct.address
                    logger.info(f"Ключ валиден, адрес: {addr}")
                except Exception as e:
                    logger.error(f"Неверный формат ключа: {e}")
                    return
            
            self.cfg.set_key(p.strip())
            self.pk = p.strip()
            logger.info("PrivKey сохранен")

    def _scan_tokens(self):
        """Запуск сканирования токенов для выбранного адреса"""
        addr = self.addr_in.text().strip()
        if blockchain_enabled and not Web3.is_address(addr):
            return logger.error("Некорректный адрес")
        self.table_res.setRowCount(0)
        threading.Thread(target=self._scan_thread, args=(addr,), daemon=True).start()

    def _scan_thread(self, addr):
        """Поток для сканирования токенов на адресе"""
        logger.info("Сканирование токенов...")
        contracts, page = set(), 1
        
        try:
            # Используем постраничную обработку для получения всех токенов
            while True:
                try:
                    params = {
                        'module': 'account',
                        'action': 'tokentx',
                        'address': addr,
                        'page': page,
                        'offset': 100,
                        'sort': 'asc'
                    }
                    result = bscscan_request(params)
                except Exception as e:
                    logger.error(f"Ошибка BscScan на странице {page}: {e}")
                    break
                    
                if not result:
                    break
                
                for tx in result:
                    contracts.add((tx['contractAddress'], tx['tokenSymbol'], int(tx['tokenDecimal'])))
                
                # Обновляем прогресс и переходим к следующей странице
                logger.info(f"Обработана страница {page}, найдено {len(contracts)} уникальных токенов")
                page += 1
                self.update_progress_signal.emit("scan", min(30, (page % 31)))
                time.sleep(1)  # Задержка 1 секунда между запросами
        
            tokens = []
            total = len(contracts)
            for i, (c, s, d) in enumerate(contracts, start=1):
                self.update_progress_signal.emit("scan", 30 + int(i / total * 70))
                try:
                    bal = int(bscscan_request({
                        'module':'account','action':'tokenbalance',
                        'contractaddress': c,'address': addr,'tag':'latest'
                    })) / 10**d
                    if bal > 0:
                        tokens.append((c, s, d, bal))
                except Exception as e:
                    logger.warning(f"Ошибка при получении баланса {s}: {e}")
                time.sleep(1)  # Задержка 1 секунда между запросами

            # Обновляем таблицу токенов
            self._update_tokens_table(tokens)
            
            self.combo = QtWidgets.QComboBox()
            for c, s, d, v in tokens:
                self.combo.addItem(f"{s} ({v})", userData=(c, d))
                
            logger.info(f"Найдено {len(tokens)} токенов c положительным балансом")
            self.update_progress_signal.emit("scan", 100)
        except Exception as e:
            logger.error(f"Ошибка при сканировании токенов: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _update_tokens_table(self, tokens):
        """Обновить таблицу токенов: список tuples (contract, symbol, decimals, balance)"""
        self.tokens_table.setRowCount(0)
        for c, s, d, v in tokens:
            row = self.tokens_table.rowCount()
            self.tokens_table.insertRow(row)
            self.tokens_table.setItem(row, 0, QtWidgets.QTableWidgetItem(s))
            self.tokens_table.setItem(row, 1, QtWidgets.QTableWidgetItem(c))
            self.tokens_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(v)))
            self.tokens_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(d)))

    def _search_txs(self):
        """Поиск транзакций по заданным параметрам"""
        sys_addr = self.sys_in.text().strip()
        if blockchain_enabled and not Web3.is_address(sys_addr):
            return logger.error("Неверный системный адрес")
        
        data = getattr(self, 'combo', None)
        if data is None or data.currentIndex() < 0:
            return logger.error("Выберите токен")
        
        contract, dec = data.currentData()
        track_individual_tx = self.track_tx_checkbox_analyze.isChecked()
        
        # Определяем режим поиска
        if self.radio_exact.isChecked():
            # Режим точной суммы
            amt = self.spin_amt.value()
            if amt <= 0:
                return logger.error("Сумма должна быть > 0")
            
            # Показываем диалог подтверждения
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение', 
                f'Искать входящие транзакции с суммой {amt} токенов на адрес {sys_addr}?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            
            if reply == QtWidgets.QMessageBox.Yes:
                threading.Thread(
                    target=self._search_thread,
                    args=(sys_addr, contract, dec, amt, track_individual_tx),
                    daemon=True
                ).start()
        else:
            # Режим диапазона сумм
            min_amt = self.spin_amt_from.value()
            max_amt = self.spin_amt_to.value()
            
            if min_amt < 0:
                return logger.error("Минимальная сумма должна быть >= 0")
            
            if max_amt <= min_amt:
                return logger.error("Максимальная сумма должна быть больше минимальной")
            
            # Показываем диалог подтверждения
            reply = QtWidgets.QMessageBox.question(
                self, 'Подтверждение', 
                f'Искать входящие транзакции с суммой от {min_amt} до {max_amt} токенов на адрес {sys_addr}?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            
            if reply == QtWidgets.QMessageBox.Yes:
                threading.Thread(
                    target=self._search_range_thread, 
                    args=(sys_addr, contract, dec, min_amt, max_amt, track_individual_tx),
                    daemon=True
                ).start()
    def _search_thread(self, sys_addr, contract, dec, amt, track_individual_tx):
        """Поток для поиска транзакций с точным значением"""
        self.table_res.setRowCount(0)
        logger.info(f"Поиск Tx для адреса {sys_addr}, токен {contract}, сумма {amt}...")
        min_tx_count = self.spin_tx_count.value()
        
        # Используем новую функцию для постраничного поиска
        try:
            self.stop_search_event.clear()  # Сбрасываем флаг остановки
            
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=sys_addr,
                token_contract=contract,
                exact_amount=amt,
                page_size=1000,
                max_pages=100,
                delay_seconds=1,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {sender: count for sender, count in sender_counter.items() 
                              if count >= min_tx_count}
            
            # Заполняем таблицу результатов
            for row, (sender, count) in enumerate(sorted(filtered_senders.items(), key=lambda x: x[1], reverse=True)):
                self.table_res.insertRow(row)
                sender_item = QtWidgets.QTableWidgetItem(sender)
                
                # Если включен режим награждения за каждую транзакцию и есть данные о транзакциях,
                # сохраняем их в данных элемента
                if track_individual_tx and sender in sender_transactions:
                    sender_item.setData(QtCore.Qt.UserRole, sender_transactions[sender])
                
                self.table_res.setItem(row, 0, sender_item)
                self.table_res.setItem(row, 1, QtWidgets.QTableWidgetItem(str(count)))

            logger.info(f"Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями")
            self.update_progress_signal.emit("scan", 100)
            
            # Сохраняем найденные транзакции в базу данных
            search_info = {
                'wallet_address': sys_addr,
                'token_contract': contract,
                'exact_amount': amt,
                'min_tx_count': min_tx_count,
                'search_time': datetime.now().isoformat()
            }
            
            for tx in matching_txs:
                add_found_transaction(tx, search_info)
            
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
                
        except Exception as e:
            logger.error(f"Ошибка при поиске транзакций: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _search_range_thread(self, sys_addr, contract, dec, min_amt, max_amt, track_individual_tx):
        """Поток для поиска транзакций в диапазоне сумм"""
        self.table_res.setRowCount(0)
        logger.info(f"Поиск Tx для адреса {sys_addr}, токен {contract}, диапазон {min_amt}-{max_amt}...")
        min_tx_count = self.spin_tx_count.value()
        
        try:
            self.stop_search_event.clear()  # Сбрасываем флаг остановки
            
            # Используем новую функцию для постраничного поиска с диапазоном
            matching_txs, sender_counter, sender_transactions = search_transactions_paginated(
                wallet_address=sys_addr,
                token_contract=contract,
                min_amount=min_amt,
                max_amount=max_amt,
                page_size=1000,
                max_pages=100,
                delay_seconds=1,
                stop_flag=self.stop_search_event,
                track_individual_tx=track_individual_tx
            )
            
            # Фильтруем отправителей по минимальному количеству транзакций
            filtered_senders = {sender: count for sender, count in sender_counter.items() 
                              if count >= min_tx_count}
            
            # Заполняем таблицу результатов
            for row, (sender, count) in enumerate(sorted(filtered_senders.items(), key=lambda x: x[1], reverse=True)):
                self.table_res.insertRow(row)
                sender_item = QtWidgets.QTableWidgetItem(sender)
                
                # Если включен режим награждения за каждую транзакцию и есть данные о транзакциях,
                # сохраняем их в данных элемента
                if track_individual_tx and sender in sender_transactions:
                    sender_item.setData(QtCore.Qt.UserRole, sender_transactions[sender])
                
                self.table_res.setItem(row, 0, sender_item)
                self.table_res.setItem(row, 1, QtWidgets.QTableWidgetItem(str(count)))

            logger.info(f"Найдено {len(filtered_senders)} отправителей с {min_tx_count}+ транзакциями")
            self.update_progress_signal.emit("scan", 100)
            
            # Сохраняем найденные транзакции в базу данных
            search_info = {
                'wallet_address': sys_addr,
                'token_contract': contract,
                'min_amount': min_amt,
                'max_amount': max_amt,
                'min_tx_count': min_tx_count,
                'search_time': datetime.now().isoformat()
            }
            
            for tx in matching_txs:
                add_found_transaction(tx, search_info)
                
            # Если нужно учитывать отдельные транзакции, сохраняем их в базу
            search_time = datetime.now().isoformat()
            if track_individual_tx and sender_transactions:
                for sender, tx_list in sender_transactions.items():
                    for tx_info in tx_list:
                        add_sender_transaction(sender, tx_info, search_time)
                
        except Exception as e:
            logger.error(f"Ошибка при поиске транзакций по диапазону: {e}")
            self.update_progress_signal.emit("scan", 100)

    def _export_csv(self):
        """Экспорт результатов в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('from_address,tx_count\n')
                for i in range(self.table_res.rowCount()):
                    address = self.table_res.item(i, 0).text()
                    count = self.table_res.item(i, 1).text() if self.table_res.columnCount() > 1 else "1"
                    f.write(f'{address},{count}\n')
            logger.info(f"Экспортировано {self.table_res.rowCount()} адресов в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте CSV: {e}")

    def _add_to_rewards_tab(self, address):
        """Добавление адреса в таблицу наград"""
        # Проверка, есть ли уже такой адрес в таблице
        for row in range(self.table_rewards.rowCount()):
            if self.table_rewards.item(row, 0).text() == address:
                logger.warning(f"Адрес {address} уже добавлен в таблицу наград")
                return
        
        # Добавляем новую строку
        row = self.table_rewards.rowCount()
        self.table_rewards.insertRow(row)
        
        # Адрес
        self.table_rewards.setItem(row, 0, QtWidgets.QTableWidgetItem(address))
        
        # PLEX - пустое значение по умолчанию
        plex_item = QtWidgets.QTableWidgetItem("")
        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 1, plex_item)
        
        # USDT - пустое значение по умолчанию
        usdt_item = QtWidgets.QTableWidgetItem("")
        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
        self.table_rewards.setItem(row, 2, usdt_item)
        
        logger.info(f"Адрес {address} добавлен в таблицу наград")
        
    def _import_from_analysis(self):
        """Импорт адресов из вкладки анализа в таблицу наград"""
        if self.table_res.rowCount() == 0:
            return logger.warning("Нет данных в таблице анализа")
            
        # Запрашиваем режим импорта, если включен режим награждения за каждую транзакцию
        import_mode = "single"
        if self.reward_per_tx_mode:
            reply = QtWidgets.QMessageBox.question(
                self, 'Режим импорта', 
                'Выберите режим импорта:\n\n'
                'Нажмите "Да" для импорта каждой транзакции (одна транзакция - одна награда)\n'
                'Нажмите "Нет" для импорта только адресов (один адрес - одна награда)',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes
            )
            import_mode = "per_tx" if reply == QtWidgets.QMessageBox.Yes else "single"
        
        # Сохраняем текущее состояние для возможности отмены
        self._save_rewards_state()
        
        count = 0
        if import_mode == "per_tx":
            # Импортируем каждую транзакцию
            for row in range(self.table_res.rowCount()):
                sender_item = self.table_res.item(row, 0)
                sender = sender_item.text()
                # Получаем данные о транзакциях из элемента
                tx_list = sender_item.data(QtCore.Qt.UserRole) if hasattr(sender_item, 'data') else None
                
                if tx_list:
                    for tx_info in tx_list:
                        r = self.table_rewards.rowCount()
                        self.table_rewards.insertRow(r)
                        
                        addr_item = QtWidgets.QTableWidgetItem(sender)
                        addr_item.setData(QtCore.Qt.UserRole, tx_info['hash'])
                        self.table_rewards.setItem(r, 0, addr_item)
                        
                        plex_item = QtWidgets.QTableWidgetItem("3")
                        plex_item.setFlags(plex_item.flags() | QtCore.Qt.ItemIsEditable)
                        self.table_rewards.setItem(r, 1, plex_item)
                        
                        usdt_item = QtWidgets.QTableWidgetItem("")
                        usdt_item.setFlags(usdt_item.flags() | QtCore.Qt.ItemIsEditable)
                        self.table_rewards.setItem(r, 2, usdt_item)
                        
                        count += 1
                else:
                    # Если нет данных о транзакциях, добавляем просто адрес
                    self._add_to_rewards_tab(sender)
                    count += 1
        else:
            # Обычный импорт - один адрес, одна награда
            for row in range(self.table_res.rowCount()):
                address = self.table_res.item(row, 0).text()
                self._add_to_rewards_tab(address)
                count += 1
        
        logger.info(f"Импортировано {count} записей из таблицы анализа")
        
        # Предлагаем сгенерировать случайные значения
        reply = QtWidgets.QMessageBox.question(
            self, 'Генерация значений', 
            'Сгенерировать значения PLEX и USDT для импортированных адресов?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.Yes
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self._randomize_plex()
            self._randomize_usdt()
    # --- Таб «История» ---
    def _tab_history(self):
        """Создание вкладки для отображения истории транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Таблица истории транзакций
        self.history_table = QtWidgets.QTableWidget(0, 6)
        self.history_table.setHorizontalHeaderLabels(['Время', 'Токен', 'Получатель', 'Сумма', 'TX Hash', 'Статус'])
        self.history_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.history_table.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Interactive)
        self.history_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.history_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.history_table.customContextMenuRequested.connect(self._history_context_menu)
        layout.addWidget(self.history_table)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        btn_refresh = QtWidgets.QPushButton('Обновить')
        btn_refresh.clicked.connect(self._refresh_history)
        buttons_layout.addWidget(btn_refresh)
        
        btn_track = QtWidgets.QPushButton('Отслеживать TX')
        btn_track.clicked.connect(self._track_tx_statuses)
        buttons_layout.addWidget(btn_track)
        
        btn_copy_all = QtWidgets.QPushButton('Копировать все хэши')
        btn_copy_all.clicked.connect(self._copy_all_tx_hashes)
        buttons_layout.addWidget(btn_copy_all)
        
        layout.addLayout(buttons_layout)
        
        # Загружаем историю при открытии вкладки
        QtCore.QTimer.singleShot(100, self._refresh_history)
        
        return w
    
    def _refresh_history(self):
        """Обновление истории транзакций из базы данных"""
        try:
            # Получаем данные из БД
            history = fetch_history()
            
            # Очищаем таблицу
            self.history_table.setRowCount(0)
            
            # Заполняем таблицу
            for row_data in history:
                row = self.history_table.rowCount()
                self.history_table.insertRow(row)
                
                # Время
                ts = datetime.fromisoformat(row_data['ts'])
                self.history_table.setItem(row, 0, QtWidgets.QTableWidgetItem(ts.strftime('%Y-%m-%d %H:%M:%S')))
                
                # Токен
                self.history_table.setItem(row, 1, QtWidgets.QTableWidgetItem(row_data['token']))
                
                # Получатель
                self.history_table.setItem(row, 2, QtWidgets.QTableWidgetItem(row_data['to_addr']))
                
                # Сумма
                self.history_table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(row_data['amount'])))
                
                # TX Hash
                tx_item = QtWidgets.QTableWidgetItem(row_data['tx'][:10] + "..." + row_data['tx'][-6:])
                tx_item.setToolTip(row_data['tx'])
                tx_item.setData(QtCore.Qt.UserRole, row_data['tx'])
                self.history_table.setItem(row, 4, tx_item)
                
                # Статус
                status_item = QtWidgets.QTableWidgetItem(row_data['status'])
                if row_data['status'] == 'success':
                    status_item.setBackground(QtGui.QColor('#004400'))
                elif row_data['status'] == 'failed':
                    status_item.setBackground(QtGui.QColor('#440000'))
                self.history_table.setItem(row, 5, status_item)
            
            logger.info(f"Загружено {self.history_table.rowCount()} записей истории")
            
        except Exception as e:
            logger.error(f"Ошибка при обновлении истории: {e}")
    
    def _history_context_menu(self, position):
        """Отображение контекстного меню для таблицы истории"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_addr = menu.addAction("Копировать адрес получателя")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        check_status = menu.addAction("Проверить статус")
        
        action = menu.exec_(self.history_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.history_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.history_table.item(row, 4).data(QtCore.Qt.UserRole)
        to_addr = self.history_table.item(row, 2).text()
        
        if action == copy_tx:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_addr:
            QtWidgets.QApplication.clipboard().setText(to_addr)
            logger.info(f"Адрес получателя скопирован: {to_addr}")
            
        elif action == open_tx:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")
            
        elif action == check_status:
            threading.Thread(
                target=self._check_tx_status,
                args=(tx_hash,),
                daemon=True
            ).start()
    
    def _check_tx_status(self, tx_hash):
        """Проверка статуса транзакции"""
        try:
            if not blockchain_enabled:
                logger.warning("Блокчейн отключен, невозможно проверить статус")
                return
                
            w3 = self.rpc.web3()
            
            logger.info(f"Проверка статуса транзакции {tx_hash}...")
            try:
                receipt = w3.eth.get_transaction_receipt(tx_hash)
                
                if receipt and receipt.get('status') == 1:
                    status = 'success'
                    logger.info(f"Транзакция {tx_hash} успешно выполнена")
                else:
                    status = 'failed'
                    logger.info(f"Транзакция {tx_hash} не выполнена")
                
                # Обновляем статус в базе данных
                update_tx_status(tx_hash, status)
                
                # Обновляем таблицу
                self._refresh_history()
                
            except Exception as e:
                logger.warning(f"Не удалось получить чек транзакции {tx_hash}: {e}")
                
        except Exception as e:
            logger.error(f"Ошибка при проверке статуса транзакции: {e}")
    
    def _track_tx_statuses(self):
        """Запуск отслеживания статусов всех транзакций в отдельном потоке"""
        if self.tx_tracker and self.tx_tracker.is_alive():
            return logger.warning("Отслеживание уже запущено")
            
        self.tx_tracker = threading.Thread(
            target=self._tx_tracker_thread,
            daemon=True
        )
        self.tx_tracker.start()
        logger.info("Запущено отслеживание статусов транзакций")
    
    def _tx_tracker_thread(self):
        """Поток для отслеживания статусов транзакций"""
        if not blockchain_enabled:
            return logger.warning("Блокчейн отключен, невозможно отслеживать статусы")
            
        try:
            w3 = self.rpc.web3()
            
            # Получаем все транзакции со статусом pending
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT tx FROM history WHERE status='pending'")
            pending_txs = [row['tx'] for row in cur.fetchall()]
            conn.close()
            
            logger.info(f"Найдено {len(pending_txs)} ожидающих транзакций для проверки")
            
            for tx_hash in pending_txs:
                try:
                    receipt = w3.eth.get_transaction_receipt(tx_hash)
                    
                    if receipt:
                        if receipt.get('status') == 1:
                            status = 'success'
                            logger.info(f"Транзакция {tx_hash} успешно выполнена")
                        else:
                            status = 'failed'
                            logger.info(f"Транзакция {tx_hash} не выполнена")
                        
                        # Обновляем статус в базе данных
                        update_tx_status(tx_hash, status)
                except Exception as e:
                    logger.warning(f"Не удалось проверить транзакцию {tx_hash}: {e}")
                
                time.sleep(0.5)  # Небольшая задержка между запросами
            
            # Обновляем таблицу
            QtCore.QMetaObject.invokeMethod(
                self, "_refresh_history", 
                QtCore.Qt.QueuedConnection
            )
            
            logger.info("Проверка статусов транзакций завершена")
            
        except Exception as e:
            logger.error(f"Ошибка при отслеживании статусов: {e}")
    def _copy_all_tx_hashes(self):
        """Копирование всех хэшей транзакций в буфер обмена"""
        try:
            tx_hashes = copy_all_transactions_hashes()
            if tx_hashes:
                QtWidgets.QApplication.clipboard().setText(tx_hashes)
                logger.info(f"Скопированы все хэши транзакций")
            else:
                logger.warning("Нет транзакций для копирования")
        except Exception as e:
            logger.error(f"Ошибка при копировании хэшей: {e}")

    # --- Таб «Найденные Tx» ---
    def _tab_found_tx(self):
        """Создание вкладки для просмотра найденных транзакций"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Таблица найденных транзакций
        self.found_tx_table = QtWidgets.QTableWidget(0, 7)
        self.found_tx_table.setHorizontalHeaderLabels([
            'Время поиска', 'От кого', 'Кому', 'Токен', 'Сумма', 'Хэш Tx', 'Время Tx'
        ])
        self.found_tx_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.found_tx_table.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Interactive)
        self.found_tx_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.found_tx_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.found_tx_table.customContextMenuRequested.connect(self._found_tx_context_menu)
        layout.addWidget(self.found_tx_table)
        
        # Кнопки действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        btn_refresh_found = QtWidgets.QPushButton('Обновить')
        btn_refresh_found.clicked.connect(self._refresh_found_tx)
        buttons_layout.addWidget(btn_refresh_found)
        
        btn_export_found = QtWidgets.QPushButton('Экспорт в CSV')
        btn_export_found.clicked.connect(self._export_found_tx)
        buttons_layout.addWidget(btn_export_found)
        
        btn_clear_found = QtWidgets.QPushButton('Очистить')
        btn_clear_found.clicked.connect(self._clear_found_tx)
        buttons_layout.addWidget(btn_clear_found)
        
        layout.addLayout(buttons_layout)
        
        # Загружаем данные при открытии вкладки
        QtCore.QTimer.singleShot(100, self._refresh_found_tx)
        
        return w
    
    def _refresh_found_tx(self):
        """Обновление таблицы найденных транзакций"""
        try:
            # Получаем данные из БД
            found_txs = fetch_found_transactions()
            
            # Очищаем таблицу
            self.found_tx_table.setRowCount(0)
            
            # Заполняем таблицу
            for tx in found_txs:
                row = self.found_tx_table.rowCount()
                self.found_tx_table.insertRow(row)
                
                # Время поиска
                ts = datetime.fromisoformat(tx['ts'])
                self.found_tx_table.setItem(row, 0, QtWidgets.QTableWidgetItem(ts.strftime('%Y-%m-%d %H:%M:%S')))
                
                # От кого
                from_item = QtWidgets.QTableWidgetItem(tx['from_addr'])
                self.found_tx_table.setItem(row, 1, from_item)
                
                # Кому
                to_item = QtWidgets.QTableWidgetItem(tx['to_addr'])
                self.found_tx_table.setItem(row, 2, to_item)
                
                # Токен
                self.found_tx_table.setItem(row, 3, QtWidgets.QTableWidgetItem(tx['token_name']))
                
                # Сумма
                self.found_tx_table.setItem(row, 4, QtWidgets.QTableWidgetItem(str(tx['amount'])))
                
                # Хэш Tx
                tx_item = QtWidgets.QTableWidgetItem(tx['tx_hash'][:10] + "..." + tx['tx_hash'][-6:])
                tx_item.setToolTip(tx['tx_hash'])
                tx_item.setData(QtCore.Qt.UserRole, tx['tx_hash'])
                self.found_tx_table.setItem(row, 5, tx_item)
                
                # Время Tx
                if tx['block_time']:
                    block_time = datetime.fromisoformat(tx['block_time'])
                    self.found_tx_table.setItem(row, 6, QtWidgets.QTableWidgetItem(block_time.strftime('%Y-%m-%d %H:%M:%S')))
                else:
                    self.found_tx_table.setItem(row, 6, QtWidgets.QTableWidgetItem("-"))
            
            logger.info(f"Загружено {self.found_tx_table.rowCount()} найденных транзакций")
            
        except Exception as e:
            logger.error(f"Ошибка при обновлении найденных транзакций: {e}")
    
    def _found_tx_context_menu(self, position):
        """Отображение контекстного меню для таблицы найденных транзакций"""
        menu = QtWidgets.QMenu()
        
        copy_tx = menu.addAction("Копировать хэш транзакции")
        copy_from = menu.addAction("Копировать адрес отправителя")
        copy_to = menu.addAction("Копировать адрес получателя")
        open_tx = menu.addAction("Открыть транзакцию в BscScan")
        add_from_to_rewards = menu.addAction("Добавить отправителя в награды")
        
        action = menu.exec_(self.found_tx_table.viewport().mapToGlobal(position))
        
        if not action:
            return
            
        selected = self.found_tx_table.selectedItems()
        if not selected:
            return
            
        row = selected[0].row()
        tx_hash = self.found_tx_table.item(row, 5).data(QtCore.Qt.UserRole)
        from_addr = self.found_tx_table.item(row, 1).text()
        to_addr = self.found_tx_table.item(row, 2).text()
        
        if action == copy_tx:
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == copy_from:
            QtWidgets.QApplication.clipboard().setText(from_addr)
            logger.info(f"Адрес отправителя скопирован: {from_addr}")
            
        elif action == copy_to:
            QtWidgets.QApplication.clipboard().setText(to_addr)
            logger.info(f"Адрес получателя скопирован: {to_addr}")
            
        elif action == open_tx:
            url = f"https://bscscan.com/tx/{tx_hash}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыта транзакция в BscScan: {tx_hash}")
            
        elif action == add_from_to_rewards:
            self._add_to_rewards_tab(from_addr)
    def _export_found_tx(self):
        """Экспорт найденных транзакций в CSV файл"""
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', '*.csv')
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('timestamp,from_addr,to_addr,token,amount,tx_hash,block_time\n')
                for row in range(self.found_tx_table.rowCount()):
                    timestamp = self.found_tx_table.item(row, 0).text()
                    from_addr = self.found_tx_table.item(row, 1).text()
                    to_addr = self.found_tx_table.item(row, 2).text()
                    token = self.found_tx_table.item(row, 3).text()
                    amount = self.found_tx_table.item(row, 4).text()
                    tx_hash = self.found_tx_table.item(row, 5).data(QtCore.Qt.UserRole)
                    block_time = self.found_tx_table.item(row, 6).text()
                    
                    f.write(f'"{timestamp}","{from_addr}","{to_addr}","{token}",{amount},"{tx_hash}","{block_time}"\n')
            
            logger.info(f"Найденные транзакции экспортированы в {path}")
        except Exception as e:
            logger.error(f"Ошибка при экспорте найденных транзакций: {e}")
    
    def _clear_found_tx(self):
        """Очистка таблицы найденных транзакций"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение', 
            'Вы уверены, что хотите очистить все найденные транзакции?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            count = clear_found_transactions()
            self.found_tx_table.setRowCount(0)
            logger.info(f"Очищено {count} найденных транзакций")

    # --- Вкладка «Массовая рассылка» ---
    def _tab_mass_distribution(self):
        """Создание вкладки для массовой рассылки токенов
        
        #MCP:MASS_DIST - Вкладка массовой рассылки
        TODO:MCP - Добавить поддержку параллельной отправки
        """
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа импорта адресов
        import_group = QtWidgets.QGroupBox("Импорт адресов получателей")
        import_layout = QtWidgets.QVBoxLayout(import_group)
        
        # Текстовое поле для вставки адресов
        self.mass_addresses_text = QtWidgets.QTextEdit()
        self.mass_addresses_text.setMaximumHeight(100)
        self.mass_addresses_text.setPlaceholderText(
            "Вставьте адреса сюда (разделенные пробелом, запятой, точкой с запятой или новой строкой)"
        )
        import_layout.addWidget(self.mass_addresses_text)
        
        # Кнопки импорта
        import_buttons = QtWidgets.QHBoxLayout()
        
        self.mass_paste_btn = QtWidgets.QPushButton("Вставить из буфера")
        self.mass_paste_btn.clicked.connect(self._mass_paste_addresses)
        import_buttons.addWidget(self.mass_paste_btn)
        
        self.mass_import_excel_btn = QtWidgets.QPushButton("Импорт из Excel")
        self.mass_import_excel_btn.clicked.connect(self._mass_import_excel)
        import_buttons.addWidget(self.mass_import_excel_btn)
        
        self.mass_add_addresses_btn = QtWidgets.QPushButton("Добавить адреса")
        self.mass_add_addresses_btn.clicked.connect(self._mass_add_addresses)
        import_buttons.addWidget(self.mass_add_addresses_btn)
        
        self.mass_clear_addresses_btn = QtWidgets.QPushButton("Очистить список")
        self.mass_clear_addresses_btn.clicked.connect(self._mass_clear_addresses)
        import_buttons.addWidget(self.mass_clear_addresses_btn)
        
        import_layout.addLayout(import_buttons)
        layout.addWidget(import_group)
        
        # Таблица адресов для рассылки
        self.mass_table = QtWidgets.QTableWidget(0, 5)
        self.mass_table.setHorizontalHeaderLabels([
            'Адрес получателя', 'Статус', 'Отправлено раз', 'Последний хэш', 'Время'
        ])
        self.mass_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        self.mass_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        self.mass_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.mass_table.customContextMenuRequested.connect(self._mass_table_context_menu)
        layout.addWidget(self.mass_table)
        
        # Статистика
        stats_group = QtWidgets.QGroupBox("Статистика")
        stats_layout = QtWidgets.QHBoxLayout(stats_group)
        
        self.mass_total_addresses_label = QtWidgets.QLabel("Всего адресов: 0")
        stats_layout.addWidget(self.mass_total_addresses_label)
        
        self.mass_unique_addresses_label = QtWidgets.QLabel("Уникальных: 0")
        stats_layout.addWidget(self.mass_unique_addresses_label)
        
        self.mass_sent_count_label = QtWidgets.QLabel("Отправлено: 0")
        stats_layout.addWidget(self.mass_sent_count_label)
        
        self.mass_error_count_label = QtWidgets.QLabel("Ошибок: 0")
        stats_layout.addWidget(self.mass_error_count_label)
        
        layout.addWidget(stats_group)
        
        # Настройки рассылки
        settings_group = QtWidgets.QGroupBox("Настройки рассылки")
        settings_layout = QtWidgets.QGridLayout(settings_group)
        
        # Выбор токена
        settings_layout.addWidget(QtWidgets.QLabel("Токен:"), 0, 0)
        self.mass_token_combo = QtWidgets.QComboBox()
        self.mass_token_combo.addItems(['PLEX', 'USDT', 'BNB', 'Другой...'])
        self.mass_token_combo.currentTextChanged.connect(self._mass_token_changed)
        settings_layout.addWidget(self.mass_token_combo, 0, 1)
        
        # Адрес контракта для произвольного токена
        self.mass_custom_token_edit = QtWidgets.QLineEdit()
        self.mass_custom_token_edit.setPlaceholderText("Адрес контракта токена (0x...)")
        self.mass_custom_token_edit.setEnabled(False)
        settings_layout.addWidget(self.mass_custom_token_edit, 0, 2)
        
        # Сумма для отправки
        settings_layout.addWidget(QtWidgets.QLabel("Сумма:"), 1, 0)
        self.mass_amount_spin = QtWidgets.QDoubleSpinBox()
        self.mass_amount_spin.setRange(0.00000001, 1000000)
        self.mass_amount_spin.setDecimals(8)
        self.mass_amount_spin.setValue(0.05)  # По умолчанию 0.05
        settings_layout.addWidget(self.mass_amount_spin, 1, 1)
        
        # Интервал между отправками
        settings_layout.addWidget(QtWidgets.QLabel("Интервал (сек):"), 2, 0)
        self.mass_interval_spin = QtWidgets.QSpinBox()
        self.mass_interval_spin.setRange(1, 600)
        self.mass_interval_spin.setValue(5)  # 5 секунд по умолчанию
        settings_layout.addWidget(self.mass_interval_spin, 2, 1)
        
        # Количество циклов
        settings_layout.addWidget(QtWidgets.QLabel("Количество циклов:"), 3, 0)
        self.mass_cycles_spin = QtWidgets.QSpinBox()
        self.mass_cycles_spin.setRange(1, 100)
        self.mass_cycles_spin.setValue(1)  # 1 цикл по умолчанию
        settings_layout.addWidget(self.mass_cycles_spin, 3, 1)
        
        # Режим рассылки
        settings_layout.addWidget(QtWidgets.QLabel("Режим:"), 4, 0)
        self.mass_mode_combo = QtWidgets.QComboBox()
        self.mass_mode_combo.addItems(['Последовательный', 'Параллельный (3 потока)', 'Параллельный (5 потоков)'])
        settings_layout.addWidget(self.mass_mode_combo, 4, 1)
        
        # Отображение текущей цены газа
        self.mass_gas_price_label = QtWidgets.QLabel(f"Цена газа: {self.cfg.get_gas_price()} Gwei")
        settings_layout.addWidget(self.mass_gas_price_label, 5, 0, 1, 2)
        
        layout.addWidget(settings_group)
        
        # Управление рассылкой
        control_group = QtWidgets.QGroupBox("Управление рассылкой")
        control_layout = QtWidgets.QVBoxLayout(control_group)
        
        # Кнопки управления
        control_buttons = QtWidgets.QHBoxLayout()
        
        self.mass_estimate_btn = QtWidgets.QPushButton("Оценить стоимость")
        self.mass_estimate_btn.clicked.connect(self._mass_estimate_cost)
        control_buttons.addWidget(self.mass_estimate_btn)
        
        self.mass_start_btn = QtWidgets.QPushButton("Начать рассылку")
        self.mass_start_btn.clicked.connect(self._mass_start_distribution)
        control_buttons.addWidget(self.mass_start_btn)
        
        self.mass_pause_btn = QtWidgets.QPushButton("Пауза")
        self.mass_pause_btn.setEnabled(False)
        self.mass_pause_btn.clicked.connect(self._mass_pause_distribution)
        control_buttons.addWidget(self.mass_pause_btn)
        
        self.mass_resume_btn = QtWidgets.QPushButton("Продолжить")
        self.mass_resume_btn.setEnabled(False)
        self.mass_resume_btn.clicked.connect(self._mass_resume_distribution)
        control_buttons.addWidget(self.mass_resume_btn)
        
        self.mass_stop_btn = QtWidgets.QPushButton("Остановить")
        self.mass_stop_btn.setEnabled(False)
        self.mass_stop_btn.clicked.connect(self._mass_stop_distribution)
        control_buttons.addWidget(self.mass_stop_btn)
        
        control_layout.addLayout(control_buttons)
        
        # Прогресс
        self.mass_progress = QtWidgets.QProgressBar()
        control_layout.addWidget(self.mass_progress)
        
        self.mass_progress_label = QtWidgets.QLabel("Готово к работе")
        control_layout.addWidget(self.mass_progress_label)
        
        layout.addWidget(control_group)
        
        # Кнопки экспорта и сохранения
        export_buttons = QtWidgets.QHBoxLayout()
        
        self.mass_save_list_btn = QtWidgets.QPushButton("Сохранить список")
        self.mass_save_list_btn.clicked.connect(self._mass_save_list)
        export_buttons.addWidget(self.mass_save_list_btn)
        
        self.mass_load_list_btn = QtWidgets.QPushButton("Загрузить список")
        self.mass_load_list_btn.clicked.connect(self._mass_load_list)
        export_buttons.addWidget(self.mass_load_list_btn)
        
        self.mass_export_results_btn = QtWidgets.QPushButton("Экспорт результатов")
        self.mass_export_results_btn.clicked.connect(self._mass_export_results)
        export_buttons.addWidget(self.mass_export_results_btn)
        
        layout.addLayout(export_buttons)
        
        # Инициализация переменных для управления рассылкой
        self.mass_distribution_active = False
        self.mass_distribution_paused = False
        self.mass_distribution_thread = None
        self.mass_current_cycle = 0
        self.mass_total_sent = 0
        self.mass_errors_count = 0
        
        return w
    
    def _mass_paste_addresses(self):
        """Вставка адресов из буфера обмена"""
        clipboard = QtWidgets.QApplication.clipboard()
        text = clipboard.text()
        if text:
            self.mass_addresses_text.setText(text)
            logger.info("Адреса вставлены из буфера обмена")
    
    def _mass_import_excel(self):
        """Импорт адресов из Excel файла"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Выберите Excel файл', '', 'Excel Files (*.xlsx *.xls);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            # Попытка импорта с использованием pandas
            try:
                import pandas as pd
                df = pd.read_excel(path)
                
                # Ищем столбец с адресами
                addresses = []
                for col in df.columns:
                    col_data = df[col].dropna().astype(str)
                    # Проверяем, похожи ли данные на адреса BSC
                    potential_addresses = [addr for addr in col_data if addr.startswith('0x') and len(addr) == 42]
                    if potential_addresses:
                        addresses.extend(potential_addresses)
                
                if addresses:
                    self.mass_addresses_text.setText(' '.join(addresses))
                    logger.info(f"Импортировано {len(addresses)} адресов из Excel")
                else:
                    logger.warning("Не найдено адресов в Excel файле")
                    
            except ImportError:
                # Если pandas не установлен, используем openpyxl напрямую
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    
                    addresses = []
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell and isinstance(cell, str) and cell.startswith('0x') and len(cell) == 42:
                                addresses.append(cell)
                    
                    if addresses:
                        self.mass_addresses_text.setText(' '.join(addresses))
                        logger.info(f"Импортировано {len(addresses)} адресов из Excel")
                    else:
                        logger.warning("Не найдено адресов в Excel файле")
                        
                except ImportError:
                    logger.error("Для импорта Excel необходимо установить pandas или openpyxl")
                    QtWidgets.QMessageBox.warning(
                        self, 'Ошибка',
                        'Для импорта Excel файлов необходимо установить библиотеку pandas или openpyxl:\n'
                        'pip install pandas openpyxl',
                        QtWidgets.QMessageBox.Ok
                    )
                    
        except Exception as e:
            logger.error(f"Ошибка при импорте Excel: {e}")
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка импорта',
                f'Не удалось импортировать адреса из Excel:\n{str(e)}',
                QtWidgets.QMessageBox.Ok
            )
    
    def _mass_add_addresses(self):
        """Добавление адресов из текстового поля в таблицу"""
        text = self.mass_addresses_text.toPlainText().strip()
        if not text:
            return logger.warning("Нет адресов для добавления")
        
        # Разбиваем текст на адреса по различным разделителям
        import re
        addresses = re.split(r'[\s,;]+', text)
        
        # Фильтруем и валидируем адреса
        valid_addresses = []
        for addr in addresses:
            addr = addr.strip()
            if addr and self._validate_address(addr):
                valid_addresses.append(addr)
        
        if not valid_addresses:
            return logger.warning("Не найдено валидных адресов")
        
        # Получаем текущие адреса из таблицы
        current_addresses = set()
        for row in range(self.mass_table.rowCount()):
            current_addresses.add(self.mass_table.item(row, 0).text())
        
        # Добавляем только уникальные адреса
        added_count = 0
        for addr in valid_addresses:
            if addr not in current_addresses:
                row = self.mass_table.rowCount()
                self.mass_table.insertRow(row)
                
                # Адрес
                self.mass_table.setItem(row, 0, QtWidgets.QTableWidgetItem(addr))
                # Статус
                self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem("Ожидание"))
                # Отправлено раз
                self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem("0"))
                # Последний хэш
                self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem("-"))
                # Время
                self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem("-"))
                
                current_addresses.add(addr)
                added_count += 1
        
        # Обновляем статистику
        self._mass_update_statistics()
        
        # Очищаем поле ввода
        self.mass_addresses_text.clear()
        
        logger.info(f"Добавлено {added_count} уникальных адресов")
    
    def _validate_address(self, address):
        """Валидация BSC адреса"""
        if blockchain_enabled:
            try:
                Web3.to_checksum_address(address)
                return True
            except:
                return False
        else:
            # Простая проверка формата без Web3
            return address.startswith('0x') and len(address) == 42
    
    def _mass_clear_addresses(self):
        """Очистка списка адресов"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            'Очистить весь список адресов?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.mass_table.setRowCount(0)
            self._mass_update_statistics()
            logger.info("Список адресов очищен")
    
    def _mass_update_statistics(self):
        """Обновление статистики"""
        total = self.mass_table.rowCount()
        unique = len(set(self.mass_table.item(row, 0).text() for row in range(total)))
        
        sent_count = 0
        error_count = 0
        
        for row in range(total):
            status = self.mass_table.item(row, 1).text()
            if status == "Успешно":
                sent_count += 1
            elif status == "Ошибка":
                error_count += 1
        
        self.mass_total_addresses_label.setText(f"Всего адресов: {total}")
        self.mass_unique_addresses_label.setText(f"Уникальных: {unique}")
        self.mass_sent_count_label.setText(f"Отправлено: {sent_count}")
        self.mass_error_count_label.setText(f"Ошибок: {error_count}")
    
    def _mass_token_changed(self, token):
        """Обработка изменения выбранного токена"""
        self.mass_custom_token_edit.setEnabled(token == "Другой...")
        if token == "Другой...":
            self.mass_custom_token_edit.setFocus()
    
    def _mass_table_context_menu(self, position):
        """Контекстное меню для таблицы адресов"""
        menu = QtWidgets.QMenu()
        
        remove_action = menu.addAction("Удалить адрес")
        copy_action = menu.addAction("Копировать адрес")
        copy_hash_action = menu.addAction("Копировать хэш транзакции")
        open_bscscan_action = menu.addAction("Открыть в BscScan")
        menu.addSeparator()
        reset_status_action = menu.addAction("Сбросить статус")
        
        action = menu.exec_(self.mass_table.viewport().mapToGlobal(position))
        
        if not action:
            return
        
        selected = self.mass_table.selectedItems()
        if not selected:
            return
        
        row = selected[0].row()
        address = self.mass_table.item(row, 0).text()
        tx_hash = self.mass_table.item(row, 3).text()
        
        if action == remove_action:
            self.mass_table.removeRow(row)
            self._mass_update_statistics()
            logger.info(f"Адрес {address} удален из списка")
            
        elif action == copy_action:
            QtWidgets.QApplication.clipboard().setText(address)
            logger.info(f"Адрес скопирован: {address}")
            
        elif action == copy_hash_action and tx_hash != "-":
            QtWidgets.QApplication.clipboard().setText(tx_hash)
            logger.info(f"Хэш транзакции скопирован: {tx_hash}")
            
        elif action == open_bscscan_action:
            url = f"https://bscscan.com/address/{address}"
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
            logger.info(f"Открыт адрес в BscScan: {address}")
            
        elif action == reset_status_action:
            self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem("Ожидание"))
            self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem("0"))
            self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem("-"))
            self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem("-"))
            self._mass_update_statistics()
            logger.info(f"Статус адреса {address} сброшен")
    
    def _mass_estimate_cost(self):
        """Оценка стоимости рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для оценки")
        
        addresses_count = self.mass_table.rowCount()
        cycles = self.mass_cycles_spin.value()
        amount = self.mass_amount_spin.value()
        token = self.mass_token_combo.currentText()
        
        total_transactions = addresses_count * cycles
        total_amount = total_transactions * amount
        
        # Оценка газа (примерная)
        estimated_gas_per_tx = 0.0003  # BNB (примерно)
        total_gas = total_transactions * estimated_gas_per_tx
        
        message = f"""
Оценка стоимости рассылки:

Адресов: {addresses_count}
Циклов: {cycles}
Сумма за транзакцию: {amount} {token}

Всего транзакций: {total_transactions}
Всего токенов: {total_amount} {token}
Примерная стоимость газа: {total_gas:.4f} BNB

Убедитесь, что на кошельке достаточно средств!
        """
        
        QtWidgets.QMessageBox.information(self, "Оценка стоимости", message)
        logger.info(f"Оценка: {total_transactions} транзакций, {total_amount} {token}, ~{total_gas:.4f} BNB газа")
    
    def _mass_start_distribution(self):
        """Начало массовой рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для рассылки")
        
        if not self.pk:
            return logger.error("Не настроен приватный ключ")
        
        # Получаем параметры
        token = self.mass_token_combo.currentText()
        if token == "Другой...":
            token_address = self.mass_custom_token_edit.text().strip()
            if not token_address or not self._validate_address(token_address):
                return logger.error("Некорректный адрес токена")
        elif token == "PLEX":
            token_address = PLEX_CONTRACT
        elif token == "USDT":
            token_address = USDT_CONTRACT
        elif token == "BNB":
            token_address = None  # Нативный токен
        
        amount = self.mass_amount_spin.value()
        interval = self.mass_interval_spin.value()
        cycles = self.mass_cycles_spin.value()
        mode = self.mass_mode_combo.currentText()
        
        # Подтверждение
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            f'Начать рассылку {amount} {token} на {self.mass_table.rowCount()} адресов?\n'
            f'Циклов: {cycles}, интервал: {interval} сек',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
        
        # Запуск рассылки
        self.mass_distribution_active = True
        self.mass_distribution_paused = False
        self.mass_current_cycle = 0
        self.mass_total_sent = 0
        self.mass_errors_count = 0
        
        # Обновление UI
        self.mass_start_btn.setEnabled(False)
        self.mass_pause_btn.setEnabled(True)
        self.mass_stop_btn.setEnabled(True)
        self.mass_progress.setValue(0)
        
        # Запуск в отдельном потоке
        self.mass_distribution_thread = threading.Thread(
            target=self._mass_distribution_worker,
            args=(token_address, token, amount, interval, cycles, mode),
            daemon=True
        )
        self.mass_distribution_thread.start()
        
        logger.info(f"Начата массовая рассылка: {amount} {token}, {cycles} циклов")
    
    def _mass_distribution_worker(self, token_address, token_name, amount, interval, cycles, mode):
        """Рабочий поток для массовой рассылки"""
        try:
            addresses = []
            for row in range(self.mass_table.rowCount()):
                addresses.append(self.mass_table.item(row, 0).text())
            
            total_transactions = len(addresses) * cycles
            completed_transactions = 0
            
            # Создаем запись в базе данных
            distribution_name = f"Массовая рассылка {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            distribution_id = add_mass_distribution(
                name=distribution_name,
                token_address=token_address or "BNB",
                token_symbol=token_name,
                amount_per_tx=amount,
                total_addresses=len(addresses),
                total_cycles=cycles,
                interval_seconds=interval
            )
            
            for cycle in range(cycles):
                if not self.mass_distribution_active:
                    break
                
                self.mass_current_cycle = cycle + 1
                logger.info(f"Начат цикл {self.mass_current_cycle} из {cycles}")
                
                for idx, address in enumerate(addresses):
                    if not self.mass_distribution_active:
                        break
                    
                    # Проверка паузы
                    while self.mass_distribution_paused and self.mass_distribution_active:
                        time.sleep(0.1)
                    
                    # Обновление прогресса
                    progress = int((completed_transactions / total_transactions) * 100)
                    QtCore.QMetaObject.invokeMethod(
                        self.mass_progress, "setValue",
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(int, progress)
                    )
                    
                    QtCore.QMetaObject.invokeMethod(
                        self.mass_progress_label, "setText",
                        QtCore.Qt.QueuedConnection,
                        QtCore.Q_ARG(str, f"Цикл {self.mass_current_cycle}/{cycles}, адрес {idx+1}/{len(addresses)}")
                    )
                    
                    # Обновление статуса в таблице
                    self._mass_update_table_status(idx, "Отправка...", None, None)
                    
                    # Отправка транзакции
                    try:
                        if token_address:  # Токен
                            tx_hash = self._send_token_transaction(address, amount, token_address)
                        else:  # BNB
                            tx_hash = self._send_bnb_transaction(address, amount)
                        
                        if tx_hash:
                            self.mass_total_sent += 1
                            sent_count = int(self.mass_table.item(idx, 2).text()) + 1
                            self._mass_update_table_status(
                                idx, "Успешно", tx_hash, 
                                datetime.now().strftime('%H:%M:%S')
                            )
                            self._mass_update_sent_count(idx, sent_count)
                            
                            # Сохраняем в базу данных
                            add_mass_distribution_item(
                                distribution_id=distribution_id,
                                address=address,
                                cycle=self.mass_current_cycle,
                                tx_hash=tx_hash,
                                status='success'
                            )
                            
                            logger.info(f"Отправлено {amount} {token_name} на {address}, tx: {tx_hash}")
                        else:
                            raise Exception("Не удалось получить хэш транзакции")
                            
                    except Exception as e:
                        self.mass_errors_count += 1
                        self._mass_update_table_status(
                            idx, "Ошибка", None,
                            datetime.now().strftime('%H:%M:%S')
                        )
                        
                        # Сохраняем ошибку в базу данных
                        add_mass_distribution_item(
                            distribution_id=distribution_id,
                            address=address,
                            cycle=self.mass_current_cycle,
                            tx_hash=None,
                            status='error',
                            error_message=str(e)
                        )
                        
                        logger.error(f"Ошибка отправки на {address}: {e}")
                    
                    completed_transactions += 1
                    
                    # Задержка между транзакциями
                    if idx < len(addresses) - 1 and self.mass_distribution_active:
                        time.sleep(interval)
                
                # Задержка между циклами
                if cycle < cycles - 1 and self.mass_distribution_active:
                    logger.info(f"Цикл {self.mass_current_cycle} завершен, ожидание перед следующим...")
                    time.sleep(interval)
            
            # Обновляем статус рассылки
            if self.mass_distribution_active:
                update_mass_distribution_status(distribution_id, 'completed')
            else:
                update_mass_distribution_status(distribution_id, 'cancelled')
            
            logger.info(f"Массовая рассылка завершена: {self.mass_total_sent} успешных, {self.mass_errors_count} ошибок")
            
        except Exception as e:
            logger.error(f"Критическая ошибка в массовой рассылке: {e}")
        finally:
            self.mass_distribution_active = False
            QtCore.QMetaObject.invokeMethod(
                self, "_mass_distribution_finished",
                QtCore.Qt.QueuedConnection
            )
    
    def _mass_update_table_status(self, row, status, tx_hash, time_str):
        """Обновление статуса в таблице"""
        def update():
            if row < self.mass_table.rowCount():
                # Статус
                status_item = QtWidgets.QTableWidgetItem(status)
                if status == "Успешно":
                    status_item.setBackground(QtGui.QColor('#004400'))
                elif status == "Ошибка":
                    status_item.setBackground(QtGui.QColor('#440000'))
                elif status == "Отправка...":
                    status_item.setBackground(QtGui.QColor('#444400'))
                self.mass_table.setItem(row, 1, status_item)
                
                # Хэш транзакции
                if tx_hash:
                    self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem(tx_hash[:10] + "..." + tx_hash[-6:]))
                    self.mass_table.item(row, 3).setToolTip(tx_hash)
                
                # Время
                if time_str:
                    self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem(time_str))
                
                self._mass_update_statistics()
        
        QtCore.QMetaObject.invokeMethod(
            self, lambda: update(),
            QtCore.Qt.QueuedConnection
        )
    
    def _mass_update_sent_count(self, row, count):
        """Обновление счетчика отправок"""
        def update():
            if row < self.mass_table.rowCount():
                self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(count)))
        
        QtCore.QMetaObject.invokeMethod(
            self, lambda: update(),
            QtCore.Qt.QueuedConnection
        )
    
    @QtCore.pyqtSlot()
    def _mass_distribution_finished(self):
        """Завершение массовой рассылки"""
        self.mass_start_btn.setEnabled(True)
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(False)
        self.mass_stop_btn.setEnabled(False)
        self.mass_progress_label.setText(f"Завершено: {self.mass_total_sent} успешных, {self.mass_errors_count} ошибок")
        
        QtWidgets.QMessageBox.information(
            self, "Рассылка завершена",
            f"Массовая рассылка завершена!\n\n"
            f"Успешно отправлено: {self.mass_total_sent}\n"
            f"Ошибок: {self.mass_errors_count}"
        )
    
    def _mass_pause_distribution(self):
        """Пауза рассылки"""
        self.mass_distribution_paused = True
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(True)
        self.mass_progress_label.setText("Рассылка приостановлена")
        logger.info("Массовая рассылка приостановлена")
    
    def _mass_resume_distribution(self):
        """Возобновление рассылки"""
        self.mass_distribution_paused = False
        self.mass_pause_btn.setEnabled(True)
        self.mass_resume_btn.setEnabled(False)
        logger.info("Массовая рассылка возобновлена")
    
    def _mass_stop_distribution(self):
        """Остановка рассылки"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            'Остановить рассылку?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.mass_distribution_active = False
            self.mass_distribution_paused = False
            logger.info("Массовая рассылка остановлена пользователем")
    
    def _mass_save_list(self):
        """Сохранение списка адресов"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет адресов для сохранения")
        
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Сохранить список адресов', '', 'JSON Files (*.json);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            data = {
                'addresses': [],
                'settings': {
                    'token': self.mass_token_combo.currentText(),
                    'custom_token': self.mass_custom_token_edit.text() if self.mass_token_combo.currentText() == "Другой..." else "",
                    'amount': self.mass_amount_spin.value(),
                    'interval': self.mass_interval_spin.value(),
                    'cycles': self.mass_cycles_spin.value()
                }
            }
            
            for row in range(self.mass_table.rowCount()):
                data['addresses'].append({
                    'address': self.mass_table.item(row, 0).text(),
                    'status': self.mass_table.item(row, 1).text(),
                    'sent_count': int(self.mass_table.item(row, 2).text()),
                    'last_hash': self.mass_table.item(row, 3).text(),
                    'last_time': self.mass_table.item(row, 4).text()
                })
            
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Список сохранен в {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении списка: {e}")
    
    def _mass_load_list(self):
        """Загрузка списка адресов"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Загрузить список адресов', '', 'JSON Files (*.json);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Очищаем текущую таблицу
            self.mass_table.setRowCount(0)
            
            # Загружаем адреса
            for addr_data in data.get('addresses', []):
                row = self.mass_table.rowCount()
                self.mass_table.insertRow(row)
                
                self.mass_table.setItem(row, 0, QtWidgets.QTableWidgetItem(addr_data['address']))
                self.mass_table.setItem(row, 1, QtWidgets.QTableWidgetItem(addr_data.get('status', 'Ожидание')))
                self.mass_table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(addr_data.get('sent_count', 0))))
                self.mass_table.setItem(row, 3, QtWidgets.QTableWidgetItem(addr_data.get('last_hash', '-')))
                self.mass_table.setItem(row, 4, QtWidgets.QTableWidgetItem(addr_data.get('last_time', '-')))
            
            # Загружаем настройки
            settings = data.get('settings', {})
            if settings.get('token'):
                self.mass_token_combo.setCurrentText(settings['token'])
            if settings.get('custom_token'):
                self.mass_custom_token_edit.setText(settings['custom_token'])
            if 'amount' in settings:
                self.mass_amount_spin.setValue(settings['amount'])
            if 'interval' in settings:
                self.mass_interval_spin.setValue(settings['interval'])
            if 'cycles' in settings:
                self.mass_cycles_spin.setValue(settings['cycles'])
            
            self._mass_update_statistics()
            logger.info(f"Загружено {self.mass_table.rowCount()} адресов из {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке списка: {e}")
    
    def _mass_export_results(self):
        """Экспорт результатов рассылки"""
        if not self.mass_table.rowCount():
            return logger.warning("Нет данных для экспорта")
        
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Экспорт результатов', '', 'CSV Files (*.csv);;All Files (*.*)'
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write('address,status,sent_count,last_hash,last_time\n')
                
                for row in range(self.mass_table.rowCount()):
                    address = self.mass_table.item(row, 0).text()
                    status = self.mass_table.item(row, 1).text()
                    sent_count = self.mass_table.item(row, 2).text()
                    last_hash = self.mass_table.item(row, 3).text()
                    last_time = self.mass_table.item(row, 4).text()
                    
                    f.write(f'"{address}","{status}",{sent_count},"{last_hash}","{last_time}"\n')
            
            logger.info(f"Результаты экспортированы в {path}")
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте результатов: {e}")
    
    def _send_token_transaction(self, to_address, amount, token_address):
        """Отправка токенов на адрес"""
        try:
            w3 = self.rpc.web3()
            account = Account.from_key(self.pk)
            
            # Конвертируем адреса в checksum формат
            to_address = w3.to_checksum_address(to_address)
            token_address = w3.to_checksum_address(token_address)
            
            # Проверяем баланс
            token_contract = w3.eth.contract(address=token_address, abi=ERC20_ABI)
            decimals = token_contract.functions.decimals().call()
            balance = token_contract.functions.balanceOf(account.address).call()
            amount_wei = int(amount * (10 ** decimals))
            
            if balance < amount_wei:
                raise Exception(f"Недостаточно токенов: {balance / (10 ** decimals)} < {amount}")
            
            # Подготовка транзакции
            nonce = w3.eth.get_transaction_count(account.address)
            # Используем настроенную цену газа вместо автоматической
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            tx = token_contract.functions.transfer(
                to_address,
                amount_wei
            ).build_transaction({
                'from': account.address,
                'nonce': nonce,
                'gas': 100000,
                'gasPrice': gas_price
            })
            
            # Подписание и отправка
            signed_tx = w3.eth.account.sign_transaction(tx, self.pk)
            tx_hash = w3.eth.send_raw_transaction(signed_tx.raw_transaction)
            
            return tx_hash.hex()
            
        except Exception as e:
            logger.error(f"Ошибка при отправке токенов: {e}")
            raise
    
    def _send_bnb_transaction(self, to_address, amount):
        """Отправка BNB на адрес"""
        try:
            w3 = self.rpc.web3()
            account = Account.from_key(self.pk)
            
            # Проверяем баланс
            balance = w3.eth.get_balance(account.address)
            amount_wei = Web3.to_wei(amount, 'ether')
            
            if balance < amount_wei:
                raise Exception(f"Недостаточно BNB: {Web3.from_wei(balance, 'ether')} < {amount}")
            
            # Подготовка транзакции
            nonce = w3.eth.get_transaction_count(account.address)
            # Используем настроенную цену газа вместо автоматической
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            tx = {
                'from': account.address,
                'to': Web3.to_checksum_address(to_address),
                'value': amount_wei,
                'nonce': nonce,
                'gas': 21000,
                'gasPrice': gas_price
            }
            
            # Подписание и отправка
            signed_tx = w3.eth.account.sign_transaction(tx, self.pk)
            tx_hash = w3.eth.send_raw_transaction(signed_tx.raw_transaction)
            
            return tx_hash.hex()
            
        except Exception as e:
            logger.error(f"Ошибка при отправке BNB: {e}")
            raise

    # --- #MCP:DS_TAB - Таб «ДС» (Дополнительные Сервисы) ---
    def _tab_ds(self):
        """Создание вкладки ДС (Дополнительные Сервисы) для работы с PLEX ONE
        
        TODO:MCP - Добавить сохранение состояния вкладки ДС
        OPTIMIZE:MCP - Кэшировать данные компонентов ДС
        REFACTOR:MCP - Вынести в отдельный модуль ds_tab
        """
        try:
            # Попытка импорта компонентов ДС
            import sys
            import os
            
            # Добавляем путь к MCP компонентам ДС
            ds_components_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "..", "Эксперимент", "МСР", "WalletSender_MCP", "ds_components"
            )
            ds_components_path = os.path.normpath(ds_components_path)
            
            if os.path.exists(ds_components_path):
                sys.path.insert(0, ds_components_path)
                
                try:
                    from ds_tab_controller import DSTabController  # type: ignore
                    
                    # Получаем BSCScan API ключ из конфигурации или устанавливаем по умолчанию
                    bscscan_api_key = ""
                    if hasattr(self, 'cfg') and self.cfg:
                        bscscan_api_key = getattr(self.cfg, 'bscscan_api_key', "")
                    
                    # Создаем контроллер ДС с передачей всех необходимых параметров
                    ds_controller = DSTabController(
                        rpc_pool=self.rpc_pool if hasattr(self, 'rpc_pool') else None,
                        config=self.cfg if hasattr(self, 'cfg') else None,
                        bscscan_api_key=bscscan_api_key
                    )
                    
                    # Подключаем сигналы
                    ds_controller.status_update.connect(self._update_ds_status)
                    ds_controller.transaction_ready.connect(self._handle_ds_transaction)
                    ds_controller.balance_alert.connect(self._handle_balance_alert)
                    ds_controller.analysis_completed.connect(self._handle_analysis_completed)
                    
                    logger.info("✅ Вкладка ДС успешно инициализирована с полным функционалом")
                    return ds_controller
                    
                except ImportError as e:
                    logger.warning(f"Ошибка импорта компонентов ДС: {e}")
                    
            else:
                logger.warning(f"Путь к компонентам ДС не найден: {ds_components_path}")
                
        except Exception as e:
            logger.error(f"Ошибка инициализации вкладки ДС: {e}")
        
        # Fallback - создаем базовую заглушку
        return self._create_ds_fallback()
    
    def _create_ds_fallback(self):
        """Создание fallback интерфейса для вкладки ДС"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Заголовок
        title_label = QtWidgets.QLabel("🚀 ДС - Дополнительные Сервисы")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #1976D2;
                padding: 20px;
                text-align: center;
            }
        """)
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Информация о PLEX ONE
        info_group = QtWidgets.QGroupBox("ℹ️ Информация о PLEX ONE")
        info_layout = QtWidgets.QFormLayout(info_group)
        
        info_layout.addRow("Название:", QtWidgets.QLabel("PLEX ONE"))
        info_layout.addRow("Символ:", QtWidgets.QLabel("PLEX"))
        info_layout.addRow("Decimals:", QtWidgets.QLabel("9"))
        info_layout.addRow("Total Supply:", QtWidgets.QLabel("12,600,000 PLEX"))
        info_layout.addRow("Контракт:", QtWidgets.QLabel("0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"))
        info_layout.addRow("Сеть:", QtWidgets.QLabel("Binance Smart Chain"))
        
        layout.addWidget(info_group)
        
        # Быстрые ссылки
        links_group = QtWidgets.QGroupBox("🔗 Быстрые ссылки")
        links_layout = QtWidgets.QVBoxLayout(links_group)
        
        # Кнопки для основных действий
        buttons_layout = QtWidgets.QHBoxLayout()
        
        # BSCScan
        bscscan_btn = QtWidgets.QPushButton("🔍 BSCScan")
        bscscan_btn.clicked.connect(lambda: self._open_url(
            "https://bscscan.com/address/0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
        ))
        buttons_layout.addWidget(bscscan_btn)
        
        # PancakeSwap
        pancake_btn = QtWidgets.QPushButton("🥞 PancakeSwap")
        pancake_btn.clicked.connect(lambda: self._open_url(
            "https://pancakeswap.finance/swap"
            "?outputCurrency=0xdf179b6cAdBC61FFD86A3D2e55f6d6e083ade6c1"
            "&inputCurrency=0x55d398326f99059fF775485246999027B3197955"
        ))
        buttons_layout.addWidget(pancake_btn)
        
        # График
        chart_btn = QtWidgets.QPushButton("📈 График")
        chart_btn.clicked.connect(lambda: self._open_url(
            "https://www.geckoterminal.com/ru/bsc/pools/0x41d9650faf3341cbf8947fd8063a1fc88dbf1889"
        ))
        buttons_layout.addWidget(chart_btn)
        
        links_layout.addLayout(buttons_layout)
        layout.addWidget(links_group)
        
        # Статус загрузки компонентов
        status_group = QtWidgets.QGroupBox("📊 Статус")
        status_layout = QtWidgets.QVBoxLayout(status_group)
        
        status_text = QtWidgets.QLabel("""
<b>Статус компонентов ДС:</b><br>
• ⚠️ Полный функционал ДС не загружен<br>
• 🔗 Базовые ссылки доступны<br>
• 📁 Проверьте путь к MCP компонентам<br>
• 🔄 Перезапустите приложение после установки
        """)
        status_text.setWordWrap(True)
        status_layout.addWidget(status_text)
        
        # Кнопка перезагрузки компонентов
        reload_btn = QtWidgets.QPushButton("🔄 Перезагрузить компоненты ДС")
        reload_btn.clicked.connect(self._reload_ds_components)
        reload_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        status_layout.addWidget(reload_btn)
        
        # Кнопка настройки API ключа
        api_btn = QtWidgets.QPushButton("🔑 Настроить BSCScan API")
        api_btn.clicked.connect(self._setup_bscscan_api)
        api_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        status_layout.addWidget(api_btn)
        
        layout.addWidget(status_group)
        
        layout.addStretch()
        
        return w
    
    def _open_url(self, url):
        """Открытие URL в браузере"""
        import webbrowser
        webbrowser.open(url)
        logger.info(f"Открыт URL: {url}")
    
    def _reload_ds_components(self):
        """Перезагрузка компонентов ДС"""
        try:
            # Можно добавить логику перезагрузки
            QtWidgets.QMessageBox.information(
                self, 'Перезагрузка компонентов',
                'Для применения изменений перезапустите приложение',
                QtWidgets.QMessageBox.Ok
            )
        except Exception as e:
            logger.error(f"Ошибка перезагрузки компонентов ДС: {e}")
    
    def _setup_bscscan_api(self):
        """Настройка BSCScan API ключа"""
        text, ok = QtWidgets.QInputDialog.getText(
            self, 'BSCScan API Key',
            'Введите ваш BSCScan API ключ:\n(Получить можно на bscscan.com)',
            QtWidgets.QLineEdit.Normal
        )
        
        if ok and text.strip():
            api_key = text.strip()
            # Сохраняем в конфигурации (можно расширить Config класс)
            logger.info(f"BSCScan API ключ установлен: {api_key[:10]}...")
            
            QtWidgets.QMessageBox.information(
                self, 'API ключ установлен',
                'BSCScan API ключ сохранен.\nПерезапустите приложение для применения изменений.',
                QtWidgets.QMessageBox.Ok
            )
    
    def _update_ds_status(self, message):
        """Обновление статуса от вкладки ДС"""
        logger.info(f"ДС статус: {message}")
    
    def _handle_ds_transaction(self, tx_data):
        """Обработка готовой транзакции от ДС"""
        tx_type = tx_data.get('type', 'unknown')
        amount = tx_data.get('amount', 0)
        logger.info(f"ДС транзакция готова: {tx_type} на сумму {amount}")
        
        # Можно добавить в очередь отправки или обработать другим способом
    
    def _handle_balance_alert(self, address, balance):
        """Обработка алерта по балансу от ДС"""
        short_address = f"{address[:6]}...{address[-4:]}"
        logger.warning(f"ДС алерт: большой баланс у {short_address}: {balance} PLEX")
    
    def _handle_analysis_completed(self, results):
        """Обработка завершения анализа торговой активности"""
        all_traders = results.get('all_traders', {})
        holders = results.get('holders', {})
        analysis_type = results.get('analysis_type', 'unknown')
        
        logger.info(f"ДС анализ завершен ({analysis_type}): {len(all_traders)} трейдеров, {len(holders)} держателей")
        
        # Можно показать уведомление или обновить статистику
        if len(holders) > 0:
            logger.info(f"Найдено {len(holders)} держателей PLEX ONE")

    # --- Таб «Настройки» ---
    def _tab_settings(self):
        """Создание вкладки для настроек приложения"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа общих настроек
        general_group = QtWidgets.QGroupBox("Общие настройки")
        general_layout = QtWidgets.QFormLayout(general_group)
        
        # Цена газа
        self.gas_price_spin = QtWidgets.QDoubleSpinBox()
        self.gas_price_spin.setRange(0.1, 500)
        self.gas_price_spin.setDecimals(1)
        self.gas_price_spin.setSingleStep(0.1)
        self.gas_price_spin.setValue(float(self.cfg.get_gas_price()))
        self.gas_price_spin.valueChanged.connect(self._update_gas_price)
        general_layout.addRow("Цена газа (Gwei):", self.gas_price_spin)
        
        # Количество повторений серии
        self.repeat_count_spin = QtWidgets.QSpinBox()
        self.repeat_count_spin.setRange(1, 100)
        self.repeat_count_spin.setValue(self.cfg.get_repeat_count())
        self.repeat_count_spin.valueChanged.connect(self._update_repeat_count)
        general_layout.addRow("Количество повторений серии:", self.repeat_count_spin)
        
        # Очистка кэша
        btn_clear_cache = QtWidgets.QPushButton("Очистить кэш")
        btn_clear_cache.clicked.connect(self._clear_cache)
        general_layout.addRow("Кэширование:", btn_clear_cache)
        
        layout.addWidget(general_group)
        
        # Группа настроек блокчейна
        blockchain_group = QtWidgets.QGroupBox("Настройки блокчейна")
        blockchain_layout = QtWidgets.QFormLayout(blockchain_group)
        
        # Статус блокчейна
        blockchain_status = "Включен" if blockchain_enabled else "Отключен"
        blockchain_layout.addRow("Статус блокчейна:", QtWidgets.QLabel(blockchain_status))
        
        # Список RPC узлов
        rpc_list = QtWidgets.QListWidget()
        for node in RPC_NODES:
            rpc_list.addItem(node)
        blockchain_layout.addRow("RPC узлы:", rpc_list)
        
        # Тест подключения
        btn_test_conn = QtWidgets.QPushButton("Проверить подключение")
        btn_test_conn.clicked.connect(self._test_blockchain_connection)
        blockchain_layout.addRow("", btn_test_conn)
        
        # Информация о кошельке
        wallet_info = QtWidgets.QLabel("Не настроен")
        if blockchain_enabled and self.pk:
            try:
                address = Web3().eth.account.from_key(self.pk).address
                wallet_info.setText(address)
            except Exception:
                wallet_info.setText("Ошибка получения адреса")
        blockchain_layout.addRow("Текущий адрес:", wallet_info)
        
        layout.addWidget(blockchain_group)
        
        # Группа информации о приложении
        about_group = QtWidgets.QGroupBox("О приложении")
        about_layout = QtWidgets.QVBoxLayout(about_group)
        
        about_text = QtWidgets.QLabel(
            "WalletSender - приложение для анализа токенов BSC и распределения вознаграждений.\n"
            "Версия: 2.0\n"
            "© 2023-2025"
        )
        about_text.setWordWrap(True)
        about_layout.addWidget(about_text)
        
        layout.addWidget(about_group)
        
        return w
    
    def _update_gas_price(self, value):
        """Обновление цены газа в конфигурации"""
        self.cfg.set_gas_price(value)
        logger.info(f"Цена газа установлена на {value} Gwei")
        # Обновляем отображение в массовой рассылке, если виджет существует
        if hasattr(self, 'mass_gas_price_label'):
            self.mass_gas_price_label.setText(f"Цена газа: {value} Gwei")
    
    def _update_repeat_count(self, value):
        """Обновление количества повторений в конфигурации"""
        self.cfg.set_repeat_count(value)
        logger.info(f"Количество повторений установлено на {value}")
    
    def _clear_cache(self):
        """Очистка кэша"""
        count = cache.clear()
        logger.info(f"Очищено {count} элементов кэша")
    
    def _test_blockchain_connection(self):
        """Проверка подключения к блокчейну"""
        if not blockchain_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Блокчейн отключен',
                'Функции блокчейна отключены. Установите необходимые библиотеки.',
                QtWidgets.QMessageBox.Ok
            )
            return
            
        # Проверяем подключение в отдельном потоке
        threading.Thread(
            target=self._test_connection_thread,
            daemon=True
        ).start()
    
    def _test_connection_thread(self):
        """Поток для проверки подключения к блокчейну"""
        try:
            logger.info("Проверка подключения к блокчейну...")
            
            w3 = self.rpc.web3()
            
            if w3.is_connected():
                block = w3.eth.block_number
                used_node = self.rpc.get_healthy_node()
                logger.info(f"Подключение успешно! Узел: {used_node}, Блок: {block}")
                
                # Отправляем сигнал для обновления UI
                self.update_status_signal.emit(f"Подключение успешно! Узел: {used_node}, Блок: {block}")
                
            else:
                logger.warning("Не удалось подключиться к блокчейну")
                self.update_status_signal.emit("Не удалось подключиться к блокчейну")
        
        except Exception as e:
            logger.error(f"Ошибка при проверке подключения: {e}")
            self.update_status_signal.emit(f"Ошибка подключения: {e}")

    def _show_successful_connection(self, details):
        """Отображение успешного подключения"""
        QtWidgets.QMessageBox.information(
            self, 'Подключение успешно',
            f"Подключение к блокчейну BSC успешно установлено\n\n{details}",
            QtWidgets.QMessageBox.Ok
        )

    def _show_failed_connection(self, error_message=""):
        """Отображение ошибки подключения"""
        message = "Не удалось подключиться к блокчейну BSC"
        if error_message:
            message += f"\n\nПричина: {error_message}"
        
        QtWidgets.QMessageBox.warning(
            self, 'Ошибка подключения',
            message,
            QtWidgets.QMessageBox.Ok
        )
                
    def _show_connection_result(self, success, message):
        """Отображение результата проверки подключения"""
        if success:
            QtWidgets.QMessageBox.information(
                self, 'Подключение успешно',
                message,
                QtWidgets.QMessageBox.Ok
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, 'Ошибка подключения',
                message,
                QtWidgets.QMessageBox.Ok
            )
            
    # --- Вкладка «Массовая рассылка» ---
    def _tab_mass_distribution(self):
        """Создание вкладки для массовой рассылки токенов"""
        w = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(w)
        
        # Группа импорта адресов
        import_group = QtWidgets.QGroupBox("Импорт адресов")
        import_layout = QtWidgets.QVBoxLayout(import_group)
        
        # Кнопки импорта
        import_buttons_layout = QtWidgets.QHBoxLayout()
        
        self.mass_import_clipboard_btn = QtWidgets.QPushButton("Из буфера обмена")
        self.mass_import_clipboard_btn.clicked.connect(self._mass_import_from_clipboard)
        import_buttons_layout.addWidget(self.mass_import_clipboard_btn)
        
        self.mass_import_excel_btn = QtWidgets.QPushButton("Из Excel")
        self.mass_import_excel_btn.clicked.connect(self._mass_import_from_excel)
        import_buttons_layout.addWidget(self.mass_import_excel_btn)
        
        self.mass_clear_addresses_btn = QtWidgets.QPushButton("Очистить список")
        self.mass_clear_addresses_btn.clicked.connect(self._mass_clear_addresses)
        import_buttons_layout.addWidget(self.mass_clear_addresses_btn)
        
        import_layout.addLayout(import_buttons_layout)
        
        # Таблица адресов
        self.mass_addresses_table = QtWidgets.QTableWidget(0, 4)
        self.mass_addresses_table.setHorizontalHeaderLabels([
            '№', 'Адрес', 'Статус', 'Прогресс'
        ])
        self.mass_addresses_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.mass_addresses_table.setSelectionBehavior(QtWidgets.QTableView.SelectRows)
        import_layout.addWidget(self.mass_addresses_table)
        
        # Информация об импорте
        self.mass_import_info = QtWidgets.QLabel("Адресов загружено: 0")
        import_layout.addWidget(self.mass_import_info)
        
        layout.addWidget(import_group)
        
        # Группа настроек рассылки
        settings_group = QtWidgets.QGroupBox("Настройки рассылки")
        settings_layout = QtWidgets.QFormLayout(settings_group)
        
        # Выбор токена
        token_layout = QtWidgets.QHBoxLayout()
        self.mass_token_group = QtWidgets.QButtonGroup(w)
        self.mass_plex_radio = QtWidgets.QRadioButton('PLEX ONE')
        self.mass_usdt_radio = QtWidgets.QRadioButton('USDT')
        self.mass_bnb_radio = QtWidgets.QRadioButton('BNB')
        self.mass_custom_radio = QtWidgets.QRadioButton('Другой токен')
        self.mass_plex_radio.setChecked(True)
        
        self.mass_token_group.addButton(self.mass_plex_radio)
        self.mass_token_group.addButton(self.mass_usdt_radio)
        self.mass_token_group.addButton(self.mass_bnb_radio)
        self.mass_token_group.addButton(self.mass_custom_radio)
        
        token_layout.addWidget(self.mass_plex_radio)
        token_layout.addWidget(self.mass_usdt_radio)
        token_layout.addWidget(self.mass_bnb_radio)
        token_layout.addWidget(self.mass_custom_radio)
        
        settings_layout.addRow("Токен:", token_layout)
        
        # Поле для кастомного токена
        self.mass_custom_token_address = QtWidgets.QLineEdit()
        self.mass_custom_token_address.setPlaceholderText("Адрес контракта токена (0x...)")
        self.mass_custom_token_address.setEnabled(False)
        self.mass_custom_radio.toggled.connect(self.mass_custom_token_address.setEnabled)
        settings_layout.addRow("Адрес токена:", self.mass_custom_token_address)
        
        # Сумма для отправки
        self.mass_amount = QtWidgets.QDoubleSpinBox()
        self.mass_amount.setRange(0.00000001, 1000000)
        self.mass_amount.setDecimals(8)
        self.mass_amount.setValue(0.05)  # По умолчанию 0.05 токена
        settings_layout.addRow("Сумма на адрес:", self.mass_amount)
        
        # Интервал между отправками
        self.mass_interval = QtWidgets.QSpinBox()
        self.mass_interval.setRange(1, 600)
        self.mass_interval.setValue(5)  # 5 секунд по умолчанию
        settings_layout.addRow("Интервал (сек):", self.mass_interval)
        
        # Количество циклов
        self.mass_cycles = QtWidgets.QSpinBox()
        self.mass_cycles.setRange(1, 100)
        self.mass_cycles.setValue(10)  # 10 циклов по умолчанию
        settings_layout.addRow("Количество циклов:", self.mass_cycles)
        
        layout.addWidget(settings_group)
        
        # Группа управления рассылкой
        control_group = QtWidgets.QGroupBox("Управление рассылкой")
        control_layout = QtWidgets.QVBoxLayout(control_group)
        
        # Кнопки управления
        control_buttons_layout = QtWidgets.QHBoxLayout()
        
        self.mass_start_btn = QtWidgets.QPushButton("Начать рассылку")
        self.mass_start_btn.clicked.connect(self._mass_start_distribution)
        control_buttons_layout.addWidget(self.mass_start_btn)
        
        self.mass_pause_btn = QtWidgets.QPushButton("Пауза")
        self.mass_pause_btn.clicked.connect(self._mass_pause_distribution)
        self.mass_pause_btn.setEnabled(False)
        control_buttons_layout.addWidget(self.mass_pause_btn)
        
        self.mass_resume_btn = QtWidgets.QPushButton("Продолжить")
        self.mass_resume_btn.clicked.connect(self._mass_resume_distribution)
        self.mass_resume_btn.setEnabled(False)
        control_buttons_layout.addWidget(self.mass_resume_btn)
        
        self.mass_stop_btn = QtWidgets.QPushButton("Остановить")
        self.mass_stop_btn.clicked.connect(self._mass_stop_distribution)
        self.mass_stop_btn.setEnabled(False)
        control_buttons_layout.addWidget(self.mass_stop_btn)
        
        control_layout.addLayout(control_buttons_layout)
        
        # Прогресс рассылки
        self.mass_progress = QtWidgets.QProgressBar()
        control_layout.addWidget(self.mass_progress)
        
        # Статус рассылки
        self.mass_status_label = QtWidgets.QLabel("Готов к рассылке")
        control_layout.addWidget(self.mass_status_label)
        
        # Детальная информация
        info_layout = QtWidgets.QHBoxLayout()
        self.mass_current_cycle_label = QtWidgets.QLabel("Цикл: 0/0")
        self.mass_sent_count_label = QtWidgets.QLabel("Отправлено: 0")
        self.mass_errors_label = QtWidgets.QLabel("Ошибок: 0")
        info_layout.addWidget(self.mass_current_cycle_label)
        info_layout.addWidget(self.mass_sent_count_label)
        info_layout.addWidget(self.mass_errors_label)
        control_layout.addLayout(info_layout)
        
        layout.addWidget(control_group)
        
        # Группа статистики
        stats_group = QtWidgets.QGroupBox("Статистика")
        stats_layout = QtWidgets.QFormLayout(stats_group)
        
        self.mass_total_sent_label = QtWidgets.QLabel("0")
        self.mass_total_amount_label = QtWidgets.QLabel("0")
        self.mass_gas_spent_label = QtWidgets.QLabel("0 BNB")
        self.mass_gas_price_label = QtWidgets.QLabel(f"Цена газа: {self.cfg.get_gas_price()} Gwei")
        
        stats_layout.addRow("Всего транзакций:", self.mass_total_sent_label)
        stats_layout.addRow("Всего отправлено:", self.mass_total_amount_label)
        stats_layout.addRow("Потрачено на газ:", self.mass_gas_spent_label)
        stats_layout.addRow("", self.mass_gas_price_label)
        
        layout.addWidget(stats_group)
        
        # Инициализация переменных состояния
        self.mass_distribution_active = False
        self.mass_distribution_paused = False
        self.mass_distribution_thread = None
        self.mass_addresses = []
        self.mass_current_distribution_id = None
        
        return w
    
    def _mass_import_from_clipboard(self):
        """Импорт адресов из буфера обмена"""
        clipboard = QtWidgets.QApplication.clipboard()
        text = clipboard.text()
        
        if not text:
            logger.warning("Буфер обмена пуст")
            return
        
        # Разделяем текст по различным разделителям
        import re
        addresses = re.split(r'[\s,;\n]+', text)
        
        # Фильтруем и валидируем адреса
        valid_addresses = []
        for addr in addresses:
            addr = addr.strip()
            if addr and blockchain_enabled:
                try:
                    if Web3.is_address(addr):
                        # Приводим к checksum формату
                        checksum_addr = Web3.to_checksum_address(addr)
                        if checksum_addr not in valid_addresses:
                            valid_addresses.append(checksum_addr)
                except Exception as e:
                    logger.warning(f"Некорректный адрес: {addr}, ошибка: {e}")
            elif addr and not blockchain_enabled:
                # Если блокчейн отключен, принимаем любые строки похожие на адреса
                if addr.startswith('0x') and len(addr) == 42:
                    if addr not in valid_addresses:
                        valid_addresses.append(addr)
        
        if valid_addresses:
            self._mass_add_addresses(valid_addresses)
            logger.info(f"Импортировано {len(valid_addresses)} адресов из буфера обмена")
        else:
            logger.warning("Не найдено корректных адресов в буфере обмена")
    
    def _mass_import_from_excel(self):
        """Импорт адресов из Excel файла"""
        if not excel_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Excel не поддерживается',
                'Для импорта из Excel необходимо установить библиотеку openpyxl:\npip install openpyxl',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Диалог выбора файла
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Выберите Excel файл', '', 'Excel Files (*.xlsx *.xls)'
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            # Читаем первый лист
            sheet = workbook.active
            
            valid_addresses = []
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        addr = cell.strip()
                        if addr and blockchain_enabled:
                            try:
                                if Web3.is_address(addr):
                                    checksum_addr = Web3.to_checksum_address(addr)
                                    if checksum_addr not in valid_addresses:
                                        valid_addresses.append(checksum_addr)
                            except Exception:
                                pass
                        elif addr and not blockchain_enabled:
                            if addr.startswith('0x') and len(addr) == 42:
                                if addr not in valid_addresses:
                                    valid_addresses.append(addr)
            
            workbook.close()
            
            if valid_addresses:
                self._mass_add_addresses(valid_addresses)
                logger.info(f"Импортировано {len(valid_addresses)} адресов из Excel файла")
            else:
                logger.warning("Не найдено корректных адресов в Excel файле")
        
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла: {e}")
            QtWidgets.QMessageBox.critical(
                self, 'Ошибка',
                f'Не удалось прочитать Excel файл:\n{str(e)}',
                QtWidgets.QMessageBox.Ok
            )
    
    def _mass_add_addresses(self, addresses):
        """Добавление адресов в таблицу"""
        # Очищаем текущий список
        self.mass_addresses_table.setRowCount(0)
        self.mass_addresses = addresses
        
        # Добавляем адреса в таблицу
        for i, addr in enumerate(addresses):
            row = self.mass_addresses_table.rowCount()
            self.mass_addresses_table.insertRow(row)
            
            # Номер
            self.mass_addresses_table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(i + 1)))
            
            # Адрес
            self.mass_addresses_table.setItem(row, 1, QtWidgets.QTableWidgetItem(addr))
            
            # Статус
            status_item = QtWidgets.QTableWidgetItem("Ожидание")
            self.mass_addresses_table.setItem(row, 2, status_item)
            
            # Прогресс
            self.mass_addresses_table.setItem(row, 3, QtWidgets.QTableWidgetItem("0/0"))
        
        self.mass_import_info.setText(f"Адресов загружено: {len(addresses)}")
    
    def _mass_clear_addresses(self):
        """Очистка списка адресов"""
        self.mass_addresses_table.setRowCount(0)
        self.mass_addresses = []
        self.mass_import_info.setText("Адресов загружено: 0")
        logger.info("Список адресов очищен")
    
    def _mass_start_distribution(self):
        """Запуск массовой рассылки"""
        if not self.mass_addresses:
            QtWidgets.QMessageBox.warning(
                self, 'Нет адресов',
                'Сначала импортируйте адреса для рассылки',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        if not blockchain_enabled:
            QtWidgets.QMessageBox.warning(
                self, 'Блокчейн отключен',
                'Функции блокчейна отключены. Проверьте настройки.',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        if not self.pk:
            QtWidgets.QMessageBox.warning(
                self, 'Кошелек не настроен',
                'Сначала настройте кошелек в настройках',
                QtWidgets.QMessageBox.Ok
            )
            return
        
        # Получаем параметры рассылки
        token_type = None
        token_address = None
        
        if self.mass_plex_radio.isChecked():
            token_type = 'PLEX ONE'
            token_address = PLEX_CONTRACT
        elif self.mass_usdt_radio.isChecked():
            token_type = 'USDT'
            token_address = USDT_CONTRACT
        elif self.mass_bnb_radio.isChecked():
            token_type = 'BNB'
            token_address = None  # Нативный токен
        elif self.mass_custom_radio.isChecked():
            custom_addr = self.mass_custom_token_address.text().strip()
            if not custom_addr or not Web3.is_address(custom_addr):
                QtWidgets.QMessageBox.warning(
                    self, 'Некорректный адрес',
                    'Введите корректный адрес токена',
                    QtWidgets.QMessageBox.Ok
                )
                return
            token_type = 'CUSTOM'
            token_address = Web3.to_checksum_address(custom_addr)
        
        amount = self.mass_amount.value()
        interval = self.mass_interval.value()
        cycles = self.mass_cycles.value()
        
        # Подтверждение
        total_transactions = len(self.mass_addresses) * cycles
        total_amount = amount * total_transactions
        
        msg = f"Вы собираетесь отправить:\n\n"
        msg += f"Токен: {token_type}\n"
        msg += f"Сумма на адрес: {amount}\n"
        msg += f"Адресов: {len(self.mass_addresses)}\n"
        msg += f"Циклов: {cycles}\n"
        msg += f"Всего транзакций: {total_transactions}\n"
        msg += f"Общая сумма: {total_amount}\n\n"
        msg += f"Продолжить?"
        
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение рассылки',
            msg,
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
        
        # Создаем запись в БД
        self.mass_current_distribution_id = add_mass_distribution(
            name=f"Рассылка {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            token_address=token_address or '0x0',
            token_symbol=token_type,
            amount_per_tx=amount,
            total_addresses=len(self.mass_addresses),
            total_cycles=cycles,
            interval_seconds=interval
        )
        
        # Запускаем рассылку
        self.mass_distribution_active = True
        self.mass_distribution_paused = False
        
        # Обновляем UI
        self.mass_start_btn.setEnabled(False)
        self.mass_pause_btn.setEnabled(True)
        self.mass_stop_btn.setEnabled(True)
        self.mass_import_clipboard_btn.setEnabled(False)
        self.mass_import_excel_btn.setEnabled(False)
        self.mass_clear_addresses_btn.setEnabled(False)
        
        # Запускаем в отдельном потоке
        self.mass_distribution_thread = threading.Thread(
            target=self._mass_distribution_worker,
            args=(token_type, token_address, amount, interval, cycles),
            daemon=True
        )
        self.mass_distribution_thread.start()
        
        logger.info(f"Запущена массовая рассылка: {token_type}, {amount} x {cycles} циклов")
    
    def _mass_distribution_worker(self, token_type, token_address, amount, interval, cycles):
        """Рабочий поток для массовой рассылки"""
        try:
            total_sent = 0
            total_errors = 0
            total_gas_spent = 0
            
            for cycle in range(1, cycles + 1):
                if not self.mass_distribution_active:
                    break
                
                # Обновляем информацию о цикле
                self.update_status_signal.emit(f"Цикл {cycle}/{cycles}")
                QtCore.QMetaObject.invokeMethod(
                    self.mass_current_cycle_label,
                    "setText",
                    QtCore.Qt.QueuedConnection,
                    QtCore.Q_ARG(str, f"Цикл: {cycle}/{cycles}")
                )
                
                for i, address in enumerate(self.mass_addresses):
                    if not self.mass_distribution_active:
                        break
                    
                    # Проверка паузы
                    while self.mass_distribution_paused and self.mass_distribution_active:
                        time.sleep(0.5)
                    
                    if not self.mass_distribution_active:
                        break
                    
                    # Обновляем статус адреса
                    self._update_address_status(i, "Отправка...", QtGui.QColor('#FFA500'))
                    
                    try:
                        # Отправляем транзакцию
                        if token_type == 'BNB':
                            tx_hash, gas_used = self._send_bnb(address, amount)
                        else:
                            tx_hash, gas_used = self._send_token(address, amount, token_address)
                        
                        if tx_hash:
                            total_sent += 1
                            total_gas_spent += gas_used
                            
                            # Сохраняем в БД
                            add_mass_distribution_item(
                                self.mass_current_distribution_id,
                                address,
                                cycle,
                                tx_hash,
                                'success'
                            )
                            
                            # Обновляем статус
                            self._update_address_status(i, "Успешно", QtGui.QColor('#00FF00'))
                            self._update_address_progress(i, cycle, cycles)
                            
                            logger.info(f"Отправлено {amount} {token_type} на {address}, цикл {cycle}")
                        else:
                            raise Exception("Не удалось получить хэш транзакции")
                    
                    except Exception as e:
                        total_errors += 1
                        
                        # Сохраняем ошибку в БД
                        add_mass_distribution_item(
                            self.mass_current_distribution_id,
                            address,
                            cycle,
                            '',
                            'error',
                            str(e)
                        )
                        
                        # Обновляем статус
                        self._update_address_status(i, f"Ошибка: {str(e)[:30]}...", QtGui.QColor('#FF0000'))
                        
                        logger.error(f"Ошибка отправки на {address}: {e}")
                    
                    # Обновляем общую статистику
                    self._update_mass_statistics(total_sent, total_errors, total_gas_spent, amount * total_sent)
                    
                    # Прогресс
                    progress = int(((cycle - 1) * len(self.mass_addresses) + i + 1) / (cycles * len(self.mass_addresses)) * 100)
                    self.update_progress_signal.emit("mass", progress)
                    
                    # Интервал между отправками
                    if i < len(self.mass_addresses) - 1 and self.mass_distribution_active:
                        time.sleep(interval)
                
                # Интервал между циклами
                if cycle < cycles and self.mass_distribution_active:
                    logger.info(f"Цикл {cycle} завершен, ожидание перед следующим циклом...")
                    time.sleep(interval * 2)  # Двойной интервал между циклами
            
            # Завершение рассылки
            if self.mass_current_distribution_id:
                status = 'completed' if self.mass_distribution_active else 'cancelled'
                update_mass_distribution_status(self.mass_current_distribution_id, status)
            
            logger.info(f"Массовая рассылка завершена. Отправлено: {total_sent}, ошибок: {total_errors}")
        
        except Exception as e:
            logger.error(f"Критическая ошибка в массовой рассылке: {e}")
        
        finally:
            # Восстанавливаем UI
            self.mass_distribution_active = False
            QtCore.QMetaObject.invokeMethod(
                self,
                "_mass_distribution_finished",
                QtCore.Qt.QueuedConnection
            )
    
    @QtCore.pyqtSlot()
    def _mass_distribution_finished(self):
        """Завершение массовой рассылки - обновление UI"""
        self.mass_start_btn.setEnabled(True)
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(False)
        self.mass_stop_btn.setEnabled(False)
        self.mass_import_clipboard_btn.setEnabled(True)
        self.mass_import_excel_btn.setEnabled(True)
        self.mass_clear_addresses_btn.setEnabled(True)
        self.mass_status_label.setText("Рассылка завершена")
    
    def _update_address_status(self, row, status, color=None):
        """Обновление статуса адреса в таблице"""
        try:
            status_item = QtWidgets.QTableWidgetItem(status)
            if color:
                status_item.setBackground(color)
            
            QtCore.QMetaObject.invokeMethod(
                self.mass_addresses_table,
                "setItem",
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(int, row),
                QtCore.Q_ARG(int, 2),
                QtCore.Q_ARG(QtWidgets.QTableWidgetItem, status_item)
            )
        except Exception as e:
            logger.error(f"Ошибка обновления статуса: {e}")
    
    def _update_address_progress(self, row, current, total):
        """Обновление прогресса адреса"""
        try:
            progress_item = QtWidgets.QTableWidgetItem(f"{current}/{total}")
            
            QtCore.QMetaObject.invokeMethod(
                self.mass_addresses_table,
                "setItem",
                QtCore.Qt.QueuedConnection,
                QtCore.Q_ARG(int, row),
                QtCore.Q_ARG(int, 3),
                QtCore.Q_ARG(QtWidgets.QTableWidgetItem, progress_item)
            )
        except Exception as e:
            logger.error(f"Ошибка обновления прогресса: {e}")
    
    def _update_mass_statistics(self, sent, errors, gas_spent, total_amount):
        """Обновление статистики рассылки"""
        QtCore.QMetaObject.invokeMethod(
            self.mass_sent_count_label,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, f"Отправлено: {sent}")
        )
        
        QtCore.QMetaObject.invokeMethod(
            self.mass_errors_label,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, f"Ошибок: {errors}")
        )
        
        QtCore.QMetaObject.invokeMethod(
            self.mass_total_sent_label,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, str(sent))
        )
        
        QtCore.QMetaObject.invokeMethod(
            self.mass_total_amount_label,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, f"{total_amount:.8f}")
        )
        
        QtCore.QMetaObject.invokeMethod(
            self.mass_gas_spent_label,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, f"{gas_spent:.8f} BNB")
        )
    
    def _mass_pause_distribution(self):
        """Приостановка массовой рассылки"""
        self.mass_distribution_paused = True
        self.mass_pause_btn.setEnabled(False)
        self.mass_resume_btn.setEnabled(True)
        self.mass_status_label.setText("Рассылка приостановлена")
        logger.info("Массовая рассылка приостановлена")
    
    def _mass_resume_distribution(self):
        """Возобновление массовой рассылки"""
        self.mass_distribution_paused = False
        self.mass_pause_btn.setEnabled(True)
        self.mass_resume_btn.setEnabled(False)
        self.mass_status_label.setText("Рассылка продолжается")
        logger.info("Массовая рассылка возобновлена")
    
    def _mass_stop_distribution(self):
        """Остановка массовой рассылки"""
        reply = QtWidgets.QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите остановить рассылку?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.mass_distribution_active = False
            self.mass_distribution_paused = False
            self.mass_status_label.setText("Рассылка остановлена")
            logger.info("Массовая рассылка остановлена пользователем")
    
    def _send_bnb(self, to_address, amount):
        """Отправка BNB"""
        try:
            w3 = self.rpc.web3()
            account = w3.eth.account.from_key(self.pk)
            
            # Получаем nonce
            nonce = w3.eth.get_transaction_count(account.address)
            
            # Подготавливаем транзакцию
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            value = w3.to_wei(amount, 'ether')
            
            tx = {
                'nonce': nonce,
                'to': to_address,
                'value': value,
                'gas': 21000,
                'gasPrice': gas_price,
                'chainId': 56  # BSC mainnet
            }
            
            # Подписываем и отправляем
            signed = account.sign_transaction(tx)
            tx_hash = w3.eth.send_raw_transaction(signed.raw_transaction)
            
            # Ждем подтверждения
            receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
            
            gas_used = receipt.gasUsed * gas_price / 10**18
            
            return tx_hash.hex(), gas_used
        
        except Exception as e:
            logger.error(f"Ошибка отправки BNB: {e}")
            raise
    
    def _send_token(self, to_address, amount, token_address):
        """Отправка токена"""
        try:
            w3 = self.rpc.web3()
            account = w3.eth.account.from_key(self.pk)
            
            # Конвертируем адреса в checksum формат
            to_address = w3.to_checksum_address(to_address)
            token_address = w3.to_checksum_address(token_address)
            
            # Получаем контракт токена
            token_contract = w3.eth.contract(address=token_address, abi=ERC20_ABI)
            
            # Получаем decimals
            decimals = token_contract.functions.decimals().call()
            
            # Конвертируем сумму
            amount_wei = int(amount * 10**decimals)
            
            # Получаем nonce
            nonce = w3.eth.get_transaction_count(account.address)
            
            # Подготавливаем транзакцию
            gas_price = w3.to_wei(self.cfg.get_gas_price(), 'gwei')
            
            # Создаем транзакцию transfer
            tx = token_contract.functions.transfer(to_address, amount_wei).build_transaction({
                'from': account.address,
                'nonce': nonce,
                'gasPrice': gas_price,
                'gas': 100000  # Примерная оценка газа
            })
            
            # Подписываем и отправляем
            signed = account.sign_transaction(tx)
            tx_hash = w3.eth.send_raw_transaction(signed.raw_transaction)
            
            # Ждем подтверждения
            receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
            
            gas_used = receipt.gasUsed * gas_price / 10**18
            
            return tx_hash.hex(), gas_used
        
        except Exception as e:
            logger.error(f"Ошибка отправки токена: {e}")
            raise

# ======== #MCP:MAIN ========
# Entry point
if __name__ == "__main__":
    """Точка входа в приложение
    
    TODO:MCP - Добавить обработку аргументов командной строки
    REFACTOR:MCP - Выделить инициализацию в отдельную функцию
    """
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
